import json
import os
import datetime
import zipfile
import re
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import Counter, defaultdict
import traceback  # Import for detailed error reporting

# Assuming telemetry is in a utils directory accessible via PYTHONPATH
# This is primarily for the UI, but good to have a dummy if processor is run standalone for tests.
try:
    from utils.telemetry import send_event  # This will be used by the UI calling this processor
except ImportError:
    print(
        "[WARN] Telemetry module not found in gstr1_processor. Telemetry for unexpected sections will be handled by UI.")


    def send_event(event_name, payload):  # Dummy function
        pass

# ----------------------- Global Variables ----------------------- #
INDIAN_FORMAT = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00" + \
                r";-##\,##\,##\,##0.00" + \
                r";-"

RED_BOLD_FONT = Font(color="FF0000", bold=True)
BOLD_FONT = Font(bold=True)

MAIN_VALUE_CONFIG = {
    "B2B,SEZ,DE": {"value_col": "Invoice Value", "id_col": "Invoice Number"},
    "CDNR": {"value_col": "Note Value", "id_col": "Note Number"},
    "EXP": {"value_col": "Total Invoice Value", "id_col": "Invoice Number"},
    "B2BA": {"value_col": "Total Invoice Value", "id_col": "Revised/Original Invoice No."},
    "CDNUR": {"value_col": "Computed Invoice Value", "id_col": "C/D Note No"},
    "B2CS": {"value_col": "Computed Invoice Value", "id_col": None},
    "B2CSA": {"value_col": "Computed Invoice Value", "id_col": None},
    "NIL": {"value_col": "Computed Invoice Value", "id_col": None},
    "HSN": {"value_col": "Computed Invoice Value", "id_col": None},
    "AT": {"value_col": "Computed Invoice Value", "id_col": None},
    "TXPD": {"value_col": "Computed Invoice Value", "id_col": None},
    "DOC": {"value_col": None, "id_col": None}
}

DETAIL_SHEET_TOTAL_COLUMNS = {
    "B2B,SEZ,DE": ["Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "CDNR": ["Note Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "EXP": ["Total Invoice Value", "Total Taxable Value", "Integrated Tax", "Cess"],
    "B2BA": ["Total Invoice Value", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "CDNUR": ["Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "B2CS": ["Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "B2CSA": ["Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "NIL": ["Computed Invoice Value", "Nil Rated Supplies", "Exempted(Other than Nil rated/non-GST supply)",
            "Non-GST Supplies"],
    "HSN": ["Quantity", "Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax",
            "Cess", "No. of Records"],
    "AT": ["Computed Invoice Value", "Gross Advance Adjusted", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "TXPD": ["Computed Invoice Value", "Gross Advance Adjusted", "Integrated Tax", "Central Tax", "State/UT Tax",
             "Cess"],
    "DOC": ["Total Number", "Cancelled", "Net Issued"]
}

SECTION_TITLES = {
    "B2B,SEZ,DE": "B2B, SEZ, DE Invoices",
    "CDNR": "CDNR - Credit/Debit Notes (Registered)",
    "B2CS": "B2CS - B2C (Others)",
    "B2CSA": "10 - Amended B2C(Others)",
    "NIL": "NIL - Nil Rated, Exempted and Non-GST Supplies",
    "EXP": "EXP - Exports Invoices (with/without payment)",
    "HSN": "HSN - HSN wise details of outward supplies",
    "B2BA": "B2BA - Amended B2B Invoices",
    "CDNUR": "CDNUR - Credit/Debit Notes (Unregistered)",
    "DOC1": "1. Invoices for outward supply",
    "DOC2": "2. Invoices for inward supply from unregistered person",
    "DOC3": "3. Revised Invoice",
    "DOC4": "4. Debit Note",
    "DOC5": "5. Credit Note",
    "DOC6": "6. Receipt voucher",
    "DOC7": "7. Payment Voucher",
    "DOC8": "8. Refund voucher",
    "DOC9": "9. Delivery Challan for job work",
    "DOC10": "10. Delivery Challan for supply on approval",
    "DOC11": "11. Delivery Challan in case of liquid gas",
    "DOC12": "12. Delivery Challan in cases other than by way of supply (excluding at S no. 9 to 11)",
    "AT": "Tax Liability (Advances Received)",
    "TXPD": "Adjustment of Advances",
    "Summary-B2B": "4A-Supplies to registered persons(other than reverse charge)-B2B Regular-Summary",
    "Summary-SEZWP-WOP": "6B-Supplies made to SEZ-SEZWP/SEZWOP Total-Summary",
    "Summary-B2CS": "7-Supplies to unregistered persons-B2CS (Others)-Summary",
    "Summary-B2CSA": "10 - Amended B2C(Others)-Summary",
    "Summary-CDNR": "9B-Credit/Debit Notes(Registered)-Summary",
    "Summary-NIL": "8-Nil Rated,exempted,non GST supplies-Summary",
    "Summary-EXPWP": "6A–Exports (with payment)-Summary",
    "Summary-EXPWOP": "6A–Exports (without payment)-Summary",
    "Summary-EXP-Total": "6A–Exports (with/without payment)-Summary",
    "Summary-B2BA Total": "9A-Amendment to Supplies made to registered persons in earlier tax period-B2B Amended total-Summary",
    "Summary-CDNUR-B2CL": "9B-Credit/Debit Notes(Unregistered)-B2CL-Summary",
    "Summary-CDNUR-EXPWP": "9B-Credit/Debit Notes(Unregistered)-EXPWP-Summary",
    "Summary-CDNUR-EXPWOP": "9B-Credit/Debit Notes(Unregistered)-EXPWOP-Summary",
    "Summary-CDNUR-TOTAL": "9B-Credit/Debit Notes(Unregistered)-CDNUR-Total-Summary",
    "Summary-HSN": "12-HSN wise outward supplies-Summary",
    "Summary-DOC": "13-Documents issued-Summary",
    "Summary-AT": "11A(1),11A(2)-Advances received-No invoice issued (tax to be added to tax liability)-Summary",
    "Summary-TXPD": "11B(1),11B(2)-Advances received in earlier tax period-Adjusted in this tax period-Summary",
    "B2B,SEZ,DE_sws": "B2B, SEZ, DE Invoices - Sorted Supplier_wise",
    "CDNR_sws": "CDNR - Credit/Debit Notes (Registered) - Sorted Supplier_wise"
}

COLUMN_HEADERS = {
    "B2B,SEZ,DE": [
        "GSTIN/UIN of Recipient", "Receiver Name", "Invoice Number", "Invoice Date",
        "Reporting Month", "Tax Type", "Invoice Value", "Place of Supply", "Reverse Charge",
        "Applicable % of Tax Rate", "Invoice Type", "E-Commerce GSTIN", "Rate", "Taxable Value",
        "Integrated Tax", "Central Tax", "State/UT Tax", "Cess", "IRN", "IRN Date", "E-Invoice Status"
    ],
    "CDNR": [
        "GSTIN/UIN of Recipient", "Receiver Name", "Note Number", "Note Date", "Reporting Month",
        "Note Type", "Place of Supply", "Reverse Charge", "Note Supply Type", "Note Value",
        "Applicable % of Tax Rate", "Rate", "Taxable Value", "Integrated Tax",
        "Central Tax", "State/UT Tax", "Cess", "IRN", "IRN Date", "E-Invoice Status"
    ],
    "B2CS": [
        "Reporting Month", "Place of Supply", "Rate", "Computed Invoice Value", "Taxable Value",
        "Integrated Tax", "Central Tax", "State/UT Tax", "Cess", "Applicable % of Tax Rate",
        "Type", "Supply Type"
    ],
    "B2CSA": [
        "Reporting Month", "Original Month", "Place of Supply", "Supply Type", "Type",
        "Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax",
        "State/UT Tax", "Cess"
    ],
    "NIL": [
        "Reporting Month", "Supply Type", "Computed Invoice Value", "Nil Rated Supplies",
        "Exempted(Other than Nil rated/non-GST supply)", "Non-GST Supplies"
    ],
    "EXP": [
        "Invoice Number", "Invoice Date", "Reporting Month", "GST payment", "Supply type",
        "Total Invoice Value", "Rate", "Total Taxable Value", "Integrated Tax", "Central Tax",
        "State/UT Tax", "Cess", "IRN", "IRN Date"
    ],
    "HSN": [
        "Reporting Month", "HSN/SAC", "No. of Records", "UQC", "Quantity",
        "Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax",
        "State/UT Tax", "Cess", "Tax Rate"
    ],
    "B2BA": [
        "GSTIN/UIN of Recipient", "Revised Invoice No.", "Revised Invoice Date", "Reporting Month",
        "Revised/Original Invoice No.", "Revised/Original Invoice Date",
        "Total Invoice Value", "Rate", "Total Taxable Value", "Integrated Tax",
        "Central Tax", "State/UT Tax", "Cess"
    ],
    "CDNUR": [
        "C/D Note No", "C/D Note Date", "Reporting Month", "Note Type", "Type", "Rate",
        "Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess",
        "IRN", "IRN Date"
    ],
    "DOC": [
        "Reporting Month", "From (Sr. No.)", "To (Sr. No.)", "Total Number", "Cancelled", "Net Issued"
    ],
    "AT": [
        "Month", "Place of Supply", "Supply Type", "Computed Invoice Value", "Gross Advance Adjusted",
        "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"
    ],
    "TXPD": [
        "Month", "Place of Supply", "Supply Type", "Computed Invoice Value", "Gross Advance Adjusted",
        "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"
    ],
    "Summary": [
        "Reporting Month", "No. of Records", "Invoice Value", "Taxable Value", "Integrated Tax",
        "Central Tax", "State/UT Tax", "Cess"
    ],
    "Summary-DOC": [
        "Reporting Month", "No. of Records", "Net Issued Documents", "Documents Issued", "Documents Cancelled"
    ]
}

COLUMN_FORMATS = {
    "B2B,SEZ,DE": {
        "Invoice Date": "DD-MM-YYYY", "Tax Type": "General", "Invoice Value": INDIAN_FORMAT,
        "Place of Supply": "#,##0", "Rate": INDIAN_FORMAT, "Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT, "IRN Date": "DD-MM-YYYY"
    },
    "CDNR": {
        "Note Date": "DD-MM-YYYY", "Note Value": INDIAN_FORMAT, "Place of Supply": "#,##0",
        "Rate": INDIAN_FORMAT, "Taxable Value": INDIAN_FORMAT, "Integrated Tax": INDIAN_FORMAT,
        "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT, "Cess": INDIAN_FORMAT,
        "IRN Date": "DD-MM-YYYY"
    },
    "B2CS": {
        "Place of Supply": "#,##0", "Rate": INDIAN_FORMAT, "Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT, "Computed Invoice Value": INDIAN_FORMAT
    },
    "B2CSA": {
        "Reporting Month": "General",
        "Original Month": "General",
        "Place of Supply": "General",
        "Supply Type": "General",
        "Type": "General",
        "Computed Invoice Value": INDIAN_FORMAT,
        "Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT,
        "Central Tax": INDIAN_FORMAT,
        "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT
    },
    "NIL": {
        "Nil Rated Supplies": INDIAN_FORMAT, "Exempted(Other than Nil rated/non-GST supply)": INDIAN_FORMAT,
        "Non-GST Supplies": INDIAN_FORMAT, "Computed Invoice Value": INDIAN_FORMAT
    },
    "EXP": {
        "Invoice Number": "General", "Invoice Date": "DD-MM-YYYY", "GST payment": "General",
        "Supply type": "General", "Total Invoice Value": INDIAN_FORMAT, "Rate": INDIAN_FORMAT,
        "Total Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT, "IRN": "General", "IRN Date": "DD-MM-YYYY"
    },
    "HSN": {
        "Reporting Month": "General", "HSN/SAC": "#,##0", "No. of Records": "#,##0", "UQC": "General",
        "Quantity": INDIAN_FORMAT, "Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT, "Computed Invoice Value": INDIAN_FORMAT, "Tax Rate": "#,##0.00"
    },
    "B2BA": {
        "GSTIN/UIN of Recipient": "General", "Revised Invoice No.": "General", "Revised Invoice Date": "DD-MM-YYYY",
        "Revised/Original Invoice No.": "General", "Revised/Original Invoice Date": "DD-MM-YYYY",
        "Total Invoice Value": INDIAN_FORMAT, "Rate": INDIAN_FORMAT, "Total Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT
    },
    "CDNUR": {
        "C/D Note Date": "DD-MM-YYYY", "Rate": INDIAN_FORMAT, "Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT,
        "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT, "Cess": INDIAN_FORMAT,
        "Computed Invoice Value": INDIAN_FORMAT,
        "IRN Date": "DD-MM-YYYY"
    },
    "DOC": {
        "Reporting Month": "General", "From (Sr. No.)": "General", "To (Sr. No.)": "General",
        "Total Number": "#,##0", "Cancelled": "#,##0", "Net Issued": "#,##0"
    },
    "AT": {
        "Month": "General", "Place of Supply": "General", "Supply Type": "General",
        "Gross Advance Adjusted": INDIAN_FORMAT, "Integrated Tax": INDIAN_FORMAT,
        "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT, "Cess": INDIAN_FORMAT,
        "Computed Invoice Value": INDIAN_FORMAT
    },
    "TXPD": {
        "Month": "General", "Place of Supply": "General", "Supply Type": "General",
        "Gross Advance Adjusted": INDIAN_FORMAT, "Integrated Tax": INDIAN_FORMAT,
        "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT, "Cess": INDIAN_FORMAT,
        "Computed Invoice Value": INDIAN_FORMAT
    },
    "Summary": {
        "Reporting Month": "General", "No. of Records": "#,##0", "Invoice Value": INDIAN_FORMAT,
        "Taxable Value": INDIAN_FORMAT,
        "Integrated Tax": INDIAN_FORMAT, "Central Tax": INDIAN_FORMAT, "State/UT Tax": INDIAN_FORMAT,
        "Cess": INDIAN_FORMAT,
        "Note Value": INDIAN_FORMAT,
        "Computed Invoice Value": INDIAN_FORMAT
    },
    "Summary-DOC": {
        "Reporting Month": "General", "No. of Records": "#,##0", "Net Issued Documents": "#,##0",
        "Documents Issued": "#,##0", "Documents Cancelled": "#,##0"
    }
}

NUMERIC_KEYS_BY_SECTION = {
    "B2B,SEZ,DE": ["Invoice Value", "Place of Supply", "Rate", "Taxable Value", "Integrated Tax", "Central Tax",
                   "State/UT Tax", "Cess"],
    "CDNR": ["Note Value", "Place of Supply", "Rate", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax",
             "Cess"],
    "B2CS": ["Place of Supply", "Rate", "Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax",
             "State/UT Tax", "Cess"],
    "B2CSA": ["Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "NIL": ["Computed Invoice Value", "Nil Rated Supplies", "Exempted(Other than Nil rated/non-GST supply)",
            "Non-GST Supplies"],
    "EXP": ["Total Invoice Value", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "HSN": ["No. of Records", "Quantity", "Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax",
            "State/UT Tax", "Cess", "Tax Rate"],
    "B2BA": ["Total Invoice Value", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "CDNUR": ["Computed Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "DOC": ["Total Number", "Cancelled", "Net Issued"],
    "AT": ["Computed Invoice Value", "Gross Advance Adjusted", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "TXPD": ["Computed Invoice Value", "Gross Advance Adjusted", "Integrated Tax", "Central Tax", "State/UT Tax",
             "Cess"],
    "Summary": ["No. of Records",
                "Invoice Value",  # Lowercase 'v' for B2B/SEZ actual data key
                "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess",
                "Note Value",  # For CDNR actual data key
                "Computed Invoice Value",  # For B2CS, B2CSA, NIL, AT, TXPD, HSN, CDNUR actual data key
                "Total Invoice Value"  # For B2BA, EXP actual data key
                ],
    "Summary-DOC": ["No. of Records", "Net Issued Documents", "Documents Issued", "Documents Cancelled"]
}


# ----------------------- Utility Functions ----------------------- #
def parse_filename(filename):
    basename = os.path.basename(filename)
    month_match = re.search(r'GSTR1_(\d{6})', basename)
    month = month_match.group(1) if month_match else None
    excl_match = re.search(r'excluding_([A-Z_]+)', basename)
    excluded = excl_match.group(1).split('_') if excl_match else []
    return month, excluded


def parse_large_filename(filename):
    base = os.path.splitext(os.path.basename(filename))[0]
    parts = base.split('_')
    for part in parts:
        if len(part) == 6 and part.isdigit():
            return part
    match = re.search(r'(\d{6})$', base)
    if match:
        return match.group(1)
    return ""


def get_tax_period(ret_str, include_year=False):
    month_map = {
        "01": "January", "02": "February", "03": "March", "04": "April",
        "05": "May", "06": "June", "07": "July", "08": "August",
        "09": "September", "10": "October", "11": "November", "12": "December"
    }
    ret_str = str(ret_str).strip() if ret_str is not None else ""

    if not ret_str or len(ret_str) < 2:
        return "Unknown"

    month_code = ret_str[:2]
    month_name = month_map.get(month_code)

    if not month_name:
        return "Unknown"

    if include_year:
        if len(ret_str) == 6 and ret_str[2:].isdigit() and len(ret_str[2:]) == 4:
            year_str = ret_str[2:]
            return f"{month_name} {year_str}"
        elif len(ret_str) == 4 and ret_str.isdigit():
            return f"{month_name} {ret_str}"
        else:
            return f"{month_name} YYYY"  # Default placeholder if year is ambiguous
    else:
        return month_name


def parse_date_string(date_str):
    if not date_str:
        return None
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d-%m-%y"):
        try:
            return datetime.datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return None


def parse_number(value, float_2dec=False, int_no_dec=False):
    if value is None or value == "":
        return 0
    try:
        num = float(value)
        if int_no_dec:
            return int(num)
        return round(num, 2) if float_2dec else num
    except (ValueError, TypeError):
        return 0


def load_json_data_from_file(file_path, is_zip=False):
    print(f"[DEBUG] Loading JSON from {file_path} (is_zip={is_zip})")
    data_list = []
    try:
        if is_zip:
            with zipfile.ZipFile(file_path, "r") as z:
                json_file_name = next((name for name in z.namelist() if name.lower().endswith(".json")), None)
                if json_file_name:
                    with z.open(json_file_name) as f:
                        file_content_bytes = f.read()
                        try:
                            file_content_str = file_content_bytes.decode('utf-8')
                        except UnicodeDecodeError:
                            print(f"[WARN] UTF-8 decoding failed for {json_file_name} in {file_path}. Trying latin-1.")
                            file_content_str = file_content_bytes.decode('latin-1', errors='replace')

                        data = json.loads(file_content_str)
                        data["month"] = get_tax_period(data.get("fp", ""))
                        data["_raw_json_content_for_snippet"] = file_content_str
                        data_list.append(data)
        else:
            with open(file_path, "r", encoding="utf-8") as f:
                file_content_str = f.read()
                data = json.loads(file_content_str)
                period_key = list(data.keys())[0] if data else ""

                file_reporting_period_str = period_key
                if period_key and isinstance(data.get(period_key), dict):
                    actual_ret_period_val = data.get(period_key, {}).get("summary", {}).get("data", {}).get(
                        "ret_period")
                    if actual_ret_period_val:
                        file_reporting_period_str = actual_ret_period_val

                data["month"] = get_tax_period(file_reporting_period_str)
                data["_raw_json_content_for_snippet"] = file_content_str
                data_list.append(data)
        print(f"[DEBUG] Loaded JSON from {file_path} successfully")
    except Exception as e:
        print(f"[DEBUG] Error loading {file_path}: {e}\n{traceback.format_exc()}")
    return data_list


# ----------------------- Extraction Functions (<500 Logic) ----------------------- #
# --- extract_b2b_entries ---
def extract_b2b_entries(data):
    print("[DEBUG] Extracting B2B,SEZ,DE section...")
    if not data:
        print("[DEBUG] Extracted B2B,SEZ,DE section...done (no data)")
        return []
    period_key = list(data.keys())[0]
    reporting_month = data.get("month", get_tax_period(period_key))

    suppliers = data.get(period_key, {}).get("sections", {}).get("B2B", {}).get("suppliers", [])
    results = []
    invoice_counts = {}
    for supplier in suppliers:
        sup_info = supplier.get("supplier", {})
        gstin = sup_info.get("ctin", "")
        trade_name = sup_info.get("trade_name", "")
        tax_type = sup_info.get("txp_typ", "")
        for inv in supplier.get("invoiceDetails", []):
            invoice_num = inv.get("inum", "").strip()
            if not invoice_num: continue
            invoice_date = parse_date_string(inv.get("idt", ""))
            invoice_value = parse_number(inv.get("val", ""), float_2dec=True)
            reverse_charge = inv.get("rchrg", "")
            invoice_type = inv.get("inv_typ", "")
            ecom_gstin = inv.get("ctin", "")
            irn = inv.get("irn", "")
            irn_date = parse_date_string(inv.get("irngendate", ""))
            e_inv_status = "Yes" if irn else ""
            top_txval = parse_number(inv.get("invtxval", 0), float_2dec=True)
            top_iamt = parse_number(inv.get("inviamt", 0), float_2dec=True)
            top_camt = parse_number(inv.get("invcamt", 0), float_2dec=True)
            top_samt = parse_number(inv.get("invsamt", 0), float_2dec=True)
            top_csamt = parse_number(inv.get("invcsamt", 0), float_2dec=True)
            nested = inv.get("invoiceDetails", [])
            if not nested or not nested[0].get("inv", []):
                results.append({
                    "GSTIN/UIN of Recipient": gstin, "Receiver Name": trade_name, "Invoice Number": invoice_num,
                    "Invoice Date": invoice_date, "Reporting Month": reporting_month, "Tax Type": tax_type,
                    "Invoice Value": invoice_value, "Place of Supply": gstin[:2] if gstin else "",
                    "Reverse Charge": reverse_charge, "Applicable % of Tax Rate": None,
                    "Invoice Type": invoice_type, "E-Commerce GSTIN": ecom_gstin, "Rate": "error",
                    "Taxable Value": top_txval, "Integrated Tax": top_iamt, "Central Tax": top_camt,
                    "State/UT Tax": top_samt, "Cess": top_csamt, "IRN": irn, "IRN Date": irn_date,
                    "E-Invoice Status": e_inv_status, "highlight": False
                })
                invoice_counts[invoice_num] = invoice_counts.get(invoice_num, 0) + 1
                continue
            inv_data = nested[0]["inv"][0]
            pos_from_item = inv_data.get("pos")
            place_of_supply = pos_from_item if pos_from_item else (gstin[:2] if gstin else "")
            for item in inv_data.get("itms", []):
                itm_det = item.get("itm_det", {})
                if not all(k in itm_det for k in ("rt", "txval")): continue
                rate = itm_det["rt"]
                txval = parse_number(itm_det["txval"], float_2dec=True)
                iamt = parse_number(itm_det.get("iamt", 0), float_2dec=True)
                camt = parse_number(itm_det.get("camt", 0), float_2dec=True)
                samt = parse_number(itm_det.get("samt", 0), float_2dec=True)
                csamt = parse_number(itm_det.get("csamt", 0), float_2dec=True)
                results.append({
                    "GSTIN/UIN of Recipient": gstin, "Receiver Name": trade_name, "Invoice Number": invoice_num,
                    "Invoice Date": invoice_date, "Reporting Month": reporting_month, "Tax Type": tax_type,
                    "Invoice Value": invoice_value, "Place of Supply": place_of_supply,
                    "Reverse Charge": reverse_charge, "Applicable % of Tax Rate": None,
                    "Invoice Type": invoice_type, "E-Commerce GSTIN": ecom_gstin, "Rate": rate,
                    "Taxable Value": txval, "Integrated Tax": iamt, "Central Tax": camt,
                    "State/UT Tax": samt, "Cess": csamt, "IRN": irn, "IRN Date": irn_date,
                    "E-Invoice Status": e_inv_status, "highlight": False
                })
                invoice_counts[invoice_num] = invoice_counts.get(invoice_num, 0) + 1
    for row in results:
        if invoice_counts.get(row["Invoice Number"], 0) > 1: row["highlight"] = True
    print("[DEBUG] Extracted B2B,SEZ,DE section...done")
    return results


# --- extract_cdnr_entries ---
def extract_cdnr_entries(data):
    print("[DEBUG] Extracting CDNR section...")
    if not data:
        print("[DEBUG] Extracted CDNR section...done (no data)")
        return []
    period_key = list(data.keys())[0]
    reporting_month = data.get("month", get_tax_period(period_key))

    suppliers = data.get(period_key, {}).get("sections", {}).get("CDNR", {}).get("suppliers", [])
    results = []
    note_counts = {}
    for supplier in suppliers:
        s_info = supplier.get("supplier", {})
        gstin = s_info.get("ctin", "")
        trade_name = s_info.get("trade_name", "")
        for note in supplier.get("invoiceDetails", []):
            note_num = note.get("nt_num", "").strip()
            if not note_num: continue
            note_date = parse_date_string(note.get("nt_dt", ""))
            note_type = note.get("ntty", "")
            note_value = -abs(parse_number(note.get("val", 0), float_2dec=True))
            reverse_charge = note.get("rchrg", "")
            supply_type = note.get("inv_typ", "")
            irn = note.get("irn", "")
            irn_date = parse_date_string(note.get("irngendate", ""))
            e_inv_status = "Yes" if irn else ""
            top_txval = -abs(parse_number(note.get("invtxval", 0), float_2dec=True))
            top_iamt = -abs(parse_number(note.get("inviamt", 0), float_2dec=True))
            top_camt = -abs(parse_number(note.get("invcamt", 0), float_2dec=True))
            top_samt = -abs(parse_number(note.get("invsamt", 0), float_2dec=True))
            top_csamt = -abs(parse_number(note.get("invcsamt", 0), float_2dec=True))
            nested = note.get("invoiceDetails", [])
            if not nested or not nested[0].get("nt", []):
                results.append({
                    "GSTIN/UIN of Recipient": gstin, "Receiver Name": trade_name, "Note Number": note_num,
                    "Note Date": note_date, "Reporting Month": reporting_month, "Note Type": note_type,
                    "Place of Supply": gstin[:2] if gstin else "", "Reverse Charge": reverse_charge,
                    "Note Supply Type": supply_type, "Note Value": note_value,
                    "Applicable % of Tax Rate": None, "Rate": "error", "Taxable Value": top_txval,
                    "Integrated Tax": top_iamt, "Central Tax": top_camt, "State/UT Tax": top_samt,
                    "Cess": top_csamt, "IRN": irn, "IRN Date": irn_date,
                    "E-Invoice Status": e_inv_status, "highlight": False
                })
                note_counts[note_num] = note_counts.get(note_num, 0) + 1
                continue
            nt_data = nested[0]["nt"][0]
            pos_from_item = nt_data.get("pos")
            place_of_supply = pos_from_item if pos_from_item else (gstin[:2] if gstin else "")
            for item in nt_data.get("itms", []):
                itm_det = item.get("itm_det", {})
                if not all(k in itm_det for k in ("rt", "txval")): continue
                rate = itm_det["rt"]
                txval = -abs(parse_number(itm_det["txval"], float_2dec=True))
                iamt = -abs(parse_number(itm_det.get("iamt", 0), float_2dec=True))
                camt = -abs(parse_number(itm_det.get("camt", 0), float_2dec=True))
                samt = -abs(parse_number(itm_det.get("samt", 0), float_2dec=True))
                csamt = -abs(parse_number(itm_det.get("csamt", 0), float_2dec=True))
                results.append({
                    "GSTIN/UIN of Recipient": gstin, "Receiver Name": trade_name, "Note Number": note_num,
                    "Note Date": note_date, "Reporting Month": reporting_month, "Note Type": note_type,
                    "Place of Supply": place_of_supply, "Reverse Charge": reverse_charge,
                    "Note Supply Type": supply_type, "Note Value": note_value,
                    "Applicable % of Tax Rate": None, "Rate": rate, "Taxable Value": txval,
                    "Integrated Tax": iamt, "Central Tax": camt, "State/UT Tax": samt,
                    "Cess": csamt, "IRN": irn, "IRN Date": irn_date,
                    "E-Invoice Status": e_inv_status, "highlight": False
                })
                note_counts[note_num] = note_counts.get(note_num, 0) + 1
    for row in results:
        if note_counts.get(row["Note Number"], 0) > 1: row["highlight"] = True
    print("[DEBUG] Extracted CDNR section...done")
    return results


# --- extract_b2cs_entries ---
def extract_b2cs_entries(data):
    print("[DEBUG] Extracting B2CS section...")
    if not data:
        print("[DEBUG] Extracted B2CS section...done (empty data)")
        return []
    period_key = list(data.keys())[0]
    reporting_month = data.get("month", get_tax_period(period_key))

    inner = data.get(period_key, {})
    b2cs_obj = inner.get("sections", {}).get("B2CS", {})
    items = b2cs_obj.get("invoiceDetails", []) if isinstance(b2cs_obj, dict) else b2cs_obj if isinstance(b2cs_obj,
                                                                                                         list) else []
    results = []
    for item in items:
        taxable_value = parse_number(item.get("invtxval", item.get("txval", "")), float_2dec=True)
        integrated_tax = parse_number(item.get("inviamt", item.get("iamt", "")), float_2dec=True)
        central_tax = parse_number(item.get("invcamt", item.get("camt", "")), float_2dec=True)
        state_ut_tax = parse_number(item.get("invsamt", item.get("samt", "")), float_2dec=True)
        cess = parse_number(item.get("invcsamt", item.get("csamt", "")), float_2dec=True)
        computed_invoice_value = taxable_value + integrated_tax + central_tax + state_ut_tax + cess
        row = {
            "Reporting Month": reporting_month,
            "Place of Supply": parse_number(item.get("pos", ""), int_no_dec=True),
            "Rate": parse_number(item.get("rt", ""), float_2dec=True),
            "Computed Invoice Value": computed_invoice_value, "Taxable Value": taxable_value,
            "Integrated Tax": integrated_tax, "Central Tax": central_tax,
            "State/UT Tax": state_ut_tax, "Cess": cess, "Applicable % of Tax Rate": None,
            "Type": item.get("typ", ""), "Supply Type": item.get("sply_ty", "")
        }
        results.append(row)
    print("[DEBUG] Extracted B2CS section...done")
    return results


# --- extract_b2csa_entries (New) ---
def extract_b2csa_entries(data):
    print("[DEBUG] Extracting B2CSA section...")
    if not data:
        print("[DEBUG] Extracted B2CSA section...done (empty data)")
        return []

    period_key = list(data.keys())[0]
    reporting_month = data.get("month", get_tax_period(period_key))

    b2csa_section_data = data.get(period_key, {}).get("sections", {}).get("B2CSA", {})
    items = b2csa_section_data.get("invoiceDetails", [])
    results = []

    for item in items:
        original_month_raw = item.get("omon", "")
        original_month_str = get_tax_period(original_month_raw, include_year=True)

        place_of_supply = item.get("pos", "")
        supply_type = item.get("sply_ty", "")
        type_val = item.get("typ", "")

        taxable_value = parse_number(item.get("invtxval", 0), float_2dec=True)
        integrated_tax = parse_number(item.get("inviamt", 0), float_2dec=True)
        central_tax = parse_number(item.get("invcamt", 0), float_2dec=True)
        state_ut_tax = parse_number(item.get("invsamt", 0), float_2dec=True)
        cess = parse_number(item.get("invcsamt", 0), float_2dec=True)

        computed_invoice_value = taxable_value + integrated_tax + central_tax + state_ut_tax + cess

        row = {
            "Reporting Month": reporting_month,
            "Original Month": original_month_str,
            "Place of Supply": place_of_supply,
            "Supply Type": supply_type,
            "Type": type_val,
            "Computed Invoice Value": computed_invoice_value,
            "Taxable Value": taxable_value,
            "Integrated Tax": integrated_tax,
            "Central Tax": central_tax,
            "State/UT Tax": state_ut_tax,
            "Cess": cess
        }
        results.append(row)

    print(f"[DEBUG] Extracted B2CSA section...done ({len(results)} items)")
    return results


# --- extract_nil_entries ---
def extract_nil_entries(data):
    print("[DEBUG] Extracting NIL section...")
    if not data:
        print("[DEBUG] Extracted NIL section...done (empty data)")
        return []
    period_key = list(data.keys())[0]
    reporting_month = data.get("month", get_tax_period(period_key))

    inner = data.get(period_key, {})
    nil_obj = inner.get("sections", {}).get("NIL", {})
    items = nil_obj.get("inv", nil_obj.get("invoiceDetails", [])) if isinstance(nil_obj,
                                                                                dict) else nil_obj if isinstance(
        nil_obj, list) else []
    results = []
    for inv_item in items:
        nil_rated = parse_number(inv_item.get("nil_amt", ""), float_2dec=True)
        exempted = parse_number(inv_item.get("expt_amt", ""), float_2dec=True)
        non_gst = parse_number(inv_item.get("ngsup_amt", ""), float_2dec=True)
        computed_invoice_value = nil_rated + exempted + non_gst
        row = {
            "Reporting Month": reporting_month, "Supply Type": inv_item.get("sply_ty", ""),
            "Computed Invoice Value": computed_invoice_value, "Nil Rated Supplies": nil_rated,
            "Exempted(Other than Nil rated/non-GST supply)": exempted, "Non-GST Supplies": non_gst
        }
        results.append(row)
    print("[DEBUG] Extracted NIL section...done")
    return results


# --- extract_exp_entries ---
def extract_exp_entries(data):
    from collections import Counter
    entries = []
    period_key = next((k for k in data.keys() if k != "month" and k != "_raw_json_content_for_snippet"), None)
    if not period_key: return []

    reporting_month = data.get("month", get_tax_period(period_key))
    exp_section = data[period_key].get("sections", {}).get("EXP", {})
    for invoice in exp_section.get("invoiceDetails", []):
        inum = invoice.get("inum", "")
        idt = parse_date_string(invoice.get("idt", ""))
        val = parse_number(invoice.get("val", 0.00), float_2dec=True)
        irn = invoice.get("irn", "")
        irn_date = parse_date_string(invoice.get("irngendate", ""))
        gst_payment = invoice.get("exp_typ", "")
        supply_type = invoice.get("srctyp", "")
        top_txval = parse_number(invoice.get("invtxval", 0), float_2dec=True)
        top_iamt = parse_number(invoice.get("inviamt", 0), float_2dec=True)
        top_camt = parse_number(invoice.get("invcamt", 0), float_2dec=True)
        top_samt = parse_number(invoice.get("invsamt", 0), float_2dec=True)
        top_csamt = parse_number(invoice.get("invcsamt", 0), float_2dec=True)
        nested_list = invoice.get("invoiceDetails", [])
        inv_array = nested_list[0].get("inv", []) if nested_list and nested_list[0] else []
        if not inv_array:
            entries.append({
                "Invoice Number": inum, "Invoice Date": idt, "Reporting Month": reporting_month,
                "GST payment": gst_payment, "Supply type": supply_type, "Total Invoice Value": val,
                "Rate": "error",
                "Total Taxable Value": top_txval, "Integrated Tax": top_iamt,
                "Central Tax": top_camt, "State/UT Tax": top_samt, "Cess": top_csamt,
                "IRN": irn, "IRN Date": irn_date, "highlight": False
            })
            continue

        for inv_item_detail in inv_array:
            for item in inv_item_detail.get("itms", []):
                itm_det = item.get("itm_det", item)
                if not all(k in itm_det for k in ("rt", "txval")): continue
                rt = itm_det["rt"]
                txval = parse_number(itm_det["txval"], float_2dec=True)
                iamt = parse_number(itm_det.get("iamt", 0.00), float_2dec=True)
                camt = parse_number(itm_det.get("camt", 0.00), float_2dec=True)
                samt = parse_number(itm_det.get("samt", 0.00), float_2dec=True)
                csamt = parse_number(itm_det.get("csamt", 0.00), float_2dec=True)
                entries.append({
                    "Invoice Number": inum, "Invoice Date": idt, "Reporting Month": reporting_month,
                    "GST payment": gst_payment, "Supply type": supply_type, "Total Invoice Value": val,
                    "Rate": rt, "Total Taxable Value": txval, "Integrated Tax": iamt,
                    "Central Tax": camt, "State/UT Tax": samt, "Cess": csamt,
                    "IRN": irn, "IRN Date": irn_date, "highlight": False
                })
    counts = Counter(row["Invoice Number"] for row in entries)
    for row in entries:
        if counts.get(row["Invoice Number"], 0) > 1: row["highlight"] = True
    return entries


# --- extract_hsn_entries ---
def extract_hsn_entries(data_list):
    print("[DEBUG] Extracting HSN section...")
    if not data_list: return []
    hsn_dict = {}
    for data_item in data_list:
        if not data_item or not isinstance(data_item, dict): continue

        reporting_month_val = data_item.get("month", "Unknown")
        hsn_items_list = []

        if "fp" in data_item:
            hsn_items_list = data_item.get("hsn", {}).get("data", [])
        else:
            period_key = next((k for k in data_item.keys() if k not in ["month", "_raw_json_content_for_snippet"]),
                              None)
            if not period_key: continue

            hsn_obj = data_item.get(period_key, {}).get("sections", {}).get("HSN", {})
            hsn_items_list = hsn_obj.get("invoiceDetails", hsn_obj.get("data", [])) if isinstance(hsn_obj, dict) else \
                hsn_obj if isinstance(hsn_obj, list) else []

        if reporting_month_val == "Unknown" and hsn_items_list:
            if "fp" in data_item:
                reporting_month_val = get_tax_period(data_item.get("fp"))
            else:
                period_key = next((k for k in data_item.keys() if k not in ["month", "_raw_json_content_for_snippet"]),
                                  None)
                if period_key:
                    rp = data_item.get(period_key, {}).get("summary", {}).get("data", {}).get("ret_period")
                    if rp: reporting_month_val = get_tax_period(rp)

        for item in hsn_items_list:
            hsn_sac = item.get("hsn_sc", "").strip()
            uqc = item.get("uqc", "").strip()
            if not hsn_sac: continue
            key = (reporting_month_val, hsn_sac, uqc)
            if key not in hsn_dict:
                hsn_dict[key] = {
                    "Reporting Month": reporting_month_val, "HSN/SAC": hsn_sac, "No. of Records": 0,
                    "UQC": uqc, "Quantity": 0, "Computed Invoice Value": 0,
                    "Taxable Value": 0, "Integrated Tax": 0, "Central Tax": 0,
                    "State/UT Tax": 0, "Cess": 0, "Tax Rate": parse_number(item.get("rt", ""), float_2dec=True)
                }
            hsn_dict[key]["No. of Records"] += int(item.get("num", 1))
            hsn_dict[key]["Quantity"] += parse_number(item.get("qty", ""), float_2dec=True)
            hsn_dict[key]["Taxable Value"] += parse_number(item.get("txval", ""), float_2dec=True)
            hsn_dict[key]["Integrated Tax"] += parse_number(item.get("iamt", ""), float_2dec=True)
            hsn_dict[key]["Central Tax"] += parse_number(item.get("camt", ""), float_2dec=True)
            hsn_dict[key]["State/UT Tax"] += parse_number(item.get("samt", ""), float_2dec=True)
            hsn_dict[key]["Cess"] += parse_number(item.get("csamt", ""), float_2dec=True)

    for key_val_hsn in hsn_dict:
        hsn_dict[key_val_hsn]["Computed Invoice Value"] = (
                hsn_dict[key_val_hsn]["Taxable Value"] + hsn_dict[key_val_hsn]["Integrated Tax"] +
                hsn_dict[key_val_hsn]["Central Tax"] + hsn_dict[key_val_hsn]["State/UT Tax"] + hsn_dict[key_val_hsn][
                    "Cess"]
        )
    results = list(hsn_dict.values())
    financial_order = ["April", "May", "June", "July", "August", "September", "October", "November", "December",
                       "January", "February", "March", "Unknown"]
    results.sort(key=lambda x: (
        financial_order.index(x["Reporting Month"]) if x["Reporting Month"] in financial_order else 999, x["HSN/SAC"]))
    print("[DEBUG] Extracted HSN section...done")
    return results


# --- extract_b2ba_entries ---
def extract_b2ba_entries(data):
    from collections import Counter
    entries = []
    period_key = next((k for k in data.keys() if k != "month" and k != "_raw_json_content_for_snippet"), None)
    if not period_key: return []

    reporting_month = data.get("month", get_tax_period(period_key))
    section = data[period_key].get("sections", {}).get("B2BA", {})
    invoice_highlight_tracker = Counter()

    for inv_wrap in section.get("invoiceDetails", []):
        inum = inv_wrap.get("inum", "")
        idt = parse_date_string(inv_wrap.get("idt", ""))
        oinum = inv_wrap.get("oinum", "")
        oidt = parse_date_string(inv_wrap.get("oidt", ""))
        val = parse_number(inv_wrap.get("val", 0), float_2dec=True)
        ctin = inv_wrap.get("ctin", "")

        top_txval = parse_number(inv_wrap.get("invtxval", 0), float_2dec=True)
        top_iamt = parse_number(inv_wrap.get("inviamt", 0), float_2dec=True)
        top_camt = parse_number(inv_wrap.get("invcamt", 0), float_2dec=True)
        top_samt = parse_number(inv_wrap.get("invsamt", 0), float_2dec=True)
        top_csamt = parse_number(inv_wrap.get("invcsamt", 0), float_2dec=True)

        nested = inv_wrap.get("invoiceDetails", [])
        inv_list = nested[0].get("inv", []) if nested and nested[0] else []

        current_invoice_item_count = 0
        if not inv_list:
            entries.append({
                "GSTIN/UIN of Recipient": ctin, "Revised Invoice No.": inum, "Revised Invoice Date": idt,
                "Reporting Month": reporting_month, "Revised/Original Invoice No.": oinum,
                "Revised/Original Invoice Date": oidt, "Total Invoice Value": val, "Rate": "error",
                "Total Taxable Value": top_txval, "Integrated Tax": top_iamt, "Central Tax": top_camt,
                "State/UT Tax": top_samt, "Cess": top_csamt, "highlight": False
            })
            invoice_highlight_tracker[oinum] += 1
            continue

        for inv_item_detail in inv_list:
            for item in inv_item_detail.get("itms", []):
                itm_det = item.get("itm_det", {})
                if not all(k in itm_det for k in ("rt", "txval")): continue
                current_invoice_item_count += 1
                rt = itm_det["rt"]
                txval = parse_number(itm_det["txval"], float_2dec=True)
                iamt = parse_number(itm_det.get("iamt", 0), float_2dec=True)
                camt = parse_number(itm_det.get("camt", 0), float_2dec=True)
                samt = parse_number(itm_det.get("samt", 0), float_2dec=True)
                csamt = parse_number(itm_det.get("csamt", 0), float_2dec=True)
                entries.append({
                    "GSTIN/UIN of Recipient": ctin, "Revised Invoice No.": inum, "Revised Invoice Date": idt,
                    "Reporting Month": reporting_month, "Revised/Original Invoice No.": oinum,
                    "Revised/Original Invoice Date": oidt, "Total Invoice Value": val,
                    "Rate": rt, "Total Taxable Value": txval, "Integrated Tax": iamt,
                    "Central Tax": camt, "State/UT Tax": samt, "Cess": csamt, "highlight": False
                })
        if current_invoice_item_count > 0:
            invoice_highlight_tracker[oinum] += current_invoice_item_count

    for row in entries:
        if invoice_highlight_tracker.get(row["Revised/Original Invoice No."], 0) > 1:
            row["highlight"] = True
    return entries


# --- extract_cdnur_entries ---
def extract_cdnur_entries(data):
    from collections import Counter
    entries = []
    note_counts = {}
    period_key = next((k for k in data.keys() if k != "month" and k != "_raw_json_content_for_snippet"), None)
    if not period_key: return []

    reporting_month = data.get("month", get_tax_period(period_key))
    cdnur_section = data[period_key].get("sections", {}).get("CDNUR", {})

    for note in cdnur_section.get("invoiceDetails", []):
        nt_num = note.get("nt_num", "")
        nt_dt = parse_date_string(note.get("nt_dt", ""))
        ntty = note.get("ntty", "")
        typ = note.get("typ", "")
        irn = note.get("irn", "")
        irn_date = parse_date_string(note.get("irngendate", ""))

        top_val_field = note.get("val")
        top_txval = -abs(parse_number(note.get("invtxval", note.get("txval", 0)), float_2dec=True))
        top_iamt = -abs(parse_number(note.get("inviamt", note.get("iamt", 0)), float_2dec=True))
        top_camt = -abs(parse_number(note.get("invcamt", note.get("camt", 0)), float_2dec=True))
        top_samt = -abs(parse_number(note.get("invsamt", note.get("samt", 0)), float_2dec=True))
        top_csamt = -abs(parse_number(note.get("invcsamt", note.get("csamt", 0)), float_2dec=True))

        if top_val_field is not None:
            top_computed_value = -abs(parse_number(top_val_field, float_2dec=True))
        else:
            _ptxval = parse_number(note.get("invtxval", note.get("txval", 0)), float_2dec=True)
            _piamt = parse_number(note.get("inviamt", note.get("iamt", 0)), float_2dec=True)
            _pcamt = parse_number(note.get("invcamt", note.get("camt", 0)), float_2dec=True)
            _psamt = parse_number(note.get("invsamt", note.get("samt", 0)), float_2dec=True)
            _pcsamt = parse_number(note.get("invcsamt", note.get("csamt", 0)), float_2dec=True)
            top_computed_value = -abs(_ptxval + _piamt + _pcamt + _psamt + _pcsamt)

        itms = []
        nested_details_level1 = note.get("invoiceDetails", [])
        if nested_details_level1 and isinstance(nested_details_level1, list) and nested_details_level1[0]:
            if "itms" in nested_details_level1[0]:
                itms = nested_details_level1[0].get("itms", [])
            else:
                nested_details_level2 = nested_details_level1[0].get("invoiceDetails", [])
                if nested_details_level2 and isinstance(nested_details_level2, list) and nested_details_level2[0]:
                    if "itms" in nested_details_level2[0]:
                        itms = nested_details_level2[0].get("itms", [])

        if not itms:
            entries.append({
                "C/D Note No": nt_num, "C/D Note Date": nt_dt, "Reporting Month": reporting_month,
                "Note Type": ntty, "Type": typ, "Rate": "error",
                "Computed Invoice Value": top_computed_value, "Taxable Value": top_txval,
                "Integrated Tax": top_iamt, "Central Tax": top_camt, "State/UT Tax": top_samt,
                "Cess": top_csamt, "IRN": irn, "IRN Date": irn_date, "highlight": False
            })
            note_counts[nt_num] = note_counts.get(nt_num, 0) + 1
            continue

        for item in itms:
            itm_det = item.get("itm_det", item)
            if not all(k in itm_det for k in ("rt", "txval")): continue
            rt = itm_det["rt"]
            txval = -abs(parse_number(itm_det["txval"], float_2dec=True))
            iamt = -abs(parse_number(itm_det.get("iamt", 0), float_2dec=True))
            camt = -abs(parse_number(itm_det.get("camt", 0), float_2dec=True))
            samt = -abs(parse_number(itm_det.get("samt", 0), float_2dec=True))
            csamt = -abs(parse_number(itm_det.get("csamt", 0), float_2dec=True))
            entries.append({
                "C/D Note No": nt_num, "C/D Note Date": nt_dt, "Reporting Month": reporting_month,
                "Note Type": ntty, "Type": typ, "Rate": rt,
                "Computed Invoice Value": top_computed_value,
                "Taxable Value": txval,
                "Integrated Tax": iamt, "Central Tax": camt, "State/UT Tax": samt,
                "Cess": csamt, "IRN": irn, "IRN Date": irn_date, "highlight": False
            })
            note_counts[nt_num] = note_counts.get(nt_num, 0) + 1

    for row in entries:
        if note_counts.get(row["C/D Note No"], 0) > 1:
            row["highlight"] = True
    return entries


# --- extract_doc_entries ---
def extract_doc_entries(data):
    print("[DEBUG] Extracting DOC section...")
    if not data: return []
    period_key = list(data.keys())[0]
    reporting_month = data.get("month", get_tax_period(period_key))
    doc_obj = data.get(period_key, {}).get("sections", {}).get("DOC", {})
    items = doc_obj.get("doc_details", doc_obj.get("invoiceDetails", [])) if isinstance(doc_obj,
                                                                                        dict) else doc_obj if isinstance(
        doc_obj, list) else []
    results = []
    for item in items:
        doc_nature_code = item.get("doc_typ", "")
        doc_nature_title = SECTION_TITLES.get(f"DOC{doc_nature_code}", f"Unknown Doc Type {doc_nature_code}")
        for doc_detail in item.get("docs", []):
            row = {
                "Reporting Month": reporting_month, "From (Sr. No.)": doc_detail.get("from", ""),
                "To (Sr. No.)": doc_detail.get("to", ""),
                "Total Number": parse_number(doc_detail.get("totnum", ""), int_no_dec=True),
                "Cancelled": parse_number(doc_detail.get("cancel", ""), int_no_dec=True),
                "Net Issued": parse_number(doc_detail.get("net_issue", ""), int_no_dec=True),
                "doc_type_title": doc_nature_title,
                "doc_type_code": f"DOC{doc_nature_code}"
            }
            results.append(row)
    print("[DEBUG] Extracted DOC section...done")
    return results


# --- extract_at_entries ---
def extract_at_entries(data):
    print("[DEBUG] Extracting AT section...")
    if not data: return []
    period_key = list(data.keys())[0]
    reporting_month = data.get("month", get_tax_period(period_key))
    at_obj = data.get(period_key, {}).get("sections", {}).get("AT", {})
    items = at_obj.get("invoiceDetails", at_obj.get("data", [])) if isinstance(at_obj, dict) else at_obj if isinstance(
        at_obj, list) else []
    results = []
    for item in items:
        gross_advance_received = parse_number(item.get("ad_amt", item.get("invadamt", "")), float_2dec=True)
        iamt = parse_number(item.get("iamt", item.get("inviamt", "")), float_2dec=True)
        camt = parse_number(item.get("camt", item.get("invcamt", "")), float_2dec=True)
        samt = parse_number(item.get("samt", item.get("invsamt", "")), float_2dec=True)
        cess = parse_number(item.get("csamt", item.get("invcsamt", "")), float_2dec=True)
        computed_invoice_value = gross_advance_received + iamt + camt + samt + cess

        row = {
            "Month": reporting_month,
            "Place of Supply": item.get("pos", ""),
            "Supply Type": item.get("sply_ty", ""),
            "Computed Invoice Value": computed_invoice_value,
            "Gross Advance Adjusted": gross_advance_received,
            "Integrated Tax": iamt, "Central Tax": camt, "State/UT Tax": samt, "Cess": cess
        }
        results.append(row)
    print("[DEBUG] Extracted AT section...done")
    return results


# --- extract_txpd_entries (covers ATADJ - Adjustment of Advances) ---
def extract_txpd_entries(data):
    print("[DEBUG] Extracting TXPD/ATADJ section...")
    if not data: return []
    period_key = list(data.keys())[0]
    reporting_month = data.get("month", get_tax_period(period_key))

    txpd_obj = data.get(period_key, {}).get("sections", {}).get("TXPD", {})
    if not txpd_obj:
        txpd_obj = data.get(period_key, {}).get("sections", {}).get("ATADJ", {})

    items = txpd_obj.get("invoiceDetails", txpd_obj.get("data", [])) if isinstance(txpd_obj,
                                                                                   dict) else txpd_obj if isinstance(
        txpd_obj, list) else []
    results = []
    for item in items:
        gross_advance_adjusted = parse_number(item.get("ad_amt", item.get("invadamt", "")), float_2dec=True)
        iamt = parse_number(item.get("iamt", item.get("inviamt", "")), float_2dec=True)
        camt = parse_number(item.get("camt", item.get("invcamt", "")), float_2dec=True)
        samt = parse_number(item.get("samt", item.get("invsamt", "")), float_2dec=True)
        cess = parse_number(item.get("csamt", item.get("invcsamt", "")), float_2dec=True)
        computed_invoice_value = gross_advance_adjusted + iamt + camt + samt + cess

        row = {
            "Month": reporting_month,
            "Place of Supply": item.get("pos", ""),
            "Supply Type": item.get("sply_ty", ""),
            "Computed Invoice Value": computed_invoice_value,
            "Gross Advance Adjusted": gross_advance_adjusted,
            "Integrated Tax": iamt, "Central Tax": camt, "State/UT Tax": samt, "Cess": cess
        }
        results.append(row)
    print("[DEBUG] Extracted TXPD/ATADJ section...done")
    return results


# ----------------------- Extraction Functions (>500 Logic) ----------------------- #
def extract_b2b_entries_large(data):
    print("[DEBUG] Extracting B2B,SEZ,DE section (large JSON)...")
    if not data or not isinstance(data, dict): return []
    reporting_month = data.get("month", get_tax_period(data.get("fp", "")))
    results = []
    invoice_highlight_tracker = Counter()

    for buyer in data.get("b2b", []):
        gstin = buyer.get("ctin", "")
        receiver_name_large = buyer.get("trdnm", buyer.get("lgnm", ""))
        for inv in buyer.get("inv", []):
            invoice_num = inv.get("inum", "").strip()
            if not invoice_num: continue

            invoice_value = parse_number(inv.get("val", ""), float_2dec=True)
            inv_typ = inv.get("inv_typ", "")
            pos = inv.get("pos", "")

            item_rates = set()
            if inv.get("itms"):
                for item_detail_obj in inv.get("itms", []):
                    current_item_data = item_detail_obj.get("itm_det", item_detail_obj)
                    if current_item_data and "rt" in current_item_data:
                        item_rates.add(parse_number(current_item_data.get("rt", ""), float_2dec=True))
            multi_rate_highlight = len(item_rates) > 1

            invoice_base = {
                "GSTIN/UIN of Recipient": gstin, "Receiver Name": receiver_name_large,
                "Invoice Number": invoice_num, "Invoice Date": parse_date_string(inv.get("idt", "")),
                "Reporting Month": reporting_month,
                "Tax Type": inv_typ,
                "Invoice Value": invoice_value, "Place of Supply": pos,
                "Reverse Charge": inv.get("rchrg", ""), "Applicable % of Tax Rate": None,
                "Invoice Type": inv_typ,
                "E-Commerce GSTIN": inv.get("etin", ""),
                "IRN": inv.get("irn", ""), "IRN Date": parse_date_string(inv.get("irngendate", "")),
                "E-Invoice Status": "Yes" if inv.get("irn") else "", "highlight": multi_rate_highlight
            }

            items_processed_for_invoice = 0
            if not inv.get("itms"):
                row = invoice_base.copy()
                row.update({
                    "Rate": "error",
                    "Taxable Value": parse_number(inv.get("txval", 0), float_2dec=True),
                    "Integrated Tax": parse_number(inv.get("iamt", 0), float_2dec=True),
                    "Central Tax": parse_number(inv.get("camt", 0), float_2dec=True),
                    "State/UT Tax": parse_number(inv.get("samt", 0), float_2dec=True),
                    "Cess": parse_number(inv.get("csamt", 0), float_2dec=True),
                })
                results.append(row)
                invoice_highlight_tracker[invoice_num] += 1
                continue

            for item_detail_obj in inv.get("itms", []):
                itm_det = item_detail_obj.get("itm_det", item_detail_obj)
                if not itm_det or "rt" not in itm_det or "txval" not in itm_det: continue
                items_processed_for_invoice += 1

                rate = parse_number(itm_det.get("rt", ""), float_2dec=True)
                row = invoice_base.copy()
                row.update({
                    "Rate": rate,
                    "Taxable Value": parse_number(itm_det.get("txval", ""), float_2dec=True),
                    "Integrated Tax": parse_number(itm_det.get("iamt", ""), float_2dec=True),
                    "Central Tax": parse_number(itm_det.get("camt", ""), float_2dec=True),
                    "State/UT Tax": parse_number(itm_det.get("samt", ""), float_2dec=True),
                    "Cess": parse_number(itm_det.get("csamt", ""), float_2dec=True),
                })
                results.append(row)

            if items_processed_for_invoice > 0:
                invoice_highlight_tracker[invoice_num] += items_processed_for_invoice
            elif not items_processed_for_invoice and inv.get("itms"):
                row_err = invoice_base.copy()
                row_err.update(
                    {"Rate": "error (no valid items)", "Taxable Value": 0, "Integrated Tax": 0, "Central Tax": 0,
                     "State/UT Tax": 0, "Cess": 0})
                results.append(row_err)
                invoice_highlight_tracker[invoice_num] += 1

    for row in results:
        if invoice_highlight_tracker.get(row["Invoice Number"], 0) > 1: row["highlight"] = True
    print("[DEBUG] Extracted B2B,SEZ,DE section (large JSON)...done")
    return results


# ----------------------- Summary Calculation Functions ----------------------- #
def safe_add(current_value, new_value):
    if isinstance(new_value, (int, float)):
        return current_value + new_value
    return current_value


def calculate_monthly_summary(
        data, date_key, value_key, taxable_key, iamt_key,
        camt_key, samt_key, cess_key, invoice_key=None, processed_months=None):
    summary = {}

    if processed_months:
        for month_name in processed_months:
            summary[month_name] = {
                value_key: 0.0,
                "taxable_value": 0.0, "integrated_tax": 0.0,
                "central_tax": 0.0, "state_ut_tax": 0.0, "Cess": 0.0,
                "unique_invoices_for_count": set() if invoice_key else None,
                "_processed_invoice_value_docs": set() if invoice_key else None
            }

    for row in data:
        month_val = row.get(date_key)
        if not isinstance(month_val, str) or month_val == "Unknown": continue

        if month_val not in summary:
            summary[month_val] = {
                value_key: 0.0,
                "taxable_value": 0.0, "integrated_tax": 0.0,
                "central_tax": 0.0, "state_ut_tax": 0.0, "Cess": 0.0,
                "unique_invoices_for_count": set() if invoice_key else None,
                "_processed_invoice_value_docs": set() if invoice_key else None
            }

        doc_val_to_add = row.get(value_key)
        if doc_val_to_add is not None:
            if invoice_key:
                inv_id = row.get(invoice_key)
                if inv_id and inv_id not in summary[month_val]["_processed_invoice_value_docs"]:
                    summary[month_val][value_key] = safe_add(summary[month_val][value_key], doc_val_to_add)
                    summary[month_val]["_processed_invoice_value_docs"].add(inv_id)
            else:
                summary[month_val][value_key] = safe_add(summary[month_val][value_key], doc_val_to_add)

        if taxable_key and row.get(taxable_key) is not None:
            summary[month_val]["taxable_value"] = safe_add(summary[month_val]["taxable_value"], row.get(taxable_key))
        if iamt_key and row.get(iamt_key) is not None:
            summary[month_val]["integrated_tax"] = safe_add(summary[month_val]["integrated_tax"], row.get(iamt_key))
        if camt_key and row.get(camt_key) is not None:
            summary[month_val]["central_tax"] = safe_add(summary[month_val]["central_tax"], row.get(camt_key))
        if samt_key and row.get(samt_key) is not None:
            summary[month_val]["state_ut_tax"] = safe_add(summary[month_val]["state_ut_tax"], row.get(samt_key))
        if cess_key and row.get(cess_key) is not None:
            summary[month_val]["Cess"] = safe_add(summary[month_val]["Cess"], row.get(cess_key))

        if invoice_key and row.get(invoice_key):
            summary[month_val]["unique_invoices_for_count"].add(row[invoice_key])

    financial_order = ["April", "May", "June", "July", "August", "September", "October", "November", "December",
                       "January", "February", "March", "Unknown"]
    result = []
    for m_iter in financial_order:
        if m_iter in summary:
            record_count = 0
            if invoice_key:
                record_count = len(summary[m_iter]["unique_invoices_for_count"])
            else:
                record_count = sum(1 for r_item in data if r_item.get(date_key) == m_iter)

            output_row = {
                "Reporting Month": m_iter,
                "No. of Records": record_count,
                value_key: round(summary[m_iter][value_key], 2),
                "Taxable Value": round(summary[m_iter]["taxable_value"], 2),
                "Integrated Tax": round(summary[m_iter]["integrated_tax"], 2),
                "Central Tax": round(summary[m_iter]["central_tax"], 2),
                "State/UT Tax": round(summary[m_iter]["state_ut_tax"], 2),
                "Cess": round(summary[m_iter]["Cess"], 2)
            }
            result.append(output_row)
    return result


# ----------------------- Excel Report Generation Helper Functions ----------------------- #
def _add_total_row_to_detail_sheet(ws, section_key, rows_data, column_headers,
                                   column_formats_for_section):
    if not rows_data:
        return

    total_row_values = defaultdict(float)
    processed_ids_for_main_value = set()

    main_value_config_for_section = MAIN_VALUE_CONFIG.get(section_key)
    main_value_col_name = None
    main_id_col_name = None

    if main_value_config_for_section:
        main_value_col_name = main_value_config_for_section.get("value_col")
        main_id_col_name = main_value_config_for_section.get("id_col")

    columns_to_sum_and_display = DETAIL_SHEET_TOTAL_COLUMNS.get(section_key, [])
    if not columns_to_sum_and_display:
        print(f"[DEBUG] No total columns defined for detail section {section_key}. Skipping total row.")
        return

    for row in rows_data:
        for col_header in columns_to_sum_and_display:
            current_value_in_row = row.get(col_header)
            if not isinstance(current_value_in_row, (int, float)):
                continue

            if col_header == main_value_col_name and main_id_col_name:
                doc_id_val = row.get(main_id_col_name)
                if doc_id_val:
                    if doc_id_val not in processed_ids_for_main_value:
                        total_row_values[col_header] += current_value_in_row
                        processed_ids_for_main_value.add(doc_id_val)
            elif (col_header == main_value_col_name and not main_id_col_name) or \
                    (col_header != main_value_col_name):
                total_row_values[col_header] += current_value_in_row

    total_row_idx = ws.max_row + 1
    first_col_written = False
    for c_idx, col_header in enumerate(column_headers, 1):
        cell = ws.cell(row=total_row_idx, column=c_idx)
        if not first_col_written:
            label_col_idx = 1

            preferred_label_cols = ["Original Month", "Receiver Name", "GSTIN/UIN of Recipient", "HSN/SAC",
                                    "C/D Note No", "Invoice Number"]
            for pref_col in preferred_label_cols:
                if pref_col in column_headers:
                    try:
                        label_col_idx = column_headers.index(pref_col) + 1
                        break
                    except ValueError:
                        continue

            if c_idx == label_col_idx:
                cell.value = "Total"
                cell.font = RED_BOLD_FONT
                first_col_written = True
                if col_header not in columns_to_sum_and_display:
                    continue

        if col_header in columns_to_sum_and_display and col_header in total_row_values:
            value = total_row_values[col_header]
            if isinstance(value, float) and col_header not in ["Rate", "Tax Rate", "Place of Supply", "No. of Records",
                                                               "Total Number", "Cancelled", "Net Issued"]:
                value = round(value, 2)
            cell.value = value
            cell.font = RED_BOLD_FONT

            current_col_formats = column_formats_for_section if column_formats_for_section else {}
            if col_header in current_col_formats:
                format_str = current_col_formats[col_header]
                if isinstance(value, (int, float)):
                    cell.number_format = format_str
        elif not cell.value:
            cell.value = ""


def _add_total_row_to_summary_sheet(ws, summary_data_list, display_columns, data_keys_map, format_map):
    if not summary_data_list:
        return

    grand_totals = defaultdict(float)
    numeric_keys_for_this_summary = NUMERIC_KEYS_BY_SECTION.get("Summary-DOC") \
        if "DOC" in ws.title else NUMERIC_KEYS_BY_SECTION.get("Summary")

    for row_data in summary_data_list:
        for display_col_name in display_columns:
            actual_data_key = data_keys_map.get(display_col_name, display_col_name)

            if actual_data_key in numeric_keys_for_this_summary:
                value_to_sum = row_data.get(actual_data_key, 0)
                if isinstance(value_to_sum, (int, float)):
                    grand_totals[actual_data_key] += value_to_sum

    total_row_idx = ws.max_row + 1

    reporting_month_col_idx = -1
    try:
        reporting_month_col_idx = display_columns.index("Reporting Month") + 1
    except ValueError:
        reporting_month_col_idx = 1

    for c_idx, display_col_name in enumerate(display_columns, 1):
        cell = ws.cell(row=total_row_idx, column=c_idx)
        if c_idx == reporting_month_col_idx:
            cell.value = "Total"
            cell.font = RED_BOLD_FONT
        else:
            actual_data_key_for_sum = data_keys_map.get(display_col_name, display_col_name)

            if actual_data_key_for_sum in grand_totals:
                value = grand_totals[actual_data_key_for_sum]
                if display_col_name not in ["No. of Records", "Net Issued Documents", "Documents Issued",
                                            "Documents Cancelled"]:
                    value = round(value, 2)

                cell.value = value
                cell.font = RED_BOLD_FONT

                current_format_map_for_summary = format_map if format_map else {}
                if display_col_name in current_format_map_for_summary:
                    cell.number_format = current_format_map_for_summary[display_col_name]
                elif actual_data_key_for_sum in current_format_map_for_summary:
                    cell.number_format = current_format_map_for_summary[actual_data_key_for_sum]
            else:
                cell.value = ""


def create_excel_report(data_dict, wb, ignore_warnings=False):
    print("[DEBUG] Creating detailed sheets...")
    for section_key, rows_data in data_dict.items():
        if section_key == "DOC": continue

        columns_that_should_have_totals = DETAIL_SHEET_TOTAL_COLUMNS.get(section_key, [])
        has_data = any(
            isinstance(r.get(k, 0), (int, float)) and r.get(k, 0) != 0
            for r in rows_data for k in columns_that_should_have_totals if k in r
        )
        if not rows_data or (not ignore_warnings and not has_data):
            print(f"[DEBUG] Skipping sheet R1-{section_key} due to no meaningful data in totalable columns.")
            continue

        sheet_name = f"R1-{section_key}"
        if sheet_name in wb.sheetnames: wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)

        current_column_headers = COLUMN_HEADERS.get(section_key, [])
        if not current_column_headers:
            print(f"[WARN] No column headers for section {section_key}. Skipping.")
            continue

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(current_column_headers))
        title_cell = ws.cell(row=1, column=1, value=SECTION_TITLES.get(section_key, section_key))
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

        for idx, col_name in enumerate(current_column_headers, start=1):
            hdr_cell = ws.cell(row=2, column=idx, value=col_name)
            hdr_cell.font = BOLD_FONT
            hdr_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            hdr_cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "B3"

        sheet_has_error_string = False
        for r_idx, r_data in enumerate(rows_data, start=3):
            is_highlight = r_data.get("highlight", False)
            for c_idx, col_name in enumerate(current_column_headers, start=1):
                val = r_data.get(col_name, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if val == "error":
                    sheet_has_error_string = True
                    cell.font = RED_BOLD_FONT
                elif is_highlight:
                    cell.font = RED_BOLD_FONT

                current_col_formats = COLUMN_FORMATS.get(section_key, {})
                if col_name in current_col_formats:
                    format_str = current_col_formats[col_name]
                    if isinstance(val, (int, float)):
                        cell.number_format = format_str
                    elif isinstance(val, datetime.date):
                        cell.number_format = format_str

        _add_total_row_to_detail_sheet(ws, section_key, rows_data, current_column_headers,
                                       COLUMN_FORMATS.get(section_key, {}))

        apply_format_and_autofit(ws, current_column_headers, 3, COLUMN_FORMATS.get(section_key, {}))

        if sheet_has_error_string:
            ws.sheet_properties.tabColor = "FF0000"
        print(f"[DEBUG] Created sheet {sheet_name}")
    print("[DEBUG] Finished creating detailed sheets")


def create_or_replace_sheet(wb, sheet_name, title_text, columns_list):
    if sheet_name in wb.sheetnames: wb.remove(wb[sheet_name])
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns_list))
    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, col_val in enumerate(columns_list, 1):
        cell = ws.cell(row=2, column=idx, value=col_val)
        cell.font = BOLD_FONT
        cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B3"
    return ws


def fill_worksheet_data(ws, columns_to_fetch, data_rows, start_row=3):
    red_font = Font(color="FF0000", bold=True)
    general_numeric_cols = {
        "No. of Records", "Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess",
        "Net Issued Documents", "Documents Issued", "Documents Cancelled", "Total Number", "Cancelled",
        "Net Issued", "Note Value", "Total Invoice Value", "Computed Invoice Value",
        "Nil Rated Supplies", "Exempted(Other than Nil rated/non-GST supply)", "Non-GST Supplies",
        "Gross Advance Adjusted", "Quantity", "Rate", "Cess"
    }
    for row_idx, row_dict in enumerate(data_rows, start_row):
        is_highlight_row_fill = row_dict.get("highlight", False)
        for col_idx, col_name_fetch in enumerate(columns_to_fetch, 1):
            val_to_write = row_dict.get(col_name_fetch, "")
            cell = ws.cell(row=row_idx, column=col_idx)

            if col_name_fetch in general_numeric_cols and isinstance(val_to_write, str):
                try:
                    val_to_write = float(val_to_write)
                except ValueError:
                    pass

            cell.value = val_to_write
            if is_highlight_row_fill: cell.font = red_font
            if val_to_write == "error": cell.font = red_font


def apply_format_and_autofit(ws, columns_for_format, start_row=3, col_format_map=None):
    for col_idx, col_name_format in enumerate(columns_for_format, 1):
        col_letter = get_column_letter(col_idx)
        header_cell_value = ws.cell(row=2, column=col_idx).value
        max_len = len(str(header_cell_value if header_cell_value is not None else col_name_format))

        for r_val_format in range(start_row, ws.max_row + 1):
            cell_to_format = ws.cell(row=r_val_format, column=col_idx)
            cell_val_apply = cell_to_format.value

            if col_format_map and col_name_format in col_format_map:
                format_str_apply = col_format_map[col_name_format]
                if isinstance(cell_val_apply, datetime.date) and format_str_apply == "DD-MM-YYYY":
                    cell_to_format.number_format = format_str_apply
                elif isinstance(cell_val_apply, (int, float)):
                    cell_to_format.number_format = format_str_apply

            if cell_val_apply is not None: max_len = max(max_len, len(str(cell_val_apply)))

        ws.column_dimensions[col_letter].width = max(15, max_len + 2)


# ----------------------- Main Processing Function ----------------------- #
def process_gstr1(small_files, large_files, excluded_sections_by_month, template_path, save_path,
                  ignore_warnings=False):
    print("[DEBUG] Starting GSTR1 processing...")

    KNOWN_SECTION_KEYS_SMALL_JSON = {
        "B2B", "CDNR", "B2CS", "B2CSA", "NIL", "EXP", "HSN",
        "B2BA", "CDNUR", "DOC", "AT", "TXPD", "ATADJ"
    }
    KNOWN_DATA_KEYS_LARGE_JSON = {"b2b", "hsn"}
    METADATA_KEYS_LARGE_JSON = {"fp", "gstin", "gt", "cur_gt", "hash", "efiled_on", "month", "rtn_typ",
                                "_raw_json_content_for_snippet"}

    combined_data = {key: [] for key in
                     ["B2B,SEZ,DE", "CDNR", "B2CS", "B2CSA", "NIL", "EXP", "HSN", "B2BA", "CDNUR", "DOC", "AT", "TXPD"]}
    all_data_list_for_hsn = []
    processed_months = set()
    unexpected_sections_details = []

    print("[DEBUG] Processing small JSON files...")
    for file_path in small_files:
        month_from_filename, excluded_from_filename = parse_filename(file_path)
        data_list_from_file = load_json_data_from_file(file_path)

        for data_item in data_list_from_file:
            raw_json_content = data_item.pop("_raw_json_content_for_snippet", None)

            all_data_list_for_hsn.append(data_item.copy())

            file_reporting_month_name = data_item.get("month", "Unknown")
            if file_reporting_month_name != "Unknown":
                processed_months.add(file_reporting_month_name)

            period_key_val = None
            potential_period_keys = [k for k in data_item.keys() if k not in ["month", "_raw_json_content_for_snippet"]]
            if potential_period_keys:
                period_key_val = potential_period_keys[0]

            if period_key_val and isinstance(data_item.get(period_key_val), dict):
                json_sections_obj = data_item[period_key_val].get("sections", {})
                if isinstance(json_sections_obj, dict):
                    for section_key_found in json_sections_obj.keys():
                        if section_key_found not in KNOWN_SECTION_KEYS_SMALL_JSON:
                            snippet = json_sections_obj.get(section_key_found)
                            try:
                                snippet_str = json.dumps(snippet, indent=2)
                                NEW_CHARACTER_LIMIT = 10000
                                if len(snippet_str) > NEW_CHARACTER_LIMIT:
                                    snippet_str = snippet_str[:NEW_CHARACTER_LIMIT] + "\n... (truncated)"
                            except TypeError:
                                snippet_str = str(snippet)[:NEW_CHARACTER_LIMIT]
                                if len(str(snippet)) > NEW_CHARACTER_LIMIT:
                                    snippet_str += "\n... (truncated)"

                            unexpected_sections_details.append({
                                "file_type": "small",
                                "filename": os.path.basename(file_path),
                                "section_name": section_key_found,
                                "reporting_month": file_reporting_month_name,
                                "snippet": snippet_str
                            })
                            print(
                                f"[PROCESSOR_INFO] Unexpected section '{section_key_found}' found in small file '{os.path.basename(file_path)}'.")

            current_exclusions = list(excluded_from_filename)
            if month_from_filename and month_from_filename in excluded_sections_by_month:
                current_exclusions.extend(
                    e for e in excluded_sections_by_month[month_from_filename] if e not in current_exclusions)

            if "B2B" not in current_exclusions: combined_data["B2B,SEZ,DE"].extend(extract_b2b_entries(data_item))
            if "CDNR" not in current_exclusions: combined_data["CDNR"].extend(extract_cdnr_entries(data_item))
            if "B2CS" not in current_exclusions: combined_data["B2CS"].extend(extract_b2cs_entries(data_item))
            if "B2CSA" not in current_exclusions: combined_data["B2CSA"].extend(extract_b2csa_entries(data_item))
            if "NIL" not in current_exclusions: combined_data["NIL"].extend(extract_nil_entries(data_item))
            if "EXP" not in current_exclusions: combined_data["EXP"].extend(extract_exp_entries(data_item))
            if "B2BA" not in current_exclusions: combined_data["B2BA"].extend(extract_b2ba_entries(data_item))
            if "CDNUR" not in current_exclusions: combined_data["CDNUR"].extend(extract_cdnur_entries(data_item))
            if "DOC" not in current_exclusions: combined_data["DOC"].extend(extract_doc_entries(data_item))
            if "AT" not in current_exclusions: combined_data["AT"].extend(extract_at_entries(data_item))
            if "TXPD" not in current_exclusions and "ATADJ" not in current_exclusions:
                combined_data["TXPD"].extend(extract_txpd_entries(data_item))
    print("[DEBUG] Finished processing small JSON files")

    print("[DEBUG] Processing large JSON files...")
    for month_key_large_file_map, (filepath_large, large_file_excluded_sections) in large_files.items():
        if not filepath_large: continue
        current_exclusions_large = list(large_file_excluded_sections)
        large_json_data_list = load_json_data_from_file(filepath_large, is_zip=True)

        for data_item_large in large_json_data_list:
            raw_json_content_large = data_item_large.pop("_raw_json_content_for_snippet", None)
            all_data_list_for_hsn.append(data_item_large.copy())

            file_reporting_month_name_large = data_item_large.get("month", "Unknown")
            if file_reporting_month_name_large != "Unknown":
                processed_months.add(file_reporting_month_name_large)

            if "B2B" not in current_exclusions_large:
                combined_data["B2B,SEZ,DE"].extend(extract_b2b_entries_large(data_item_large))
    print("[DEBUG] Finished processing large JSON files")

    print("[DEBUG] Aggregating HSN data from all files...")
    combined_data["HSN"] = extract_hsn_entries(all_data_list_for_hsn)
    print("[DEBUG] HSN data aggregation completed")

    has_data_processed = any(combined_data[section] for section in combined_data if section != "DOC") or combined_data[
        "HSN"]
    if not has_data_processed and not ignore_warnings:
        if not combined_data.get("DOC"):
            if unexpected_sections_details:
                print("[WARN] No data extracted for standard sections, but unexpected sections were found.")
            else:
                raise ValueError("No data found in provided JSON files for any standard section.")
        print("[WARN] No data found for main sections or HSN. Only DOC sheets might be generated if they have data.")
    print("[DEBUG] Data validation completed (basic check for any data)")

    print("[DEBUG] Sorting data...")
    financial_order_sort = ["April", "May", "June", "July", "August", "September", "October", "November", "December",
                            "January", "February", "March", "Unknown"]
    date_sort_key_map = {"B2B,SEZ,DE": "Invoice Date", "CDNR": "Note Date", "EXP": "Invoice Date",
                         "B2BA": "Revised Invoice Date", "CDNUR": "C/D Note Date"}
    month_sort_key_map = {
        "B2B,SEZ,DE": "Reporting Month", "CDNR": "Reporting Month", "B2CS": "Reporting Month",
        "B2CSA": "Reporting Month", "NIL": "Reporting Month", "EXP": "Reporting Month",
        "HSN": "Reporting Month", "B2BA": "Reporting Month", "CDNUR": "Reporting Month",
        "DOC": "Reporting Month", "AT": "Month", "TXPD": "Month"
    }

    for section, data_rows_sort in combined_data.items():
        if not data_rows_sort: continue

        primary_sort_key_func = lambda x_sort: (
            financial_order_sort.index(x_sort.get(month_sort_key_map.get(section, "Reporting Month"), "Unknown"))
            if x_sort.get(month_sort_key_map.get(section, "Reporting Month"),
                          "Unknown") in financial_order_sort else 999
        )

        secondary_sort_key_name = date_sort_key_map.get(section)

        if secondary_sort_key_name:
            data_rows_sort.sort(key=lambda x_sort_sec: (
                primary_sort_key_func(x_sort_sec),
                x_sort_sec.get(secondary_sort_key_name) or datetime.date.max
            ))
        elif section == "HSN":
            data_rows_sort.sort(key=lambda x_hsn: (
                primary_sort_key_func(x_hsn),
                x_hsn.get("HSN/SAC", "")
            ))
        elif section == "DOC":
            data_rows_sort.sort(key=lambda x_doc: (
                primary_sort_key_func(x_doc),
                x_doc.get("doc_type_title", "")
            ))
        elif section == "B2CSA":
            data_rows_sort.sort(key=lambda x_b2csa: (
                primary_sort_key_func(x_b2csa),
                x_b2csa.get("Original Month", ""),
                x_b2csa.get("Place of Supply", "")
            ))
        else:
            data_rows_sort.sort(key=primary_sort_key_func)

    print("[DEBUG] Data sorting completed")

    print("[DEBUG] Initializing workbook...")
    wb = load_workbook(template_path) if template_path and os.path.exists(template_path) else Workbook()
    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1: wb.remove(wb["Sheet"])
    print("[DEBUG] Workbook initialized")

    create_excel_report(combined_data, wb, ignore_warnings)

    print("[DEBUG] Generating document-specific sheets...")
    doc_data_from_combined = combined_data.get("DOC", [])
    for i in range(1, 13):
        doc_code_filter = f"DOC{i}"
        doc_type_title_filter = SECTION_TITLES.get(doc_code_filter, f"Unknown Doc Type {i}")
        filtered_rows_doc = [row for row in doc_data_from_combined if row.get("doc_type_code") == doc_code_filter]

        numeric_cols_doc_check = NUMERIC_KEYS_BY_SECTION.get("DOC", [])
        has_meaningful_doc_data_check = any(
            isinstance(row.get(key, 0), (int, float)) and row.get(key, 0) != 0
            for row in filtered_rows_doc for key in numeric_cols_doc_check if key in row
        )
        if not filtered_rows_doc or (not ignore_warnings and not has_meaningful_doc_data_check):
            continue
        sheet_name_doc = f"R1-{doc_code_filter}"
        ws_doc_specific = create_or_replace_sheet(wb, sheet_name_doc, doc_type_title_filter, COLUMN_HEADERS["DOC"])
        fill_worksheet_data(ws_doc_specific, COLUMN_HEADERS["DOC"], filtered_rows_doc)
        _add_total_row_to_detail_sheet(ws_doc_specific, "DOC", filtered_rows_doc, COLUMN_HEADERS["DOC"],
                                       COLUMN_FORMATS.get("DOC", {}))
        apply_format_and_autofit(ws_doc_specific, COLUMN_HEADERS["DOC"], col_format_map=COLUMN_FORMATS.get("DOC", {}))
        print(f"[DEBUG] Created sheet {sheet_name_doc}")
    print("[DEBUG] Finished generating document-specific sheets")

    print("[DEBUG] Generating supplier-wise sorted sheets...")
    for section_sws_key in ["CDNR", "B2B,SEZ,DE"]:
        rows_sws_data = combined_data.get(section_sws_key, [])
        columns_to_total_for_sws = DETAIL_SHEET_TOTAL_COLUMNS.get(section_sws_key, [])
        has_meaningful_sws_data_check = any(
            isinstance(r.get(k, 0), (int, float)) and r.get(k, 0) != 0
            for r in rows_sws_data for k in columns_to_total_for_sws if k in r
        )
        if not rows_sws_data or (not ignore_warnings and not has_meaningful_sws_data_check):
            continue

        sort_keys_sws = ("Receiver Name", "GSTIN/UIN of Recipient")
        if section_sws_key == "CDNR":
            sort_keys_sws = ("Receiver Name", "GSTIN/UIN of Recipient", "Note Number")

        sorted_rows_sws_data = sorted(rows_sws_data, key=lambda x_sws: tuple(
            str(x_sws.get(k_sws, "")).strip().lower() for k_sws in sort_keys_sws
        ))

        sheet_name_sws_val = f"R1-{section_sws_key}_sws"
        title_sws = SECTION_TITLES.get(f"{section_sws_key}_sws", section_sws_key + " Supplier Wise")
        ws_supplier_wise = create_or_replace_sheet(wb, sheet_name_sws_val, title_sws, COLUMN_HEADERS[section_sws_key])
        fill_worksheet_data(ws_supplier_wise, COLUMN_HEADERS[section_sws_key], sorted_rows_sws_data)
        _add_total_row_to_detail_sheet(ws_supplier_wise, section_sws_key, sorted_rows_sws_data,
                                       COLUMN_HEADERS[section_sws_key],
                                       COLUMN_FORMATS.get(section_sws_key, {}))
        apply_format_and_autofit(ws_supplier_wise, COLUMN_HEADERS[section_sws_key],
                                 col_format_map=COLUMN_FORMATS.get(section_sws_key, {}))

        sheet_has_error_sws = any(
            r.get(c) == "error" for r in sorted_rows_sws_data for c in COLUMN_HEADERS[section_sws_key])
        if sheet_has_error_sws:
            ws_supplier_wise.sheet_properties.tabColor = "FF0000"
        print(f"[DEBUG] Created sheet {sheet_name_sws_val}")
    print("[DEBUG] Finished generating supplier-wise sorted sheets")

    print("[DEBUG] Generating summary sheets...")
    summary_display_columns = COLUMN_HEADERS["Summary"]
    summary_display_columns_note_type = [col if col != "Invoice Value" else "Note Value" for col in
                                         summary_display_columns]

    summary_ws_list_final = []
    summary_numeric_check_keys_list = NUMERIC_KEYS_BY_SECTION["Summary"]

    base_data_keys_map = {col: col for col in summary_display_columns}

    # B2B Regular Summary
    value_key_for_calc_b2b = "Invoice Value"
    b2b_regular_data = [row for row in combined_data.get("B2B,SEZ,DE", []) if row.get("Tax Type") in ["R", "NT", "CO"]]
    b2b_summary_original = calculate_monthly_summary(b2b_regular_data, "Reporting Month",
                                                     value_key=value_key_for_calc_b2b,
                                                     taxable_key="Taxable Value", iamt_key="Integrated Tax",
                                                     camt_key="Central Tax", samt_key="State/UT Tax", cess_key="Cess",
                                                     invoice_key="Invoice Number", processed_months=processed_months)
    if b2b_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in b2b_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_b2b_sum = create_or_replace_sheet(wb, "R1-Summary-B2B", SECTION_TITLES["Summary-B2B"],
                                             summary_display_columns)
        b2b_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_b2b, 0)} for row in
                                     b2b_summary_original]
        fill_worksheet_data(ws_b2b_sum, summary_display_columns, b2b_summary_display_ready)
        current_total_keys_map = {**base_data_keys_map, "Invoice Value": value_key_for_calc_b2b}
        _add_total_row_to_summary_sheet(ws_b2b_sum, b2b_summary_original, summary_display_columns,
                                        current_total_keys_map,
                                        COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_b2b_sum)

    # SEZ Summary
    value_key_for_calc_sez = "Invoice Value"
    sez_data_filter = [row for row in combined_data.get("B2B,SEZ,DE", []) if
                       row.get("Tax Type") in ["SEZWP", "SEZWOP", "SEWP", "SEWOP", "SEZ", "DE"]]
    sez_summary_original = calculate_monthly_summary(sez_data_filter, "Reporting Month",
                                                     value_key=value_key_for_calc_sez,
                                                     taxable_key="Taxable Value", iamt_key="Integrated Tax",
                                                     camt_key="Central Tax", samt_key="State/UT Tax", cess_key="Cess",
                                                     invoice_key="Invoice Number", processed_months=processed_months)
    if sez_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in sez_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_sez_sum = create_or_replace_sheet(wb, "R1-Summary-SEZWP-WOP", SECTION_TITLES["Summary-SEZWP-WOP"],
                                             summary_display_columns)
        sez_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_sez, 0)} for row in
                                     sez_summary_original]
        fill_worksheet_data(ws_sez_sum, summary_display_columns, sez_summary_display_ready)
        current_total_keys_map = {**base_data_keys_map, "Invoice Value": value_key_for_calc_sez}
        _add_total_row_to_summary_sheet(ws_sez_sum, sez_summary_original, summary_display_columns,
                                        current_total_keys_map,
                                        COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_sez_sum)

    # B2CS Summary
    value_key_for_calc_b2cs = "Computed Invoice Value"
    b2cs_summary_original = calculate_monthly_summary(combined_data.get("B2CS", []), "Reporting Month",
                                                      value_key=value_key_for_calc_b2cs,
                                                      taxable_key="Taxable Value", iamt_key="Integrated Tax",
                                                      camt_key="Central Tax", samt_key="State/UT Tax", cess_key="Cess",
                                                      invoice_key=None,
                                                      processed_months=processed_months)
    if b2cs_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in b2cs_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_b2cs_sum = create_or_replace_sheet(wb, "R1-Summary-B2CS", SECTION_TITLES["Summary-B2CS"],
                                              summary_display_columns)
        b2cs_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_b2cs, 0)} for row in
                                      b2cs_summary_original]
        fill_worksheet_data(ws_b2cs_sum, summary_display_columns, b2cs_summary_display_ready)
        current_total_keys_map = {**base_data_keys_map, "Invoice Value": value_key_for_calc_b2cs}
        _add_total_row_to_summary_sheet(ws_b2cs_sum, b2cs_summary_original, summary_display_columns,
                                        current_total_keys_map,
                                        COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_b2cs_sum)

    # B2CSA Summary
    value_key_for_calc_b2csa = "Computed Invoice Value"
    b2csa_summary_original = calculate_monthly_summary(
        combined_data.get("B2CSA", []), "Reporting Month",
        value_key=value_key_for_calc_b2csa,
        taxable_key="Taxable Value",
        iamt_key="Integrated Tax",
        camt_key="Central Tax",
        samt_key="State/UT Tax",
        cess_key="Cess",
        invoice_key=None,
        processed_months=processed_months
    )
    if b2csa_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in b2csa_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_b2csa_sum = create_or_replace_sheet(wb, "R1-Summary-B2CSA", SECTION_TITLES["Summary-B2CSA"],
                                               summary_display_columns)
        b2csa_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_b2csa, 0)} for row in
                                       b2csa_summary_original]
        fill_worksheet_data(ws_b2csa_sum, summary_display_columns, b2csa_summary_display_ready)
        current_total_keys_map = {**base_data_keys_map, "Invoice Value": value_key_for_calc_b2csa}
        _add_total_row_to_summary_sheet(ws_b2csa_sum, b2csa_summary_original, summary_display_columns,
                                        current_total_keys_map,
                                        COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_b2csa_sum)

    # CDNR Summary
    value_key_for_calc_cdnr = "Note Value"
    cdnr_summary_original = calculate_monthly_summary(combined_data.get("CDNR", []), "Reporting Month",
                                                      value_key=value_key_for_calc_cdnr,
                                                      taxable_key="Taxable Value", iamt_key="Integrated Tax",
                                                      camt_key="Central Tax", samt_key="State/UT Tax", cess_key="Cess",
                                                      invoice_key="Note Number", processed_months=processed_months)
    if cdnr_summary_original and (ignore_warnings or any(
            r.get(value_key_for_calc_cdnr, 0) != 0 or r.get("Taxable Value", 0) != 0 for r in cdnr_summary_original
    )):
        ws_cdnr_sum = create_or_replace_sheet(wb, "R1-Summary-CDNR", SECTION_TITLES["Summary-CDNR"],
                                              summary_display_columns_note_type)
        fill_worksheet_data(ws_cdnr_sum, summary_display_columns_note_type, cdnr_summary_original)
        current_total_keys_map = {col: col for col in summary_display_columns_note_type}
        _add_total_row_to_summary_sheet(ws_cdnr_sum, cdnr_summary_original, summary_display_columns_note_type,
                                        current_total_keys_map, COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_cdnr_sum)

    # NIL Summary
    value_key_for_calc_nil = "Computed Invoice Value"
    nil_summary_data_raw = combined_data.get("NIL", [])
    nil_summary_for_calc = [
        {
            "Reporting Month": r["Reporting Month"],
            "No. of Records": 1,
            value_key_for_calc_nil: r["Computed Invoice Value"],
            "Taxable Value": r["Computed Invoice Value"],
            "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0
        } for r in nil_summary_data_raw
    ]
    nil_summary_original = calculate_monthly_summary(nil_summary_for_calc, "Reporting Month",
                                                     value_key=value_key_for_calc_nil,
                                                     taxable_key="Taxable Value",
                                                     iamt_key="Integrated Tax", camt_key="Central Tax",
                                                     samt_key="State/UT Tax", cess_key="Cess",
                                                     invoice_key=None, processed_months=processed_months)

    if nil_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in nil_summary_original for k in [value_key_for_calc_nil, "Taxable Value"] if
            k != "No. of Records")):
        ws_nil_sum = create_or_replace_sheet(wb, "R1-Summary-NIL", SECTION_TITLES["Summary-NIL"],
                                             summary_display_columns)
        nil_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_nil, 0)} for row in
                                     nil_summary_original]
        fill_worksheet_data(ws_nil_sum, summary_display_columns, nil_summary_display_ready)
        current_total_keys_map = {**base_data_keys_map, "Invoice Value": value_key_for_calc_nil}
        _add_total_row_to_summary_sheet(ws_nil_sum, nil_summary_original, summary_display_columns,
                                        current_total_keys_map, COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_nil_sum)

    # AT Summary
    value_key_for_calc_at = "Computed Invoice Value"
    at_summary_original = calculate_monthly_summary(combined_data.get("AT", []), "Month",
                                                    value_key=value_key_for_calc_at,
                                                    taxable_key="Gross Advance Adjusted", iamt_key="Integrated Tax",
                                                    camt_key="Central Tax", samt_key="State/UT Tax", cess_key="Cess",
                                                    invoice_key=None, processed_months=processed_months)
    if at_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in at_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_at_sum = create_or_replace_sheet(wb, "R1-Summary-AT", SECTION_TITLES["Summary-AT"], summary_display_columns)
        at_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_at, 0)} for row in
                                    at_summary_original]
        fill_worksheet_data(ws_at_sum, summary_display_columns, at_summary_display_ready)
        current_total_keys_map = {**base_data_keys_map, "Invoice Value": value_key_for_calc_at}
        _add_total_row_to_summary_sheet(ws_at_sum, at_summary_original, summary_display_columns, current_total_keys_map,
                                        COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_at_sum)

    # TXPD (ATADJ) Summary
    value_key_for_calc_txpd = "Computed Invoice Value"
    txpd_summary_original = calculate_monthly_summary(combined_data.get("TXPD", []), "Month",
                                                      value_key=value_key_for_calc_txpd,
                                                      taxable_key="Gross Advance Adjusted", iamt_key="Integrated Tax",
                                                      camt_key="Central Tax", samt_key="State/UT Tax", cess_key="Cess",
                                                      invoice_key=None, processed_months=processed_months)
    if txpd_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in txpd_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_txpd_sum = create_or_replace_sheet(wb, "R1-Summary-TXPD", SECTION_TITLES["Summary-TXPD"],
                                              summary_display_columns)
        txpd_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_txpd, 0)} for row in
                                      txpd_summary_original]
        fill_worksheet_data(ws_txpd_sum, summary_display_columns, txpd_summary_display_ready)
        current_total_keys_map = {**base_data_keys_map, "Invoice Value": value_key_for_calc_txpd}
        _add_total_row_to_summary_sheet(ws_txpd_sum, txpd_summary_original, summary_display_columns,
                                        current_total_keys_map,
                                        COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_txpd_sum)

    # HSN Summary
    value_key_for_calc_hsn = "Computed Invoice Value"
    hsn_summary_original = calculate_monthly_summary(combined_data.get("HSN", []), "Reporting Month",
                                                     value_key=value_key_for_calc_hsn,
                                                     taxable_key="Taxable Value", iamt_key="Integrated Tax",
                                                     camt_key="Central Tax", samt_key="State/UT Tax", cess_key="Cess",
                                                     invoice_key=None,
                                                     processed_months=processed_months)
    if hsn_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in hsn_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_hsn_sum = create_or_replace_sheet(wb, "R1-Summary-HSN", SECTION_TITLES["Summary-HSN"],
                                             summary_display_columns)
        hsn_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_hsn, 0)} for row in
                                     hsn_summary_original]
        fill_worksheet_data(ws_hsn_sum, summary_display_columns, hsn_summary_display_ready)
        current_total_keys_map = {**base_data_keys_map, "Invoice Value": value_key_for_calc_hsn}
        _add_total_row_to_summary_sheet(ws_hsn_sum, hsn_summary_original, summary_display_columns,
                                        current_total_keys_map,
                                        COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_hsn_sum)

    # DOC Summary
    doc_summary_cols_list = COLUMN_HEADERS["Summary-DOC"]
    doc_summary_numeric_check_keys_list = NUMERIC_KEYS_BY_SECTION["Summary-DOC"]
    doc_summary_data_list = []
    if doc_data_from_combined or processed_months:
        month_dict_doc = {month_iter: {"doc_types_count": 0, "net_issued": 0, "total_num": 0, "cancelled": 0} for
                          month_iter in processed_months}

        temp_doc_agg = {}
        for row_doc_agg in doc_data_from_combined:
            month_doc_agg = row_doc_agg["Reporting Month"]
            if month_doc_agg not in temp_doc_agg:
                temp_doc_agg[month_doc_agg] = {"unique_doc_types": set(), "net_issued": 0, "total_num": 0,
                                               "cancelled": 0}

            if month_doc_agg not in month_dict_doc:
                month_dict_doc[month_doc_agg] = {"doc_types_count": 0, "net_issued": 0, "total_num": 0, "cancelled": 0}

            temp_doc_agg[month_doc_agg]["unique_doc_types"].add(row_doc_agg["doc_type_title"])
            temp_doc_agg[month_doc_agg]["net_issued"] += row_doc_agg.get("Net Issued", 0)
            temp_doc_agg[month_doc_agg]["total_num"] += row_doc_agg.get("Total Number", 0)
            temp_doc_agg[month_doc_agg]["cancelled"] += row_doc_agg.get("Cancelled", 0)

        for month_iter_fill in processed_months:
            if month_iter_fill in temp_doc_agg:
                month_dict_doc[month_iter_fill]["doc_types_count"] = len(
                    temp_doc_agg[month_iter_fill]["unique_doc_types"])
                month_dict_doc[month_iter_fill]["net_issued"] = temp_doc_agg[month_iter_fill]["net_issued"]
                month_dict_doc[month_iter_fill]["total_num"] = temp_doc_agg[month_iter_fill]["total_num"]
                month_dict_doc[month_iter_fill]["cancelled"] = temp_doc_agg[month_iter_fill]["cancelled"]

        doc_summary_data_list = [{"Reporting Month": month_iter_data,
                                  "No. of Records": values_data["doc_types_count"],
                                  "Net Issued Documents": values_data["net_issued"],
                                  "Documents Issued": values_data["total_num"],
                                  "Documents Cancelled": values_data["cancelled"]}
                                 for month_iter_data, values_data in month_dict_doc.items() if
                                 month_iter_data != "Unknown"]

        doc_summary_data_list.sort(
            key=lambda x_doc_sum: financial_order_sort.index(x_doc_sum["Reporting Month"]) if x_doc_sum[
                                                                                                  "Reporting Month"] in financial_order_sort else 999)

        if doc_summary_data_list and (ignore_warnings or any(
                r.get(k, 0) != 0 for r in doc_summary_data_list for k in doc_summary_numeric_check_keys_list if
                k != "No. of Records")):
            ws_doc_summary_final = create_or_replace_sheet(wb, "R1-Summary-DOC", SECTION_TITLES["Summary-DOC"],
                                                           doc_summary_cols_list)
            fill_worksheet_data(ws_doc_summary_final, doc_summary_cols_list, doc_summary_data_list)
            doc_summary_keys_for_summation = {col: col for col in doc_summary_cols_list}
            _add_total_row_to_summary_sheet(ws_doc_summary_final, doc_summary_data_list, doc_summary_cols_list,
                                            doc_summary_keys_for_summation, COLUMN_FORMATS["Summary-DOC"])

    # B2BA Summary
    value_key_for_calc_b2ba = "Total Invoice Value"
    b2ba_summary_original = calculate_monthly_summary(combined_data.get("B2BA", []), "Reporting Month",
                                                      value_key=value_key_for_calc_b2ba,
                                                      taxable_key="Total Taxable Value", iamt_key="Integrated Tax",
                                                      camt_key="Central Tax", samt_key="State/UT Tax", cess_key="Cess",
                                                      invoice_key="Revised/Original Invoice No.",
                                                      processed_months=processed_months)
    if b2ba_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in b2ba_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_b2ba_sum = create_or_replace_sheet(wb, "R1-Summary-B2BA Total", SECTION_TITLES["Summary-B2BA Total"],
                                              summary_display_columns)
        b2ba_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_b2ba, 0)} for row in
                                      b2ba_summary_original]
        fill_worksheet_data(ws_b2ba_sum, summary_display_columns, b2ba_summary_display_ready)
        current_total_keys_map = {**base_data_keys_map, "Invoice Value": value_key_for_calc_b2ba}
        _add_total_row_to_summary_sheet(ws_b2ba_sum, b2ba_summary_original, summary_display_columns,
                                        current_total_keys_map,
                                        COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_b2ba_sum)

    # EXP Summaries (WPAY, WOPAY, Total)
    value_key_for_calc_exp = "Total Invoice Value"
    exp_data_all = combined_data.get("EXP", [])
    expwp_data_filter = [row for row in exp_data_all if row.get("GST payment") == "WPAY"]
    expwp_summary_original = calculate_monthly_summary(expwp_data_filter, "Reporting Month",
                                                       value_key=value_key_for_calc_exp,
                                                       taxable_key="Total Taxable Value", iamt_key="Integrated Tax",
                                                       camt_key="Central Tax", samt_key="State/UT Tax", cess_key="Cess",
                                                       invoice_key="Invoice Number", processed_months=processed_months)
    if expwp_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in expwp_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_expwp_sum = create_or_replace_sheet(wb, "R1-Summary-EXPWP", SECTION_TITLES["Summary-EXPWP"],
                                               summary_display_columns)
        expwp_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_exp, 0)} for row in
                                       expwp_summary_original]
        fill_worksheet_data(ws_expwp_sum, summary_display_columns, expwp_summary_display_ready)
        current_total_keys_map = {**base_data_keys_map, "Invoice Value": value_key_for_calc_exp}
        _add_total_row_to_summary_sheet(ws_expwp_sum, expwp_summary_original, summary_display_columns,
                                        current_total_keys_map,
                                        COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_expwp_sum)

    expwop_data_filter = [row for row in exp_data_all if row.get("GST payment") == "WOPAY"]
    expwop_summary_original = calculate_monthly_summary(expwop_data_filter, "Reporting Month",
                                                        value_key=value_key_for_calc_exp,
                                                        taxable_key="Total Taxable Value", iamt_key="Integrated Tax",
                                                        camt_key="Central Tax", samt_key="State/UT Tax",
                                                        cess_key="Cess",
                                                        invoice_key="Invoice Number", processed_months=processed_months)
    if expwop_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in expwop_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_expwop_sum = create_or_replace_sheet(wb, "R1-Summary-EXPWOP", SECTION_TITLES["Summary-EXPWOP"],
                                                summary_display_columns)
        expwop_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_exp, 0)} for row in
                                        expwop_summary_original]
        fill_worksheet_data(ws_expwop_sum, summary_display_columns, expwop_summary_display_ready)
        _add_total_row_to_summary_sheet(ws_expwop_sum, expwop_summary_original, summary_display_columns,
                                        current_total_keys_map,
                                        COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_expwop_sum)

    exp_total_summary_original = calculate_monthly_summary(exp_data_all, "Reporting Month",
                                                           value_key=value_key_for_calc_exp,
                                                           taxable_key="Total Taxable Value", iamt_key="Integrated Tax",
                                                           camt_key="Central Tax", samt_key="State/UT Tax",
                                                           cess_key="Cess",
                                                           invoice_key="Invoice Number", processed_months=processed_months)
    if exp_total_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in exp_total_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_exp_total_sum = create_or_replace_sheet(wb, "R1-Summary-EXP-Total", SECTION_TITLES["Summary-EXP-Total"],
                                                   summary_display_columns)
        exp_total_summary_display_ready = [{**row, "Invoice Value": row.get(value_key_for_calc_exp, 0)} for row in
                                           exp_total_summary_original]
        fill_worksheet_data(ws_exp_total_sum, summary_display_columns, exp_total_summary_display_ready)
        _add_total_row_to_summary_sheet(ws_exp_total_sum, exp_total_summary_original, summary_display_columns,
                                        current_total_keys_map, COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_exp_total_sum)

    # CDNUR Summaries (B2CL, EXPWP, EXPWOP, Total)
    value_key_for_calc_cdnur = "Computed Invoice Value"
    cdnur_data_all = combined_data.get("CDNUR", [])

    cdnur_b2cl_filter = [row for row in cdnur_data_all if row.get("Type") == "B2CL"]
    cdnur_b2cl_summary_original = calculate_monthly_summary(cdnur_b2cl_filter, "Reporting Month",
                                                            value_key=value_key_for_calc_cdnur,
                                                            taxable_key="Taxable Value", iamt_key="Integrated Tax",
                                                            camt_key="Central Tax", samt_key="State/UT Tax",
                                                            cess_key="Cess",
                                                            invoice_key="C/D Note No",
                                                            processed_months=processed_months)
    if cdnur_b2cl_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in cdnur_b2cl_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_cdnur_b2cl = create_or_replace_sheet(wb, "R1-Summary-CDNUR-B2CL", SECTION_TITLES["Summary-CDNUR-B2CL"],
                                                summary_display_columns_note_type)
        cdnur_b2cl_display_ready = [{**row, "Note Value": row.get(value_key_for_calc_cdnur, 0)} for row in
                                    cdnur_b2cl_summary_original]
        fill_worksheet_data(ws_cdnur_b2cl, summary_display_columns_note_type, cdnur_b2cl_display_ready)
        current_total_keys_map = {col: col for col in summary_display_columns_note_type}
        current_total_keys_map["Note Value"] = value_key_for_calc_cdnur
        _add_total_row_to_summary_sheet(ws_cdnur_b2cl, cdnur_b2cl_summary_original, summary_display_columns_note_type,
                                        current_total_keys_map, COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_cdnur_b2cl)

    cdnur_expwp_filter = [row for row in cdnur_data_all if row.get("Type") == "EXPWP"]
    cdnur_expwp_summary_original = calculate_monthly_summary(cdnur_expwp_filter, "Reporting Month",
                                                             value_key=value_key_for_calc_cdnur,
                                                             taxable_key="Taxable Value", iamt_key="Integrated Tax",
                                                             camt_key="Central Tax", samt_key="State/UT Tax",
                                                             cess_key="Cess",
                                                             invoice_key="C/D Note No",
                                                             processed_months=processed_months)
    if cdnur_expwp_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in cdnur_expwp_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_cdnur_expwp = create_or_replace_sheet(wb, "R1-Summary-CDNUR-EXPWP", SECTION_TITLES["Summary-CDNUR-EXPWP"],
                                                 summary_display_columns_note_type)
        cdnur_expwp_display_ready = [{**row, "Note Value": row.get(value_key_for_calc_cdnur, 0)} for row in
                                     cdnur_expwp_summary_original]
        fill_worksheet_data(ws_cdnur_expwp, summary_display_columns_note_type, cdnur_expwp_display_ready)
        _add_total_row_to_summary_sheet(ws_cdnur_expwp, cdnur_expwp_summary_original, summary_display_columns_note_type,
                                        current_total_keys_map, COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_cdnur_expwp)

    cdnur_expwop_filter = [row for row in cdnur_data_all if row.get("Type") == "EXPWOP"]
    cdnur_expwop_summary_original = calculate_monthly_summary(cdnur_expwop_filter, "Reporting Month",
                                                              value_key=value_key_for_calc_cdnur,
                                                              taxable_key="Taxable Value", iamt_key="Integrated Tax",
                                                              camt_key="Central Tax", samt_key="State/UT Tax",
                                                              cess_key="Cess",
                                                              invoice_key="C/D Note No",
                                                              processed_months=processed_months)
    if cdnur_expwop_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in cdnur_expwop_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_cdnur_expwop = create_or_replace_sheet(wb, "R1-Summary-CDNUR-EXPWOP", SECTION_TITLES["Summary-CDNUR-EXPWOP"],
                                                  summary_display_columns_note_type)
        cdnur_expwop_display_ready = [{**row, "Note Value": row.get(value_key_for_calc_cdnur, 0)} for row in
                                      cdnur_expwop_summary_original]
        fill_worksheet_data(ws_cdnur_expwop, summary_display_columns_note_type, cdnur_expwop_display_ready)
        _add_total_row_to_summary_sheet(ws_cdnur_expwop, cdnur_expwop_summary_original,
                                        summary_display_columns_note_type,
                                        current_total_keys_map, COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_cdnur_expwop)

    cdnur_total_summary_original = calculate_monthly_summary(cdnur_data_all, "Reporting Month",
                                                             value_key=value_key_for_calc_cdnur,
                                                             taxable_key="Taxable Value", iamt_key="Integrated Tax",
                                                             camt_key="Central Tax", samt_key="State/UT Tax",
                                                             cess_key="Cess",
                                                             invoice_key="C/D Note No",
                                                             processed_months=processed_months)
    if cdnur_total_summary_original and (ignore_warnings or any(
            r.get(k, 0) != 0 for r in cdnur_total_summary_original for k in summary_numeric_check_keys_list if
            k != "No. of Records")):
        ws_cdnur_total = create_or_replace_sheet(wb, "R1-Summary-CDNUR-TOTAL", SECTION_TITLES["Summary-CDNUR-TOTAL"],
                                                 summary_display_columns_note_type)
        cdnur_total_display_ready = [{**row, "Note Value": row.get(value_key_for_calc_cdnur, 0)} for row in
                                     cdnur_total_summary_original]
        fill_worksheet_data(ws_cdnur_total, summary_display_columns_note_type, cdnur_total_display_ready)
        _add_total_row_to_summary_sheet(ws_cdnur_total, cdnur_total_summary_original, summary_display_columns_note_type,
                                        current_total_keys_map, COLUMN_FORMATS["Summary"])
        summary_ws_list_final.append(ws_cdnur_total)

    print("[DEBUG] Applying formatting to summary sheets (after adding totals)...")
    for ws_summary_iter in summary_ws_list_final:
        current_display_cols_for_format = summary_display_columns  # Default
        title_check = ws_summary_iter.title.upper()
        if "CDNR" in title_check or "CDNUR" in title_check:
            current_display_cols_for_format = summary_display_columns_note_type

        apply_format_and_autofit(ws_summary_iter, current_display_cols_for_format,
                                 col_format_map=COLUMN_FORMATS.get("Summary", {}))
    print("[DEBUG] Finished formatting summary sheets")

    print(f"[DEBUG] Saving workbook to {save_path}...")
    try:
        wb.save(save_path)
        print(f"[DEBUG] Workbook saved successfully to {save_path}")
    except Exception as e_save:
        print(f"[ERROR] Failed to save workbook: {e_save}")
        detailed_error_info = traceback.format_exc()
        print(detailed_error_info)
        raise
    print("[DEBUG] GSTR1 processing completed")
    return wb, unexpected_sections_details


if __name__ == '__main__':
    import traceback

    print("GSTR1 Processor script loaded. To run, call process_gstr1() with appropriate arguments.")
