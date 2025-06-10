# gstr2b_processor.py

import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os, datetime
import traceback  # Import for detailed error reporting

# Assuming telemetry is a module you have for sending events.
# If not, you might need to comment out or adjust these lines.
try:
    from utils.telemetry import send_event
except ImportError:
    print(
        "[WARN] Telemetry module not found in gstr2b_processor. Telemetry will be handled by UI if not available here.")


    def send_event(event_type, payload):  # Dummy function
        print(f"[TELEMETRY_STUB] Event: {event_type}, Payload: {payload}")

# ----------------------- Global Variables for Totals ----------------------- #
INDIAN_FORMAT_GSTR2B = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00" + \
                       r";-##\,##\,##\,##0.00" + \
                       r";-"
RED_BOLD_FONT = Font(bold=True, color="FF0000")
BOLD_FONT = Font(bold=True)  # Standard bold font

# Configuration for summing main document values uniquely in detail sheets
GSTR2B_MAIN_VALUE_CONFIG = {
    "2B-B2B": {"value_col": "Invoice Value", "id_col": "Invoice Number"},
    "2B-B2BA": {"value_col": "Invoice Value", "id_col": "Invoice Number"},
    "2B-B2B(ITC_Rej)": {"value_col": "Invoice Value", "id_col": "Invoice Number"},
    "2B-CDNR": {"value_col": "Note Value", "id_col": "Note Number"},
    "2B-IMPG": {"value_col": "Calculated Invoice Value", "id_col": "Bill of Entry Number"},
    "2B-B2BA(cum)": {"value_col": "Calculated Invoice Value", "id_col": None}  # Direct sum for this cumulative sheet
}

# Columns to sum in the total row of detail sheets
GSTR2B_DETAIL_SHEET_TOTAL_COLUMNS = {
    "2B-B2B": ["Invoice Value", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "2B-B2BA": ["Invoice Value", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "2B-B2BA(cum)": ["Total Documents", "Calculated Invoice Value", "Total Taxable Value", "Integrated Tax",
                     "Central Tax", "State/UT Tax", "Cess"],
    "2B-CDNR": ["Note Value", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    "2B-IMPG": ["Calculated Invoice Value", "Taxable Value", "Integrated Tax", "Cess"],
    "2B-B2B(ITC_Rej)": ["Invoice Value", "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"]
}
# Add _sws variations to detail sheet total columns, they share the same columns for totals
for key in list(GSTR2B_DETAIL_SHEET_TOTAL_COLUMNS.keys()):  # Iterate over a copy of keys
    if not key.endswith("_sws"):
        sws_key = key + "_sws"
        GSTR2B_DETAIL_SHEET_TOTAL_COLUMNS[sws_key] = GSTR2B_DETAIL_SHEET_TOTAL_COLUMNS[key]
        if key in GSTR2B_MAIN_VALUE_CONFIG:
            GSTR2B_MAIN_VALUE_CONFIG[sws_key] = GSTR2B_MAIN_VALUE_CONFIG[key]

# Columns to sum in the total row of summary sheets
GSTR2B_SUMMARY_SHEET_TOTAL_COLUMNS = ["Invoice Value", "Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax",
                                      "Cess"]

# ----------------------- Allowed Sections ----------------------- #
# Defines expected child keys for specific parent nodes within the "data" object of GSTR-2B JSON
ALLOWED_DOC_SUBSECTIONS = {
    "docdata": {"b2b", "b2ba", "cdnr", "impg"},  # Added "isd" as a common one, can be adjusted
    "docRejdata": {"b2b"},  # Added "b2ba", "cdnra" based on some specs, can be adjusted
    "cpsumm": {"b2b", "b2ba", "cdnr"}  # Cumulative Period Summary for B2BA
    # Other top-level keys in "data" like "rtnprd", "isdsum", "itcsum" are handled implicitly by extractors or ignored if not processed.
}


# ----------------------- Utility Functions ----------------------- #
def get_tax_period(ret_period_str):
    """
    Parses a return period string (expected "MMYYYY") and returns the month name.
    """
    month_map = {
        "01": "January", "02": "February", "03": "March", "04": "April",
        "05": "May", "06": "June", "07": "July", "08": "August",
        "09": "September", "10": "October", "11": "November", "12": "December"
    }
    # Ensure ret_period_str is a string and has at least 2 characters for month part
    if isinstance(ret_period_str, str) and len(ret_period_str) >= 2:
        month_code = ret_period_str[:2]
        return month_map.get(month_code, "Unknown")
    return "Unknown"


def parse_number(val, float_2dec=False, int_no_dec=False):
    try:
        num = float(val)
        if int_no_dec:
            return int(num)
        if float_2dec:
            return round(num, 2)
        return num
    except (ValueError, TypeError):
        return 0


def get_numeric_value(item, key):
    val = item.get(key, 0)
    if isinstance(val, str):
        val = val.strip()
    return parse_number(val, float_2dec=True)


# ----------------------- Extraction Functions ----------------------- #
# (extract_b2b, extract_b2ba, extract_b2ba_cum, extract_cdnr, extract_impg, extract_b2b_itc_rej remain the same)
def extract_b2b(data, filing_period):
    """Extract B2B section data from GSTR-2B JSON (ITC Accepted)."""
    b2b_data = data.get("data", {}).get("docdata", {}).get("b2b", [])
    if not isinstance(b2b_data, list): b2b_data = []
    extracted_data = {"2B-B2B": []}
    for supplier in b2b_data:
        if not isinstance(supplier, dict): continue
        ctin = supplier.get("ctin", "")
        trdnm = supplier.get("trdnm", "")
        invoices = supplier.get("inv", [])
        if not isinstance(invoices, list): invoices = [invoices] if isinstance(invoices, dict) else []
        for inv in invoices:
            if not isinstance(inv, dict): continue
            base_row = {
                "GSTIN/UIN of Supplier": ctin, "Trade/Legal Name": trdnm, "Invoice Number": inv.get("inum", ""),
                "Invoice Type": inv.get("typ", ""), "Invoice Date": inv.get("dt", ""),
                "Invoice Value": get_numeric_value(inv, "val"),
                "Place of Supply": parse_number(inv.get("pos", "0"), int_no_dec=True),
                "Supply Attract Reverse Charge": inv.get("rev", ""),
                "GSTR Period": get_tax_period(supplier.get("supprd", "")),
                "GSTR Filing Date": supplier.get("supfildt", ""),
                "GSTR Filing Period": filing_period,  # This is the recipient's GSTR-2B period
                "ITC Availability": inv.get("itcavl", ""), "Reason": inv.get("rsn", ""),
                "Source": inv.get("srctyp", "")
            }
            items = inv.get("items", [])
            if not isinstance(items, list): items = []
            if items:
                for item in items:
                    if not isinstance(item, dict): continue
                    row = base_row.copy()
                    row.update({
                        "Invoice Part": item.get("num", ""), "Tax Rate": get_numeric_value(item, "rt"),
                        "Total Taxable Value": get_numeric_value(item, "txval"),
                        "Integrated Tax": get_numeric_value(item, "igst"),
                        "Central Tax": get_numeric_value(item, "cgst"),
                        "State/UT Tax": get_numeric_value(item, "sgst"),
                        "Cess": get_numeric_value(item, "cess")
                    })
                    extracted_data["2B-B2B"].append(row)
            else:
                row = base_row.copy()
                row.update({
                    "Invoice Part": "", "Tax Rate": "",
                    "Total Taxable Value": get_numeric_value(inv, "txval"),  # Fallback if no items
                    "Integrated Tax": get_numeric_value(inv, "igst"),
                    "Central Tax": get_numeric_value(inv, "cgst"),
                    "State/UT Tax": get_numeric_value(inv, "sgst"),
                    "Cess": get_numeric_value(inv, "cess")
                })
                extracted_data["2B-B2B"].append(row)
    return extracted_data


def extract_b2ba(data, filing_period):
    """Extract B2BA section data from GSTR-2B JSON."""
    b2ba_data = data.get("data", {}).get("docdata", {}).get("b2ba", [])
    if not isinstance(b2ba_data, list): b2ba_data = []
    extracted_data = {"2B-B2BA": []}
    for supplier in b2ba_data:
        if not isinstance(supplier, dict): continue
        ctin = supplier.get("ctin", "")
        invoices = supplier.get("inv", [])
        if not isinstance(invoices, list): invoices = [invoices] if isinstance(invoices, dict) else []
        for inv in invoices:
            if not isinstance(inv, dict): continue
            base_row = {
                "GSTIN/UIN of Supplier": ctin, "Trade/Legal Name": supplier.get("trdnm", ""),
                "Original Invoice Number": inv.get("oinum", ""), "Original Invoice Date": inv.get("oidt", ""),
                "Invoice Number": inv.get("inum", ""), "Invoice Date": inv.get("dt", ""),
                "Invoice Type": inv.get("typ", ""), "Invoice Value": get_numeric_value(inv, "val"),
                "Place of Supply": parse_number(inv.get("pos", "0"), int_no_dec=True),
                "Supply Attract Reverse Charge": inv.get("rev", ""),
                "GSTR Period": get_tax_period(supplier.get("supprd", "")),
                "GSTR Filing Date": supplier.get("supfildt", ""),
                "GSTR Filing Period": filing_period,
                "ITC Availability": inv.get("itcavl", ""), "Reason": inv.get("rsn", "")
            }
            items = inv.get("items", [])
            if not isinstance(items, list): items = []
            if items:
                for item in items:
                    if not isinstance(item, dict): continue
                    row = base_row.copy()
                    row.update({
                        "Invoice Part": item.get("num", ""), "Tax Rate": get_numeric_value(item, "rt"),
                        "Total Taxable Value": get_numeric_value(item, "txval"),
                        "Integrated Tax": get_numeric_value(item, "igst"),
                        "Central Tax": get_numeric_value(item, "cgst"),
                        "State/UT Tax": get_numeric_value(item, "sgst"),
                        "Cess": get_numeric_value(item, "cess")
                    })
                    extracted_data["2B-B2BA"].append(row)
            else:  # Fallback if no items
                row = base_row.copy()
                row.update({
                    "Invoice Part": "", "Tax Rate": "",
                    "Total Taxable Value": get_numeric_value(inv, "txval"),
                    "Integrated Tax": get_numeric_value(inv, "igst"),
                    "Central Tax": get_numeric_value(inv, "cgst"),
                    "State/UT Tax": get_numeric_value(inv, "sgst"),
                    "Cess": get_numeric_value(inv, "cess")
                })
                extracted_data["2B-B2BA"].append(row)
    return extracted_data


def extract_b2ba_cum(data, filing_period):
    """Extract B2BA cumulative section data from GSTR-2B JSON."""
    b2ba_cum_data = data.get("data", {}).get("cpsumm", {}).get("b2ba", [])
    if not isinstance(b2ba_cum_data, list): b2ba_cum_data = []
    extracted_data = {"2B-B2BA(cum)": []}
    for supplier in b2ba_cum_data:
        if not isinstance(supplier, dict): continue
        row = {
            "GSTIN/UIN of Supplier": supplier.get("ctin", ""),
            "Trade/Legal Name": supplier.get("trdnm", ""),
            "Total Documents": parse_number(supplier.get("ttldocs", 0), int_no_dec=True),
            "Total Taxable Value": get_numeric_value(supplier, "txval"),
            "Integrated Tax": get_numeric_value(supplier, "igst"),
            "Central Tax": get_numeric_value(supplier, "cgst"),
            "State/UT Tax": get_numeric_value(supplier, "sgst"),
            "Cess": get_numeric_value(supplier, "cess"),
            "GSTR Period": get_tax_period(supplier.get("supprd", "")),
            "GSTR Filing Date": supplier.get("supfildt", ""),
            "GSTR Filing Period": filing_period
        }
        row["Calculated Invoice Value"] = sum(get_numeric_value(row, k) for k in
                                              ["Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax",
                                               "Cess"])
        extracted_data["2B-B2BA(cum)"].append(row)
    return extracted_data


def extract_cdnr(data, filing_period):
    """Extract CDNR section data from GSTR-2B JSON."""
    cdnr_data = data.get("data", {}).get("docdata", {}).get("cdnr", [])
    if not isinstance(cdnr_data, list): cdnr_data = []
    extracted_data = {"2B-CDNR": []}
    for supplier in cdnr_data:
        if not isinstance(supplier, dict): continue
        ctin = supplier.get("ctin", "")
        notes = supplier.get("nt", [])
        if not isinstance(notes, list): notes = [notes] if isinstance(notes, dict) else []
        for note in notes:
            if not isinstance(note, dict): continue
            base_row = {
                "GSTIN/UIN of Supplier": ctin, "Trade/Legal Name": supplier.get("trdnm", ""),
                "Note Number": note.get("ntnum", ""), "Note Type": note.get("typ", ""),
                "Note Supply Type": note.get("suptyp", ""), "Note Date": note.get("dt", ""),
                "Note Value": get_numeric_value(note, "val"),
                "Place of Supply": parse_number(note.get("pos", "0"), int_no_dec=True),
                "Supply Attract Reverse Charge": note.get("rev", ""),
                "GSTR Period": get_tax_period(supplier.get("supprd", "")),
                "GSTR Filing Date": supplier.get("supfildt", ""),
                "GSTR Filing Period": filing_period,
                "ITC Availability": note.get("itcavl", ""), "Reason": note.get("rsn", "")
            }
            items = note.get("items", [])
            if not isinstance(items, list): items = []
            if items:
                for item in items:
                    if not isinstance(item, dict): continue
                    row = base_row.copy()
                    row.update({
                        "Note Part": item.get("num", ""), "Tax Rate": get_numeric_value(item, "rt"),
                        "Total Taxable Value": get_numeric_value(item, "txval"),
                        "Integrated Tax": get_numeric_value(item, "igst"),
                        "Central Tax": get_numeric_value(item, "cgst"),
                        "State/UT Tax": get_numeric_value(item, "sgst"),
                        "Cess": get_numeric_value(item, "cess")
                    })
                    extracted_data["2B-CDNR"].append(row)
            else:  # Fallback if no items
                row = base_row.copy()
                row.update({
                    "Note Part": "", "Tax Rate": "",
                    "Total Taxable Value": get_numeric_value(note, "txval"),
                    "Integrated Tax": get_numeric_value(note, "igst"),
                    "Central Tax": get_numeric_value(note, "cgst"),
                    "State/UT Tax": get_numeric_value(note, "sgst"),
                    "Cess": get_numeric_value(note, "cess")
                })
                extracted_data["2B-CDNR"].append(row)
    return extracted_data


def extract_impg(data, filing_period):
    """Extract IMPG section data from GSTR-2B JSON."""
    impg_data = data.get("data", {}).get("docdata", {}).get("impg", [])
    if not isinstance(impg_data, list): impg_data = []
    extracted_data = {"2B-IMPG": []}
    for entry in impg_data:
        if not isinstance(entry, dict): continue
        row = {
            "ICEGATE Reference Date": entry.get("refdt", ""), "Port Code": entry.get("portcode", ""),
            "Bill of Entry Number": entry.get("boenum", ""), "Bill of Entry Date": entry.get("boedt", ""),
            "Taxable Value": get_numeric_value(entry, "txval"),
            "Integrated Tax": get_numeric_value(entry, "igst"), "Cess": get_numeric_value(entry, "cess"),
            "Record Date": entry.get("recdt", ""), "GSTR Filing Period": filing_period,
            "Amended (Yes)": entry.get("isamd", "")
        }
        row["Calculated Invoice Value"] = sum(
            get_numeric_value(row, k) for k in ["Taxable Value", "Integrated Tax", "Cess"])
        extracted_data["2B-IMPG"].append(row)
    return extracted_data


def extract_b2b_itc_rej(data, filing_period):
    """Extract B2B (ITC Rejected) section data from GSTR-2B JSON."""
    # This typically comes from docRejdata -> b2b
    b2b_rej_data = data.get("data", {}).get("docRejdata", {}).get("b2b", [])
    if not isinstance(b2b_rej_data, list): b2b_rej_data = []
    extracted_data = {"2B-B2B(ITC_Rej)": []}
    for supplier in b2b_rej_data:
        if not isinstance(supplier, dict): continue
        ctin = supplier.get("ctin", "")
        invoices = supplier.get("inv", [])
        if not isinstance(invoices, list): invoices = [invoices] if isinstance(invoices, dict) else []
        for inv in invoices:
            if not isinstance(inv, dict): continue
            base_row = {
                "GSTIN/UIN of Supplier": ctin, "Trade/Legal Name": supplier.get("trdnm", ""),
                "Invoice Number": inv.get("inum", ""), "Invoice Type": inv.get("typ", ""),
                "Invoice Date": inv.get("dt", ""), "Invoice Value": get_numeric_value(inv, "val"),
                "Place of Supply": parse_number(inv.get("pos", "0"), int_no_dec=True),
                "GSTR Period": get_tax_period(supplier.get("supprd", "")),
                "GSTR Filing Date": supplier.get("supfildt", ""),
                "GSTR Filing Period": filing_period, "Source": inv.get("srctyp", "")
                # ITC Availability and Reason are typically not in rejected data in the same way as accepted.
            }
            items = inv.get("items", [])
            if not isinstance(items, list): items = []
            if items:
                for item in items:
                    if not isinstance(item, dict): continue
                    row = base_row.copy()
                    row.update({
                        "Invoice Part": item.get("num", ""), "Tax Rate": get_numeric_value(item, "rt"),
                        "Total Taxable Value": get_numeric_value(item, "txval"),
                        "Integrated Tax": get_numeric_value(item, "igst"),
                        "Central Tax": get_numeric_value(item, "cgst"),
                        "State/UT Tax": get_numeric_value(item, "sgst"),
                        "Cess": get_numeric_value(item, "cess")
                    })
                    extracted_data["2B-B2B(ITC_Rej)"].append(row)
            else:  # Fallback if no items
                row = base_row.copy()
                row.update({
                    "Invoice Part": "", "Tax Rate": "",
                    "Total Taxable Value": get_numeric_value(inv, "txval"),
                    "Integrated Tax": get_numeric_value(inv, "igst"),
                    "Central Tax": get_numeric_value(inv, "cgst"),
                    "State/UT Tax": get_numeric_value(inv, "sgst"),
                    "Cess": get_numeric_value(inv, "cess")
                })
                extracted_data["2B-B2B(ITC_Rej)"].append(row)
    return extracted_data


# ----------------------- Helper Function for Detail Sheet Totals (Unchanged) ----------------------- #
def _add_total_row_to_gstr2b_detail_sheet(ws, sheet_key, rows_data, column_headers_for_sheet, column_formats_for_sheet):
    if not rows_data: return

    total_row_idx = ws.max_row + 1
    totals = {}
    columns_to_sum_for_current_sheet = GSTR2B_DETAIL_SHEET_TOTAL_COLUMNS.get(sheet_key, [])
    main_value_conf = GSTR2B_MAIN_VALUE_CONFIG.get(sheet_key, {})
    main_value_col = main_value_conf.get("value_col")
    id_col = main_value_conf.get("id_col")

    # Sum main value column (e.g., Invoice Value, Note Value) uniquely by ID
    if main_value_col and main_value_col in columns_to_sum_for_current_sheet:
        if id_col:
            summed_ids_for_main_val = set()
            current_sum_main_val = 0
            for row_item in rows_data:
                doc_id = row_item.get(id_col)
                # Use a tuple of (doc_id, value) if values can change for the same ID across rows but should be summed once per ID
                # For simple "sum once per ID", just the doc_id is enough if the value is consistent or taken from one instance.
                # The current logic sums the value from the first encountered unique ID.
                # If values differ for the same ID across rows (e.g. multi-rate invoice lines sharing one invoice value),
                # this ensures the invoice value is added only once for that ID.
                if doc_id and doc_id not in summed_ids_for_main_val:  # Simpler: count value for first occurrence of ID
                    current_sum_main_val += parse_number(row_item.get(main_value_col, 0), float_2dec=True)
                    summed_ids_for_main_val.add(doc_id)
                elif not doc_id:  # For entries without a unique ID, sum them directly
                    current_sum_main_val += parse_number(row_item.get(main_value_col, 0), float_2dec=True)
            totals[main_value_col] = current_sum_main_val
        else:  # If no id_col, sum all occurrences (e.g., for B2BA(cum))
            totals[main_value_col] = sum(parse_number(r.get(main_value_col, 0), float_2dec=True) for r in rows_data)

    # Sum other specified numeric columns
    for col_header in columns_to_sum_for_current_sheet:
        if col_header == main_value_col:  # Already handled
            continue
        totals[col_header] = sum(parse_number(r.get(col_header, 0), float_2dec=True) for r in rows_data)

    # Write the total row
    # Determine where to put "Total" label, typically first or second column
    label_col_index = 1  # Default to first column
    if len(column_headers_for_sheet) > 1 and "Trade/Legal Name" in column_headers_for_sheet:
        try:
            label_col_index = column_headers_for_sheet.index("Trade/Legal Name") + 1
        except ValueError:
            pass  # Keep default if not found
    elif len(column_headers_for_sheet) > 1 and "GSTIN/UIN of Supplier" in column_headers_for_sheet:
        try:
            label_col_index = column_headers_for_sheet.index("GSTIN/UIN of Supplier") + 1
        except ValueError:
            pass

    ws.cell(row=total_row_idx, column=label_col_index).value = "Total"
    ws.cell(row=total_row_idx, column=label_col_index).font = RED_BOLD_FONT
    ws.cell(row=total_row_idx, column=label_col_index).alignment = Alignment(horizontal="left")

    for col_idx, header_name in enumerate(column_headers_for_sheet, start=1):
        if col_idx == label_col_index and ws.cell(row=total_row_idx, column=col_idx).value == "Total":
            continue  # Skip overwriting "Total" if it's in a data column

        cell = ws.cell(row=total_row_idx, column=col_idx)
        if header_name in totals:
            cell.value = totals[header_name]
            cell.font = RED_BOLD_FONT
            fmt = column_formats_for_sheet.get(header_name, INDIAN_FORMAT_GSTR2B if isinstance(totals[header_name], (
            int, float)) else "General")
            cell.number_format = fmt
        elif not cell.value:  # Only fill if empty, to not overwrite "Total" if it landed in a non-summed col
            cell.value = ""


# ----------------------- Helper Function for Summary Sheet Totals (Unchanged) ----------------------- #
def _add_total_row_to_gstr2b_summary_sheet(ws, summary_sheet_data, display_column_headers, summary_column_formats_map):
    if not summary_sheet_data: return

    total_row_idx = ws.max_row + 1
    grand_totals = {}

    # Summing based on GSTR2B_SUMMARY_SHEET_TOTAL_COLUMNS definition
    for col_header in GSTR2B_SUMMARY_SHEET_TOTAL_COLUMNS:  # These are the data keys
        grand_totals[col_header] = sum(
            parse_number(month_data.get(col_header, 0), float_2dec=True) for month_data in summary_sheet_data)

    # Writing totals based on display_column_headers
    for col_idx, header_name_display in enumerate(display_column_headers, start=1):
        cell = ws.cell(row=total_row_idx, column=col_idx)
        if header_name_display == "Month":  # Label for the total row
            cell.value = "Total"
            cell.font = RED_BOLD_FONT
            cell.alignment = Alignment(horizontal="left")
        elif header_name_display in grand_totals:  # Check if the display header is a key in grand_totals
            total_val = grand_totals[header_name_display]
            cell.value = total_val
            cell.font = RED_BOLD_FONT
            # Apply number format based on the display header name
            fmt = summary_column_formats_map.get(header_name_display, INDIAN_FORMAT_GSTR2B)
            cell.number_format = fmt
        else:
            # This case might occur if display_column_headers has items not in GSTR2B_SUMMARY_SHEET_TOTAL_COLUMNS
            # or if a key in GSTR2B_SUMMARY_SHEET_TOTAL_COLUMNS is not in display_column_headers.
            # Ensure consistency or handle as needed. For now, blank.
            cell.value = ""


# ----------------------- Summary Generation (Unchanged) ----------------------- #
def create_summary_sheets(wb, combined_data):
    print("[DEBUG] Creating summary sheets...")
    summary_data_output = {
        "2B-Summary-B2B_not RC(ITC_Avl)": [], "2B-Summary-B2B_RC(ITC_Avl)": [],
        "2B-Summary-B2BA_not RC(ITC_Avl)": [], "2B-Summary-B2BA_RC(ITC_Avl)": [],
        "2B-Summary-B2BA_cum(ITC_Avl)": [],
        "2B-Summary-CDNR_DN(ITC_Avl)": [], "2B-Summary-CDNR_CN(ITC_Avl)": [], "2B-Summary-CDNR_RC(ITC_Avl)": [],
        "2B-Summary-IMPG(ITC_Avl)": [], "2B-Summary-IMPGA(ITC_Avl)": [],
        "2B-Summary-B2B(ITC_Rej)": []
    }
    financial_order = ["April", "May", "June", "July", "August", "September",
                       "October", "November", "December", "January", "February", "March"]

    all_filing_periods = set()
    for _, rows in combined_data.items():
        for row in rows:
            filing_period = row.get("GSTR Filing Period")  # This is recipient's GSTR-2B period
            if filing_period and isinstance(filing_period, str):
                all_filing_periods.add(filing_period)

    # Ensure "Unknown" is handled if present, though ideally all should be valid.
    all_filing_periods_sorted = sorted(list(all_filing_periods),
                                       key=lambda x: financial_order.index(x) if x in financial_order else (
                                                   len(financial_order) + (1 if x == "Unknown" else 0)))

    monthly_template = {col: 0 for col in GSTR2B_SUMMARY_SHEET_TOTAL_COLUMNS}
    agg_summaries = {key: {month: {"Month": month, **monthly_template.copy()} for month in all_filing_periods_sorted}
                     for key in summary_data_output.keys()}

    # B2B Summaries
    b2b_rows = combined_data.get("2B-B2B", [])
    unique_inv_b2b_not_rc = {}  # To sum invoice value once per invoice: {filing_period: {inv_num: inv_val}}
    unique_inv_b2b_rc = {}
    for r in b2b_rows:
        fp = r.get("GSTR Filing Period")
        if not fp or fp not in all_filing_periods_sorted: continue
        inv_val = get_numeric_value(r, "Invoice Value")
        inv_num = r.get("Invoice Number")
        target_summary_key = None
        unique_set_for_inv_val = None

        if r.get("Supply Attract Reverse Charge") == "N":
            target_summary_key = "2B-Summary-B2B_not RC(ITC_Avl)"
            unique_set_for_inv_val = unique_inv_b2b_not_rc
        elif r.get("Supply Attract Reverse Charge") == "Y":
            target_summary_key = "2B-Summary-B2B_RC(ITC_Avl)"
            unique_set_for_inv_val = unique_inv_b2b_rc

        if not target_summary_key: continue

        target_month_summary = agg_summaries[target_summary_key][fp]
        # Add invoice value only once per unique invoice number for this filing period
        if inv_num not in unique_set_for_inv_val.setdefault(fp, set()):
            target_month_summary["Invoice Value"] += inv_val
            unique_set_for_inv_val[fp].add(inv_num)

        target_month_summary["Taxable Value"] += get_numeric_value(r, "Total Taxable Value")
        target_month_summary["Integrated Tax"] += get_numeric_value(r, "Integrated Tax")
        target_month_summary["Central Tax"] += get_numeric_value(r, "Central Tax")
        target_month_summary["State/UT Tax"] += get_numeric_value(r, "State/UT Tax")
        target_month_summary["Cess"] += get_numeric_value(r, "Cess")

    # B2BA Summaries
    b2ba_rows = combined_data.get("2B-B2BA", [])
    unique_inv_b2ba_not_rc = {}
    unique_inv_b2ba_rc = {}
    for r in b2ba_rows:
        fp = r.get("GSTR Filing Period")
        if not fp or fp not in all_filing_periods_sorted: continue
        inv_val = get_numeric_value(r, "Invoice Value")
        inv_num = r.get("Invoice Number")  # Amended invoice number
        target_summary_key = None
        unique_set_for_inv_val = None

        if r.get("Supply Attract Reverse Charge") == "N":
            target_summary_key = "2B-Summary-B2BA_not RC(ITC_Avl)"
            unique_set_for_inv_val = unique_inv_b2ba_not_rc
        elif r.get("Supply Attract Reverse Charge") == "Y":
            target_summary_key = "2B-Summary-B2BA_RC(ITC_Avl)"
            unique_set_for_inv_val = unique_inv_b2ba_rc

        if not target_summary_key: continue

        target_month_summary = agg_summaries[target_summary_key][fp]
        if inv_num not in unique_set_for_inv_val.setdefault(fp, set()):
            target_month_summary["Invoice Value"] += inv_val
            unique_set_for_inv_val[fp].add(inv_num)

        target_month_summary["Taxable Value"] += get_numeric_value(r, "Total Taxable Value")
        target_month_summary["Integrated Tax"] += get_numeric_value(r, "Integrated Tax")
        target_month_summary["Central Tax"] += get_numeric_value(r, "Central Tax")
        target_month_summary["State/UT Tax"] += get_numeric_value(r, "State/UT Tax")
        target_month_summary["Cess"] += get_numeric_value(r, "Cess")

    # B2BA Cumulative Summary
    b2bacum_rows = combined_data.get("2B-B2BA(cum)", [])
    for r in b2bacum_rows:  # This is already aggregated by supplier, so invoice value is direct sum
        fp = r.get("GSTR Filing Period")
        if not fp or fp not in all_filing_periods_sorted: continue
        target_month_summary = agg_summaries["2B-Summary-B2BA_cum(ITC_Avl)"][fp]
        target_month_summary["Invoice Value"] += get_numeric_value(r,
                                                                   "Calculated Invoice Value")  # Already a sum for this entry
        target_month_summary["Taxable Value"] += get_numeric_value(r, "Total Taxable Value")
        target_month_summary["Integrated Tax"] += get_numeric_value(r, "Integrated Tax")
        target_month_summary["Central Tax"] += get_numeric_value(r, "Central Tax")
        target_month_summary["State/UT Tax"] += get_numeric_value(r, "State/UT Tax")
        target_month_summary["Cess"] += get_numeric_value(r, "Cess")

    # CDNR Summaries
    cdnr_rows = combined_data.get("2B-CDNR", [])
    unique_notes_cn = {}
    unique_notes_dn = {}
    unique_notes_rc = {}
    for r in cdnr_rows:
        fp = r.get("GSTR Filing Period")
        if not fp or fp not in all_filing_periods_sorted: continue
        note_val = get_numeric_value(r, "Note Value")
        note_num = r.get("Note Number")
        note_type = r.get("Note Type", "").upper()
        target_key = None
        unique_set_for_note_val = None

        if r.get("Supply Attract Reverse Charge") == "Y":  # RC CDNs/DNs
            target_key = "2B-Summary-CDNR_RC(ITC_Avl)"
            unique_set_for_note_val = unique_notes_rc
        elif note_type == "C":  # Credit Notes (non-RC)
            target_key = "2B-Summary-CDNR_CN(ITC_Avl)"
            unique_set_for_note_val = unique_notes_cn
        elif note_type == "D":  # Debit Notes (non-RC)
            target_key = "2B-Summary-CDNR_DN(ITC_Avl)"
            unique_set_for_note_val = unique_notes_dn

        if not target_key: continue

        target_month_summary = agg_summaries[target_key][fp]
        if note_num not in unique_set_for_note_val.setdefault(fp, set()):
            target_month_summary[
                "Invoice Value"] += note_val  # Using "Invoice Value" as the generic column name for value
            unique_set_for_note_val[fp].add(note_num)

        target_month_summary["Taxable Value"] += get_numeric_value(r, "Total Taxable Value")
        target_month_summary["Integrated Tax"] += get_numeric_value(r, "Integrated Tax")
        target_month_summary["Central Tax"] += get_numeric_value(r, "Central Tax")
        target_month_summary["State/UT Tax"] += get_numeric_value(r, "State/UT Tax")
        target_month_summary["Cess"] += get_numeric_value(r, "Cess")

    # IMPG Summaries
    impg_rows = combined_data.get("2B-IMPG", [])
    unique_boe_impg = {}  # For non-amended
    unique_boe_impga = {}  # For amended
    for r in impg_rows:
        fp = r.get("GSTR Filing Period")
        if not fp or fp not in all_filing_periods_sorted: continue
        boe_val = get_numeric_value(r, "Calculated Invoice Value")
        boe_num = r.get("Bill of Entry Number")
        is_amended = r.get("Amended (Yes)", "N").upper() == "Y"
        target_key = None
        unique_set_for_boe_val = None

        if is_amended:
            target_key = "2B-Summary-IMPGA(ITC_Avl)"
            unique_set_for_boe_val = unique_boe_impga
        else:
            target_key = "2B-Summary-IMPG(ITC_Avl)"
            unique_set_for_boe_val = unique_boe_impg

        target_month_summary = agg_summaries[target_key][fp]
        if boe_num not in unique_set_for_boe_val.setdefault(fp, set()):
            target_month_summary["Invoice Value"] += boe_val  # Using "Invoice Value" as the generic column name
            unique_set_for_boe_val[fp].add(boe_num)

        target_month_summary["Taxable Value"] += get_numeric_value(r, "Taxable Value")
        target_month_summary["Integrated Tax"] += get_numeric_value(r, "Integrated Tax")
        # IMPG only has IGST and Cess, CGST/SGST should be 0.
        target_month_summary["Cess"] += get_numeric_value(r, "Cess")

    # B2B ITC Rejected Summary
    b2b_rej_rows = combined_data.get("2B-B2B(ITC_Rej)", [])
    unique_inv_b2b_rej = {}
    for r in b2b_rej_rows:
        fp = r.get("GSTR Filing Period")
        if not fp or fp not in all_filing_periods_sorted: continue
        target_month_summary = agg_summaries["2B-Summary-B2B(ITC_Rej)"][fp]
        inv_val = get_numeric_value(r, "Invoice Value")
        inv_num = r.get("Invoice Number")

        if inv_num not in unique_inv_b2b_rej.setdefault(fp, set()):
            target_month_summary["Invoice Value"] += inv_val
            unique_inv_b2b_rej[fp].add(inv_num)

        target_month_summary["Taxable Value"] += get_numeric_value(r, "Total Taxable Value")
        target_month_summary["Integrated Tax"] += get_numeric_value(r, "Integrated Tax")
        target_month_summary["Central Tax"] += get_numeric_value(r, "Central Tax")
        target_month_summary["State/UT Tax"] += get_numeric_value(r, "State/UT Tax")
        target_month_summary["Cess"] += get_numeric_value(r, "Cess")

    # Finalize summary_data_output from aggregated_summaries
    for key, monthly_data_dict in agg_summaries.items():
        summary_data_output[key] = [data for month, data in sorted(monthly_data_dict.items(),
                                                                   key=lambda item: all_filing_periods_sorted.index(
                                                                       item[0])) if
                                    data["Month"] in all_filing_periods_sorted]

    section_titles_summary = {
        "2B-Summary-B2B_not RC(ITC_Avl)": "3. ITC Available - PART A1 - Supplies other than Reverse charge - B2B Invoices (IMS) - Summary",
        "2B-Summary-B2B_RC(ITC_Avl)": "3. ITC Available - PART A3 - Supplies liable for Reverse charge - B2B Invoices - Summary",
        "2B-Summary-B2BA_not RC(ITC_Avl)": "3. ITC Available - PART A1 - Supplies other than Reverse charge - B2BA Invoices (IMS) - Summary",
        "2B-Summary-B2BA_RC(ITC_Avl)": "3. ITC Available - PART A3 - Supplies liable for Reverse charge - B2BA Invoices (IMS) - Summary",
        "2B-Summary-B2BA_cum(ITC_Avl)": "3. ITC Available - PART A1&A3 - All Supplies including those liable for Reverse charge - B2BA(cum) Invoices (IMS) - Summary",
        "2B-Summary-CDNR_DN(ITC_Avl)": "3. ITC Available - PART A1 - B2B Debit Notes - Summary",
        # Debit Notes reduce ITC for supplier, increase for recipient (if original was also availed)
        "2B-Summary-CDNR_CN(ITC_Avl)": "3. ITC Available - PART B1 - B2B Credit Notes (IMS) - Summary",
        # Credit Notes increase ITC for supplier, reduce for recipient
        "2B-Summary-CDNR_RC(ITC_Avl)": "3. ITC Available - PART B1 - B2B Credit/Debit Notes (Reverse charge) - Summary",
        "2B-Summary-IMPG(ITC_Avl)": "3. ITC Available - PART A4 - Import of goods from overseas - Summary",
        "2B-Summary-IMPGA(ITC_Avl)": "3. ITC Available - PART A4 - Import of goods from overseas (Amendment) - Summary",
        "2B-Summary-B2B(ITC_Rej)": "6. ITC Rejected - PART A1 - Supplies other than Reverse charge - B2B Invoices (IMS) - Summary",
    }
    # Consistent column headers for all summary sheets
    summary_display_headers = ["Month"] + GSTR2B_SUMMARY_SHEET_TOTAL_COLUMNS

    # Column formats for summary sheets
    summary_column_formats = {"Month": "General"}
    for col in GSTR2B_SUMMARY_SHEET_TOTAL_COLUMNS:
        summary_column_formats[col] = INDIAN_FORMAT_GSTR2B

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    title_font_style = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")

    def sheet_has_valid_data(rows_list, numeric_cols_check):
        if not rows_list: return False
        for r_item in rows_list:
            for header_item in numeric_cols_check:  # Check against actual data keys
                if header_item in r_item and parse_number(r_item.get(header_item, 0)) != 0:
                    return True
        return False

    for sheet_key_name, monthly_data_list in summary_data_output.items():
        # Determine relevant numeric columns for this specific summary type
        # For IMPG, CGST and SGST are not applicable.
        current_numeric_check_cols = list(GSTR2B_SUMMARY_SHEET_TOTAL_COLUMNS)
        if "IMPG" in sheet_key_name:
            current_numeric_check_cols = [col for col in current_numeric_check_cols if
                                          col not in ["Central Tax", "State/UT Tax"]]

        if not sheet_has_valid_data(monthly_data_list, current_numeric_check_cols):
            print(
                f"[DEBUG] Skipping summary sheet {sheet_key_name} due to no valid (non-zero) data in relevant columns.")
            continue

        if sheet_key_name in wb.sheetnames: wb.remove(wb[sheet_key_name])
        ws = wb.create_sheet(sheet_key_name)
        ws.freeze_panes = "B3"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(summary_display_headers))
        title_cell = ws.cell(row=1, column=1, value=section_titles_summary.get(sheet_key_name, sheet_key_name))
        title_cell.font = title_font_style
        title_cell.alignment = center_alignment

        for col_idx, col_name_val in enumerate(summary_display_headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=col_name_val)
            cell.font = BOLD_FONT
            cell.fill = header_fill
            cell.alignment = center_alignment

        for row_idx, data_item in enumerate(monthly_data_list, start=3):
            for col_idx, col_name_val in enumerate(summary_display_headers, start=1):
                cell_value = data_item.get(col_name_val, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                if col_name_val in summary_column_formats:
                    cell.number_format = summary_column_formats[col_name_val]

        _add_total_row_to_gstr2b_summary_sheet(ws, monthly_data_list, summary_display_headers, summary_column_formats)

        for col_idx, header_val in enumerate(summary_display_headers, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for r_num in range(2, ws.max_row + 1):  # Start from header row for width calculation
                cell_val_str = str(ws.cell(row=r_num, column=col_idx).value or "")
                max_len = max(max_len, len(cell_val_str))
            ws.column_dimensions[col_letter].width = max(15, max_len + 2)  # Min width of 15

    print("[DEBUG] Finished creating summary sheets")


# ----------------------- Excel Report Generation (Unchanged) ----------------------- #
def create_excel_report(data_dict, save_path, template_path=None):
    if template_path and os.path.exists(template_path):
        wb = load_workbook(template_path)
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1 and wb.sheetnames[
            0] == "Sheet":  # Remove default sheet only if it's the only one
            wb.remove(wb["Sheet"])

    section_titles_detail = {
        "2B-B2B": "Taxable inward supplies received from registered person : 2B-B2B",
        "2B-B2BA": "Amendments to previously filed invoices by supplier : 2B-B2BA",
        "2B-B2BA(cum)": "Amendments to previously filed invoices by supplier : 2B-B2BA (Cumulative)",
        "2B-CDNR": "Debit/Credit notes(Original) : 2B-CDNR",
        "2B-IMPG": "Import of Goods : 2B-IMPG",
        "2B-B2B(ITC_Rej)": "Taxable inward supplies received from registered person : 2B-B2B (ITC Rejected)",
        "2B-B2B_sws": "Taxable inward supplies received from registered person : 2B-B2B - Sorted Supplier_wise",
        "2B-CDNR_sws": "Debit/Credit notes(Original) : 2B-CDNR - Sorted Supplier_wise"
    }

    column_headers_detail = {
        "2B-B2B": ["GSTIN/UIN of Supplier", "Trade/Legal Name", "Invoice Number", "Invoice Part", "Invoice Type",
                   "Tax Rate", "Invoice Date", "Invoice Value", "Place of Supply", "Supply Attract Reverse Charge",
                   "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess", "GSTR Period",
                   "GSTR Filing Date", "GSTR Filing Period", "ITC Availability", "Reason", "Source"],
        "2B-B2BA": ["GSTIN/UIN of Supplier", "Trade/Legal Name", "Original Invoice Number", "Original Invoice Date",
                    "Invoice Number", "Invoice Part", "Invoice Date", "Invoice Type", "Tax Rate", "Invoice Value",
                    "Place of Supply", "Supply Attract Reverse Charge", "Total Taxable Value", "Integrated Tax",
                    "Central Tax", "State/UT Tax", "Cess", "GSTR Period",
                    "GSTR Filing Date", "GSTR Filing Period", "ITC Availability", "Reason"],
        "2B-B2BA(cum)": ["GSTIN/UIN of Supplier", "Trade/Legal Name", "Total Documents", "Calculated Invoice Value",
                         "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess", "GSTR Period",
                         "GSTR Filing Date", "GSTR Filing Period"],
        "2B-CDNR": ["GSTIN/UIN of Supplier", "Trade/Legal Name", "Note Number", "Note Part", "Note Type", "Tax Rate",
                    "Note Supply Type", "Note Date", "Note Value", "Place of Supply", "Supply Attract Reverse Charge",
                    "Total Taxable Value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess", "GSTR Period",
                    "GSTR Filing Date", "GSTR Filing Period", "ITC Availability", "Reason"],
        "2B-IMPG": ["ICEGATE Reference Date", "Port Code", "Bill of Entry Number", "Bill of Entry Date",
                    "Calculated Invoice Value", "Taxable Value", "Integrated Tax", "Cess", "Record Date",
                    "GSTR Filing Period", "Amended (Yes)"],
        "2B-B2B(ITC_Rej)": ["GSTIN/UIN of Supplier", "Trade/Legal Name", "Invoice Number", "Invoice Part", "Invoice Type",
                            "Tax Rate", "Invoice Date", "Invoice Value", "Place of Supply",
                            # "Supply Attract Reverse Charge", # Typically not applicable or shown for rejected
                            "Total Taxable Value", "Integrated Tax", "Central Tax",
                            "State/UT Tax", "Cess", "GSTR Period", "GSTR Filing Date", "GSTR Filing Period", "Source"]
    }
    for key in list(column_headers_detail.keys()):  # Create _sws headers
        if key in ["2B-B2B", "2B-CDNR"]:  # Only for these two as per original logic
            column_headers_detail[key + "_sws"] = column_headers_detail[key]

    column_formats_map_detail = {
        "2B-B2B": {"Invoice Date": "dd-mm-yyyy", "Invoice Value": INDIAN_FORMAT_GSTR2B,
                   "GSTR Filing Date": "dd-mm-yyyy",
                   "Total Taxable Value": INDIAN_FORMAT_GSTR2B, "Integrated Tax": INDIAN_FORMAT_GSTR2B,
                   "Central Tax": INDIAN_FORMAT_GSTR2B, "State/UT Tax": INDIAN_FORMAT_GSTR2B,
                   "Cess": INDIAN_FORMAT_GSTR2B, "Tax Rate": "#,##0.00", "Place of Supply": "0"},
        "2B-B2BA": {"Original Invoice Date": "dd-mm-yyyy", "Invoice Date": "dd-mm-yyyy",
                    "GSTR Filing Date": "dd-mm-yyyy",
                    "Invoice Value": INDIAN_FORMAT_GSTR2B, "Total Taxable Value": INDIAN_FORMAT_GSTR2B,
                    "Integrated Tax": INDIAN_FORMAT_GSTR2B, "Central Tax": INDIAN_FORMAT_GSTR2B,
                    "State/UT Tax": INDIAN_FORMAT_GSTR2B, "Cess": INDIAN_FORMAT_GSTR2B, "Tax Rate": "#,##0.00",
                    "Place of Supply": "0"},
        "2B-B2BA(cum)": {"Total Documents": "#,##0", "Calculated Invoice Value": INDIAN_FORMAT_GSTR2B,
                         "GSTR Filing Date": "dd-mm-yyyy",
                         "Total Taxable Value": INDIAN_FORMAT_GSTR2B, "Integrated Tax": INDIAN_FORMAT_GSTR2B,
                         "Central Tax": INDIAN_FORMAT_GSTR2B, "State/UT Tax": INDIAN_FORMAT_GSTR2B,
                         "Cess": INDIAN_FORMAT_GSTR2B},
        "2B-CDNR": {"Note Date": "dd-mm-yyyy", "Note Value": INDIAN_FORMAT_GSTR2B, "GSTR Filing Date": "dd-mm-yyyy",
                    "Total Taxable Value": INDIAN_FORMAT_GSTR2B, "Integrated Tax": INDIAN_FORMAT_GSTR2B,
                    "Central Tax": INDIAN_FORMAT_GSTR2B, "State/UT Tax": INDIAN_FORMAT_GSTR2B,
                    "Cess": INDIAN_FORMAT_GSTR2B, "Tax Rate": "#,##0.00", "Place of Supply": "0"},
        "2B-IMPG": {"ICEGATE Reference Date": "dd-mm-yyyy", "Bill of Entry Date": "dd-mm-yyyy",
                    "Record Date": "dd-mm-yyyy",
                    "Calculated Invoice Value": INDIAN_FORMAT_GSTR2B,
                    "Taxable Value": INDIAN_FORMAT_GSTR2B, "Integrated Tax": INDIAN_FORMAT_GSTR2B,
                    "Cess": INDIAN_FORMAT_GSTR2B, "Port Code": "@"},  # Treat port code as text
        "2B-B2B(ITC_Rej)": {"Invoice Date": "dd-mm-yyyy", "Invoice Value": INDIAN_FORMAT_GSTR2B,
                            "GSTR Filing Date": "dd-mm-yyyy",
                            "Total Taxable Value": INDIAN_FORMAT_GSTR2B, "Integrated Tax": INDIAN_FORMAT_GSTR2B,
                            "Central Tax": INDIAN_FORMAT_GSTR2B, "State/UT Tax": INDIAN_FORMAT_GSTR2B,
                            "Cess": INDIAN_FORMAT_GSTR2B, "Tax Rate": "#,##0.00", "Place of Supply": "0"}
    }
    for key in list(column_formats_map_detail.keys()):  # Create _sws formats
        if key in ["2B-B2B", "2B-CDNR"]:
            column_formats_map_detail[key + "_sws"] = column_formats_map_detail[key]

    header_fill_style = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    title_font_main = Font(bold=True, size=12)
    center_align_style = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)  # Added wrap_text for headers

    def write_detail_sheet(sheet_name, data_rows, headers_list_for_sheet):
        # Check if data_rows is empty or if all relevant numeric fields are zero
        if not data_rows:
            print(f"[DEBUG] No data for sheet {sheet_name}. Skipping sheet generation.")
            return

        # Check for meaningful data before creating sheet
        columns_to_check_for_data = GSTR2B_DETAIL_SHEET_TOTAL_COLUMNS.get(sheet_name, [])
        has_meaningful_data = False
        if columns_to_check_for_data:  # If specific columns are defined for totals
            for r_item_check in data_rows:
                for col_check in columns_to_check_for_data:
                    if parse_number(r_item_check.get(col_check, 0)) != 0:
                        has_meaningful_data = True
                        break
                if has_meaningful_data:
                    break
        else:  # If no specific columns, assume any row means data (e.g., if totals are not applicable)
            has_meaningful_data = bool(data_rows)

        if not has_meaningful_data:
            print(f"[DEBUG] No meaningful (non-zero) data in totalable columns for sheet {sheet_name}. Skipping.")
            return

        if sheet_name in wb.sheetnames: wb.remove(wb[sheet_name])
        ws = wb.create_sheet(sheet_name)
        ws.freeze_panes = "B3"

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers_list_for_sheet))
        title_cell = ws.cell(row=1, column=1, value=section_titles_detail.get(sheet_name, sheet_name))
        title_cell.font = title_font_main
        title_cell.alignment = center_align_style

        for idx, h_val in enumerate(headers_list_for_sheet, start=1):
            c = ws.cell(row=2, column=idx, value=h_val)
            c.font = BOLD_FONT
            c.fill = header_fill_style
            c.alignment = center_align_style

        current_formats_for_sheet = column_formats_map_detail.get(sheet_name, {})
        for r_idx, r_item_data in enumerate(data_rows, start=3):
            for c_idx, h_val in enumerate(headers_list_for_sheet, start=1):
                val_to_write = r_item_data.get(h_val, "")
                cell = ws.cell(row=r_idx, column=c_idx, value=val_to_write)
                fmt_str = current_formats_for_sheet.get(h_val)
                if fmt_str:
                    # Apply date format if value is a date string that needs to be treated as date
                    if fmt_str == "dd-mm-yyyy" and isinstance(val_to_write, str):
                        try:  # Attempt to parse if it's a valid date string, otherwise keep as string
                            # This assumes dates are already in "DD-MM-YYYY" from JSON or parsed before this stage.
                            # If they are actual datetime objects, openpyxl handles them.
                            # For strings, direct format might not work unless Excel interprets them.
                            # It's better to convert to datetime objects before writing if possible.
                            # For now, just applying format.
                            pass  # Dates should ideally be datetime objects for formatting to reliably work.
                        except ValueError:
                            pass
                    cell.number_format = fmt_str

        # Add total row if applicable for this sheet type
        if sheet_name in GSTR2B_DETAIL_SHEET_TOTAL_COLUMNS:
            _add_total_row_to_gstr2b_detail_sheet(ws, sheet_name, data_rows, headers_list_for_sheet,
                                                  current_formats_for_sheet)

        for c_idx, header_val_width in enumerate(headers_list_for_sheet, start=1):
            col_letter = get_column_letter(c_idx)
            max_len = len(str(header_val_width))  # Start with header length
            for r_num in range(3, ws.max_row + 1):  # Check data rows
                cell_val_str = str(ws.cell(row=r_num, column=c_idx).value or "")
                max_len = max(max_len, len(cell_val_str))
            # Set a minimum width, e.g., 12, or a bit more than max_len
            ws.column_dimensions[col_letter].width = max(12, max_len + 2)

        print(f"[DEBUG] Created detail sheet: {sheet_name}")

    for key_name, list_of_rows in data_dict.items():
        if key_name in column_headers_detail:  # Ensure headers are defined
            write_detail_sheet(key_name, list_of_rows, column_headers_detail[key_name])

    # Supplier-wise sorted sheets
    if "2B-B2B" in data_dict and data_dict["2B-B2B"]:
        sws_b2b_data = list(data_dict["2B-B2B"])  # Make a copy to sort

        # Robust date parsing for sort key
        def b2b_sort_key_func(r_item):
            date_str = r_item.get("Invoice Date", "")
            dt_obj = datetime.datetime.min
            if date_str and isinstance(date_str, str):
                try:
                    dt_obj = datetime.datetime.strptime(date_str, "%d-%m-%Y")
                except ValueError:  # Handle other potential date formats if necessary or log warning
                    pass
            return (r_item.get("Trade/Legal Name", "").lower(), dt_obj)

        sws_b2b_data.sort(key=b2b_sort_key_func)
        if sws_b2b_data: write_detail_sheet("2B-B2B_sws", sws_b2b_data, column_headers_detail["2B-B2B_sws"])

    if "2B-CDNR" in data_dict and data_dict["2B-CDNR"]:
        sws_cdnr_data = list(data_dict["2B-CDNR"])  # Make a copy

        def cdnr_sort_key_func(r_item):
            date_str = r_item.get("Note Date", "")
            dt_obj = datetime.datetime.min
            if date_str and isinstance(date_str, str):
                try:
                    dt_obj = datetime.datetime.strptime(date_str, "%d-%m-%Y")
                except ValueError:
                    pass
            return (r_item.get("Trade/Legal Name", "").lower(), dt_obj)

        sws_cdnr_data.sort(key=cdnr_sort_key_func)
        if sws_cdnr_data: write_detail_sheet("2B-CDNR_sws", sws_cdnr_data, column_headers_detail["2B-CDNR_sws"])

    create_summary_sheets(wb, data_dict)  # Create summary sheets

    try:
        wb.save(save_path)
        print(f"[DEBUG] Workbook saved successfully to {save_path}")
        return f"✅ Successfully saved GSTR-2B Excel report: {save_path}"
    except Exception as e_save:
        print(f"[ERROR] Failed to save GSTR-2B workbook: {e_save}\n{traceback.format_exc()}")
        # Raise the error to be caught by the main processing function for UI display
        raise PermissionError(
            f"Failed to save the file. It might be open or you don't have permissions: {save_path}. Error: {e_save}")


# ----------------------- Main Processing Function ----------------------- #
def process_gstr2b(json_files, template_path, save_path):
    print("[DEBUG] Starting GSTR-2B processing...")
    combined_data = {
        "2B-B2B": [], "2B-B2BA": [], "2B-B2BA(cum)": [],
        "2B-CDNR": [], "2B-IMPG": [], "2B-B2B(ITC_Rej)": []
    }
    # json_payloads = {} # Not strictly needed if processing one by one
    unexpected_subsections_details = []  # MODIFIED: For new telemetry

    for file_path in json_files:
        basename = os.path.basename(file_path)
        print(f"[DEBUG] Processing GSTR-2B file: {basename}")
        try:
            with open(file_path, "r", encoding="utf-8") as f:
                payload = json.load(f)
        except Exception as e:
            error_msg = f"Failed to load or parse JSON from {basename}: {e}"
            print(f"[ERROR] {error_msg}")
            send_event("error", {"module": "gstr2b_processor", "file": basename, "error": error_msg,
                                 "traceback": traceback.format_exc()})
            # Decide if one bad file should stop all processing or just be skipped.
            # For now, skip this file and continue.
            unexpected_subsections_details.append({  # Log as a processing issue for this file
                "file_type": "gstr2b_json_load_error",
                "filename": basename,
                "section_path": "N/A",
                "reporting_month": "N/A",
                "snippet": f"Error loading file: {error_msg}"
            })
            continue  # Skip to the next file

        # json_payloads[basename] = payload # Store if needed for later cross-file analysis (not currently used)
        data_node_main = payload.get("data", {})

        # Get reporting period for this file (recipient's GSTR-2B period)
        # "rtnprd" is typically "MMYYYY" for GSTR-2B
        gstr2b_return_period_raw = data_node_main.get("rtnprd", "")
        current_file_filing_period_str = get_tax_period(gstr2b_return_period_raw)  # This will be "Month Name"

        # --- MODIFIED: Unexpected subsection detection ---
        for parent_key, allowed_children_set in ALLOWED_DOC_SUBSECTIONS.items():
            parent_node_content = data_node_main.get(parent_key)  # e.g., data_node_main['docdata']
            if parent_node_content and isinstance(parent_node_content, dict):
                observed_children_keys = set(parent_node_content.keys())
                unexpected_children = observed_children_keys - allowed_children_set

                for child_key in unexpected_children:
                    snippet_data = parent_node_content.get(child_key)
                    try:
                        snippet_str = json.dumps(snippet_data, indent=2)
                        # Consistent character limit with GSTR-1
                        NEW_CHARACTER_LIMIT = 10000
                        if len(snippet_str) > NEW_CHARACTER_LIMIT:
                            snippet_str = snippet_str[:NEW_CHARACTER_LIMIT] + "\n... (truncated)"
                    except TypeError:  # Fallback for non-serializable data
                        snippet_str = str(snippet_data)[:NEW_CHARACTER_LIMIT]
                        if len(str(snippet_data)) > NEW_CHARACTER_LIMIT:
                            snippet_str += "\n... (truncated)"

                    unexpected_subsections_details.append({
                        "file_type": "gstr2b_json",
                        "filename": basename,
                        "section_path": f"data.{parent_key}.{child_key}",  # More specific path
                        "reporting_month": current_file_filing_period_str,  # Month name of the GSTR-2B
                        "raw_period": gstr2b_return_period_raw,  # "MMYYYY"
                        "snippet": snippet_str
                    })
                    print(
                        f"[PROCESSOR_INFO] Unexpected subsection 'data.{parent_key}.{child_key}' found in GSTR-2B file '{basename}'.")
        # --- End of unexpected subsection detection ---

        # Standard data extraction
        # Pass current_file_filing_period_str to extractors for the "GSTR Filing Period" column
        for extractor_function in (
                extract_b2b, extract_b2ba, extract_b2ba_cum,
                extract_cdnr, extract_impg, extract_b2b_itc_rej
        ):
            try:
                # Pass current_file_filing_period_str which is the month name of the GSTR-2B itself
                section_data_dict = extractor_function(payload, current_file_filing_period_str)
                for key_from_extractor, rows_from_extractor in section_data_dict.items():
                    combined_data.setdefault(key_from_extractor, []).extend(rows_from_extractor)
            except Exception as e_extract:
                print(
                    f"[ERROR] Failed during extraction with {extractor_function.__name__} for file {basename}: {e_extract}\n{traceback.format_exc()}")
                send_event("error",
                           {"module": "gstr2b_processor", "function": extractor_function.__name__, "file": basename,
                            "error": str(e_extract), "traceback": traceback.format_exc()})

    # Sort data within each section (e.g., by date)
    # This part assumes dates are strings; robust parsing is good.
    def parse_date_robustly(date_val_str, date_fmt="%d-%m-%Y"):
        if not date_val_str or not isinstance(date_val_str, str): return datetime.datetime.min
        try:
            return datetime.datetime.strptime(date_val_str, date_fmt)
        except ValueError:  # Add more formats if needed or handle dd/mm/yyyy etc.
            try:  # Try another common format
                return datetime.datetime.strptime(date_val_str, "%Y-%m-%d")
            except ValueError:
                return datetime.datetime.min  # Fallback for unparseable dates

    date_field_mapping = {
        "2B-B2B": "Invoice Date", "2B-B2BA": "Invoice Date", "2B-B2B(ITC_Rej)": "Invoice Date",
        "2B-CDNR": "Note Date", "2B-IMPG": "Bill of Entry Date",  # Assuming YYYYMMDD or parseable
        "2B-B2BA(cum)": "GSTR Filing Date"  # This might be DD-MM-YYYY
    }

    for data_key, sort_date_field in date_field_mapping.items():
        if data_key in combined_data and combined_data[data_key]:  # Check if key exists and has data
            # For B2BA(cum), "GSTR Filing Date" is the primary date-like field.
            date_format_to_use = "%d-%m-%Y"  # Default
            if data_key == "2B-IMPG" and sort_date_field == "Bill of Entry Date":  # BOE date might be YYYYMMDD
                pass  # Keep default, parse_date_robustly will try common formats.

            combined_data[data_key].sort(
                key=lambda r_sort: parse_date_robustly(r_sort.get(sort_date_field, ""), date_format_to_use))

    report_generation_msg = create_excel_report(combined_data, save_path,
                                                template_path)  # This can raise PermissionError

    # --- MODIFIED: Telemetry and final message construction ---
    base_message_for_ui = report_generation_msg  # "✅ Successfully saved..."

    completion_payload = {
        "output_file": save_path,
        "file_count": len(json_files),
        "template_used": bool(template_path),
        "status": "success"
        # "message" will be built based on unexpected_subsections_details for telemetry
    }

    if unexpected_subsections_details:
        completion_payload["unexpected_sections_found"] = True
        completion_payload["unexpected_section_details"] = unexpected_subsections_details  # List of dicts with snippets

        manager_alert_parts = ["Unexpected subsections encountered in GSTR-2B processing:"]
        for detail in unexpected_subsections_details:
            manager_alert_parts.append(
                f"File: '{detail['filename']}', Path: '{detail['section_path']}', Period: {detail.get('reporting_month', detail.get('raw_period', 'N/A'))}."
            )
        # For telemetry, the full detail is in unexpected_section_details.
        # For a readable message, we just summarize.
        completion_payload["manager_alert_summary_unexpected_sections"] = " ".join(manager_alert_parts)
        completion_payload[
            "message"] = f"Processing completed with warnings about unexpected subsections. Details: {completion_payload['manager_alert_summary_unexpected_sections']}"

        # The UI will get the base_message_for_ui and the unexpected_subsections_details separately.
        # The processor's log/telemetry message can be more verbose.
    else:
        completion_payload["message"] = "Processing completed successfully without any unexpected subsections found."

    send_event("gstr2b_complete", completion_payload)
    print(f"[DEBUG] GSTR-2B processing telemetry: {completion_payload['message']}")

    return base_message_for_ui, unexpected_subsections_details  # MODIFIED return