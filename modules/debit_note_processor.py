import tkinter as tk  # Kept for messagebox, consider removing if UI fully decouples
from tkinter import messagebox  # Used for error popups directly in processor
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import os
import logging
import math

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Constants ---
INDIAN_NUMBER_FORMAT = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"

# Standardized Fixed Headers for Credit/Debit Notes (Shared with credit_note_processor)
NOTE_FIXED_HEADERS = [
    'GSTIN/UIN of Recipient',  # For Debit Notes, this would be Supplier's GSTIN
    'Receiver Name',  # For Debit Notes, this would be Supplier's Name
    'Branch',
    'Note Number',
    'Note Date',
    'Note Ref. No.',
    'Note Ref. Date',
    'Note Value',
    'Taxable Value',
    'Integrated Tax',
    'Central Tax',
    'State/UT Tax',
    'Cess',
    'Round Off'
]

# Headers to exclude from the 'Total' row calculation (Shared)
NOTE_EXCLUDE_FROM_TOTAL_HEADERS = [
    'GSTIN/UIN of Recipient',
    'Receiver Name',
    'Branch',
    'Note Number',
    'Note Date',
    'Note Ref. No.',
    'Note Ref. Date',
]

# Headers that are strictly textual (Shared)
NOTE_STRICTLY_TEXTUAL_HEADERS = [
    'GSTIN/UIN of Recipient',
    'Receiver Name',
    'Branch',
    'Note Number',
    'Note Ref. No.'
]

# Headers that are numeric but excluded from Cr/Dr suffix logic (Shared)
NOTE_NUMERIC_NO_CRDR_SUFFIX_CHECK = [
    'Note Value'
]

# Sheet titles for DEBIT NOTE processing
DEBIT_SECTION_TITLES = [
    ("DEBIT-Total", "Debit Note Register - Total"),
    ("DEBIT-Total_sws", "Debit Note Register - Total - Supplier wise"),  # Changed from Receiver to Supplier
    ("DEBIT-Summary-Total", "Debit Note Register - Total - Summary"),
]


# --- Helper Functions (Identical to credit_note_processor, can be shared if in a utils module) ---
def find_header_row(worksheet):
    logging.debug(f"Searching for header row in sheet: {worksheet.title}")
    for row in worksheet.iter_rows():
        first_cell_val = row[0].value
        if isinstance(first_cell_val, str) and "date" in first_cell_val.lower():
            logging.debug(f"Header row potentially found at Excel row: {row[0].row}")
            return row[0].row
    logging.warning(f"Header row containing 'Date' not found in sheet: {worksheet.title}")
    return None


def create_or_replace_sheet(wb, sheet_name, title_text, columns):
    logging.info(f"Creating or replacing sheet: {sheet_name}")
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1)
    cell.value = title_text
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, col in enumerate(columns, start=1):
        header_cell = ws.cell(row=2, column=idx, value=col)
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B3"
    return ws


def apply_format_and_autofit(ws, columns, start_row=3, col_format_map=None):
    logging.debug(f"Applying format and autofit for sheet: {ws.title}")
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        if col_format_map and col_name in col_format_map:
            for row_num in range(start_row, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = col_format_map[col_name]
        for row_num in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = max(15, max_len + 2)
    logging.debug(f"Finished applying format and autofit for sheet: {ws.title}")


def add_total_row(ws, columns, start_row, end_row):
    logging.debug(f"Adding total row for sheet: {ws.title}")
    total_row_data = ['Total'] + [''] * (len(columns) - 1)
    for col_idx in range(1, len(columns) + 1):
        if col_idx == 1: continue
        header_name = columns[col_idx - 1]
        if header_name in NOTE_EXCLUDE_FROM_TOTAL_HEADERS:
            continue
        total = 0
        has_numeric_data = False
        try:
            for row in range(start_row, end_row + 1):
                value = ws.cell(row=row, column=col_idx).value
                if isinstance(value, (int, float)) and not (math.isnan(value) or math.isinf(value)):
                    total += float(value)
                    has_numeric_data = True
            if has_numeric_data:
                total_row_data[col_idx - 1] = total
        except Exception as e:
            total_row_data[col_idx - 1] = "Error"
            logging.error(f"Error totaling column '{header_name}': {e}")
    row_num = end_row + 1
    ws.append(total_row_data)
    for col_idx, value in enumerate(total_row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = Font(bold=True, color="FF0000")
        if isinstance(value, (int, float)):
            cell.number_format = INDIAN_NUMBER_FORMAT
    logging.debug(f"Finished adding total row for sheet: {ws.title}")


def safe_note_value_conversion(value, header_standard_name, cell_obj=None, note_type="debit"):
    if header_standard_name in NOTE_STRICTLY_TEXTUAL_HEADERS:
        return str(value).strip() if value is not None else ''
    if value is None: return 0.0

    if header_standard_name in NOTE_NUMERIC_NO_CRDR_SUFFIX_CHECK:  # e.g. 'Note Value'
        try:
            numeric_val = float(value)
            return -abs(numeric_val) if note_type == "credit" else abs(numeric_val)
        except (ValueError, TypeError):
            logging.warning(f"Could not convert '{value}' to float for '{header_standard_name}'. Returning 0.0.")
            return 0.0

    numeric_value = 0.0
    is_string_value = isinstance(value, str)
    value_str_stripped = str(value).strip() if is_string_value else ""

    if not value_str_stripped and not isinstance(value, (int, float)): return 0.0

    try:
        numeric_value = float(value)
    except (ValueError, TypeError):
        if is_string_value:
            if value_str_stripped.endswith(' Cr'):
                try:
                    numeric_value = float(value_str_stripped[:-3])
                except ValueError:
                    return value_str_stripped  # Return original if "123 Cr" parse fails
            elif value_str_stripped.endswith(' Dr'):
                try:
                    numeric_value = float(value_str_stripped[:-3])
                except ValueError:
                    return value_str_stripped  # Return original if "123 Dr" parse fails
            else:  # Not Cr/Dr, but failed direct float
                return value_str_stripped  # Return as text
        else:  # Not a string, not float/int directly
            return str(value)  # Return string representation

    # Apply primary sign logic based on note_type for general numeric fields
    if header_standard_name not in ['Round Off'] + NOTE_STRICTLY_TEXTUAL_HEADERS + NOTE_NUMERIC_NO_CRDR_SUFFIX_CHECK:
        if note_type == "credit":
            numeric_value = -abs(numeric_value)
        elif note_type == "debit":
            numeric_value = abs(numeric_value)  # Ensure positive for debit note main components

    # Cell format based sign adjustment (primarily for 'Round Off' or if suffixes weren't present)
    if cell_obj and cell_obj.number_format and (
            header_standard_name == 'Round Off' or not value_str_stripped.endswith((' Cr', ' Dr'))):
        format_str = str(cell_obj.number_format)
        if 'Dr' in format_str and numeric_value < 0:
            numeric_value = abs(numeric_value)  # If format is Dr and value is negative, make it positive
        elif 'Cr' in format_str and numeric_value > 0:
            # Only make it negative if it's a credit note,
            # OR if it's a debit note AND the header is 'Round Off' (where Cr might mean negative round-off)
            if note_type == "credit":
                numeric_value = -abs(numeric_value)
            elif note_type == "debit" and header_standard_name == 'Round Off':
                numeric_value = -abs(numeric_value)
            # For other debit note components (Taxable Value, Taxes etc.), if numeric_value is already positive,
            # a 'Cr' in cell format should not flip its sign to negative. It remains positive.
    return numeric_value


# --- Main Processing Function ---
def process_debit_note_data(input_files, template_file=None, existing_wb=None):
    logging.info("Starting debit note data processing")
    all_debit_note_data = []

    source_key_to_standard_header = {
        'date': 'Note Date',
        'particulars': 'Receiver Name',  # For Debit Note, this is Supplier
        'voucher type': 'Voucher Type',  # Not in fixed, but can be used for logic if needed
        'voucher no': 'Note Number',
        'voucher no.': 'Note Number',
        'gstin/uin': 'GSTIN/UIN of Recipient',  # Supplier's GSTIN
        'gross total': 'Note Value',
        'value': 'Taxable Value',
        'taxable amount': 'Taxable Value',
        'igst': 'Integrated Tax',
        'cgst': 'Central Tax',
        'sgst': 'State/UT Tax',
        'state/ut tax': 'State/UT Tax',
        'cess': 'Cess',
        'round off': 'Round Off',
        'voucher ref. no.': 'Note Ref. No.',
        'voucher ref no.': 'Note Ref. No.',
        'voucher ref. no': 'Note Ref. No.',
        'voucher ref no': 'Note Ref. No.',
        'voucher ref. date': 'Note Ref. Date',
        'voucher ref date': 'Note Ref. Date',
        'supplier invoice no.': 'Note Ref. No.',  # Common ref for DN
        'supplier invoice no': 'Note Ref. No.',
        'supplier invoice date': 'Note Ref. Date',
    }
    note_fixed_headers_lower = {h.lower() for h in NOTE_FIXED_HEADERS}
    extra_headers_set = set()

    logging.debug("Starting file processing loop for debit note data")
    for filepath, branch_key in input_files:
        logging.info(f"Processing debit note file: {filepath} with branch key: {branch_key}")
        try:
            wb = load_workbook(filepath, data_only=True)
            sheet_to_process_name = "Purchase Register"
            ws = None
            if sheet_to_process_name in wb.sheetnames:
                ws = wb[sheet_to_process_name]
            elif "Debit Note Register" in wb.sheetnames:
                ws = wb["Debit Note Register"]
                logging.info(f"Found and using 'Debit Note Register' sheet in {filepath}")
            else:
                logging.warning(f"'{sheet_to_process_name}' or 'Debit Note Register' not found. Trying alternatives.")
                possible_sheets = [s_name for s_name in wb.sheetnames if
                                   "register" in s_name.lower() or "debit" in s_name.lower()]
                if possible_sheets:
                    ws = wb[possible_sheets[0]]
                    logging.info(f"Using first suitable sheet found: {ws.title}")
                else:
                    ws = wb.active
                    logging.warning(f"No specific sheet found. Falling back to active sheet: {ws.title}")
            if not ws:
                logging.error(f"No suitable sheet in {filepath}")
                messagebox.showerror("Sheet Error", f"Could not find a suitable sheet in {os.path.basename(filepath)}.")
                continue

            logging.info(f"Processing sheet: {ws.title} from file {filepath}")
            header_row_num = find_header_row(ws)
            if not header_row_num:
                logging.error(f"Header row not found in {filepath} (sheet: {ws.title})")
                messagebox.showerror("Header Error",
                                     f"Header row not found in {os.path.basename(filepath)} (sheet: {ws.title}).")
                continue

            original_headers_from_sheet = [cell.value for cell in ws[header_row_num] if cell.value is not None]
            original_headers_lower_map = {str(h).lower(): h for h in original_headers_from_sheet if h}

            for original_header in original_headers_from_sheet:
                if original_header is None: continue
                header_l = str(original_header).lower()
                is_mapped = header_l in source_key_to_standard_header
                is_fixed = False
                if is_mapped:
                    standard_h = source_key_to_standard_header[header_l]
                    if standard_h.lower() in note_fixed_headers_lower: is_fixed = True
                elif header_l in note_fixed_headers_lower:
                    is_fixed = True
                if not is_mapped and not is_fixed:
                    extra_headers_set.add(original_header)

            logging.debug(f"Original headers from '{filepath}': {original_headers_from_sheet}")
            logging.debug(f"Extra headers identified so far: {extra_headers_set}")

            for row_idx, row_cells_tuple in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=False),
                                                      start=header_row_num + 1):
                row_data_orig_case = {}
                current_row_cells_map = {}
                for i, cell_obj in enumerate(row_cells_tuple):
                    if i < len(original_headers_from_sheet):
                        header_val_sheet = original_headers_from_sheet[i]
                        row_data_orig_case[header_val_sheet] = cell_obj.value
                        current_row_cells_map[header_val_sheet] = cell_obj

                raw_note_date_value = row_data_orig_case.get(original_headers_lower_map.get('date'))
                particulars_value = str(
                    row_data_orig_case.get(original_headers_lower_map.get('particulars', ''), '')).lower()
                if not raw_note_date_value or 'grand total' in particulars_value:
                    continue

                processed_row_data = {'Branch': branch_key}
                try:  # Note Date
                    if isinstance(raw_note_date_value, str):
                        try:
                            processed_row_data['Note Date'] = datetime.datetime.strptime(raw_note_date_value,
                                                                                         '%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            processed_row_data['Note Date'] = datetime.datetime.strptime(raw_note_date_value.split()[0],
                                                                                         '%d-%m-%Y')
                    elif isinstance(raw_note_date_value, datetime.datetime):
                        processed_row_data['Note Date'] = raw_note_date_value
                    elif isinstance(raw_note_date_value, (int, float)):
                        processed_row_data['Note Date'] = datetime.datetime.fromordinal(
                            datetime.datetime(1900, 1, 1).toordinal() + int(raw_note_date_value) - 2)
                    else:
                        raise ValueError("Invalid date type")
                except Exception as e_date:
                    logging.debug(
                        f"Row {row_idx}: Skipping due to 'Note Date' parsing error: {raw_note_date_value}, Error: {e_date}")
                    continue

                raw_note_ref_date_value = row_data_orig_case.get(original_headers_lower_map.get('voucher ref. date'))
                if not raw_note_ref_date_value:
                    raw_note_ref_date_value = row_data_orig_case.get(
                        original_headers_lower_map.get('supplier invoice date'))

                if raw_note_ref_date_value:
                    try:
                        if isinstance(raw_note_ref_date_value, str):
                            try:
                                processed_row_data['Note Ref. Date'] = datetime.datetime.strptime(
                                    raw_note_ref_date_value, '%Y-%m-%d %H:%M:%S')
                            except ValueError:
                                processed_row_data['Note Ref. Date'] = datetime.datetime.strptime(
                                    raw_note_ref_date_value.split()[0], '%d-%m-%Y')
                        elif isinstance(raw_note_ref_date_value, datetime.datetime):
                            processed_row_data['Note Ref. Date'] = raw_note_ref_date_value
                        elif isinstance(raw_note_ref_date_value, (int, float)):
                            processed_row_data['Note Ref. Date'] = datetime.datetime.fromordinal(
                                datetime.datetime(1900, 1, 1).toordinal() + int(raw_note_ref_date_value) - 2)
                        else:
                            processed_row_data['Note Ref. Date'] = str(raw_note_ref_date_value)
                    except Exception as e_ref_date:
                        logging.warning(
                            f"Row {row_idx}: Could not parse 'Note Ref. Date': {raw_note_ref_date_value}, Error: {e_ref_date}. Storing as string.")
                        processed_row_data['Note Ref. Date'] = str(raw_note_ref_date_value)
                else:
                    processed_row_data['Note Ref. Date'] = ''

                for original_header, original_value in row_data_orig_case.items():
                    if original_header is None: continue
                    header_lower = str(original_header).lower()
                    standard_header = source_key_to_standard_header.get(header_lower)

                    if standard_header:
                        if standard_header not in ['Note Date', 'Note Ref. Date']:
                            cell_obj = current_row_cells_map.get(original_header)
                            processed_row_data[standard_header] = safe_note_value_conversion(original_value,
                                                                                             standard_header, cell_obj,
                                                                                             "debit")
                    elif original_header in extra_headers_set:
                        cell_obj = current_row_cells_map.get(original_header)
                        processed_row_data[original_header] = safe_note_value_conversion(original_value,
                                                                                         original_header, cell_obj,
                                                                                         "debit")

                for fixed_h in NOTE_FIXED_HEADERS:
                    if fixed_h not in processed_row_data:
                        processed_row_data[fixed_h] = '' if fixed_h in NOTE_STRICTLY_TEXTUAL_HEADERS or fixed_h in [
                            'Note Date', 'Note Ref. Date'] else 0.0
                for extra_h in extra_headers_set:
                    if extra_h not in processed_row_data:
                        processed_row_data[extra_h] = safe_note_value_conversion(None, extra_h, None, "debit")
                all_debit_note_data.append(processed_row_data)
        except Exception as e_file:
            logging.error(f"Error processing file {filepath}: {e_file}", exc_info=True)
            messagebox.showerror("File Processing Error",
                                 f"Error processing file {os.path.basename(filepath)}: {e_file}")
        logging.info(f"Finished processing debit note file: {filepath}")

    logging.debug("Finished file processing loop for debit notes")
    final_headers_list = NOTE_FIXED_HEADERS + sorted(list(extra_headers_set))
    logging.info(f"Final headers for debit note detail sheets: {final_headers_list}")

    sort_key_notes = lambda x: (x.get('Note Date', datetime.datetime.min) if isinstance(x.get('Note Date'),
                                                                                        datetime.datetime) else datetime.datetime.min,
                                str(x.get('Note Number', '')))
    all_debit_note_data.sort(key=sort_key_notes)

    def sort_key_supplier_wise(row):
        supplier_name = str(row.get('Receiver Name', '')).strip().lower()
        date_val = row.get('Note Date', datetime.datetime.min)
        note_num_val = str(row.get('Note Number', ''))
        if supplier_name in ['cash', '(cancelled )'] or not supplier_name: return (
        1, supplier_name, date_val, note_num_val)
        return (0, supplier_name, date_val, note_num_val)

    all_debit_note_data_sws = sorted(all_debit_note_data, key=sort_key_supplier_wise)

    if existing_wb is not None:
        output_wb = existing_wb
    elif template_file:
        output_wb = load_workbook(template_file)
    else:
        output_wb = Workbook()
        if 'Sheet' in output_wb.sheetnames and len(output_wb.sheetnames) == 1: del output_wb['Sheet']

    detail_col_format_map = {h: INDIAN_NUMBER_FORMAT for h in NOTE_FIXED_HEADERS if
                             h not in NOTE_STRICTLY_TEXTUAL_HEADERS + ['Note Date', 'Note Ref. Date']}
    detail_col_format_map['Note Date'] = 'DD-MM-YYYY'
    detail_col_format_map['Note Ref. Date'] = 'DD-MM-YYYY'
    for eh in extra_headers_set:
        if eh not in detail_col_format_map: detail_col_format_map[eh] = INDIAN_NUMBER_FORMAT

    detail_sheets_to_create = [
        ("DEBIT-Total", all_debit_note_data),
        ("DEBIT-Total_sws", all_debit_note_data_sws),
    ]
    for sheet_key, data_list in detail_sheets_to_create:
        if not data_list: logging.info(f"No data for detail sheet: {sheet_key}, skipping."); continue
        display_title = next((t for n, t in DEBIT_SECTION_TITLES if n == sheet_key), sheet_key)
        ws = create_or_replace_sheet(output_wb, sheet_key, display_title, final_headers_list)
        data_start_row = 3
        for row_item in data_list:
            row_values = [row_item.get(h, '') if not (
                        h in ['Note Date', 'Note Ref. Date'] and isinstance(row_item.get(h), datetime.datetime)) \
                              else row_item.get(h).strftime('%d-%m-%Y') \
                          for h in final_headers_list]
            ws.append(row_values)
        add_total_row(ws, final_headers_list, data_start_row, ws.max_row)
        apply_format_and_autofit(ws, final_headers_list, col_format_map=detail_col_format_map, start_row=data_start_row)

    summary_headers = ['Month', 'No. of Notes', 'Note Value', 'Taxable Value', 'Integrated Tax', 'Central Tax',
                       'State/UT Tax', 'Cess']
    summary_col_format_map = {h: INDIAN_NUMBER_FORMAT for h in summary_headers if h not in ['Month', 'No. of Notes']}
    months_order = ['April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December', 'January',
                    'February', 'March']
    summary_sheets_to_create = [("DEBIT-Summary-Total", all_debit_note_data)]

    for sheet_key, data_list_summary in summary_sheets_to_create:
        if not data_list_summary: logging.info(f"No data for summary sheet: {sheet_key}, skipping."); continue
        display_title_summary = next((t for n, t in DEBIT_SECTION_TITLES if n == sheet_key), sheet_key)
        ws_summary = create_or_replace_sheet(output_wb, sheet_key, display_title_summary, summary_headers)
        monthly_summary_aggr = {}
        unique_notes_by_month = {}
        for row_item_sum in data_list_summary:
            date_obj_sum = row_item_sum.get('Note Date')
            note_num_sum = str(row_item_sum.get('Note Number', '')).strip()
            if not isinstance(date_obj_sum, datetime.datetime) or not note_num_sum: continue
            month_name_sum = months_order[(date_obj_sum.month - 4 + 12) % 12]
            if month_name_sum not in monthly_summary_aggr:
                monthly_summary_aggr[month_name_sum] = {h: 0.0 for h in summary_headers if
                                                        h not in ['Month', 'No. of Notes']}
                monthly_summary_aggr[month_name_sum]['count'] = 0
                unique_notes_by_month[month_name_sum] = set()
            if note_num_sum not in unique_notes_by_month[month_name_sum]:
                monthly_summary_aggr[month_name_sum]['count'] += 1
                unique_notes_by_month[month_name_sum].add(note_num_sum)

            numeric_summary_fields = [h for h in summary_headers if h not in ['Month', 'No. of Notes']]
            for field_to_sum in numeric_summary_fields:
                val_add = row_item_sum.get(field_to_sum, 0.0)
                if isinstance(val_add, (int, float)) and not (math.isnan(val_add) or math.isinf(val_add)):
                    monthly_summary_aggr[month_name_sum][field_to_sum] += float(val_add)

        summary_data_start_row = 3;
        rows_added_sum = 0
        for month_iter in months_order:
            if month_iter in monthly_summary_aggr:
                month_data_iter = monthly_summary_aggr[month_iter]
                row_to_append_sum = [month_iter, month_data_iter.get('count', 0)] + \
                                    [month_data_iter.get(h, 0.0) for h in summary_headers if
                                     h not in ['Month', 'No. of Notes']]
                ws_summary.append(row_to_append_sum)
                rows_added_sum += 1
        if rows_added_sum > 0:
            total_row_sum_vals = ['Total', sum(monthly_summary_aggr[m].get('count', 0) for m in monthly_summary_aggr)]
            for h_idx_sum in range(2, len(summary_headers)):
                col_h_sum = summary_headers[h_idx_sum]
                total_row_sum_vals.append(
                    sum(monthly_summary_aggr[m].get(col_h_sum, 0.0) for m in monthly_summary_aggr))
            ws_summary.append(total_row_sum_vals)
            total_row_num_on_sheet = summary_data_start_row + rows_added_sum
            for c_idx_sum, val_iter in enumerate(total_row_sum_vals, 1):
                cell_sum_total = ws_summary.cell(row=total_row_num_on_sheet, column=c_idx_sum)
                cell_sum_total.font = Font(bold=True, color="FF0000")
                if isinstance(val_iter, (int, float)) and summary_headers[c_idx_sum - 1] not in ['Month',
                                                                                                 'No. of Notes']:
                    cell_sum_total.number_format = INDIAN_NUMBER_FORMAT
        apply_format_and_autofit(ws_summary, summary_headers, col_format_map=summary_col_format_map,
                                 start_row=summary_data_start_row)

    logging.info("Completed debit note data processing")
    return output_wb
