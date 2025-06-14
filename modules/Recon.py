import re
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment  # MODIFICATION: Added Alignment
from openpyxl.utils import get_column_letter
import logging
import os
import traceback

# Attempt to import telemetry
try:
    from utils.telemetry import send_event

    logging.info(f"{os.path.basename(__file__)}: Successfully imported send_event from utils.telemetry")
except ImportError as e_import:
    logging.warning(
        f"{os.path.basename(__file__)}: Could not import send_event from utils.telemetry (error: {e_import}). Telemetry will be disabled for this module.")
    print(f"[WARN] {os.path.basename(__file__)}: Failed to import telemetry from utils.telemetry. Error: {e_import}")


    def send_event(event_name, payload):  # Dummy function
        pass

# -----------------------------------------------------------------------------
# Constants & Styles
# -----------------------------------------------------------------------------
EPSILON = 0.01
MISSING_FILL = PatternFill(start_color='FFC7CE', fill_type='solid')  # light red
DIFF_FILL = PatternFill(start_color='FFEB9C', fill_type='solid')  # light orange
NO_FILL = PatternFill()  # clear
INDIAN_NUM_FORMAT = '#,##,##0.00'  # NEW: Indian number format


# -----------------------------------------------------------------------------
# Helper Functions
# -----------------------------------------------------------------------------
def normalize_key(x):
    """Strip non‐alphanumeric, lowercase."""
    return re.sub(r'[^A-Za-z0-9]', '', str(x) or '').lower()


def get_numeric(sheet, r, c):
    """Safely return float value or 0.0 if cell is None or not a number."""
    if r is None or c is None or sheet is None:
        return 0.0
    v = sheet.cell(row=r, column=c).value
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def get_header_col(sheet, header_name, sheet_title_for_warning=""):
    """Case‐insensitive lookup of header in row 2. Returns None if not found."""
    if not sheet:
        return None
    h = header_name.strip().lower()
    if sheet.max_row >= 2:
        for cell in sheet[2]:  # Assuming headers are in row 2
            if cell.value and str(cell.value).strip().lower() == h:
                return cell.column
    return None


def cleanup_sheet(sheet):
    """Remove existing '* Diff' columns and clear fills from row 3 onwards."""
    if not sheet:
        return

    to_del = []
    if sheet.max_row >= 2:
        for cell in sheet[2]:
            if isinstance(cell.value, str) and cell.value.strip().lower().endswith(' diff'):
                to_del.append(cell.column)

    for col_idx in sorted(to_del, reverse=True):
        sheet.delete_cols(col_idx)

    for row_idx in range(3, sheet.max_row + 1):
        for col_idx in range(1, sheet.max_column + 1):
            if sheet.max_row >= row_idx and sheet.max_column >= col_idx:
                current_cell = sheet.cell(row=row_idx, column=col_idx)
                if current_cell.value is not None or \
                        (current_cell.has_style and current_cell.fill != NO_FILL):
                    current_cell.fill = NO_FILL


# NEW: Autofit columns function
def autofit_sheet_columns(sheet, start_row=2):
    """
    Autofits columns in the given sheet.
    Skips rows before start_row (typically row 1 is a main title).
    Considers header in row 2 for width calculation.
    """
    if not sheet:
        return
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        # Check header in row 2 for this column
        # (Assuming row 1 is title, row 2 is actual headers)
        header_cell_value_row2 = sheet.cell(row=2, column=column_cells[0].column).value
        if header_cell_value_row2:
            max_length = max(max_length, len(str(header_cell_value_row2)))

        # Iterate from start_row (default 2) onwards for cell content
        for cell in column_cells:
            if cell.row < start_row:
                continue
            try:
                if cell.value is not None:
                    cell_value_str = str(cell.value)
                    # A simple heuristic for length, could be expanded for formatted numbers
                    current_len = len(cell_value_str)
                    if cell.number_format and "0.00" in cell.number_format and isinstance(cell.value, (int, float)):
                        # For formatted numbers, estimate width based on formatted string
                        # This is a simplification; true formatted width is complex
                        formatted_val_str = f"{cell.value:{INDIAN_NUM_FORMAT.replace('₹ ', '')}}"  # Use a generic format for length
                        current_len = len(formatted_val_str)
                    max_length = max(max_length, current_len)
            except:
                pass

        if max_length > 0:
            # Add a little padding, ensure a minimum width if only header exists
            adjusted_width = max(max_length + 2, len(str(header_cell_value_row2)) + 2 if header_cell_value_row2 else 8)
            sheet.column_dimensions[column_letter].width = adjusted_width
        elif header_cell_value_row2:  # If only header has content
            sheet.column_dimensions[column_letter].width = len(str(header_cell_value_row2)) + 2


def prepare_sale_sheets_infrastructure(sale_sheets_list, numeric_headers, inv_header, gstin_header,
                                       modified_sheets_for_autofit):  # MODIFICATION: Added modified_sheets_for_autofit
    sale_data_indices = {}
    sale_diff_indices = {}
    sale_map = {}

    if not sale_sheets_list:
        return sale_data_indices, sale_diff_indices, sale_map

    for sht in sale_sheets_list:
        if not sht: continue
        title = sht.title
        sale_data_indices[title] = {}
        sale_diff_indices[title] = {}
        modified_sheets_for_autofit.add(sht)  # MODIFICATION: Track sheet

        cols_to_insert_info = []
        for header_name in numeric_headers:
            col_idx = get_header_col(sht, header_name, title)
            if col_idx:
                cols_to_insert_info.append((col_idx, header_name, False))  # field_key, is_gstin_diff
            else:
                messagebox.showwarning("Missing Header",
                                       f"Header '{header_name}' not found in SALE sheet '{title}'. Diff column for it will be skipped.")
                send_event("recon_warning", {"type": "missing_header", "sheet": title, "header": header_name,
                                             "context": "sale_sheets_infrastructure"})

        gstin_col_idx = get_header_col(sht, gstin_header, title)
        if gstin_col_idx:
            cols_to_insert_info.append((gstin_col_idx, 'gstin', True))  # field_key, is_gstin_diff
        else:
            messagebox.showwarning("Missing Header",
                                   f"Header '{gstin_header}' (for GSTIN) not found in SALE sheet '{title}'. GSTIN/UIN Diff column will be skipped.")
            send_event("recon_warning", {"type": "missing_header", "sheet": title, "header": gstin_header,
                                         "context": "sale_sheets_infrastructure_gstin"})

        cols_to_insert_info.sort(key=lambda item: item[0], reverse=True)

        for original_col_idx, field_key, is_gstin_diff_col in cols_to_insert_info:
            sht.insert_cols(original_col_idx + 1)
            diff_column_label = 'GSTIN/UIN Diff' if is_gstin_diff_col else f"{field_key} Diff"
            header_cell = sht.cell(row=2, column=original_col_idx + 1)
            header_cell.value = diff_column_label
            header_cell.font = Font(bold=True)

        sale_data_indices[title]['inv'] = get_header_col(sht, inv_header, title)
        sale_data_indices[title]['gstin'] = get_header_col(sht, gstin_header, title)
        sale_diff_indices[title]['gstin'] = get_header_col(sht, 'GSTIN/UIN Diff',
                                                           title)  # This should find the newly added one

        for f_hdr in numeric_headers:
            sale_data_indices[title][f_hdr] = get_header_col(sht, f_hdr, title)
            sale_diff_indices[title][f_hdr] = get_header_col(sht, f"{f_hdr} Diff", title)

        inv_col_idx_map = sale_data_indices[title].get('inv')
        if inv_col_idx_map:
            for r_idx in range(3, sht.max_row + 1):
                raw_invoice_val = sht.cell(row=r_idx, column=inv_col_idx_map).value
                if not raw_invoice_val or str(raw_invoice_val).strip().lower() == 'total':
                    continue
                norm_key = normalize_key(raw_invoice_val)
                if norm_key not in sale_map:
                    sale_map[norm_key] = (sht, r_idx)
        else:
            messagebox.showwarning("Critical Header Missing",
                                   f"Invoice header '{inv_header}' not found in '{title}'. This sheet cannot be mapped for SALE reconciliation.")
            send_event("recon_error", {"type": "critical_header_missing", "sheet": title, "header": inv_header,
                                       "context": "sale_sheets_infrastructure_invoice"})

    return sale_data_indices, sale_diff_indices, sale_map


def add_totals_to_sheet(sheet_obj, data_indices_dict, diff_indices_dict, key_header_name_in_data_indices,
                        id_field_name_in_diff_indices='gstin',
                        modified_sheets_for_autofit=None):  # MODIFICATION: Added modified_sheets_for_autofit
    if not sheet_obj or not data_indices_dict or not diff_indices_dict:
        return

    if modified_sheets_for_autofit is not None:  # MODIFICATION
        modified_sheets_for_autofit.add(sheet_obj)

    id_col_for_total_label = data_indices_dict.get(key_header_name_in_data_indices)
    if not id_col_for_total_label:  # Fallback if key header (like invoice number) is missing
        # Try to find the first column that is not a diff column for the "Total" label
        first_data_col = None
        for col_idx_loop in range(1, sheet_obj.max_column + 1):
            header_val = sheet_obj.cell(row=2, column=col_idx_loop).value
            if header_val and not str(header_val).endswith(" Diff"):
                first_data_col = col_idx_loop
                break
        id_col_for_total_label = first_data_col if first_data_col else 1

    data_rows_for_sum = [
        r for r in range(3, sheet_obj.max_row + 1)
        if sheet_obj.cell(row=r, column=id_col_for_total_label).value and \
           str(sheet_obj.cell(row=r, column=id_col_for_total_label).value).strip().lower() != 'total'
    ]
    last_data_r = 2
    total_r = 3
    if data_rows_for_sum:
        last_data_r = max(data_rows_for_sum)
        total_r = last_data_r + 1

    if total_r > sheet_obj.max_row + 1 and sheet_obj.max_row >= 2:  # If total row would be beyond current max_row + 1
        total_r = sheet_obj.max_row + 1
    elif sheet_obj.max_row < 2:  # Handles empty or header-only sheet
        total_r = 3

    total_r = max(3, total_r)  # Ensure total row is at least 3
    last_data_r = max(2, last_data_r)  # Ensure last data row is at least 2 for formula

    # MODIFICATION: Remove "Total" writing
    # total_label_cell = sheet_obj.cell(row=total_r, column=id_col_for_total_label)
    # total_label_cell.value = "Total"
    # total_label_cell.font = Font(bold=True)

    for field_key, diff_col_for_sum_idx in diff_indices_dict.items():
        if not diff_col_for_sum_idx: continue
        # Apply sum only to numeric diff columns, not GSTIN diff
        if field_key != id_field_name_in_diff_indices and field_key != 'gstin':  # MODIFICATION: also check for 'gstin' explicitly
            col_letter_sum = get_column_letter(diff_col_for_sum_idx)
            formula_start_row = 3
            formula_end_row = max(formula_start_row, last_data_r)
            formula = f"=SUM({col_letter_sum}{formula_start_row}:{col_letter_sum}{formula_end_row})"
            sum_cell = sheet_obj.cell(row=total_r, column=diff_col_for_sum_idx)
            sum_cell.value = formula
            sum_cell.fill = DIFF_FILL
            sum_cell.font = Font(bold=True, color="00008B")
            sum_cell.number_format = INDIAN_NUM_FORMAT  # MODIFICATION: Apply Indian number format to totals
            # Ensure the "Total" label cell (if it were to exist) is cleared or styled if needed
            # For now, we just don't write "Total"


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main():
    send_event("recon_started", {})
    root = tk.Tk()
    root.withdraw()

    modified_sheets_for_autofit = set()  # NEW: Set to track modified sheets

    infile = filedialog.askopenfilename(
        title="Select workbook",
        filetypes=[("Excel files", "*.xlsx;*.xlsm"), ("All files", "*.*")]
    )
    if not infile:
        send_event("recon_cancelled", {"stage": "input_file_selection"})
        return
    outfile = filedialog.asksaveasfilename(
        title="Save as", defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]
    )
    if not outfile:
        send_event("recon_cancelled", {"stage": "output_file_selection", "input_file_selected": bool(infile)})
        return

    logging.info(f"Reconciliation: Input file: {infile}, Output file: {outfile}")
    send_event("recon_files_selected",
               {"input_file_name": os.path.basename(infile), "output_file_name": os.path.basename(outfile)})

    try:
        wb = load_workbook(infile)
    except Exception as e:
        messagebox.showerror("File Load Error", f"Could not load the workbook '{infile}'.\nError: {e}")
        send_event("recon_error",
                   {"type": "file_load_error", "filename": os.path.basename(infile), "error_message": str(e),
                    "traceback": traceback.format_exc()})
        return

    sale_b2b_sh = wb['SALE-B2B'] if 'SALE-B2B' in wb.sheetnames else None
    sale_others_sh = wb['SALE-Others'] if 'SALE-Others' in wb.sheetnames else None
    r1_main_sh = wb['R1-B2B,SEZ,DE'] if 'R1-B2B,SEZ,DE' in wb.sheetnames else None
    r1_exp_sh = wb['R1-EXP'] if 'R1-EXP' in wb.sheetnames else None
    r1_b2ba_sh = wb['R1-B2BA'] if 'R1-B2BA' in wb.sheetnames else None
    gstr2b_b2b_sh = wb['2B-B2B'] if '2B-B2B' in wb.sheetnames else None
    pur_total_sh = wb['PUR-Total'] if 'PUR-Total' in wb.sheetnames else None
    r1_cdnr_sh = wb['R1-CDNR'] if 'R1-CDNR' in wb.sheetnames else None
    credit_r_sh = wb['CREDIT-R'] if 'CREDIT-R' in wb.sheetnames else None

    any_reconciliation_performed = False
    s1_executed, s2_executed, s3_executed, s4_executed, s5_executed = False, False, False, False, False

    sheets_to_cleanup = [s for s in
                         [sale_b2b_sh, sale_others_sh, r1_main_sh, r1_exp_sh, r1_b2ba_sh, gstr2b_b2b_sh, pur_total_sh,
                          r1_cdnr_sh, credit_r_sh] if s]
    if not sheets_to_cleanup:
        messagebox.showinfo("No Sheets Found", "No relevant sheets found in the workbook to perform any operations.")
        send_event("recon_aborted", {"reason": "no_relevant_sheets"})
        return

    for sht in sheets_to_cleanup:
        cleanup_sheet(sht)
        modified_sheets_for_autofit.add(sht)  # MODIFICATION: Add cleaned sheets

    sale_sheets_loaded = [s for s in [sale_b2b_sh, sale_others_sh] if s]
    sale_data_indices, sale_diff_indices, sale_map = {}, {}, {}
    numeric_sale_headers = ['Invoice Value', 'Taxable Value', 'Integrated Tax',
                            'Central Tax', 'State/UT Tax', 'Cess']
    invoice_header_sale = 'Invoice Number'
    gstin_header_sale = 'GSTIN/UIN of Recipient'

    if sale_sheets_loaded:
        sale_data_indices, sale_diff_indices, sale_map = prepare_sale_sheets_infrastructure(
            sale_sheets_loaded, numeric_sale_headers, invoice_header_sale, gstin_header_sale,
            modified_sheets_for_autofit  # MODIFICATION: Pass set
        )

    # --- Section 1: SALE ↔ R1-B2B,SEZ,DE ---
    if sale_sheets_loaded and r1_main_sh and sale_map:
        modified_sheets_for_autofit.add(r1_main_sh)  # MODIFICATION
        section1_headers_ok = True
        r1_data_indices = {}
        try:
            r1_data_indices['inv'] = get_header_col(r1_main_sh, invoice_header_sale, r1_main_sh.title)
            r1_data_indices['gstin'] = get_header_col(r1_main_sh, gstin_header_sale, r1_main_sh.title)
            if not r1_data_indices['inv'] or not r1_data_indices['gstin']:
                messagebox.showwarning("Header Error (Section 1)",
                                       f"Essential headers ('{invoice_header_sale}' or '{gstin_header_sale}') not found in '{r1_main_sh.title}'. Section 1 will be skipped.")
                send_event("recon_warning", {"type": "missing_header_section1", "sheet": r1_main_sh.title,
                                             "context": "essential_headers"})
                section1_headers_ok = False
            if section1_headers_ok:
                for f_hdr in numeric_sale_headers:
                    col = get_header_col(r1_main_sh, f_hdr, r1_main_sh.title)
                    if not col:
                        messagebox.showwarning("Header Warning (Section 1)",
                                               f"Numeric header '{f_hdr}' not found in '{r1_main_sh.title}'. Comparisons for this field will be skipped.")
                        send_event("recon_warning",
                                   {"type": "missing_header_section1", "sheet": r1_main_sh.title, "header": f_hdr})
                    r1_data_indices[f_hdr] = col
        except Exception as e:
            messagebox.showwarning("Header Error (Section 1)",
                                   f"An error occurred while getting headers for '{r1_main_sh.title}': {e}. Section 1 will be skipped.")
            send_event("recon_error",
                       {"type": "header_processing_section1", "sheet": r1_main_sh.title, "error_message": str(e)})
            section1_headers_ok = False

        if section1_headers_ok and r1_data_indices.get('inv'):
            any_reconciliation_performed = True
            s1_executed = True
            r1_map = {}
            inv_col_r1 = r1_data_indices['inv']
            for r_idx in range(3, r1_main_sh.max_row + 1):
                raw_invoice_val = r1_main_sh.cell(row=r_idx, column=inv_col_r1).value
                if not raw_invoice_val or str(raw_invoice_val).strip().lower() == 'total':
                    continue
                norm_key = normalize_key(raw_invoice_val)
                r1_map.setdefault(norm_key, []).append(r_idx)

            differences_s1 = {}
            common_invoices = set(sale_map.keys()) & set(r1_map.keys())

            for inv_key in common_invoices:
                sale_sheet_obj, sale_row = sale_map[inv_key]
                modified_sheets_for_autofit.add(sale_sheet_obj)  # MODIFICATION
                r1_rows = r1_map[inv_key]
                current_sdi = sale_data_indices.get(sale_sheet_obj.title, {})
                differences_s1[inv_key] = {}

                r1_gstin_col = r1_data_indices.get('gstin')
                sale_gstin_col = current_sdi.get('gstin')

                if r1_gstin_col and sale_gstin_col and r1_rows:
                    r1_gstin_val = normalize_key(r1_main_sh.cell(row=r1_rows[0], column=r1_gstin_col).value)
                    sale_gstin_val = normalize_key(sale_sheet_obj.cell(row=sale_row, column=sale_gstin_col).value)
                    if r1_gstin_val != sale_gstin_val:
                        r1_main_sh.cell(row=r1_rows[0], column=r1_gstin_col).fill = DIFF_FILL
                        sale_sheet_obj.cell(row=sale_row, column=sale_gstin_col).fill = DIFF_FILL
                        differences_s1[inv_key]['gstin'] = r1_gstin_val

                for field_name in numeric_sale_headers:
                    r1_field_col = r1_data_indices.get(field_name)
                    sale_field_col = current_sdi.get(field_name)

                    if r1_field_col and sale_field_col and r1_rows:
                        sale_value = get_numeric(sale_sheet_obj, sale_row, sale_field_col)
                        r1_value_sum = get_numeric(r1_main_sh, r1_rows[0],
                                                   r1_field_col) if field_name == 'Invoice Value' else sum(
                            get_numeric(r1_main_sh, r1_r, r1_field_col) for r1_r in r1_rows)

                        diff = r1_value_sum - sale_value
                        if abs(diff) >= EPSILON:
                            for r1_r_highlight in r1_rows:
                                r1_main_sh.cell(row=r1_r_highlight, column=r1_field_col).fill = DIFF_FILL
                            sale_sheet_obj.cell(row=sale_row, column=sale_field_col).fill = DIFF_FILL
                            differences_s1[inv_key][field_name] = diff

            unmatched_sale_invoices_s1 = {k: v for k, v in sale_map.items() if k not in common_invoices}
            unmatched_r1_invoices_s1 = {k: v for k, v in r1_map.items() if k not in common_invoices}
            r1_gstin_col_f, r1_inv_col_f = r1_data_indices.get('gstin'), r1_data_indices.get('inv')

            if r1_gstin_col_f and r1_inv_col_f:
                for r1_key_f, r1_rows_f in list(unmatched_r1_invoices_s1.items()):
                    if not r1_rows_f: continue
                    r1_gstin_val_f = normalize_key(r1_main_sh.cell(row=r1_rows_f[0], column=r1_gstin_col_f).value)
                    r1_numeric_values_f, all_r1_fields_present = {}, True
                    for field_f in numeric_sale_headers:
                        r1_f_col = r1_data_indices.get(field_f)
                        if not r1_f_col: all_r1_fields_present = False; break
                        r1_numeric_values_f[field_f] = get_numeric(r1_main_sh, r1_rows_f[0],
                                                                   r1_f_col) if field_f == 'Invoice Value' else sum(
                            get_numeric(r1_main_sh, r1_r_f, r1_f_col) for r1_r_f in r1_rows_f)
                    if not all_r1_fields_present: continue

                    for sale_key_f, (sale_sheet_f, sale_row_f) in list(unmatched_sale_invoices_s1.items()):
                        modified_sheets_for_autofit.add(sale_sheet_f)  # MODIFICATION
                        current_sdi_f = sale_data_indices.get(sale_sheet_f.title, {})
                        sale_gstin_col_f, sale_inv_col_f = current_sdi_f.get('gstin'), current_sdi_f.get('inv')
                        if not sale_gstin_col_f or not sale_inv_col_f: continue
                        sale_gstin_val_f = normalize_key(
                            sale_sheet_f.cell(row=sale_row_f, column=sale_gstin_col_f).value)

                        if r1_gstin_val_f == sale_gstin_val_f:
                            all_numeric_match = True
                            for field_f in numeric_sale_headers:
                                sale_f_col = current_sdi_f.get(field_f)
                                if not sale_f_col or r1_numeric_values_f.get(field_f) is None or abs(
                                        r1_numeric_values_f[field_f] - get_numeric(sale_sheet_f, sale_row_f,
                                                                                   sale_f_col)) >= EPSILON:
                                    all_numeric_match = False;
                                    break
                            if all_numeric_match:
                                for r1_r_h in r1_rows_f: r1_main_sh.cell(row=r1_r_h,
                                                                         column=r1_inv_col_f).fill = DIFF_FILL
                                sale_sheet_f.cell(row=sale_row_f, column=sale_inv_col_f).fill = DIFF_FILL
                                unmatched_r1_invoices_s1.pop(r1_key_f, None)
                                unmatched_sale_invoices_s1.pop(sale_key_f, None)
                                break
            for sale_sh_obj_miss, sale_r_miss in unmatched_sale_invoices_s1.values():
                modified_sheets_for_autofit.add(sale_sh_obj_miss)  # MODIFICATION
                for c_idx in range(1, sale_sh_obj_miss.max_column + 1): sale_sh_obj_miss.cell(row=sale_r_miss,
                                                                                              column=c_idx).fill = MISSING_FILL
            for r1_rows_miss_list in unmatched_r1_invoices_s1.values():
                for r1_r_miss in r1_rows_miss_list:
                    for c_idx in range(1, r1_main_sh.max_column + 1): r1_main_sh.cell(row=r1_r_miss,
                                                                                      column=c_idx).fill = MISSING_FILL

            for inv_key_diff, diff_data in differences_s1.items():
                if inv_key_diff not in sale_map: continue
                sale_sh_obj_inline, sale_r_inline = sale_map[inv_key_diff]
                modified_sheets_for_autofit.add(sale_sh_obj_inline)  # MODIFICATION
                current_sddi = sale_diff_indices.get(sale_sh_obj_inline.title, {})
                for field, diff_value in diff_data.items():
                    diff_col_idx = current_sddi.get(field)
                    if diff_col_idx:
                        diff_cell = sale_sh_obj_inline.cell(row=sale_r_inline, column=diff_col_idx)
                        diff_cell.value = diff_value
                        diff_cell.font = Font(bold=True, color="00008B")
                        if field != 'gstin':  # MODIFICATION: Apply number format to numeric diffs
                            diff_cell.number_format = INDIAN_NUM_FORMAT
        else:
            if not (sale_sheets_loaded and r1_main_sh):
                messagebox.showwarning("Skipping Section 1", "Required sheets for SALE ↔ R1-B2B,SEZ,DE not found.")
                send_event("recon_skipped_section", {"section": 1, "reason": "required_sheets_not_found"})

    # --- Section 2: R1-EXP ↔ SALE ---
    if r1_exp_sh and sale_sheets_loaded and sale_map:
        modified_sheets_for_autofit.add(r1_exp_sh)  # MODIFICATION
        map_exp_to_sale = {'Invoice Number': invoice_header_sale, 'Total Invoice Value': 'Invoice Value',
                           'Total Taxable Value': 'Taxable Value', 'Integrated Tax': 'Integrated Tax', 'Cess': 'Cess'}
        exp_data_indices, exp_headers_ok = {}, True
        try:
            for hdr_exp in map_exp_to_sale.keys():
                col = get_header_col(r1_exp_sh, hdr_exp, r1_exp_sh.title)
                if not col and hdr_exp == 'Invoice Number':
                    exp_headers_ok = False;
                    messagebox.showwarning("Header Warning (Section 2)",
                                           f"Critical header '{hdr_exp}' not found in '{r1_exp_sh.title}'. Section 2 skipped.")
                elif not col:
                    messagebox.showwarning("Header Warning (Section 2)",
                                           f"Header '{hdr_exp}' not found in '{r1_exp_sh.title}'.")
                exp_data_indices[hdr_exp] = col
        except Exception as e:
            exp_headers_ok = False;
            messagebox.showwarning("Header Error (Section 2)",
                                   f"Error getting headers for '{r1_exp_sh.title}': {e}. Section 2 skipped.")

        if exp_headers_ok and exp_data_indices.get('Invoice Number'):
            any_reconciliation_performed = True;
            s2_executed = True
            exp_invoice_map = {}
            exp_inv_col = exp_data_indices['Invoice Number']
            for r_exp in range(3, r1_exp_sh.max_row + 1):
                raw_inv_exp = r1_exp_sh.cell(row=r_exp, column=exp_inv_col).value
                if not raw_inv_exp or str(raw_inv_exp).strip().lower() == 'total': continue
                exp_invoice_map.setdefault(normalize_key(raw_inv_exp), []).append(r_exp)

            for inv_key_exp, r_exp_list in exp_invoice_map.items():
                if inv_key_exp in sale_map:
                    sale_sh_common, sale_r_common = sale_map[inv_key_exp]
                    modified_sheets_for_autofit.add(sale_sh_common)  # MODIFICATION
                    current_sdi_exp = sale_data_indices.get(sale_sh_common.title, {})
                    current_sddi_exp = sale_diff_indices.get(sale_sh_common.title, {})
                    for exp_hdr, sale_hdr_map_val in map_exp_to_sale.items():
                        if exp_hdr == 'Invoice Number': continue
                        exp_val_col, sale_val_col, sale_diff_col = exp_data_indices.get(exp_hdr), current_sdi_exp.get(
                            sale_hdr_map_val), current_sddi_exp.get(sale_hdr_map_val)
                        if exp_val_col and sale_val_col:
                            val_sale = get_numeric(sale_sh_common, sale_r_common, sale_val_col)
                            val_exp = sum(get_numeric(r1_exp_sh, r_exp_c, exp_val_col) for r_exp_c in r_exp_list)
                            diff_exp_sale = val_exp - val_sale
                            if abs(diff_exp_sale) >= EPSILON:
                                for r_exp_fill in r_exp_list: r1_exp_sh.cell(row=r_exp_fill,
                                                                             column=exp_val_col).fill = DIFF_FILL
                                sale_sh_common.cell(row=sale_r_common, column=sale_val_col).fill = DIFF_FILL
                                if sale_diff_col:
                                    diff_cell = sale_sh_common.cell(row=sale_r_common, column=sale_diff_col)
                                    diff_cell.value = diff_exp_sale
                                    diff_cell.font = Font(bold=True, color="00008B")
                                    diff_cell.number_format = INDIAN_NUM_FORMAT  # MODIFICATION
                else:
                    for r_exp_miss in r_exp_list:
                        for c_idx in range(1, r1_exp_sh.max_column + 1): r1_exp_sh.cell(row=r_exp_miss,
                                                                                        column=c_idx).fill = MISSING_FILL
        else:
            if not (r1_exp_sh and sale_sheets_loaded and sale_map):
                messagebox.showwarning("Skipping Section 2",
                                       "Required sheets/data for R1-EXP ↔ SALE not found or critical headers missing.")
                send_event("recon_skipped_section", {"section": 2, "reason": "required_sheets_or_headers_missing"})

    # --- Section 3: R1-B2BA ↔ SALE ---
    if r1_b2ba_sh and sale_sheets_loaded and sale_map:
        modified_sheets_for_autofit.add(r1_b2ba_sh)  # MODIFICATION
        map_b2ba_to_sale = {'GSTIN/UIN of Recipient': gstin_header_sale, 'Revised Invoice No.': invoice_header_sale,
                            'Total Invoice Value': 'Invoice Value', 'Total Taxable Value': 'Taxable Value',
                            'Integrated Tax': 'Integrated Tax', 'Central Tax': 'Central Tax',
                            'State/UT Tax': 'State/UT Tax', 'Cess': 'Cess'}
        b2ba_data_indices, b2ba_headers_ok = {}, True
        try:
            for hdr_b2ba in map_b2ba_to_sale.keys():
                col = get_header_col(r1_b2ba_sh, hdr_b2ba, r1_b2ba_sh.title)
                if not col and (hdr_b2ba == 'Revised Invoice No.' or hdr_b2ba == 'GSTIN/UIN of Recipient'):
                    b2ba_headers_ok = False;
                    messagebox.showwarning("Header Warning (Section 3)",
                                           f"Critical header '{hdr_b2ba}' not found in '{r1_b2ba_sh.title}'. Section 3 skipped.")
                elif not col:
                    messagebox.showwarning("Header Warning (Section 3)",
                                           f"Header '{hdr_b2ba}' not found in '{r1_b2ba_sh.title}'.")
                b2ba_data_indices[hdr_b2ba] = col
        except Exception as e:
            b2ba_headers_ok = False;
            messagebox.showwarning("Header Error (Section 3)",
                                   f"Error getting headers for '{r1_b2ba_sh.title}': {e}. Section 3 skipped.")

        if b2ba_headers_ok and b2ba_data_indices.get('Revised Invoice No.'):
            any_reconciliation_performed = True;
            s3_executed = True
            b2ba_invoice_map = {}
            b2ba_inv_col = b2ba_data_indices['Revised Invoice No.']
            for r_b2ba in range(3, r1_b2ba_sh.max_row + 1):
                raw_inv_b2ba = r1_b2ba_sh.cell(row=r_b2ba, column=b2ba_inv_col).value
                if not raw_inv_b2ba or str(raw_inv_b2ba).strip().lower() == 'total': continue
                b2ba_invoice_map.setdefault(normalize_key(raw_inv_b2ba), []).append(r_b2ba)

            for inv_key_b2ba, r_b2ba_list in b2ba_invoice_map.items():
                if inv_key_b2ba in sale_map:
                    sale_sh_common, sale_r_common = sale_map[inv_key_b2ba]
                    modified_sheets_for_autofit.add(sale_sh_common)  # MODIFICATION
                    current_sdi_b2ba = sale_data_indices.get(sale_sh_common.title, {})
                    current_sddi_b2ba = sale_diff_indices.get(sale_sh_common.title, {})

                    b2ba_gstin_hdr, sale_gstin_key = 'GSTIN/UIN of Recipient', 'gstin'  # sale_gstin_key is 'gstin'
                    b2ba_gstin_col, sale_gstin_col, sale_gstin_diff_col = b2ba_data_indices.get(
                        b2ba_gstin_hdr), current_sdi_b2ba.get(sale_gstin_key), current_sddi_b2ba.get(sale_gstin_key)
                    if b2ba_gstin_col and sale_gstin_col and r_b2ba_list:
                        val_gstin_b2ba = normalize_key(r1_b2ba_sh.cell(row=r_b2ba_list[0], column=b2ba_gstin_col).value)
                        val_gstin_sale = normalize_key(
                            sale_sh_common.cell(row=sale_r_common, column=sale_gstin_col).value)
                        if val_gstin_b2ba != val_gstin_sale:
                            r1_b2ba_sh.cell(row=r_b2ba_list[0], column=b2ba_gstin_col).fill = DIFF_FILL
                            sale_sh_common.cell(row=sale_r_common, column=sale_gstin_col).fill = DIFF_FILL
                            if sale_gstin_diff_col:
                                diff_cell = sale_sh_common.cell(row=sale_r_common, column=sale_gstin_diff_col)
                                diff_cell.value = val_gstin_b2ba  # This is a GSTIN string, not a number
                                diff_cell.font = Font(bold=True, color="00008B")
                                # No number format for GSTIN diff

                    for b2ba_hdr, sale_hdr_map_val in map_b2ba_to_sale.items():
                        if b2ba_hdr in ['GSTIN/UIN of Recipient', 'Revised Invoice No.']: continue
                        b2ba_val_col, sale_val_col, sale_diff_col = b2ba_data_indices.get(
                            b2ba_hdr), current_sdi_b2ba.get(sale_hdr_map_val), current_sddi_b2ba.get(sale_hdr_map_val)
                        if b2ba_val_col and sale_val_col:
                            val_sale_num = get_numeric(sale_sh_common, sale_r_common, sale_val_col)
                            val_b2ba_num = sum(
                                get_numeric(r1_b2ba_sh, r_b2ba_c, b2ba_val_col) for r_b2ba_c in r_b2ba_list)
                            diff_b2ba_sale = val_b2ba_num - val_sale_num
                            if abs(diff_b2ba_sale) >= EPSILON:
                                for r_b2ba_fill in r_b2ba_list: r1_b2ba_sh.cell(row=r_b2ba_fill,
                                                                                column=b2ba_val_col).fill = DIFF_FILL
                                sale_sh_common.cell(row=sale_r_common, column=sale_val_col).fill = DIFF_FILL
                                if sale_diff_col:
                                    diff_cell = sale_sh_common.cell(row=sale_r_common, column=sale_diff_col)
                                    diff_cell.value = diff_b2ba_sale
                                    diff_cell.font = Font(bold=True, color="00008B")
                                    diff_cell.number_format = INDIAN_NUM_FORMAT  # MODIFICATION
                else:
                    for r_b2ba_miss in r_b2ba_list:
                        for c_idx in range(1, r1_b2ba_sh.max_column + 1): r1_b2ba_sh.cell(row=r_b2ba_miss,
                                                                                          column=c_idx).fill = MISSING_FILL
        else:
            if not (r1_b2ba_sh and sale_sheets_loaded and sale_map):
                messagebox.showwarning("Skipping Section 3",
                                       "Required sheets/data for R1-B2BA ↔ SALE not found or critical headers missing.")
                send_event("recon_skipped_section", {"section": 3, "reason": "required_sheets_or_headers_missing"})

    if sale_sheets_loaded and (s1_executed or s2_executed or s3_executed):
        for sale_sh in sale_sheets_loaded:
            if sale_sh and sale_sh.title in sale_data_indices and sale_sh.title in sale_diff_indices:
                add_totals_to_sheet(sale_sh, sale_data_indices[sale_sh.title], sale_diff_indices[sale_sh.title], 'inv',
                                    'gstin', modified_sheets_for_autofit)  # MODIFICATION: Pass set

    # --- Section 4: 2B-B2B ↔ PUR-Total ---
    pur_data_indices_s4, pur_diff_indices_s4 = {}, {}
    if gstr2b_b2b_sh and pur_total_sh:
        modified_sheets_for_autofit.add(gstr2b_b2b_sh)  # MODIFICATION
        modified_sheets_for_autofit.add(pur_total_sh)  # MODIFICATION

        map_2b_to_pur = {'GSTIN/UIN of Supplier': 'GSTIN/UIN of Supplier', 'Invoice Number': 'Supplier Invoice Number',
                         'Invoice Value': 'Invoice Value', 'Total Taxable Value': 'Taxable Value',
                         'Integrated Tax': 'Integrated Tax', 'Central Tax': 'Central Tax',
                         'State/UT Tax': 'State/UT Tax', 'Cess': 'Cess'}
        s4_headers_ok = True

        pur_cols_to_insert_s4 = []
        pur_gstin_hdr_s4_orig = map_2b_to_pur['GSTIN/UIN of Supplier']
        orig_col_gstin_s4 = get_header_col(pur_total_sh, pur_gstin_hdr_s4_orig, pur_total_sh.title)
        # For Section 4, GSTIN diff column label is 'GSTIN/UIN Diff'
        gstin_diff_label_s4 = "GSTIN/UIN Diff"
        if orig_col_gstin_s4:
            pur_cols_to_insert_s4.append(
                (orig_col_gstin_s4, gstin_diff_label_s4, 'gstin_key'))  # field_key_for_diff_indices
        else:
            messagebox.showwarning("PUR-Total Header",
                                   f"Header '{pur_gstin_hdr_s4_orig}' not found in '{pur_total_sh.title}'. {gstin_diff_label_s4} for PUR skipped.")

        for b2b_hdr, pur_hdr in map_2b_to_pur.items():
            if b2b_hdr not in ['GSTIN/UIN of Supplier', 'Invoice Number']:  # Numeric headers
                orig_col_num_s4 = get_header_col(pur_total_sh, pur_hdr, pur_total_sh.title)
                if orig_col_num_s4:
                    pur_cols_to_insert_s4.append((orig_col_num_s4, f"{pur_hdr} Diff", b2b_hdr))  # b2b_hdr is field_key
                else:
                    messagebox.showwarning("PUR-Total Header",
                                           f"Header '{pur_hdr}' not found in '{pur_total_sh.title}'. Diff column for it skipped.")

        pur_cols_to_insert_s4.sort(key=lambda item: item[0], reverse=True)
        for orig_col, diff_lbl, _field_key in pur_cols_to_insert_s4:
            pur_total_sh.insert_cols(orig_col + 1)
            hdr_cell_s4 = pur_total_sh.cell(row=2, column=orig_col + 1)
            hdr_cell_s4.value = diff_lbl;
            hdr_cell_s4.font = Font(bold=True)

        b2b_data_indices_s4 = {}
        try:
            for b2b_key, pur_val_hdr in map_2b_to_pur.items():
                col_2b = get_header_col(gstr2b_b2b_sh, b2b_key, gstr2b_b2b_sh.title)
                col_pur = get_header_col(pur_total_sh, pur_val_hdr, pur_total_sh.title)
                if not col_2b or not col_pur:
                    critical_missing = (b2b_key == 'Invoice Number' or b2b_key == 'GSTIN/UIN of Supplier')
                    msg = f"Header '{b2b_key if not col_2b else pur_val_hdr}' not found in '{gstr2b_b2b_sh.title if not col_2b else pur_total_sh.title}'."
                    messagebox.showwarning("Header Warning (Section 4)",
                                           msg + (" Section 4 may be skipped." if critical_missing else ""))
                    if critical_missing: s4_headers_ok = False
                b2b_data_indices_s4[b2b_key] = col_2b
                pur_data_indices_s4[b2b_key] = col_pur  # This stores original data col indices for PUR

            # Populate pur_diff_indices_s4 by looking up the *newly added* diff column headers
            pur_diff_indices_s4['gstin_key'] = get_header_col(pur_total_sh, gstin_diff_label_s4, pur_total_sh.title)
            for b2b_hdr, pur_hdr in map_2b_to_pur.items():
                if b2b_hdr not in ['GSTIN/UIN of Supplier', 'Invoice Number']:
                    diff_col_label = f"{pur_hdr} Diff"
                    col_pur_diff = get_header_col(pur_total_sh, diff_col_label, pur_total_sh.title)
                    if not col_pur_diff: messagebox.showwarning("Diff Header Error (S4)",
                                                                f"Diff column '{diff_col_label}' not found in {pur_total_sh.title}.")
                    pur_diff_indices_s4[b2b_hdr] = col_pur_diff  # b2b_hdr is the field_key

        except Exception as e:
            s4_headers_ok = False;
            messagebox.showwarning("Header Error (S4)",
                                   f"Error processing headers for S4: {e}. S4 skipped.")

        if s4_headers_ok and b2b_data_indices_s4.get('Invoice Number') and pur_data_indices_s4.get('Invoice Number'):
            any_reconciliation_performed = True;
            s4_executed = True
            b2b_inv_map_s4, pur_inv_map_s4 = {}, {}
            inv_2b_col, inv_pur_col = b2b_data_indices_s4['Invoice Number'], pur_data_indices_s4['Invoice Number']

            for r_2b in range(3, gstr2b_b2b_sh.max_row + 1):
                raw_inv = gstr2b_b2b_sh.cell(row=r_2b, column=inv_2b_col).value
                if raw_inv and str(raw_inv).strip().lower() != 'total': b2b_inv_map_s4.setdefault(
                    normalize_key(raw_inv), []).append(r_2b)
            for r_pur in range(3, pur_total_sh.max_row + 1):
                raw_inv = pur_total_sh.cell(row=r_pur, column=inv_pur_col).value
                if raw_inv and str(raw_inv).strip().lower() != 'total':
                    key_pur = normalize_key(raw_inv)
                    if key_pur not in pur_inv_map_s4: pur_inv_map_s4[key_pur] = r_pur

            differences_s4 = {}
            common_inv_s4 = set(b2b_inv_map_s4.keys()) & set(pur_inv_map_s4.keys())

            for inv_key in common_inv_s4:
                pur_r, b2b_r_list = pur_inv_map_s4[inv_key], b2b_inv_map_s4[inv_key]
                current_diffs = {}
                b2b_gstin_hdr_key = 'GSTIN/UIN of Supplier'  # This is the key for b2b_data_indices_s4 and pur_data_indices_s4

                b2b_gstin_col = b2b_data_indices_s4.get(b2b_gstin_hdr_key)
                pur_gstin_col = pur_data_indices_s4.get(b2b_gstin_hdr_key)  # This is original GSTIN col in PUR

                if b2b_gstin_col and pur_gstin_col and b2b_r_list:
                    val_2b_gstin = normalize_key(gstr2b_b2b_sh.cell(row=b2b_r_list[0], column=b2b_gstin_col).value)
                    val_pur_gstin = normalize_key(pur_total_sh.cell(row=pur_r, column=pur_gstin_col).value)
                    if val_2b_gstin != val_pur_gstin:
                        gstr2b_b2b_sh.cell(row=b2b_r_list[0], column=b2b_gstin_col).fill = DIFF_FILL
                        pur_total_sh.cell(row=pur_r, column=pur_gstin_col).fill = DIFF_FILL
                        current_diffs['gstin_key'] = val_2b_gstin  # Store the 2B value in diff

                for b2b_hdr_key_numeric, _ in map_2b_to_pur.items():  # b2b_hdr_key_numeric is the field_key
                    if b2b_hdr_key_numeric in ['GSTIN/UIN of Supplier', 'Invoice Number']: continue
                    b2b_val_col = b2b_data_indices_s4.get(b2b_hdr_key_numeric)
                    pur_val_col = pur_data_indices_s4.get(b2b_hdr_key_numeric)  # Original PUR data col
                    if b2b_val_col and pur_val_col and b2b_r_list:
                        val_pur = get_numeric(pur_total_sh, pur_r, pur_val_col)
                        val_2b = get_numeric(gstr2b_b2b_sh, b2b_r_list[0],
                                             b2b_val_col) if b2b_hdr_key_numeric == 'Invoice Value' else sum(
                            get_numeric(gstr2b_b2b_sh, r, b2b_val_col) for r in b2b_r_list)
                        diff = val_2b - val_pur
                        if abs(diff) >= EPSILON:
                            for r_fill in b2b_r_list: gstr2b_b2b_sh.cell(row=r_fill,
                                                                         column=b2b_val_col).fill = DIFF_FILL
                            pur_total_sh.cell(row=pur_r, column=pur_val_col).fill = DIFF_FILL
                            current_diffs[b2b_hdr_key_numeric] = diff
                if current_diffs: differences_s4[inv_key] = current_diffs

            unmatched_2b_s4 = {k: v for k, v in b2b_inv_map_s4.items() if k not in common_inv_s4}
            unmatched_pur_s4 = {k: v for k, v in pur_inv_map_s4.items() if k not in common_inv_s4}
            b2b_gstin_f, pur_gstin_f = b2b_data_indices_s4.get('GSTIN/UIN of Supplier'), pur_data_indices_s4.get(
                'GSTIN/UIN of Supplier')
            b2b_inv_f, pur_inv_f = b2b_data_indices_s4.get('Invoice Number'), pur_data_indices_s4.get('Invoice Number')

            if b2b_gstin_f and pur_gstin_f and b2b_inv_f and pur_inv_f:
                for b2b_k_f, b2b_rs_f in list(unmatched_2b_s4.items()):
                    if not b2b_rs_f: continue
                    b2b_gstin_val = normalize_key(gstr2b_b2b_sh.cell(row=b2b_rs_f[0], column=b2b_gstin_f).value)
                    b2b_nums_f, all_2b_f_ok = {}, True
                    for b2b_h_f, _ in map_2b_to_pur.items():
                        if b2b_h_f in ['GSTIN/UIN of Supplier', 'Invoice Number']: continue
                        b2b_c_f = b2b_data_indices_s4.get(b2b_h_f)
                        if not b2b_c_f: all_2b_f_ok = False; break
                        b2b_nums_f[b2b_h_f] = get_numeric(gstr2b_b2b_sh, b2b_rs_f[0],
                                                          b2b_c_f) if b2b_h_f == 'Invoice Value' else sum(
                            get_numeric(gstr2b_b2b_sh, r_f, b2b_c_f) for r_f in b2b_rs_f)
                    if not all_2b_f_ok: continue
                    for pur_k_f, pur_r_f in list(unmatched_pur_s4.items()):
                        pur_gstin_val = normalize_key(pur_total_sh.cell(row=pur_r_f, column=pur_gstin_f).value)
                        if b2b_gstin_val == pur_gstin_val:
                            all_match_f = True
                            for b2b_h_f, _ in map_2b_to_pur.items():
                                if b2b_h_f in ['GSTIN/UIN of Supplier', 'Invoice Number']: continue
                                pur_c_f = pur_data_indices_s4.get(b2b_h_f)
                                if not pur_c_f or b2b_nums_f.get(b2b_h_f) is None or abs(
                                        b2b_nums_f[b2b_h_f] - get_numeric(pur_total_sh, pur_r_f, pur_c_f)) >= EPSILON:
                                    all_match_f = False;
                                    break
                            if all_match_f:
                                for r_b2b_h in b2b_rs_f: gstr2b_b2b_sh.cell(row=r_b2b_h,
                                                                            column=b2b_inv_f).fill = DIFF_FILL
                                pur_total_sh.cell(row=pur_r_f, column=pur_inv_f).fill = DIFF_FILL
                                unmatched_2b_s4.pop(b2b_k_f, None);
                                unmatched_pur_s4.pop(pur_k_f, None);
                                break
            for b2b_rs_m in unmatched_2b_s4.values():
                for r_2b_m in b2b_rs_m:
                    for c_idx in range(1, gstr2b_b2b_sh.max_column + 1): gstr2b_b2b_sh.cell(row=r_2b_m,
                                                                                            column=c_idx).fill = MISSING_FILL
            for pur_r_m in unmatched_pur_s4.values():
                for c_idx in range(1, pur_total_sh.max_column + 1): pur_total_sh.cell(row=pur_r_m,
                                                                                      column=c_idx).fill = MISSING_FILL

            for inv_key_d, diff_details in differences_s4.items():
                if inv_key_d not in pur_inv_map_s4: continue
                pur_r_inline = pur_inv_map_s4[inv_key_d]
                for f_key, diff_val in diff_details.items():  # f_key is 'gstin_key' or a b2b_hdr_key_numeric
                    diff_col_idx = pur_diff_indices_s4.get(f_key)
                    if diff_col_idx:
                        cell = pur_total_sh.cell(row=pur_r_inline, column=diff_col_idx)
                        cell.value = diff_val
                        cell.font = Font(bold=True, color="00008B")
                        if f_key != 'gstin_key':  # MODIFICATION: Apply format to numeric diffs
                            cell.number_format = INDIAN_NUM_FORMAT

            # Pass 'Supplier Invoice Number' as key_header_name for PUR-Total totals
            key_header_for_pur_totals = 'Invoice Number'  # This is the key in pur_data_indices_s4 for 'Supplier Invoice Number'
            if s4_executed and pur_data_indices_s4.get(key_header_for_pur_totals):
                add_totals_to_sheet(pur_total_sh, pur_data_indices_s4, pur_diff_indices_s4, key_header_for_pur_totals,
                                    'gstin_key', modified_sheets_for_autofit)  # MODIFICATION: Pass set
        else:
            if not (gstr2b_b2b_sh and pur_total_sh):
                messagebox.showwarning("Skipping Section 4", "Required sheets for 2B-B2B ↔ PUR-Total not found.")
                send_event("recon_skipped_section", {"section": 4, "reason": "required_sheets_not_found"})

    # --- Section 5: R1-CDNR ↔ CREDIT-R ---
    credit_r_data_indices_s5, credit_r_diff_indices_s5 = {}, {}
    if r1_cdnr_sh and credit_r_sh:
        modified_sheets_for_autofit.add(r1_cdnr_sh)  # MODIFICATION
        modified_sheets_for_autofit.add(credit_r_sh)  # MODIFICATION

        r1_cdnr_note_header = "Note Number"
        credit_r_note_header = "Note Number"
        common_gstin_header_s5 = 'GSTIN/UIN of Recipient'  # Original header name in CREDIT-R for GSTIN

        map_cdnr_to_credit_r_numeric = {  # cdnr_hdr_key : credit_r_original_hdr_name
            "Note Value": "Note Value",
            "Taxable Value": "Taxable Value",
            "Integrated Tax": "Integrated Tax",
            "Central Tax": "Central Tax",
            "State/UT Tax": "State/UT Tax",
            "Cess": "Cess"
        }
        s5_headers_ok = True

        credit_r_cols_to_insert_s5 = []
        # MODIFICATION: Correct GSTIN Diff label for CREDIT-R
        gstin_diff_label_s5 = "GSTIN/UIN Diff"
        original_col_idx_gstin_s5 = get_header_col(credit_r_sh, common_gstin_header_s5, credit_r_sh.title)
        if original_col_idx_gstin_s5:
            credit_r_cols_to_insert_s5.append(
                (original_col_idx_gstin_s5, gstin_diff_label_s5, 'gstin'))  # field_key is 'gstin'
        else:
            messagebox.showwarning("CREDIT-R Header",
                                   f"Header '{common_gstin_header_s5}' not found in '{credit_r_sh.title}'. {gstin_diff_label_s5} for CREDIT-R skipped.")

        for cdnr_hdr_s5_key, credit_r_hdr_s5_orig in map_cdnr_to_credit_r_numeric.items():
            original_col_idx_num_s5 = get_header_col(credit_r_sh, credit_r_hdr_s5_orig, credit_r_sh.title)
            if original_col_idx_num_s5:
                # Diff label uses the original header name from CREDIT-R
                credit_r_cols_to_insert_s5.append((original_col_idx_num_s5, f"{credit_r_hdr_s5_orig} Diff",
                                                   cdnr_hdr_s5_key))  # field_key is cdnr_hdr_s5_key
            else:
                messagebox.showwarning("CREDIT-R Header",
                                       f"Header '{credit_r_hdr_s5_orig}' not found in '{credit_r_sh.title}'. Diff column for it skipped.")

        credit_r_cols_to_insert_s5.sort(key=lambda item: item[0], reverse=True)
        for original_col_idx, diff_label_to_write, _field_key in credit_r_cols_to_insert_s5:
            credit_r_sh.insert_cols(original_col_idx + 1)
            header_cell_s5 = credit_r_sh.cell(row=2, column=original_col_idx + 1)
            header_cell_s5.value = diff_label_to_write
            header_cell_s5.font = Font(bold=True)

        r1_cdnr_data_indices_s5 = {}
        try:
            # R1-CDNR data indices (original columns)
            r1_cdnr_data_indices_s5['note_num'] = get_header_col(r1_cdnr_sh, r1_cdnr_note_header, r1_cdnr_sh.title)
            r1_cdnr_data_indices_s5['gstin'] = get_header_col(r1_cdnr_sh, common_gstin_header_s5,
                                                              r1_cdnr_sh.title)  # R1-CDNR also has 'GSTIN/UIN of Recipient'
            if not r1_cdnr_data_indices_s5['note_num'] or not r1_cdnr_data_indices_s5['gstin']:
                messagebox.showwarning("Header Error (Section 5)",
                                       f"Essential headers for '{r1_cdnr_sh.title}' not found. Section 5 skipped.")
                s5_headers_ok = False

            if s5_headers_ok:
                for cdnr_hdr_s5_key in map_cdnr_to_credit_r_numeric.keys():  # cdnr_hdr_s5_key is the header name in R1-CDNR
                    col = get_header_col(r1_cdnr_sh, cdnr_hdr_s5_key, r1_cdnr_sh.title)
                    if not col: messagebox.showwarning("Header Warning (S5)",
                                                       f"Header '{cdnr_hdr_s5_key}' not found in '{r1_cdnr_sh.title}'.")
                    r1_cdnr_data_indices_s5[cdnr_hdr_s5_key] = col

            # CREDIT-R data indices (original columns)
            if s5_headers_ok:
                credit_r_data_indices_s5['note_num'] = get_header_col(credit_r_sh, credit_r_note_header,
                                                                      credit_r_sh.title)
                credit_r_data_indices_s5['gstin'] = get_header_col(credit_r_sh, common_gstin_header_s5,
                                                                   # Original GSTIN header in CREDIT-R
                                                                   credit_r_sh.title)
                if not credit_r_data_indices_s5['note_num'] or not credit_r_data_indices_s5['gstin']:
                    messagebox.showwarning("Header Error (Section 5)",
                                           f"Essential headers for '{credit_r_sh.title}' not found. Section 5 skipped.")
                    s5_headers_ok = False

            # CREDIT-R data and diff indices
            if s5_headers_ok:
                for cdnr_hdr_s5_key, credit_r_hdr_s5_orig in map_cdnr_to_credit_r_numeric.items():
                    credit_r_data_indices_s5[cdnr_hdr_s5_key] = get_header_col(credit_r_sh, credit_r_hdr_s5_orig,
                                                                               # Original data col in CREDIT-R
                                                                               credit_r_sh.title)
                    # Diff column index in CREDIT-R (uses original CREDIT-R header name for "Diff" part)
                    credit_r_diff_indices_s5[cdnr_hdr_s5_key] = get_header_col(credit_r_sh,
                                                                               f"{credit_r_hdr_s5_orig} Diff",
                                                                               credit_r_sh.title)
                # GSTIN Diff column index in CREDIT-R
                credit_r_diff_indices_s5['gstin'] = get_header_col(credit_r_sh, gstin_diff_label_s5,
                                                                   # Use the corrected label
                                                                   credit_r_sh.title)

        except Exception as e:
            messagebox.showwarning("Header Error (Section 5)",
                                   f"An error occurred processing headers for Section 5: {e}. Section 5 skipped.")
            s5_headers_ok = False

        if s5_headers_ok and r1_cdnr_data_indices_s5.get('note_num') and credit_r_data_indices_s5.get('note_num'):
            any_reconciliation_performed = True;
            s5_executed = True
            r1_cdnr_map_s5 = {}
            cdnr_note_col = r1_cdnr_data_indices_s5['note_num']
            for r_cdnr in range(3, r1_cdnr_sh.max_row + 1):
                raw_note = r1_cdnr_sh.cell(row=r_cdnr, column=cdnr_note_col).value
                if raw_note and str(raw_note).strip().lower() != 'total':
                    r1_cdnr_map_s5.setdefault(normalize_key(raw_note), []).append(r_cdnr)

            credit_r_map_s5 = {}
            credit_note_col = credit_r_data_indices_s5['note_num']
            for r_credit in range(3, credit_r_sh.max_row + 1):
                raw_note = credit_r_sh.cell(row=r_credit, column=credit_note_col).value
                if raw_note and str(raw_note).strip().lower() != 'total':
                    key_credit = normalize_key(raw_note)
                    if key_credit not in credit_r_map_s5: credit_r_map_s5[key_credit] = r_credit

            differences_s5 = {}
            common_notes_s5 = set(r1_cdnr_map_s5.keys()) & set(credit_r_map_s5.keys())

            for note_key in common_notes_s5:
                credit_r_row = credit_r_map_s5[note_key]
                cdnr_rows = r1_cdnr_map_s5[note_key]
                current_diffs_s5 = {}

                cdnr_gstin_col = r1_cdnr_data_indices_s5.get('gstin')
                credit_gstin_col = credit_r_data_indices_s5.get('gstin')  # Original GSTIN col in CREDIT-R
                if cdnr_gstin_col and credit_gstin_col and cdnr_rows:
                    val_cdnr_gstin = normalize_key(r1_cdnr_sh.cell(row=cdnr_rows[0], column=cdnr_gstin_col).value)
                    val_credit_gstin = normalize_key(credit_r_sh.cell(row=credit_r_row, column=credit_gstin_col).value)
                    if val_cdnr_gstin != val_credit_gstin:
                        r1_cdnr_sh.cell(row=cdnr_rows[0], column=cdnr_gstin_col).fill = DIFF_FILL
                        credit_r_sh.cell(row=credit_r_row, column=credit_gstin_col).fill = DIFF_FILL
                        current_diffs_s5['gstin'] = val_cdnr_gstin  # Store R1-CDNR GSTIN in diff

                for cdnr_hdr_key_num, _credit_r_hdr_orig in map_cdnr_to_credit_r_numeric.items():
                    cdnr_val_col = r1_cdnr_data_indices_s5.get(cdnr_hdr_key_num)  # R1-CDNR data col
                    credit_val_col = credit_r_data_indices_s5.get(
                        cdnr_hdr_key_num)  # CREDIT-R data col (using cdnr_hdr_key_num as key)

                    if cdnr_val_col and credit_val_col and cdnr_rows:
                        val_credit = get_numeric(credit_r_sh, credit_r_row, credit_val_col)
                        val_cdnr = get_numeric(r1_cdnr_sh, cdnr_rows[0],
                                               cdnr_val_col) if cdnr_hdr_key_num == "Note Value" else sum(
                            get_numeric(r1_cdnr_sh, r_s5, cdnr_val_col) for r_s5 in cdnr_rows)
                        diff = val_cdnr - val_credit
                        if abs(diff) >= EPSILON:
                            for r_fill in cdnr_rows: r1_cdnr_sh.cell(row=r_fill, column=cdnr_val_col).fill = DIFF_FILL
                            credit_r_sh.cell(row=credit_r_row, column=credit_val_col).fill = DIFF_FILL
                            current_diffs_s5[cdnr_hdr_key_num] = diff
                if current_diffs_s5: differences_s5[note_key] = current_diffs_s5

            unmatched_cdnr_s5 = {k: v for k, v in r1_cdnr_map_s5.items() if k not in common_notes_s5}
            unmatched_credit_s5 = {k: v for k, v in credit_r_map_s5.items() if k not in common_notes_s5}
            cdnr_gstin_f, credit_gstin_f = r1_cdnr_data_indices_s5.get('gstin'), credit_r_data_indices_s5.get('gstin')
            cdnr_note_f, credit_note_f = r1_cdnr_data_indices_s5.get('note_num'), credit_r_data_indices_s5.get(
                'note_num')

            if cdnr_gstin_f and credit_gstin_f and cdnr_note_f and credit_note_f:
                for cdnr_k_f, cdnr_rs_f in list(unmatched_cdnr_s5.items()):
                    if not cdnr_rs_f: continue
                    cdnr_gstin_val = normalize_key(r1_cdnr_sh.cell(row=cdnr_rs_f[0], column=cdnr_gstin_f).value)
                    cdnr_nums_f, all_cdnr_ok_f = {}, True
                    for c_h_f, _ in map_cdnr_to_credit_r_numeric.items():
                        c_c_f = r1_cdnr_data_indices_s5.get(c_h_f)
                        if not c_c_f: all_cdnr_ok_f = False; break
                        cdnr_nums_f[c_h_f] = get_numeric(r1_cdnr_sh, cdnr_rs_f[0],
                                                         c_c_f) if c_h_f == "Note Value" else sum(
                            get_numeric(r1_cdnr_sh, r_f_s5, c_c_f) for r_f_s5 in cdnr_rs_f)
                    if not all_cdnr_ok_f: continue
                    for credit_k_f, credit_r_f in list(unmatched_credit_s5.items()):
                        credit_gstin_val = normalize_key(credit_r_sh.cell(row=credit_r_f, column=credit_gstin_f).value)
                        if cdnr_gstin_val == credit_gstin_val:
                            all_match_s5_f = True
                            for c_h_f, _ in map_cdnr_to_credit_r_numeric.items():
                                cr_c_f = credit_r_data_indices_s5.get(c_h_f)  # Using c_h_f (cdnr key)
                                if not cr_c_f or cdnr_nums_f.get(c_h_f) is None or abs(
                                        cdnr_nums_f[c_h_f] - get_numeric(credit_r_sh, credit_r_f, cr_c_f)) >= EPSILON:
                                    all_match_s5_f = False;
                                    break
                            if all_match_s5_f:
                                for r_cdnr_h_f in cdnr_rs_f: r1_cdnr_sh.cell(row=r_cdnr_h_f,
                                                                             column=cdnr_note_f).fill = DIFF_FILL
                                credit_r_sh.cell(row=credit_r_f, column=credit_note_f).fill = DIFF_FILL
                                unmatched_cdnr_s5.pop(cdnr_k_f, None);
                                unmatched_credit_s5.pop(credit_k_f, None);
                                break

            for cdnr_rs_m in unmatched_cdnr_s5.values():
                for r_m in cdnr_rs_m:
                    for c_idx in range(1, r1_cdnr_sh.max_column + 1): r1_cdnr_sh.cell(row=r_m,
                                                                                      column=c_idx).fill = MISSING_FILL
            for credit_r_m_val in unmatched_credit_s5.values():
                for c_idx in range(1, credit_r_sh.max_column + 1): credit_r_sh.cell(row=credit_r_m_val,
                                                                                    column=c_idx).fill = MISSING_FILL

            for note_key_d, diff_details in differences_s5.items():
                if note_key_d not in credit_r_map_s5: continue
                credit_r_inline_row = credit_r_map_s5[note_key_d]
                for f_key_cdnr, diff_val in diff_details.items():  # f_key_cdnr is cdnr_hdr_key or 'gstin'
                    diff_col_idx = credit_r_diff_indices_s5.get(f_key_cdnr)
                    if diff_col_idx:
                        cell = credit_r_sh.cell(row=credit_r_inline_row, column=diff_col_idx)
                        cell.value = diff_val
                        cell.font = Font(bold=True, color="00008B")
                        if f_key_cdnr != 'gstin':  # MODIFICATION: Apply format to numeric diffs
                            cell.number_format = INDIAN_NUM_FORMAT

            if s5_executed and credit_r_data_indices_s5.get('note_num'):
                add_totals_to_sheet(credit_r_sh, credit_r_data_indices_s5, credit_r_diff_indices_s5, 'note_num',
                                    'gstin', modified_sheets_for_autofit)  # MODIFICATION: Pass set
        else:
            if not (r1_cdnr_sh and credit_r_sh):
                messagebox.showwarning("Skipping Section 5", "Required sheets R1-CDNR or CREDIT-R not found.")
                send_event("recon_skipped_section", {"section": 5, "reason": "required_sheets_not_found"})

    # --- Autofit columns for all modified sheets ---
    # MODIFICATION: Call autofit for all sheets that were modified
    if any_reconciliation_performed:
        logging.info(f"Autofitting columns for {len(modified_sheets_for_autofit)} modified sheets.")
        for sheet_to_autofit in modified_sheets_for_autofit:
            if sheet_to_autofit:  # Ensure sheet object exists
                logging.info(f"Autofitting sheet: {sheet_to_autofit.title}")
                autofit_sheet_columns(sheet_to_autofit)

    # --- Save & Done ---
    if not any_reconciliation_performed:
        messagebox.showerror("No Reconciliation Performed",
                             "No reconciliation could be performed. Please check if the required sheets and headers are present in the workbook.")
        send_event("recon_completed", {"status": "no_recon_performed",
                                       "sections_executed": {"s1": s1_executed, "s2": s2_executed, "s3": s3_executed,
                                                             "s4": s4_executed, "s5": s5_executed}})
        return

    try:
        wb.save(outfile)
        messagebox.showinfo("Done", f"Reconciliation complete (if any sections ran) and saved to:\n{outfile}")
        send_event("recon_completed", {"status": "success", "output_file": os.path.basename(outfile),
                                       "sections_executed": {"s1": s1_executed, "s2": s2_executed, "s3": s3_executed,
                                                             "s4": s4_executed, "s5": s5_executed}})
    except PermissionError:
        messagebox.showerror("Save Error",
                             f"Could not save to '{outfile}'.\nPlease ensure the file is not open and you have write permissions.")
        send_event("recon_error", {"type": "save_permission_error", "filename": os.path.basename(outfile),
                                   "traceback": traceback.format_exc()})
    except Exception as e:
        messagebox.showerror("Save Error", f"An unexpected error occurred while saving:\n{e}")
        send_event("recon_error",
                   {"type": "save_unexpected_error", "filename": os.path.basename(outfile), "error_message": str(e),
                    "traceback": traceback.format_exc()})


if __name__ == '__main__':
    if not logging.getLogger().hasHandlers():
        log_dir_recon = os.path.join(os.path.expanduser("~"), "GSTProcessorLogs")
        os.makedirs(log_dir_recon, exist_ok=True)
        log_file_recon = os.path.join(log_dir_recon, "recon_script.log")
        try:
            logging.basicConfig(level=logging.INFO,
                                format='%(asctime)s - Recon - %(levelname)s - %(message)s',
                                filename=log_file_recon,
                                filemode='a')
        except Exception:
            logging.basicConfig(level=logging.INFO,
                                format='%(asctime)s - Recon - %(levelname)s - %(message)s')

    logging.info("Recon.py script started directly or called.")
    main()
    logging.info("Recon.py script finished.")

