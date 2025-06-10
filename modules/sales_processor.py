import tkinter as tk # Kept for messagebox, consider removing if UI fully decouples
from tkinter import messagebox # Used for error popups directly in processor
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import os
import logging
import math

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Sheet titles with desired order
SECTION_TITLES = [
    ("SALE-Total", "Sales Register - Total"),
    ("SALE-Total_sws", "Sales Register - Total - Receiver wise"),
    ("SALE-B2B", "Sales Register - B2B only"),
    ("SALE-B2C", "Sales Register - B2C only"),
    ("SALE-Others", "Sales Register - Others"),
    ("SALE-Summary-Total", "Sales Register - Total - Summary"),
    ("SALE-Summary-B2B", "Sales Register - B2B only - Summary"),
    ("SALE-Summary-B2C", "Sales Register - B2C only - Summary"),
    ("SALE-Summary-Others", "Sales Register - Others - Summary"),
]

# Define headers to exclude from total calculation (using standardized header names)
EXCLUDE_FROM_TOTAL_HEADERS = [
    'GSTIN/UIN of Recipient',
    'Receiver Name',
    'Branch',
    'Invoice Number',
    'Invoice Date',
    'Invoice Type',
    'Voucher Ref. No' # This will be an extra header if present
]

# Columns to exclude from Cr/Dr and format checks (using standardized header names)
# These are generally text or date columns.
EXCLUDE_HEADERS_FROM_CRDR_CHECK = [
    'GSTIN/UIN of Recipient',
    'Receiver Name',
    'Branch', # Branch is primarily text
    'Invoice Number',
    'Invoice Date', # Date is handled separately before safe_float_conversion
    'Invoice Type',
    'Voucher Ref. No', # This will be an extra header if present, typically text
    'Invoice Value' # Invoice Value is numeric but excluded from Cr/Dr logic, handled as direct numeric
]

# Number format for Indian numbering system
INDIAN_NUMBER_FORMAT = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"

# Headers that are strictly textual and should be an empty string if None/empty in source
# Branch is also textual but typically assigned directly from branch_key.
# Invoice Date is handled before this function.
STRICTLY_TEXTUAL_HEADERS = ['GSTIN/UIN of Recipient', 'Receiver Name', 'Invoice Number', 'Invoice Type', 'Voucher Ref. No', 'Branch']


def find_header_row(worksheet):
    logging.debug(f"Searching for header row in sheet: {worksheet.title}")
    for row in worksheet.iter_rows():
        # Case-insensitive check for 'Date' in the first cell
        if isinstance(row[0].value, str) and row[0].value.strip().lower() == "date":
            logging.debug(f"Header row found at row: {row[0].row}")
            return row[0].row
    logging.warning(f"Header row starting with 'Date' not found in sheet: {worksheet.title}")
    return None


def get_financial_year(date):
    if date is None:
        return None
    if date.month >= 4:
        return date.year
    return date.year - 1


def create_or_replace_sheet(wb, sheet_name, title_text, columns):
    logging.info(f"Creating or replacing sheet: {sheet_name}")
    if sheet_name in wb.sheetnames:
        logging.debug(f"Sheet '{sheet_name}' already exists, deleting.")
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1)
    cell.value = title_text
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, col in enumerate(columns, start=1):
        header_cell = ws.cell(row=2, column=idx, value=col)
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B3"
    logging.info(f"Finished creating sheet: {sheet_name}")
    return ws


def apply_format_and_autofit(ws, columns, start_row=3, col_format_map=None):
    logging.debug(f"Applying format and autofit for sheet: {ws.title}")
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        if col_format_map and col_name in col_format_map:
            for row_num in range(start_row, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = col_format_map[col_name]
        for row_num in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row_num, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = max(15, max_len + 2)
    logging.debug(f"Finished applying format and autofit for sheet: {ws.title}")


def add_total_row(ws, columns, start_row, end_row):
    logging.debug(f"Adding total row for sheet: {ws.title}")
    total_row_data = ['Total'] + [''] * (len(columns) - 1)

    for col_idx in range(1, len(columns) + 1):
        if col_idx == 1:
            continue

        header_name = columns[col_idx - 1] # This is the standardized header name
        if header_name in EXCLUDE_FROM_TOTAL_HEADERS:
            logging.debug(f"Excluding column '{header_name}' from total.")
            continue

        total = 0
        has_numeric_data = False
        try:
            for row in range(start_row, end_row + 1):
                value = ws.cell(row=row, column=col_idx).value
                if isinstance(value, (int, float)) and not (math.isnan(value) or math.isinf(value)):
                    total += float(value)
                    has_numeric_data = True
                elif value is not None and str(value).strip() != '':
                    logging.debug(f"Non-numeric value encountered in column '{header_name}' at row {row}: {value}")

            if has_numeric_data:
                total_row_data[col_idx - 1] = total
        except Exception as e:
            logging.error(f"Error totaling column '{header_name}': {e}")
            total_row_data[col_idx - 1] = "Error"

    row_num = end_row + 1
    ws.append(total_row_data)
    for col_idx, value in enumerate(total_row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = Font(bold=True, color="FF0000")
        if isinstance(value, (int, float)):
            cell.number_format = INDIAN_NUMBER_FORMAT # Use global Indian format for totals
    logging.debug(f"Finished adding total row for sheet: {ws.title}")


def safe_float_conversion(value, header, cell=None):
    # header is the standardized header name

    # Explicit handling for purely textual headers.
    # If value is None or empty string, it becomes ''. Otherwise, it's str(value).
    if header in STRICTLY_TEXTUAL_HEADERS:
        if value is None:
            return ''
        val_str = str(value).strip()
        return val_str # Return the stripped string (could be empty if original was spaces)

    # For all other headers, if value is None, it should become 0.0 for numeric processing
    if value is None:
        return 0.0

    # Handling for headers excluded from Cr/Dr logic but might be numeric (e.g., 'Invoice Value')
    # or other textual headers not in STRICTLY_TEXTUAL_HEADERS but in EXCLUDE_HEADERS_FROM_CRDR_CHECK
    if header in EXCLUDE_HEADERS_FROM_CRDR_CHECK:
        # This implies header is NOT in STRICTLY_TEXTUAL_HEADERS but IS in EXCLUDE_HEADERS_FROM_CRDR_CHECK
        # Example: 'Invoice Value'
        str_val_stripped = str(value).strip()
        if not str_val_stripped: # If effectively empty after stripping (e.g., "", "  ")
            return 0.0 # For numeric fields like 'Invoice Value', empty should be 0.0
        try:
            return float(value) # Try to convert to float
        except (ValueError, TypeError):
            # If it's a non-empty string that can't be float, return it as is.
            logging.debug(f"Could not convert value to float for excluded header '{header}': {value}. Returning as string.")
            return str_val_stripped # Return the original non-empty string

    # Default numeric conversion for fields NOT in EXCLUDE_HEADERS_FROM_CRDR_CHECK
    # (e.g., 'Taxable Value', 'Integrated Tax', 'Round Off', and other numeric extra headers)
    numeric_value = 0.0
    try:
        numeric_value = float(value) # Directly try to convert to float
    except (ValueError, TypeError):
        if isinstance(value, str):
            value_str = value.strip()
            if not value_str: # Empty string becomes 0.0
                numeric_value = 0.0
            elif value_str.endswith(' Cr'): # For Sales, Cr is positive
                try: numeric_value = float(value_str[:-3])
                except ValueError: numeric_value = 0.0
            elif value_str.endswith(' Dr'): # For Sales, Dr is negative
                try: numeric_value = -float(value_str[:-3])
                except ValueError: numeric_value = 0.0
            else: # Try to convert string directly if no Cr/Dr
                try: numeric_value = float(value_str)
                except ValueError: numeric_value = 0.0 # If still not convertible
        else:
            numeric_value = 0.0 # Non-string, non-float types become 0.0

    # Apply number format based logic (Dr/Cr in cell format)
    if cell and cell.number_format and isinstance(numeric_value, (int, float)):
        format_str = str(cell.number_format)
        if 'Dr' in format_str and numeric_value > 0: # Sales: Dr format & positive value -> make negative
            numeric_value *= -1
        elif 'Cr' in format_str and numeric_value < 0: # Sales: Cr format & negative value -> make positive
             numeric_value = abs(numeric_value)
    return numeric_value


def process_excel_data(input_files, template_file=None, existing_wb=None):
    logging.info("Starting sales data processing")
    all_data = []

    # Define the mapping from lowercase source header keys to the desired standardized header names (case-sensitive)
    source_key_to_standard_header = {
        'gstin/uin': 'GSTIN/UIN of Recipient',
        'particulars': 'Receiver Name',
        'voucher no': 'Invoice Number', # Covers "Voucher No"
        'voucher no.': 'Invoice Number',# Covers "Voucher No."
        'date': 'Invoice Date',
        'gross total': 'Invoice Value',
        'voucher type': 'Invoice Type',
        'value': 'Taxable Value', # Standard Tally export for taxable amount
        'taxable amount': 'Taxable Value', # Common alternative
        'igst': 'Integrated Tax',
        'cgst': 'Central Tax',
        'sgst': 'State/UT Tax',
        'cess': 'Cess',
        'round off': 'Round Off', # Covers "Round Off" and "ROUND OFF" due to lowercase key
        'voucher ref. no': 'Voucher Ref. No',
        'voucher ref. no.': 'Voucher Ref. No'
    }

    # Fixed headers that are expected in the output (case-sensitive)
    fixed_headers = [
        'GSTIN/UIN of Recipient',
        'Receiver Name',
        'Branch',
        'Invoice Number',
        'Invoice Date',
        'Invoice Type',
        'Invoice Value',
        'Taxable Value',
        'Integrated Tax',
        'Central Tax',
        'State/UT Tax',
        'Cess',
        'Round Off'
    ]
    fixed_headers_lower = {h.lower() for h in fixed_headers}
    extra_headers_set = set() # To collect unique extra headers (stores original casing)

    logging.debug("Starting file processing loop")
    for filepath, branch_key in input_files:
        logging.info(f"Processing sales file: {filepath} with branch_key: {branch_key}")
        try:
            wb = load_workbook(filepath, data_only=True)
            if len(wb.sheetnames) > 1:
                if "Sales Register" in wb.sheetnames:
                    ws = wb["Sales Register"]
                else:
                    logging.error(f"'Sales Register' sheet not found in {filepath}")
                    messagebox.showerror("Sheet Not Found", f"'Sales Register' sheet not found in {os.path.basename(filepath)}.\nPlease ensure the sheet name is correct.")
                    continue # Skip this file
            else:
                ws = wb.active
                if not ws:
                    logging.error(f"No active sheet in {filepath}")
                    messagebox.showerror("Sheet Not Found", f"No active sheet found in {os.path.basename(filepath)}.")
                    continue # Skip this file

            header_row_num = find_header_row(ws)
            if not header_row_num:
                logging.error(f"Header row not found in {filepath}")
                messagebox.showerror("Header Not Found", f"Header row starting with 'Date' not found in {os.path.basename(filepath)}.")
                continue # Skip this file

            original_headers_from_sheet = [cell.value for cell in ws[header_row_num] if cell.value is not None]
            original_headers_lower_map = {}
            original_cells_lower_map = {}
            for col_idx, header_val in enumerate(original_headers_from_sheet):
                if header_val is not None:
                    header_lower = str(header_val).lower()
                    original_headers_lower_map[header_lower] = header_val
                    original_cells_lower_map[header_lower] = ws.cell(row=header_row_num, column=col_idx + 1)

            logging.debug(f"Original headers from '{filepath}': {original_headers_from_sheet}")

            for original_header_val in original_headers_from_sheet:
                if original_header_val is not None:
                    header_l = str(original_header_val).lower()
                    is_mapped = header_l in source_key_to_standard_header
                    is_fixed = False
                    if is_mapped:
                        standard_h = source_key_to_standard_header[header_l]
                        if standard_h.lower() in fixed_headers_lower:
                            is_fixed = True
                    elif header_l in fixed_headers_lower:
                        is_fixed = True
                    if not is_mapped and not is_fixed:
                        extra_headers_set.add(original_header_val)

            logging.debug(f"Current extra_headers_set (original case): {extra_headers_set}")

            for row_idx, row_cells_tuple in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=False), start=header_row_num + 1):
                row_data_orig_values = {}
                current_row_cells_map = {}
                for i, cell_obj in enumerate(row_cells_tuple):
                    if i < len(original_headers_from_sheet):
                        header = original_headers_from_sheet[i]
                        row_data_orig_values[header] = cell_obj.value
                        current_row_cells_map[header] = cell_obj

                date_val = None
                original_date_header_key = None
                if 'date' in original_headers_lower_map:
                    original_date_header_key = original_headers_lower_map['date']
                    date_val = row_data_orig_values.get(original_date_header_key)

                particulars_val = None
                if 'particulars' in original_headers_lower_map:
                    original_particulars_header_key = original_headers_lower_map['particulars']
                    particulars_val = row_data_orig_values.get(original_particulars_header_key, '')

                if not date_val or (isinstance(particulars_val, str) and 'grand total' in particulars_val.lower()):
                    logging.debug(f"Skipping row {row_idx}: no Date or contains 'Grand Total'")
                    continue

                if isinstance(date_val, str):
                    try: date_val = datetime.datetime.strptime(date_val, '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        try: date_val = datetime.datetime.strptime(date_val.split()[0], '%d-%m-%Y')
                        except ValueError:
                            logging.debug(f"Skipping row {row_idx} due to date parsing error: {date_val}")
                            continue
                elif not isinstance(date_val, datetime.datetime):
                    logging.debug(f"Skipping row {row_idx} due to invalid date type: {type(date_val)}")
                    continue
                if original_date_header_key:
                    row_data_orig_values[original_date_header_key] = date_val

                processed_row_data = {}
                processed_row_data['Branch'] = branch_key

                for original_header, original_value in row_data_orig_values.items():
                    if original_header is None: continue
                    header_lower = str(original_header).lower()
                    cell_for_conversion = current_row_cells_map.get(original_header)
                    standard_header = None
                    if header_lower in source_key_to_standard_header:
                        standard_header = source_key_to_standard_header[header_lower]
                    elif header_lower in fixed_headers_lower:
                        standard_header = original_header
                    elif original_header in extra_headers_set:
                        standard_header = original_header

                    if standard_header:
                        if standard_header == 'Invoice Date':
                             processed_row_data[standard_header] = date_val
                        else:
                             processed_row_data[standard_header] = safe_float_conversion(original_value, standard_header, cell_for_conversion)

                for fh in fixed_headers:
                    if fh not in processed_row_data:
                        processed_row_data[fh] = ''
                for eh in extra_headers_set:
                    if eh not in processed_row_data:
                        original_eh_value = row_data_orig_values.get(eh)
                        if original_eh_value is not None:
                             cell_for_eh_conversion = current_row_cells_map.get(eh)
                             processed_row_data[eh] = safe_float_conversion(original_eh_value, eh, cell_for_eh_conversion)
                        else:
                             processed_row_data[eh] = ''
                all_data.append(processed_row_data)
        except Exception as e:
            logging.error(f"Error processing file {filepath}: {e}", exc_info=True)
            messagebox.showerror("File Processing Error", f"Error processing file {os.path.basename(filepath)}: {e}\n\nPlease check the logs for more details.")
        logging.info(f"Finished processing sales file: {filepath}")

    logging.debug("Finished file processing loop")
    extra_headers_list = sorted(list(extra_headers_set))
    final_headers = fixed_headers + extra_headers_list
    logging.info(f"Final headers for detail sheets: {final_headers}")

    data_by_type = {"B2B": [], "B2C": [], "Others": []}
    for row_data in all_data:
        inv_type = str(row_data.get('Invoice Type', '')).strip()
        if inv_type == "B2B": data_by_type["B2B"].append(row_data)
        elif inv_type == "B2C": data_by_type["B2C"].append(row_data)
        else:
            data_by_type["Others"].append(row_data)
            if inv_type not in ["B2B", "B2C", ""]:
                logging.debug(f"Row categorized as 'Others' due to Invoice Type: '{inv_type}' from Invoice: {row_data.get('Invoice Number')}")

    for key, data_list in data_by_type.items():
        logging.debug(f"Sorting data for category: {key}")
        data_list.sort(key=lambda x: (x.get('Invoice Date', datetime.datetime.min) if isinstance(x.get('Invoice Date'), datetime.datetime) else datetime.datetime.min,
                                      str(x.get('Invoice Number', ''))))
    all_data.sort(key=lambda x: (x.get('Invoice Date', datetime.datetime.min) if isinstance(x.get('Invoice Date'), datetime.datetime) else datetime.datetime.min,
                                 str(x.get('Invoice Number', ''))))
    def sort_key_sws(row):
        receiver = str(row.get('Receiver Name', '')).strip().lower()
        date = row.get('Invoice Date', datetime.datetime.min)
        inv_num = str(row.get('Invoice Number', ''))
        if receiver in ['cash', '(cancelled )'] or not receiver: return (1, receiver, date, inv_num)
        return (0, receiver, date, inv_num)
    all_data_sws = sorted(all_data, key=sort_key_sws)
    logging.debug("Completed data sorting")

    if existing_wb is not None: output_wb = existing_wb
    elif template_file: output_wb = load_workbook(template_file)
    else:
        output_wb = Workbook()
        if 'Sheet' in output_wb.sheetnames and len(output_wb.sheetnames) == 1: del output_wb['Sheet']

    col_format_map = {
        'Invoice Date': 'DD-MM-YYYY',
        'Invoice Value': INDIAN_NUMBER_FORMAT, 'Taxable Value': INDIAN_NUMBER_FORMAT,
        'Integrated Tax': INDIAN_NUMBER_FORMAT, 'Central Tax': INDIAN_NUMBER_FORMAT,
        'State/UT Tax': INDIAN_NUMBER_FORMAT, 'Cess': INDIAN_NUMBER_FORMAT, 'Round Off': INDIAN_NUMBER_FORMAT
    }
    for eh in extra_headers_list:
        if eh not in col_format_map: col_format_map[eh] = INDIAN_NUMBER_FORMAT

    sheets_to_create = [
        ("SALE-Total", all_data), ("SALE-Total_sws", all_data_sws),
        ("SALE-B2B", data_by_type.get("B2B", [])), ("SALE-B2C", data_by_type.get("B2C", [])),
        ("SALE-Others", data_by_type.get("Others", [])) ]

    for sheet_name_key, data_list_for_sheet in sheets_to_create:
        if not data_list_for_sheet:
            logging.info(f"No data for sheet: {sheet_name_key}, skipping creation.")
            continue
        logging.info(f"Starting population of sheet: {sheet_name_key}")
        display_title = next((t for n, t in SECTION_TITLES if n == sheet_name_key), sheet_name_key)
        ws = create_or_replace_sheet(output_wb, sheet_name_key, display_title, final_headers)
        current_start_row = 3
        for row_data_item in data_list_for_sheet:
            row_values_to_append = []
            for header_name in final_headers:
                value = row_data_item.get(header_name, '')
                if header_name == 'Invoice Date' and isinstance(value, datetime.datetime):
                    value = value.strftime('%d-%m-%Y')
                row_values_to_append.append(value)
            ws.append(row_values_to_append)
        add_total_row(ws, final_headers, current_start_row, ws.max_row)
        apply_format_and_autofit(ws, final_headers, col_format_map=col_format_map, start_row=current_start_row)
        logging.info(f"Completed population of sheet: {sheet_name_key}")

    summary_sheets_data_map = [
        ("SALE-Summary-Total", all_data), ("SALE-Summary-B2B", data_by_type.get("B2B", [])),
        ("SALE-Summary-B2C", data_by_type.get("B2C", [])), ("SALE-Summary-Others", data_by_type.get("Others", [])) ]
    summary_headers = ['Month', 'No. of Records', 'Invoice Value', 'Taxable Value', 'Integrated Tax',
                       'Central Tax', 'State/UT Tax', 'Cess']
    summary_col_format_map = {h: INDIAN_NUMBER_FORMAT for h in summary_headers if h not in ['Month', 'No. of Records']}
    months_order = ['April', 'May', 'June', 'July', 'August', 'September',
                    'October', 'November', 'December', 'January', 'February', 'March']

    for sheet_name_key, data_for_summary in summary_sheets_data_map:
        if not data_for_summary:
            logging.info(f"No data for summary sheet: {sheet_name_key}, skipping creation.")
            continue
        logging.info(f"Starting population of summary sheet: {sheet_name_key}")
        display_title = next((t for n, t in SECTION_TITLES if n == sheet_name_key), sheet_name_key)
        ws_summary = create_or_replace_sheet(output_wb, sheet_name_key, display_title, summary_headers)
        monthly_summary_data = {}
        unique_invoices_by_month = {}
        for row_item in data_for_summary:
            date_obj = row_item.get('Invoice Date')
            invoice_num_val = str(row_item.get('Invoice Number', '')).strip()
            if not isinstance(date_obj, datetime.datetime) or not invoice_num_val:
                logging.debug(f"Skipping summary calculation for row due to missing date/invoice: {row_item.get('Invoice Number')}")
                continue
            month_index = (date_obj.month - 4 + 12) % 12
            month_name = months_order[month_index]
            if month_name not in monthly_summary_data:
                monthly_summary_data[month_name] = {h: 0.0 for h in summary_headers if h not in ['Month']}
                monthly_summary_data[month_name]['count'] = 0
                unique_invoices_by_month[month_name] = set()
            if invoice_num_val not in unique_invoices_by_month[month_name]:
                monthly_summary_data[month_name]['count'] += 1
                unique_invoices_by_month[month_name].add(invoice_num_val)
            fields_to_sum_in_summary = ['Invoice Value', 'Taxable Value', 'Integrated Tax', 'Central Tax', 'State/UT Tax', 'Cess']
            for field_key in fields_to_sum_in_summary:
                value_to_add = row_item.get(field_key, 0.0)
                if isinstance(value_to_add, (int, float)) and not (math.isnan(value_to_add) or math.isinf(value_to_add)):
                    monthly_summary_data[month_name][field_key] += float(value_to_add)
                elif value_to_add is not None and str(value_to_add).strip() not in ['', '0', '0.0']:
                     logging.warning(f"Non-numeric or invalid value '{value_to_add}' for '{field_key}' in summary for invoice '{invoice_num_val}', treating as 0.")

        summary_start_row = 3
        rows_added_to_summary = 0
        for month_n in months_order:
            if month_n in monthly_summary_data:
                month_data = monthly_summary_data[month_n]
                ws_summary.append([ month_n, month_data['count'], month_data['Invoice Value'], month_data['Taxable Value'],
                    month_data['Integrated Tax'], month_data['Central Tax'], month_data['State/UT Tax'], month_data['Cess'] ])
                rows_added_to_summary += 1
        if rows_added_to_summary > 0:
            summary_total_row_values = ['Total'] + [0.0] * (len(summary_headers) - 1)
            for col_idx_summary in range(1, len(summary_headers)):
                current_col_total = 0.0
                for r_idx in range(summary_start_row, summary_start_row + rows_added_to_summary):
                    cell_val = ws_summary.cell(row=r_idx, column=col_idx_summary + 1).value
                    if isinstance(cell_val, (int, float)) and not (math.isnan(cell_val) or math.isinf(cell_val)):
                        current_col_total += float(cell_val)
                summary_total_row_values[col_idx_summary] = current_col_total
            ws_summary.append(summary_total_row_values)
            total_row_number_on_sheet = summary_start_row + rows_added_to_summary
            for c_idx, val in enumerate(summary_total_row_values, start=1):
                total_cell = ws_summary.cell(row=total_row_number_on_sheet, column=c_idx)
                total_cell.font = Font(bold=True, color="FF0000")
                if isinstance(val, (int, float)) and summary_headers[c_idx - 1] not in ['Month', 'No. of Records']:
                    total_cell.number_format = INDIAN_NUMBER_FORMAT
        apply_format_and_autofit(ws_summary, summary_headers, col_format_map=summary_col_format_map, start_row=summary_start_row)
        logging.info(f"Completed population of summary sheet: {sheet_name_key}")

    logging.info("Completed sales data processing")
    return output_wb