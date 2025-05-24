import json
import os
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from collections import defaultdict  # For _add_total_row_to_gstr3b_sheet
import datetime  # Added for datetime operations if any (though not directly used in this snippet, good practice)
import traceback  # For detailed error reporting (though not used in this snippet, good practice for main script)

# ----------------------- Global Variables ----------------------- #
INDIAN_FORMAT_GSTR3B = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-##\,##\,##\,##0.00;-"
RED_BOLD_FONT = Font(color="FF0000", bold=True)
BOLD_FONT = Font(bold=True)


# ----------------------- Utility Functions ----------------------- #
def get_tax_period(ret_period):
    # print(f"[DEBUG] Parsing tax period from ret_period: {ret_period}")
    month_map = {
        "01": "January", "02": "February", "03": "March", "04": "April",
        "05": "May", "06": "June", "07": "July", "08": "August",
        "09": "September", "10": "October", "11": "November", "12": "December"
    }
    if ret_period and len(ret_period) >= 2:
        result = month_map.get(ret_period[:2], "Unknown")
        # print(f"[DEBUG] Tax period parsed: {result}")
        return result
    # print("[DEBUG] Tax period parsing failed, returning 'Unknown'")
    return "Unknown"


def get_tax_period_from_date(trandate):
    # print(f"[DEBUG] Parsing tax period from date: {trandate}")
    try:
        parts = trandate.split("-")
        if len(parts) == 3:
            month = parts[1]
            month_map = {
                "01": "January", "02": "February", "03": "March", "04": "April",
                "05": "May", "06": "June",
                "07": "July", "08": "August", "09": "September",
                "10": "October", "11": "November", "12": "December"
            }
            result = month_map.get(month, "Unknown")
            # print(f"[DEBUG] Tax period from date parsed: {result}")
            return result
    except Exception as e:
        # print(f"[DEBUG] Error parsing tax period from date: {e}")
        pass
    # print("[DEBUG] Tax period from date parsing failed, returning 'Unknown'")
    return "Unknown"


def parse_number(val, float_2dec=False, int_no_dec=False):
    # print(f"[DEBUG] Parsing number: value={val}, float_2dec={float_2dec}, int_no_dec={int_no_dec}")
    try:
        num = float(val)
        if int_no_dec:
            result = int(num)
            # print(f"[DEBUG] Parsed as integer: {result}")
            return result
        if float_2dec:
            result = round(num, 2)
            # print(f"[DEBUG] Parsed as float with 2 decimals: {result}")
            return result
        # print(f"[DEBUG] Parsed as float: {num}")
        return num
    except (ValueError, TypeError):
        # print(f"[DEBUG] Number parsing failed, returning 0: {e}")
        return 0


def get_numeric_value(item, key):
    # print(f"[DEBUG] Getting numeric value for key: {key}")
    val = item.get(key, 0)
    if isinstance(val, str):
        val = val.strip()
        # print(f"[DEBUG] Stripped string value: {val}")
    result = parse_number(val, float_2dec=True)
    # print(f"[DEBUG] Numeric value for {key}: {result}")
    return result


# ----------------------- Extraction Functions ----------------------- #
def extract_section_3_1(data):
    # print("[DEBUG] Extracting section 3.1 data...")
    r3b_data = data.get("data", {}).get("r3b", {})
    sup_details = r3b_data.get("sup_details", {})
    ret_period = r3b_data.get("ret_period", "")
    tax_period = get_tax_period(ret_period)
    section_mapping = {
        "OSUP-Detail": "osup_det", "OSUP-Zero": "osup_zero",
        "OSUP-Nil,Exmp": "osup_nil_exmp", "ISUP-Rev": "isup_rev",
        "OSUP-NonGST": "osup_nongst",
    }
    extracted_data = {key: [] for key in section_mapping}
    for sheet_name, json_key in section_mapping.items():
        section_data = sup_details.get(json_key, {})
        if not isinstance(section_data, dict): section_data = {}
        row = {
            "Tax Period": tax_period,
            "Total taxable value": get_numeric_value(section_data, "txval"),
            "Integrated Tax": get_numeric_value(section_data, "iamt"),
            "Central Tax": get_numeric_value(section_data, "camt"),
            "State/UT Tax": get_numeric_value(section_data, "samt"),
            "Cess": get_numeric_value(section_data, "csamt")
        }
        extracted_data[sheet_name].append(row)
    for sheet_name in section_mapping:
        if not extracted_data[sheet_name]:
            extracted_data[sheet_name].append({
                "Tax Period": tax_period, "Total taxable value": 0.00, "Integrated Tax": 0.00,
                "Central Tax": 0.00, "State/UT Tax": 0.00, "Cess": 0.00
            })
    return extracted_data


def extract_section_3_2(data):
    # print("[DEBUG] Extracting section 3.2 data...")
    r3b_data = data.get("data", {}).get("r3b", {})
    inter_sup = r3b_data.get("inter_sup", {})
    ret_period = r3b_data.get("ret_period", "")
    tax_period = get_tax_period(ret_period)
    section_mapping = {
        "InterSUP-Unreg": "unreg_details", "InterSUP-Comp": "comp_details", "InterSUP-UIN": "uin_details",
    }
    extracted_data = {key: [] for key in section_mapping}
    for sheet_name, json_key in section_mapping.items():
        section_data = inter_sup.get(json_key, [])
        if not isinstance(section_data, list):
            section_data = [section_data] if isinstance(section_data, dict) else []
        for item in section_data:
            if not isinstance(item, dict): continue
            row = {
                "Tax Period": tax_period,
                "Total taxable value": get_numeric_value(item, "txval"),
                "Integrated Tax": get_numeric_value(item, "iamt")
            }
            extracted_data[sheet_name].append(row)
    for sheet_name in section_mapping:
        if not extracted_data[sheet_name]:
            extracted_data[sheet_name].append({
                "Tax Period": tax_period, "Total taxable value": 0.00, "Integrated Tax": 0.00
            })
    return extracted_data


def extract_section_4(data):
    # print("[DEBUG] Extracting section 4 data...")
    r3b_data = data.get("data", {}).get("r3b", {})
    itc_elg = r3b_data.get("itc_elg", {})
    ret_period = r3b_data.get("ret_period", "")
    tax_period = get_tax_period(ret_period)
    extracted_data = {
        "ITC-Available": [], "ITC-avl-IMPG": [], "ITC-avl-IMPS": [], "ITC-avl-ISRC": [],
        "ITC-avl-ISD": [], "ITC-avl-OTH": [], "ITC-Reversed": [], "ITC-rev-RUL": [],
        "ITC-rev-OTH": [], "Net-ITC": [], "ITC-Ineligible": [], "ITC-inelg-RUL": [], "ITC-inelg-OTH": [],
    }
    itc_avl = itc_elg.get("itc_avl", [])
    if not isinstance(itc_avl, list): itc_avl = [itc_avl] if isinstance(itc_avl, dict) else []
    summary_avl = {}
    for item in itc_avl:
        if not isinstance(item, dict): continue
        sub_type = str(item.get("ty", "")).strip()
        row = {"Tax Period": tax_period, "Integrated Tax": get_numeric_value(item, "iamt"),
               "Central Tax": get_numeric_value(item, "camt"), "State/UT Tax": get_numeric_value(item, "samt"),
               "Cess": get_numeric_value(item, "csamt")}
        key = f"ITC-avl-{sub_type}"
        if key in extracted_data: extracted_data[key].append(row)
        if tax_period not in summary_avl: summary_avl[tax_period] = {"Tax Period": tax_period, "Integrated Tax": 0,
                                                                     "Central Tax": 0, "State/UT Tax": 0, "Cess": 0}
        summary_avl[tax_period]["Integrated Tax"] += row["Integrated Tax"];
        summary_avl[tax_period]["Central Tax"] += row["Central Tax"];
        summary_avl[tax_period]["State/UT Tax"] += row["State/UT Tax"];
        summary_avl[tax_period]["Cess"] += row["Cess"]
    for sp in summary_avl.values(): extracted_data["ITC-Available"].append(sp)

    itc_rev = itc_elg.get("itc_rev", [])
    if not isinstance(itc_rev, list): itc_rev = [itc_rev] if isinstance(itc_rev, dict) else []
    summary_rev = {}
    for item in itc_rev:
        if not isinstance(item, dict): continue
        sub_type = str(item.get("ty", "")).strip()
        row = {"Tax Period": tax_period, "Integrated Tax": get_numeric_value(item, "iamt"),
               "Central Tax": get_numeric_value(item, "camt"), "State/UT Tax": get_numeric_value(item, "samt"),
               "Cess": get_numeric_value(item, "csamt")}
        key = f"ITC-rev-{sub_type}"
        if key in extracted_data: extracted_data[key].append(row)
        if tax_period not in summary_rev: summary_rev[tax_period] = {"Tax Period": tax_period, "Integrated Tax": 0,
                                                                     "Central Tax": 0, "State/UT Tax": 0, "Cess": 0}
        summary_rev[tax_period]["Integrated Tax"] += row["Integrated Tax"];
        summary_rev[tax_period]["Central Tax"] += row["Central Tax"];
        summary_rev[tax_period]["State/UT Tax"] += row["State/UT Tax"];
        summary_rev[tax_period]["Cess"] += row["Cess"]
    for sp in summary_rev.values(): extracted_data["ITC-Reversed"].append(sp)

    itc_net = itc_elg.get("itc_net", {})
    if not isinstance(itc_net, dict): itc_net = {}
    extracted_data["Net-ITC"].append({"Tax Period": tax_period, "Integrated Tax": get_numeric_value(itc_net, "iamt"),
                                      "Central Tax": get_numeric_value(itc_net, "camt"),
                                      "State/UT Tax": get_numeric_value(itc_net, "samt"),
                                      "Cess": get_numeric_value(itc_net, "csamt")})

    itc_inelg = itc_elg.get("itc_inelg", [])
    if not isinstance(itc_inelg, list): itc_inelg = [itc_inelg] if isinstance(itc_inelg, dict) else []
    summary_inelg = {}
    for item in itc_inelg:
        if not isinstance(item, dict): continue
        sub_type = str(item.get("ty", "")).strip()
        row = {"Tax Period": tax_period, "Integrated Tax": get_numeric_value(item, "iamt"),
               "Central Tax": get_numeric_value(item, "camt"), "State/UT Tax": get_numeric_value(item, "samt"),
               "Cess": get_numeric_value(item, "csamt")}
        key = f"ITC-inelg-{sub_type}"
        if key in extracted_data: extracted_data[key].append(row)
        if tax_period not in summary_inelg: summary_inelg[tax_period] = {"Tax Period": tax_period, "Integrated Tax": 0,
                                                                         "Central Tax": 0, "State/UT Tax": 0, "Cess": 0}
        summary_inelg[tax_period]["Integrated Tax"] += row["Integrated Tax"];
        summary_inelg[tax_period]["Central Tax"] += row["Central Tax"];
        summary_inelg[tax_period]["State/UT Tax"] += row["State/UT Tax"];
        summary_inelg[tax_period]["Cess"] += row["Cess"]
    for sp in summary_inelg.values(): extracted_data["ITC-Ineligible"].append(sp)

    for sheet_name_val in extracted_data:  # Renamed sheet_name to sheet_name_val
        if not extracted_data[sheet_name_val]:
            extracted_data[sheet_name_val].append(
                {"Tax Period": tax_period, "Integrated Tax": 0.00, "Central Tax": 0.00, "State/UT Tax": 0.00,
                 "Cess": 0.00})
    return extracted_data


def extract_section_5_1(data):
    # print("[DEBUG] Extracting section 5.1 data...")
    r3b_data = data.get("data", {}).get("r3b", {})
    intr_ltfee = r3b_data.get("intr_ltfee", {})
    ret_period = r3b_data.get("ret_period", "")
    tax_period = get_tax_period(ret_period)
    extracted_data = {"INTR-paid": [], "Late-fee": []}
    intr_details = intr_ltfee.get("intr_details", {})
    if isinstance(intr_details, dict):
        intr_details = [intr_details]
    elif not isinstance(intr_details, list):
        intr_details = []
    for item in intr_details:
        if not isinstance(item, dict): continue
        extracted_data["INTR-paid"].append({"Tax Period": tax_period, "Integrated Tax": get_numeric_value(item, "iamt"),
                                            "Central Tax": get_numeric_value(item, "camt"),
                                            "State/UT Tax": get_numeric_value(item, "samt"),
                                            "Cess": get_numeric_value(item, "csamt")})
    ltfee_details = intr_ltfee.get("ltfee_details", {})
    if isinstance(ltfee_details, dict):
        ltfee_details = [ltfee_details]
    elif not isinstance(ltfee_details, list):
        ltfee_details = []
    for item in ltfee_details:
        if not isinstance(item, dict): continue
        extracted_data["Late-fee"].append({"Tax Period": tax_period, "Integrated Tax": get_numeric_value(item, "iamt"),
                                           "Central Tax": get_numeric_value(item, "camt"),
                                           "State/UT Tax": get_numeric_value(item, "samt"),
                                           "Cess": get_numeric_value(item, "csamt")})
    for sheet_name_val in extracted_data:  # Renamed sheet_name
        if not extracted_data[sheet_name_val]:
            extracted_data[sheet_name_val].append(
                {"Tax Period": tax_period, "Integrated Tax": 0.00, "Central Tax": 0.00, "State/UT Tax": 0.00,
                 "Cess": 0.00})
    return extracted_data


def extract_section_6(data):
    # print("[DEBUG] Extracting section 6 data...")
    r3b_data = data.get("data", {}).get("r3b", {})
    tt_val = r3b_data.get("tt_val", {})
    ret_period = r3b_data.get("ret_period", "")
    tax_period = get_tax_period(ret_period)
    extracted_data = {"Tax-Pay": []}
    row = {"Tax Period": tax_period, "Tax-by-ITC": get_numeric_value(tt_val, "tt_itc_pd"),
           "Tax-by-Cash": get_numeric_value(tt_val, "tt_csh_pd")}
    extracted_data["Tax-Pay"].append(row)
    if not extracted_data["Tax-Pay"] or all(val == 0.00 for key, val in row.items() if key != "Tax Period"):
        extracted_data["Tax-Pay"] = [{"Tax Period": tax_period, "Tax-by-ITC": 0.00, "Tax-by-Cash": 0.00}]
    return extracted_data


def extract_section_6_1(data):
    # print("[DEBUG] Extracting section 6.1 data...")
    tax_data = data.get("taxpayable", {}).get("data", {}).get("returnsDbCdredList", {})
    ret_period = data.get("data", {}).get("r3b", {}).get("ret_period", "")
    tax_period = get_tax_period(ret_period)
    extracted_data = {
        "6.1a1": [], "6.1a2": [], "6.1a3": [], "6.1a5": [], "6.1a6": [], "6.1a7": [],
        "6.1a41": [], "6.1a42": [], "6.1a43": [], "6.1a44": [],
        "6.1b1": [], "6.1b2": [], "6.1b3": [], "6.1b5": [], "6.1b6": [], "6.1b7": [],
        "6.1b41": [], "6.1b42": [], "6.1b43": [], "6.1b44": []
    }
    if not tax_data:
        default_row = {"Tax Period": tax_period, "Integrated Tax": 0.00, "Central Tax": 0.00, "State/UT Tax": 0.00,
                       "Cess": 0.00}
        for key in extracted_data: extracted_data[key].append(default_row)
    else:
        def extract_tax_fields(item, field):
            return {"Tax Period": tax_period, "Integrated Tax": get_numeric_value(item.get("igst", {}), field),
                    "Central Tax": get_numeric_value(item.get("cgst", {}), field),
                    "State/UT Tax": get_numeric_value(item.get("sgst", {}), field),
                    "Cess": get_numeric_value(item.get("cess", {}), field)}

        # Corrected loop: unpack 4 items
        for section_list_key, trancd_expected, target_sheet_keys_list, json_field_names_list in [
            ("tax_pay", 30002, ["6.1a1", "6.1a6", "6.1a7"], ["tx", "intr", "fee"]),
            ("tax_pay", 30003, ["6.1b1", "6.1b6", "6.1b7"], ["tx", "intr", "fee"]),
            ("tax_paid.pd_by_nls", 30002, ["6.1a2"], ["tx"]),
            ("tax_paid.pd_by_nls", 30003, ["6.1b2"], ["tx"]),
            ("net_tax_pay", 30002, ["6.1a3"], ["tx"]),
            ("net_tax_pay", 30003, ["6.1b3"], ["tx"]),
            ("tax_paid.pd_by_cash", 30002, ["6.1a5"], ["tx"]),
            ("tax_paid.pd_by_cash", 30003, ["6.1b5"], ["tx"])
        ]:
            current_data_list = tax_data
            for part in section_list_key.split('.'):
                current_data_list = current_data_list.get(part, [])  # Use [] as default for .get()
            if not isinstance(current_data_list, list):
                current_data_list = [current_data_list] if isinstance(current_data_list, dict) else []

            for item in current_data_list:
                if not isinstance(item, dict): continue
                if item.get("trancd") == trancd_expected:
                    for i, sheet_key_61 in enumerate(target_sheet_keys_list):
                        # Use the corresponding json_field_name
                        json_field_to_extract = json_field_names_list[i]
                        extracted_data[sheet_key_61].append(extract_tax_fields(item, json_field_to_extract))

        pd_by_itc = tax_data.get("tax_paid", {}).get("pd_by_itc", [])
        if not isinstance(pd_by_itc, list): pd_by_itc = [pd_by_itc] if isinstance(pd_by_itc, dict) else []
        for item in pd_by_itc:
            if not isinstance(item, dict): continue
            trancd = item.get("trancd")
            if trancd not in [30002, 30003]: continue
            rows_itc = {
                pfx: {"Tax Period": tax_period, "Integrated Tax": 0, "Central Tax": 0, "State/UT Tax": 0, "Cess": 0} for
                pfx in ["igst", "cgst", "sgst", "cess"]}
            for key, val in item.items():
                if key in ["debit_id", "liab_id", "trancd", "trandate"]: continue
                parts = key.split("_");
                value = parse_number(val, float_2dec=True)
                if len(parts) == 3 and parts[2] == "amt":
                    prefix, middle, _ = parts
                    if middle.lower() in rows_itc and prefix.lower() in ["igst", "cgst", "sgst", "cess"]:
                        tax_map = {"igst": "Integrated Tax", "cgst": "Central Tax", "sgst": "State/UT Tax",
                                   "cess": "Cess"}
                        rows_itc[middle.lower()][tax_map[prefix.lower()]] += value
            prefix_61 = "6.1a4" if trancd == 30002 else "6.1b4"
            extracted_data[prefix_61 + "1"].append(rows_itc["igst"])
            extracted_data[prefix_61 + "2"].append(rows_itc["cgst"])
            extracted_data[prefix_61 + "3"].append(rows_itc["sgst"])
            extracted_data[prefix_61 + "4"].append(rows_itc["cess"])

        default_row = {"Tax Period": tax_period, "Integrated Tax": 0.00, "Central Tax": 0.00, "State/UT Tax": 0.00,
                       "Cess": 0.00}
        for key in extracted_data:
            if not extracted_data[key]: extracted_data[key].append(default_row)
    return extracted_data


# ----------------------- Excel Report Generation ----------------------- #

def _add_total_row_to_gstr3b_sheet(ws, rows_data, column_headers_list, column_formats_map):
    """Adds a formatted total row to a GSTR-3B sheet."""
    if not rows_data:
        return

    column_totals = defaultdict(float)
    tax_period_column_name = "Tax Period"

    for data_row_dict in rows_data:
        for header_name in column_headers_list:
            if header_name == tax_period_column_name:
                continue

            value = data_row_dict.get(header_name)
            if isinstance(value, (int, float)):
                column_totals[header_name] += value

    total_row_idx = ws.max_row + 1

    label_col_idx = 1
    try:
        label_col_idx = column_headers_list.index(tax_period_column_name) + 1
    except ValueError:
        print(
            f"[WARN] '{tax_period_column_name}' not found in headers for sheet {ws.title}. Using col 1 for 'Total' label.")

    for c_idx, header_name in enumerate(column_headers_list, 1):
        cell = ws.cell(row=total_row_idx, column=c_idx)
        if c_idx == label_col_idx:
            cell.value = "Total"
            cell.font = RED_BOLD_FONT
        else:
            total_value = column_totals.get(header_name, 0.0)  # Default to 0.0 if not summed

            if isinstance(total_value, float):
                total_value = round(total_value, 2)

            cell.value = total_value
            cell.font = RED_BOLD_FONT

            current_col_formats_map = column_formats_map if column_formats_map else {}
            if header_name in current_col_formats_map:
                format_str = current_col_formats_map[header_name]
                if isinstance(total_value, (int, float)):
                    cell.number_format = format_str
            # else: # If no specific format, Excel will use general or previous cell's format
            #     pass


def create_excel_report(data_dict, save_path, template_path=None):
    print("[DEBUG] Starting Excel report creation...")
    if template_path and os.path.exists(template_path):
        print(f"[DEBUG] Loading template from {template_path}")
        wb = load_workbook(template_path)
    else:
        print("[DEBUG] Creating new workbook")
        wb = Workbook()
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:  # Check if only default sheet exists
            wb.remove(wb["Sheet"])
            print("[DEBUG] Removed default 'Sheet'")

    section_titles_map = {  # Renamed section_titles to section_titles_map
        "OSUP-Detail": "3.1A - Outward taxable supplies (other than zero rated, nil rated and exempted)",
        "OSUP-Zero": "3.1B - Outward taxable supplies (zero rated)",
        "OSUP-Nil,Exmp": "3.1C - Other outward supplies (nil rated, exempted)",
        "ISUP-Rev": "3.1D - Inward supplies (liable to reverse charge)",
        "OSUP-NonGST": "3.1E - Non-GST outward supplies",
        "InterSUP-Unreg": "3.2A - Inter state - supplies made to Unregistered Persons",
        "InterSUP-Comp": "3.2B - Inter state - Supplies made to Composition Taxable Persons",
        "InterSUP-UIN": "3.2C - Inter state - Supplies made to UIN holders",
        "ITC-Available": "4A - ITC Available (whether in full or part)",
        "ITC-avl-IMPG": "4A1 - ITC Available_Import of goods",
        "ITC-avl-IMPS": "4A2 - ITC Available_Import of services",
        "ITC-avl-ISRC": "4A3 - ITC Available_Inward supplies liable to reverse charge (other than 1 & 2 above)",
        "ITC-avl-ISD": "4A4 - ITC Available_Inward supplies from ISD",
        "ITC-avl-OTH": "4A5 - ITC Available_All other ITC",
        "ITC-Reversed": "4B - ITC Reversed",
        "ITC-rev-RUL": "4B1 - ITC Reversed_As per rules 38,42 & 43 of CGST Rules and section17(5)",
        "ITC-rev-OTH": "4B2 - ITC Reversed_Others",
        "Net-ITC": "4C - Net ITC Available (A-B)",
        "ITC-Ineligible": "4D - Other Details",
        "ITC-inelg-RUL": "4D1 - Other Details_ITC reclaimed which was reversed under Table 4B2 in earlier tax period",
        "ITC-inelg-OTH": "4D2 - Other Details_Ineligible ITC under section 16(4) & ITC restricted due to PoS rules",
        "INTR-paid": "5.1A - Interest Paid",
        "Late-fee": "5.1B - Late fee",
        "Tax-Pay": "6 - Payment of Tax",
        "6.1a1": "6.1A1 - Payment of Tax - Total Tax Payable (Other than reverse charge)",
        "6.1a2": "6.1A2 - Payment of Tax - Adjustment of Negative Liability (Other than reverse charge)",
        "6.1a3": "6.1A3 - Payment of Tax - Net Tax Payable (Other than reverse charge)",
        "6.1a5": "6.1A5 - Payment of Tax - Tax Paid in Cash (Other than reverse charge)",
        "6.1a6": "6.1A6 - Payment of Tax - Interest Paid by Cash (Other than reverse charge)",
        "6.1a7": "6.1A7 - Payment of Tax - Late Fee Paid by Cash (Other than reverse charge)",
        "6.1a41": "6.1A41 - Payment of Tax - Tax Paid by ITC - IGST (Other than reverse charge)",
        "6.1a42": "6.1A42 - Payment of Tax - Tax Paid by ITC - CGST (Other than reverse charge)",
        "6.1a43": "6.1A43 - Payment of Tax - Tax Paid by ITC - SGST (Other than reverse charge)",
        "6.1a44": "6.1A44 - Payment of Tax - Tax Paid by ITC - Cess (Other than reverse charge)",
        "6.1b1": "6.1B1 - Payment of Tax - Total Tax Payable (Reverse charge)",
        "6.1b2": "6.1B2 - Payment of Tax - Adjustment of Negative Liability (Reverse charge)",
        "6.1b3": "6.1B3 - Payment of Tax - Net Tax Payable (Reverse charge)",
        "6.1b5": "6.1B5 - Payment of Tax - Tax Paid in Cash (Reverse charge)",
        "6.1b6": "6.1B6 - Payment of Tax - Interest Paid by Cash (Reverse charge)",
        "6.1b7": "6.1B7 - Payment of Tax - Late Fee Paid by Cash (Reverse charge)",
        "6.1b41": "6.1B41 - Payment of Tax - Tax Paid by ITC - IGST (Reverse charge)",
        "6.1b42": "6.1B42 - Payment of Tax - Tax Paid by ITC - CGST (Reverse charge)",
        "6.1b43": "6.1B43 - Payment of Tax - Tax Paid by ITC - SGST (Reverse charge)",
        "6.1b44": "6.1B44 - Payment of Tax - Tax Paid by ITC - Cess (Reverse charge)",
    }

    sheet_names_map = {  # Renamed sheet_names to sheet_names_map
        "OSUP-Detail": "3B-OSUP-Detail", "OSUP-Zero": "3B-OSUP-Zero", "OSUP-Nil,Exmp": "3B-OSUP-Nil,Exmp",
        "ISUP-Rev": "3B-ISUP-Rev", "OSUP-NonGST": "3B-OSUP-NonGST", "InterSUP-Unreg": "3B-InterSUP-Unreg",
        "InterSUP-Comp": "3B-InterSUP-Comp", "InterSUP-UIN": "3B-InterSUP-UIN", "ITC-Available": "3B-ITC-Available",
        "ITC-avl-IMPG": "3B-ITC-avl-IMPG", "ITC-avl-IMPS": "3B-ITC-avl-IMPS", "ITC-avl-ISRC": "3B-ITC-avl-ISRC",
        "ITC-avl-ISD": "3B-ITC-avl-ISD", "ITC-avl-OTH": "3B-ITC-avl-OTH", "ITC-Reversed": "3B-ITC-Reversed",
        "ITC-rev-RUL": "3B-ITC-rev-RUL", "ITC-rev-OTH": "3B-ITC-rev-OTH", "Net-ITC": "3B-Net-ITC",
        "ITC-Ineligible": "3B-ITC-Ineligible", "ITC-inelg-RUL": "3B-ITC-inelg-RUL", "ITC-inelg-OTH": "3B-ITC-inelg-OTH",
        "INTR-paid": "3B-INTR-paid", "Late-fee": "3B-Late-fee", "Tax-Pay": "3B-Tax-Pay",
        "6.1a1": "3B-TaxPay_TotTax-OthRC", "6.1a2": "3B-TaxPay_AdjNL-OthRC", "6.1a3": "3B-TaxPay_NetTax-OthRC",
        "6.1a5": "3B-TaxPay_pdby_Cash-OthRC", "6.1a6": "3B-TaxPay_Int_pdby_Cash-OthRC",
        "6.1a7": "3B-TaxPay_LateFee_pdby_Cash-OthRC",
        "6.1a41": "3B-TaxPay_ITC_IGST-OthRC", "6.1a42": "3B-TaxPay_ITC_CGST-OthRC",
        "6.1a43": "3B-TaxPay_ITC_SGST-OthRC",
        "6.1a44": "3B-TaxPay_ITC_Cess-OthRC", "6.1b1": "3B-TaxPay_TotTax-RC", "6.1b2": "3B-TaxPay_AdjNL-RC",
        "6.1b3": "3B-TaxPay_NetTax-RC", "6.1b5": "3B-TaxPay_pdby_Cash-RC", "6.1b6": "3B-TaxPay_Int_pdby_Cash-RC",
        "6.1b7": "3B-TaxPay_LateFee_pdby_Cash-RC", "6.1b41": "3B-TaxPay_ITC_IGST-RC", "6.1b42": "3B-TaxPay_ITC_CGST-RC",
        "6.1b43": "3B-TaxPay_ITC_SGST-RC", "6.1b44": "3B-TaxPay_ITC_Cess-RC",
    }

    column_headers_map = {  # Renamed column_headers to column_headers_map
        "OSUP-Detail": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "OSUP-Zero": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "OSUP-Nil,Exmp": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ISUP-Rev": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "OSUP-NonGST": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "InterSUP-Unreg": ["Tax Period", "Total taxable value", "Integrated Tax"],
        "InterSUP-Comp": ["Tax Period", "Total taxable value", "Integrated Tax"],
        "InterSUP-UIN": ["Tax Period", "Total taxable value", "Integrated Tax"],
        "ITC-Available": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-avl-IMPG": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-avl-IMPS": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-avl-ISRC": ["Tax Period", "Total taxable value", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-avl-ISD": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-avl-OTH": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-Reversed": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-rev-RUL": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-rev-OTH": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "Net-ITC": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-Ineligible": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-inelg-RUL": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "ITC-inelg-OTH": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "INTR-paid": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "Late-fee": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "Tax-Pay": ["Tax Period", "Tax-by-ITC", "Tax-by-Cash"],
        "6.1a1": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a2": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a3": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a5": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a6": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a7": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a41": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a42": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a43": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1a44": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b1": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b2": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b3": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b5": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b6": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b7": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b41": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b42": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b43": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
        "6.1b44": ["Tax Period", "Integrated Tax", "Central Tax", "State/UT Tax", "Cess"],
    }

    column_formats_map = {  # Renamed column_formats to column_formats_map
        "OSUP-Detail": {"Tax Period": "General", "Total taxable value": INDIAN_FORMAT_GSTR3B,
                        "Integrated Tax": INDIAN_FORMAT_GSTR3B, "Central Tax": INDIAN_FORMAT_GSTR3B,
                        "State/UT Tax": INDIAN_FORMAT_GSTR3B, "Cess": INDIAN_FORMAT_GSTR3B},
        "OSUP-Zero": {"Tax Period": "General", "Total taxable value": INDIAN_FORMAT_GSTR3B,
                      "Integrated Tax": INDIAN_FORMAT_GSTR3B, "Central Tax": INDIAN_FORMAT_GSTR3B,
                      "State/UT Tax": INDIAN_FORMAT_GSTR3B, "Cess": INDIAN_FORMAT_GSTR3B},
        "OSUP-Nil,Exmp": {"Tax Period": "General", "Total taxable value": INDIAN_FORMAT_GSTR3B,
                          "Integrated Tax": INDIAN_FORMAT_GSTR3B, "Central Tax": INDIAN_FORMAT_GSTR3B,
                          "State/UT Tax": INDIAN_FORMAT_GSTR3B, "Cess": INDIAN_FORMAT_GSTR3B},
        "ISUP-Rev": {"Tax Period": "General", "Total taxable value": INDIAN_FORMAT_GSTR3B,
                     "Integrated Tax": INDIAN_FORMAT_GSTR3B, "Central Tax": INDIAN_FORMAT_GSTR3B,
                     "State/UT Tax": INDIAN_FORMAT_GSTR3B, "Cess": INDIAN_FORMAT_GSTR3B},
        "OSUP-NonGST": {"Tax Period": "General", "Total taxable value": INDIAN_FORMAT_GSTR3B,
                        "Integrated Tax": INDIAN_FORMAT_GSTR3B, "Central Tax": INDIAN_FORMAT_GSTR3B,
                        "State/UT Tax": INDIAN_FORMAT_GSTR3B, "Cess": INDIAN_FORMAT_GSTR3B},
        "InterSUP-Unreg": {"Tax Period": "General", "Total taxable value": INDIAN_FORMAT_GSTR3B,
                           "Integrated Tax": INDIAN_FORMAT_GSTR3B},
        "InterSUP-Comp": {"Tax Period": "General", "Total taxable value": INDIAN_FORMAT_GSTR3B,
                          "Integrated Tax": INDIAN_FORMAT_GSTR3B},
        "InterSUP-UIN": {"Tax Period": "General", "Total taxable value": INDIAN_FORMAT_GSTR3B,
                         "Integrated Tax": INDIAN_FORMAT_GSTR3B},
        "ITC-Available": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                          "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                          "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-avl-IMPG": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                         "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                         "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-avl-IMPS": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                         "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                         "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-avl-ISRC": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                         "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                         "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-avl-ISD": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                        "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                        "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-avl-OTH": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                        "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                        "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-Reversed": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                         "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                         "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-rev-RUL": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                        "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                        "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-rev-OTH": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                        "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                        "Cess": INDIAN_FORMAT_GSTR3B},
        "Net-ITC": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                    "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                    "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-Ineligible": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                           "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                           "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-inelg-RUL": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                          "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                          "Cess": INDIAN_FORMAT_GSTR3B},
        "ITC-inelg-OTH": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                          "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                          "Cess": INDIAN_FORMAT_GSTR3B},
        "INTR-paid": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                      "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                      "Cess": INDIAN_FORMAT_GSTR3B},
        "Late-fee": {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                     "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                     "Cess": INDIAN_FORMAT_GSTR3B},
        "Tax-Pay": {"Tax Period": "General", "Tax-by-ITC": INDIAN_FORMAT_GSTR3B, "Tax-by-Cash": INDIAN_FORMAT_GSTR3B},
    }
    # Add formats for 6.1 sheets
    for i in ["a1", "a2", "a3", "a5", "a6", "a7", "a41", "a42", "a43", "a44", "b1", "b2", "b3", "b5", "b6", "b7", "b41",
              "b42", "b43", "b44"]:
        key_61 = f"6.1{i}"
        column_formats_map[key_61] = {"Tax Period": "General", "Integrated Tax": INDIAN_FORMAT_GSTR3B,
                                      "Central Tax": INDIAN_FORMAT_GSTR3B, "State/UT Tax": INDIAN_FORMAT_GSTR3B,
                                      "Cess": INDIAN_FORMAT_GSTR3B}

    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    title_font_style = Font(bold=True, size=12)  # Renamed title_font
    # bold_font is already global

    center_alignment = Alignment(horizontal="center", vertical="center")

    def sheet_has_valid_data(rows, numeric_headers):
        # print(f"[DEBUG] Checking if sheet has valid data, numeric_headers: {numeric_headers}")
        if not rows: return False
        for row_dict in rows:  # Iterate through list of dicts
            for header in numeric_headers:
                value = row_dict.get(header, 0)
                if isinstance(value, (int, float)) and value != 0:
                    # print(f"[DEBUG] Found non-zero value for {header}: {value}")
                    return True
        # print("[DEBUG] No non-zero numeric values found, skipping sheet")
        return False

    print("[DEBUG] Generating sheets...")
    for sheet_key, sheet_rows_data in data_dict.items():  # Renamed rows to sheet_rows_data
        # print(f"[DEBUG] Processing sheet: {sheet_key}, rows: {len(sheet_rows_data)}")

        current_headers = column_headers_map.get(sheet_key)  # Renamed column_headers to column_headers_map
        if not current_headers:
            print(f"[WARN] No headers defined for sheet key {sheet_key}, skipping sheet.")
            continue

        # Determine numeric headers for has_valid_data check, excluding "Tax Period"
        numeric_headers_for_check = [h for h in current_headers if h != "Tax Period"]
        if not sheet_has_valid_data(sheet_rows_data, numeric_headers_for_check):
            print(f"[DEBUG] Skipping sheet {sheet_key} due to no non-zero numeric data")
            continue

        new_sheet_name = sheet_names_map.get(sheet_key, "3B-" + sheet_key)  # Use sheet_names_map
        # print(f"[DEBUG] Creating sheet: {new_sheet_name}")
        if new_sheet_name in wb.sheetnames:
            wb.remove(wb[new_sheet_name])
            # print(f"[DEBUG] Removed existing sheet: {new_sheet_name}")

        ws = wb.create_sheet(new_sheet_name)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(current_headers))
        title_cell_ws = ws.cell(row=1, column=1)  # Renamed title_cell
        title_cell_ws.value = section_titles_map[sheet_key]  # Use section_titles_map
        title_cell_ws.font = title_font_style  # Use title_font_style
        title_cell_ws.alignment = center_alignment
        # print(f"[DEBUG] Set sheet title: {section_titles_map[sheet_key]}")

        for col_idx, col_name_iter in enumerate(current_headers, start=1):  # Renamed col_name
            cell = ws.cell(row=2, column=col_idx, value=col_name_iter)
            cell.font = BOLD_FONT  # Use global BOLD_FONT
            cell.fill = header_fill
            cell.alignment = center_alignment
        # print(f"[DEBUG] Added headers: {current_headers}")

        for r_idx, r_data_dict in enumerate(sheet_rows_data, start=3):  # Renamed row_idx, row_data
            for c_idx, col_name_iter_data in enumerate(current_headers, start=1):  # Renamed col_idx, col_name
                cell_value_data = r_data_dict.get(col_name_iter_data, "")  # Renamed cell_value
                cell = ws.cell(row=r_idx, column=c_idx, value=cell_value_data)
                current_col_formats_map = column_formats_map.get(sheet_key, {})  # Renamed column_formats
                if col_name_iter_data in current_col_formats_map:
                    cell.number_format = current_col_formats_map[col_name_iter_data]
        # print(f"[DEBUG] Populated data rows for {new_sheet_name}")

        # Add total row
        _add_total_row_to_gstr3b_sheet(ws, sheet_rows_data, current_headers, column_formats_map.get(sheet_key))

        for col_idx, col_name_width in enumerate(current_headers, start=1):  # Renamed col_name
            col_letter = get_column_letter(col_idx)
            max_length = len(str(col_name_width))  # Ensure col_name_width is string for len()
            # Iterate up to ws.max_row which now includes the total row
            for r_num in range(2, ws.max_row + 1):  # Renamed row to r_num
                cell_content = ws.cell(row=r_num, column=col_idx).value  # Renamed row to r_num
                if cell_content is not None:
                    max_length = max(max_length, len(str(cell_content)))
            ws.column_dimensions[col_letter].width = max(15, max_length + 2)
        # print(f"[DEBUG] Adjusted column widths for {new_sheet_name}")

        print(f"[DEBUG] Created sheet: {new_sheet_name}")

    print(f"[DEBUG] Saving workbook to {save_path}...")
    wb.save(save_path)
    print(f"[DEBUG] Workbook saved successfully to {save_path}")
    print("[DEBUG] Finished Excel report creation")
    return f"✅ Successfully saved Excel file: {save_path}"


def process_gstr3b(json_files, template_path, save_path):
    print("[DEBUG] Starting GSTR-3B processing...")
    combined_data = {
        "OSUP-Detail": [], "OSUP-Zero": [], "OSUP-Nil,Exmp": [], "ISUP-Rev": [], "OSUP-NonGST": [],
        "InterSUP-Unreg": [], "InterSUP-Comp": [], "InterSUP-UIN": [],
        "ITC-Available": [], "ITC-avl-IMPG": [], "ITC-avl-IMPS": [], "ITC-avl-ISRC": [],
        "ITC-avl-ISD": [], "ITC-avl-OTH": [], "ITC-Reversed": [], "ITC-rev-RUL": [],
        "ITC-rev-OTH": [], "Net-ITC": [], "ITC-Ineligible": [], "ITC-inelg-RUL": [], "ITC-inelg-OTH": [],
        "INTR-paid": [], "Late-fee": [], "Tax-Pay": [],
        "6.1a1": [], "6.1a2": [], "6.1a3": [], "6.1a5": [], "6.1a6": [], "6.1a7": [],
        "6.1a41": [], "6.1a42": [], "6.1a43": [], "6.1a44": [],
        "6.1b1": [], "6.1b2": [], "6.1b3": [], "6.1b5": [], "6.1b6": [], "6.1b7": [],
        "6.1b41": [], "6.1b42": [], "6.1b43": [], "6.1b44": []
    }
    # print(f"[DEBUG] Initialized combined_data with keys: {list(combined_data.keys())}")

    print("[DEBUG] Processing JSON files...")
    for file_path in json_files:
        print(f"[DEBUG] Loading JSON file: {file_path}")
        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            # print(f"[DEBUG] Loaded JSON data from {file_path}")

            extraction_functions = [
                extract_section_3_1, extract_section_3_2, extract_section_4,
                extract_section_5_1, extract_section_6, extract_section_6_1
            ]
            for func in extraction_functions:
                extracted_part = func(data)
                for key, rows in extracted_part.items():
                    combined_data[key].extend(rows)
                    # print(f"[DEBUG] Extended {key} with {len(rows)} rows")

    print("[DEBUG] Finished processing JSON files")

    print("[DEBUG] Sorting data...")
    financial_order = ["April", "May", "June", "July", "August", "September", "October", "November", "December",
                       "January", "February", "March", "Unknown"]
    for key_sort in combined_data:  # Renamed key to key_sort
        # print(f"[DEBUG] Sorting section: {key_sort}")
        # Ensure all items in the list are dictionaries and have "Tax Period"
        valid_rows_for_sorting = [x for x in combined_data[key_sort] if isinstance(x, dict) and "Tax Period" in x]
        combined_data[key_sort] = sorted(
            valid_rows_for_sorting,
            key=lambda x: financial_order.index(x["Tax Period"]) if x["Tax Period"] in financial_order else len(
                financial_order)
        )
        # print(f"[DEBUG] Sorted {key_sort} with {len(combined_data[key_sort])} rows")
    print("[DEBUG] Data sorting completed")

    print("[DEBUG] Generating Excel report...")
    result = create_excel_report(combined_data, save_path, template_path)
    print("[DEBUG] GSTR-3B processing completed")
    return result

