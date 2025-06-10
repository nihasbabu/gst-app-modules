import tkinter as tk  # Kept for messagebox, consider removing if UI fully decouples
from tkinter import filedialog, messagebox  # Used for error popups directly in processor
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import os
# import pyperclip # Removed as UI/main block is removed
import logging
import math

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Sheet titles with desired order
SECTION_TITLES = [
    ("PUR-Total", "Purchase Register - Total"),
    ("PUR-Total_sws", "Purchase Register - Total - Supplier wise"),
    ("PUR-Summary-Total", "Purchase Register - Total - Summary"),
]

# Define headers to exclude from total calculation (case-sensitive, matches final_headers)
EXCLUDE_FROM_TOTAL_HEADERS = [
    'GSTIN/UIN of Supplier',
    'Supplier Name',
    'Branch',
    'Supplier Invoice Number',
    'Invoice Date',
    'Supplier Invoice Date',
    'Invoice Type',
    'Voucher Number',
    'Voucher Ref. No'  # This will be an extra header if present
]

# Columns to exclude from Cr/Dr and format checks (these are generally text or date columns)
# This list should contain all fixed headers EXCEPT 'Round Off' (case-sensitive)
EXCLUDE_HEADERS_FROM_CRDR_CHECK = [
    'GSTIN/UIN of Supplier',
    'Supplier Name',
    'Branch',
    'Supplier Invoice Number',
    'Invoice Date',
    'Supplier Invoice Date',
    'Invoice Type',
    'Voucher Number',
    'Invoice Value',  # Invoice Value is excluded from Cr/Dr check
    'Taxable Value',  # Taxable Value is excluded from Cr/Dr check
    'Integrated Tax',  # These are direct numeric values
    'Central Tax',
    'State/UT Tax',
    'Cess',
    'Voucher Ref. No'  # This will be an extra header if present
]

# Number format for Indian numbering system
INDIAN_NUMBER_FORMAT = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"


def find_header_row(worksheet):
    logging.info("Searching for header row starting with 'Date'")
    for row in worksheet.iter_rows():
        if isinstance(row[0].value, str) and row[0].value.strip().lower() == "date":
            logging.info(f"Found header row at row {row[0].row}")
            return row[0].row
    logging.warning(f"Header row not found in sheet: {worksheet.title}")
    return None


def get_financial_year(date):
    if date is None:
        return None
    if date.month >= 4:
        return date.year
    return date.year - 1


def create_or_replace_sheet(wb, sheet_name, title_text, columns):
    logging.info(f"Creating or replacing sheet: {sheet_name}")
    if sheet_name in wb.sheetnames:
        logging.info(f"Sheet {sheet_name} exists, removing it")
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1)
    cell.value = title_text
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, col in enumerate(columns, start=1):
        header_cell = ws.cell(row=2, column=idx, value=col)
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B3"
    return ws


def apply_format_and_autofit(ws, columns, start_row=3, col_format_map=None):
    logging.info(f"Applying formats and autofitting columns for sheet: {ws.title}")
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        if col_format_map and col_name in col_format_map:
            for row in range(start_row, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = col_format_map[col_name]
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = max(15, max_len + 2)  # Added padding
    logging.info("Finished applying formats and autofitting")


def add_total_row(ws, columns, start_row, end_row):  # Removed INDIAN_NUMBER_FORMAT arg, use global
    logging.info(f"Adding total row for sheet: {ws.title}, from row {start_row} to {end_row}")
    total_row_data = ['Total'] + [''] * (len(columns) - 1)

    for col_idx in range(1, len(columns) + 1):
        if col_idx == 1:
            continue
        header_name = columns[col_idx - 1]  # Standardized header name
        if header_name in EXCLUDE_FROM_TOTAL_HEADERS:
            logging.debug(f"Excluding column '{header_name}' from total.")
            continue

        total = 0
        has_numeric_data = False
        try:
            for row in range(start_row, end_row + 1):
                value = ws.cell(row=row, column=col_idx).value
                if isinstance(value, (int, float)) and not (math.isnan(value) or math.isinf(value)):
                    total += float(value)
                    has_numeric_data = True
                elif value is not None and str(value).strip() != '':
                    logging.debug(f"Non-numeric value encountered in column '{header_name}' at row {row}: {value}")
            if has_numeric_data:
                total_row_data[col_idx - 1] = total
        except Exception as e:
            logging.error(f"Error totaling column '{header_name}': {e}")
            total_row_data[col_idx - 1] = "Error"

    row_num = end_row + 1
    logging.debug(f"Writing total row at row {row_num}")
    for col_idx, value in enumerate(total_row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = Font(bold=True, color="FF0000")
        if isinstance(value, (int, float)):
            cell.number_format = INDIAN_NUMBER_FORMAT  # Use global format
    logging.info("Total row added successfully")


def safe_float_conversion(value, header, cell=None):  # header is standardized
    if value is None:
        return 0.0

    if header in EXCLUDE_HEADERS_FROM_CRDR_CHECK:
        try:
            float_value = float(value) if str(value).strip() != '' and value is not None else 0.0
            if isinstance(value,
                          str) and value.strip() != '' and float_value == 0.0 and value.strip() != '0' and value.strip() != '0.0':
                return value
            return float_value
        except (ValueError, TypeError):
            logging.debug(f"Could not convert value to float for header '{header}' (standard conversion): {value}")
            return value

    numeric_value = 0.0
    try:
        numeric_value = float(value)
    except (ValueError, TypeError):
        if isinstance(value, str):
            value_str = value.strip()
            if value_str.endswith(' Cr'):  # For Purchase, Cr means negative
                try:
                    numeric_value = -float(value_str[:-3])
                except ValueError:
                    logging.debug(f"Could not convert Cr string to float for header '{header}': {value}")
                    numeric_value = 0.0
            elif value_str.endswith(' Dr'):  # For Purchase, Dr means positive
                try:
                    numeric_value = float(value_str[:-3])
                except ValueError:
                    logging.debug(f"Could not convert Dr string to float for header '{header}': {value}")
                    numeric_value = 0.0
            else:
                try:
                    numeric_value = float(value_str) if value_str else 0.0
                except ValueError:
                    logging.debug(f"Could not convert string to float for header '{header}': {value}")
                    numeric_value = 0.0
        else:
            logging.debug(f"Unexpected value type for conversion for header '{header}': {type(value)}")
            numeric_value = 0.0

    if cell and cell.number_format and isinstance(numeric_value, (int, float)):
        format_str = str(cell.number_format)
        if 'Dr' in format_str and numeric_value < 0:  # Purchase: Dr format but negative value -> make positive
            numeric_value = abs(numeric_value)
        elif 'Cr' in format_str and numeric_value > 0:  # Purchase: Cr format but positive value -> make negative
            numeric_value = -abs(numeric_value)
    return numeric_value


def process_purchase_data(input_files, template_file=None, existing_wb=None):
    logging.info(f"Starting purchase processing with {len(input_files)} input files")
    all_data = []

    source_key_to_standard_header = {
        'gstin/uin': 'GSTIN/UIN of Supplier',
        'particulars': 'Supplier Name',
        'supplier invoice no.': 'Supplier Invoice Number',
        'supplier invoice no': 'Supplier Invoice Number',
        'date': 'Invoice Date',  # This is the transaction date from Tally
        'supplier invoice date': 'Supplier Invoice Date',  # Actual invoice date from supplier
        'voucher type': 'Invoice Type',
        'voucher no.': 'Voucher Number',
        'voucher no': 'Voucher Number',
        'gross total': 'Invoice Value',  # Tally's Gross Total
        'invoice value': 'Invoice Value',  # Alternative common name
        'igst': 'Integrated Tax',
        'cgst': 'Central Tax',
        'sgst': 'State/UT Tax',
        'cess': 'Cess',
        'round off': 'Round Off',
        'voucher ref. no': 'Voucher Ref. No',
        'voucher ref. no.': 'Voucher Ref. No',
        # 'value', 'purchase gst', 'purchase igst' are handled for 'Taxable Value'
    }

    fixed_headers = [
        'GSTIN/UIN of Supplier',
        'Supplier Name',
        'Branch',
        'Supplier Invoice Number',
        'Invoice Date',  # Transaction Date
        'Supplier Invoice Date',  # Actual Supplier Bill Date
        'Invoice Type',
        'Voucher Number',
        'Invoice Value',
        'Taxable Value',
        'Integrated Tax',
        'Central Tax',
        'State/UT Tax',
        'Cess',
        'Round Off'
    ]
    fixed_headers_lower = {h.lower() for h in fixed_headers}
    extra_headers_set = set()

    logging.info("Starting file processing loop for purchase data")
    for filepath, branch_key in input_files:
        logging.info(f"Processing file: {filepath} with branch_key: {branch_key}")
        try:
            wb = load_workbook(filepath, data_only=True)  # data_only=True for values
            logging.info(f"Loaded workbook: {filepath}")

            if len(wb.sheetnames) > 1:
                if "Purchase Register" in wb.sheetnames:
                    ws = wb["Purchase Register"]
                else:
                    logging.error(f"'Purchase Register' sheet not found in {filepath}")
                    messagebox.showerror("Sheet Not Found",
                                         f"'Purchase Register' sheet not found in {os.path.basename(filepath)}.")
                    continue
            else:
                ws = wb.active
                if not ws:
                    logging.error(f"No active sheet in {filepath}")
                    messagebox.showerror("Sheet Not Found", f"No active sheet found in {os.path.basename(filepath)}.")
                    continue

            header_row_num = find_header_row(ws)
            if not header_row_num:
                logging.error(f"Header row not found in {filepath}")
                messagebox.showerror("Header Not Found", f"Header row not found in {os.path.basename(filepath)}.")
                continue

            original_headers_from_sheet = [cell.value for cell in ws[header_row_num] if cell.value is not None]
            original_headers_lower_map = {}  # Maps lowercase original header to its original casing
            original_cells_lower_map = {}  # Maps lowercase original header to header cell object
            for col_idx, header_val in enumerate(original_headers_from_sheet):
                if header_val is not None:
                    header_lower = str(header_val).lower()
                    original_headers_lower_map[header_lower] = header_val
                    original_cells_lower_map[header_lower] = ws.cell(row=header_row_num, column=col_idx + 1)

            logging.debug(f"Original headers from '{filepath}': {original_headers_from_sheet}")

            # Identify extra headers
            for original_header_val in original_headers_from_sheet:
                if original_header_val is not None:
                    header_l = str(original_header_val).lower()
                    is_mapped = header_l in source_key_to_standard_header
                    is_taxable_value_source = header_l in ['value', 'purchase gst', 'purchase igst']
                    is_fixed = False
                    if is_mapped:
                        standard_h = source_key_to_standard_header[header_l]
                        if standard_h.lower() in fixed_headers_lower: is_fixed = True
                    elif header_l in fixed_headers_lower:
                        is_fixed = True

                    if not is_mapped and not is_taxable_value_source and not is_fixed:
                        extra_headers_set.add(original_header_val)  # Add original casing

            logging.debug(f"Current extra_headers_set (original case): {extra_headers_set}")

            for row_idx, row_cells_tuple in enumerate(ws.iter_rows(min_row=header_row_num + 1),
                                                      start=header_row_num + 1):
                row_data_orig_values = {}  # Maps original header string to value for current row
                current_row_cells_map = {}  # Maps original header string to cell object for current row
                for i, cell_obj in enumerate(row_cells_tuple):
                    if i < len(original_headers_from_sheet):
                        header = original_headers_from_sheet[i]
                        row_data_orig_values[header] = cell_obj.value
                        current_row_cells_map[header] = cell_obj

                # Get date value using case-insensitive lookup on original headers
                date_val = None
                original_date_header_key = None
                if 'date' in original_headers_lower_map:  # Transaction date
                    original_date_header_key = original_headers_lower_map['date']
                    date_val = row_data_orig_values.get(original_date_header_key)

                particulars_val = None
                if 'particulars' in original_headers_lower_map:
                    original_particulars_header_key = original_headers_lower_map['particulars']
                    particulars_val = row_data_orig_values.get(original_particulars_header_key, '')

                if not date_val or (isinstance(particulars_val, str) and 'grand total' in particulars_val.lower()):
                    logging.debug(f"Skipping row {row_idx}: no Date or contains 'Grand Total'")
                    continue

                # Convert transaction date
                if isinstance(date_val, str):
                    try:
                        date_val = datetime.datetime.strptime(date_val, '%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        try:
                            date_val = datetime.datetime.strptime(date_val.split()[0], '%d-%m-%Y')  # Tally format
                        except ValueError:
                            logging.debug(f"Skipping row {row_idx} due to transaction date parsing error: {date_val}")
                            continue
                elif not isinstance(date_val, datetime.datetime):
                    logging.debug(f"Skipping row {row_idx} due to invalid transaction date type: {type(date_val)}")
                    continue
                if original_date_header_key: row_data_orig_values[original_date_header_key] = date_val

                processed_row_data = {}
                processed_row_data['Branch'] = branch_key

                # Process each original header from the sheet for the current row
                for original_header, original_value in row_data_orig_values.items():
                    if original_header is None: continue
                    header_lower = str(original_header).lower()
                    cell = current_row_cells_map.get(original_header)

                    # Skip Taxable Value source columns here; they are handled separately
                    if header_lower in ['value', 'purchase gst', 'purchase igst']:
                        continue

                    standard_header = None
                    if header_lower in source_key_to_standard_header:
                        standard_header = source_key_to_standard_header[header_lower]
                    elif header_lower in fixed_headers_lower:  # e.g. "Branch" if it was in source
                        standard_header = original_header  # Use original casing
                    elif original_header in extra_headers_set:
                        standard_header = original_header  # Use original casing

                    if standard_header:
                        if standard_header == 'Invoice Date':  # Transaction Date
                            processed_row_data[standard_header] = date_val
                        # Special handling for Supplier Invoice Date if it's different from transaction date
                        elif standard_header == 'Supplier Invoice Date':
                            sup_inv_date_val = original_value
                            if isinstance(sup_inv_date_val, str):
                                try:
                                    sup_inv_date_val = datetime.datetime.strptime(sup_inv_date_val, '%Y-%m-%d %H:%M:%S')
                                except ValueError:
                                    try:
                                        sup_inv_date_val = datetime.datetime.strptime(sup_inv_date_val.split()[0],
                                                                                      '%d-%m-%Y')
                                    except ValueError:
                                        logging.warning(
                                            f"Could not parse Supplier Invoice Date '{sup_inv_date_val}', keeping as string.")
                            elif not isinstance(sup_inv_date_val, datetime.datetime) and sup_inv_date_val is not None:
                                logging.warning(
                                    f"Supplier Invoice Date '{sup_inv_date_val}' is not a recognized date format, keeping as is.")
                            processed_row_data[standard_header] = sup_inv_date_val
                        else:
                            processed_row_data[standard_header] = safe_float_conversion(original_value, standard_header,
                                                                                        cell)

                # Taxable Value Extraction Logic (Priority: Value -> PURCHASE GST -> PURCHASE IGST)
                taxable_value_source_val = None
                taxable_value_source_cell = None

                tv_keys_priority = ['value', 'purchase gst', 'purchase igst']
                for tv_key_lower in tv_keys_priority:
                    if tv_key_lower in original_headers_lower_map:
                        original_tv_header = original_headers_lower_map[tv_key_lower]
                        val_from_source = row_data_orig_values.get(original_tv_header)
                        cell_from_source = current_row_cells_map.get(original_tv_header)
                        converted_val = safe_float_conversion(val_from_source, 'Taxable Value',
                                                              cell_from_source)  # Check with 'Taxable Value' as target
                        if (isinstance(converted_val, (int, float)) and converted_val != 0) or \
                                (isinstance(converted_val, str) and converted_val.strip() not in ['', '0', '0.0']):
                            taxable_value_source_val = val_from_source
                            taxable_value_source_cell = cell_from_source
                            logging.debug(f"Using '{original_tv_header}' for Taxable Value: {taxable_value_source_val}")
                            break  # Found a non-zero source

                processed_row_data['Taxable Value'] = safe_float_conversion(taxable_value_source_val, 'Taxable Value',
                                                                            taxable_value_source_cell)

                for header in fixed_headers:
                    if header not in processed_row_data:
                        processed_row_data[header] = ''
                for eh_header in extra_headers_set:
                    if eh_header not in processed_row_data:  # If extra header wasn't in this row's source
                        original_eh_value = row_data_orig_values.get(eh_header)
                        if original_eh_value is not None:
                            cell_for_eh = current_row_cells_map.get(eh_header)
                            processed_row_data[eh_header] = safe_float_conversion(original_eh_value, eh_header,
                                                                                  cell_for_eh)
                        else:
                            processed_row_data[eh_header] = ''

                all_data.append(processed_row_data)
                logging.debug(f"Processed row {row_idx} and added to all_data")

        except Exception as e:
            logging.error(f"Error processing file {filepath}: {e}", exc_info=True)
            messagebox.showerror("File Processing Error", f"Error processing file {os.path.basename(filepath)}: {e}")
        logging.info(f"Finished processing file: {filepath}")

    logging.info("Finished file processing loop for purchase data")

    extra_headers_list = sorted(list(extra_headers_set))
    final_headers = fixed_headers + extra_headers_list
    logging.info(f"Final headers for purchase detail sheets: {final_headers}")
    logging.info(f"Identified unique extra headers (original case): {extra_headers_list}")
    logging.info(f"Total records processed across all files: {len(all_data)}")

    # Data Structuring and Sorting
    data_by_type = {"PUR-Total": all_data}

    def sort_key_pur_total(row):  # Sort by Supplier Invoice Date (actual bill date)
        date = row.get('Supplier Invoice Date')  # Use Supplier Invoice Date for primary sort
        if not isinstance(date, datetime.datetime):  # Fallback to Transaction Date if SID is bad
            date = row.get('Invoice Date')
        if isinstance(date, datetime.datetime):
            fy = get_financial_year(date)
            return (fy, date.month if date.month >= 4 else date.month + 12, date.day,
                    str(row.get('Supplier Invoice Number', '')))
        return (0, 0, 0, str(row.get('Supplier Invoice Number', '')))

    logging.info("Sorting data for PUR-Total by Supplier Invoice Date (Financial Year)")
    data_by_type["PUR-Total"].sort(key=sort_key_pur_total)

    def sort_key_sws(row):  # Supplier Wise Sort
        supplier = str(row.get('Supplier Name', '')).lower()
        # Use Supplier Invoice Date for sorting within supplier
        date = row.get('Supplier Invoice Date')
        if not isinstance(date, datetime.datetime): date = row.get('Invoice Date')  # Fallback
        if not isinstance(date, datetime.datetime): date = datetime.datetime.min

        if supplier in ['cash', '(cancelled )'] or not supplier:
            return (1, supplier, date, str(row.get('Supplier Invoice Number', '')))
        return (0, supplier, date, str(row.get('Supplier Invoice Number', '')))

    logging.info("Sorting data for PUR-Total_sws by Supplier Name and Supplier Invoice Date")
    data_by_type["PUR-Total_sws"] = sorted(all_data, key=sort_key_sws)

    # Workbook Creation
    if existing_wb is not None:
        output_wb = existing_wb
    elif template_file:
        output_wb = load_workbook(template_file)
    else:
        output_wb = Workbook()
        if 'Sheet' in output_wb.sheetnames and len(output_wb.sheetnames) == 1: del output_wb['Sheet']

    col_format_map = {
        'Invoice Date': 'DD-MM-YYYY',
        'Supplier Invoice Date': 'DD-MM-YYYY',
        'Invoice Value': INDIAN_NUMBER_FORMAT,
        'Taxable Value': INDIAN_NUMBER_FORMAT,
        'Integrated Tax': INDIAN_NUMBER_FORMAT,
        'Central Tax': INDIAN_NUMBER_FORMAT,
        'State/UT Tax': INDIAN_NUMBER_FORMAT,
        'Cess': INDIAN_NUMBER_FORMAT,
        'Round Off': INDIAN_NUMBER_FORMAT
    }
    for eh in extra_headers_list:  # Format extra headers too
        if eh not in col_format_map: col_format_map[eh] = INDIAN_NUMBER_FORMAT

    sheets_to_create = [
        ("PUR-Total", data_by_type.get("PUR-Total", [])),
        ("PUR-Total_sws", data_by_type.get("PUR-Total_sws", [])),
    ]

    for sheet_name, data in sheets_to_create:
        if not data:  # Skip empty sheets
            logging.info(f"No data for sheet: {sheet_name}, skipping creation.")
            continue
        logging.info(f"Processing sheet: {sheet_name} with {len(data)} records")
        title = next(t for n, t in SECTION_TITLES if n == sheet_name)
        ws = create_or_replace_sheet(output_wb, sheet_name, title, final_headers)
        start_row_data = 3
        for row_data in data:
            row_values = []
            for header in final_headers:  # Use final_headers for order and inclusion
                value = row_data.get(header, '')
                if header in ['Invoice Date', 'Supplier Invoice Date'] and isinstance(value, datetime.datetime):
                    value = value.strftime('%d-%m-%Y')
                row_values.append(value)
            ws.append(row_values)
        add_total_row(ws, final_headers, start_row_data, ws.max_row)
        apply_format_and_autofit(ws, final_headers, col_format_map=col_format_map, start_row=start_row_data)
        logging.info(f"Created sheet {sheet_name} with {len(data)} records")

    # Summary Sheet Processing
    summary_sheet_name = "PUR-Summary-Total"
    summary_data_source = all_data  # Summary uses all data
    if not summary_data_source:
        logging.info(f"No data for summary sheet: {summary_sheet_name}, skipping creation.")
    else:
        logging.info(f"Processing summary sheet: {summary_sheet_name}")
        summary_title = next(t for n, t in SECTION_TITLES if n == summary_sheet_name)
        summary_headers = ['Month', 'No. of Records', 'Invoice Value', 'Taxable Value', 'Integrated Tax',
                           'Central Tax', 'State/UT Tax', 'Cess']
        summary_col_format_map = {h: INDIAN_NUMBER_FORMAT for h in summary_headers if
                                  h not in ['Month', 'No. of Records']}

        ws_summary = create_or_replace_sheet(output_wb, summary_sheet_name, summary_title, summary_headers)
        monthly_summary_aggr = {}
        unique_invoices_by_month = {}  # To count unique invoices per month
        months_order = ['April', 'May', 'June', 'July', 'August', 'September',
                        'October', 'November', 'December', 'January', 'February', 'March']

        for row in summary_data_source:
            # Use Supplier Invoice Date for month grouping in summary
            date_for_summary = row.get('Supplier Invoice Date')
            if not isinstance(date_for_summary, datetime.datetime):  # Fallback
                date_for_summary = row.get('Invoice Date')

            inv_num_for_summary = str(row.get('Supplier Invoice Number', '')).strip()  # Key for uniqueness

            if not isinstance(date_for_summary, datetime.datetime) or not inv_num_for_summary:
                logging.debug("Skipping summary row: no valid Supplier Invoice Date/Number or Transaction Date.")
                continue

            month_idx = (date_for_summary.month - 4 + 12) % 12  # April is 0
            month_name = months_order[month_idx]

            if month_name not in monthly_summary_aggr:
                monthly_summary_aggr[month_name] = {h: 0.0 for h in summary_headers if h not in ['Month']}
                monthly_summary_aggr[month_name]['count'] = 0  # For 'No. of Records'
                unique_invoices_by_month[month_name] = set()

            if inv_num_for_summary not in unique_invoices_by_month[month_name]:
                monthly_summary_aggr[month_name]['count'] += 1
                unique_invoices_by_month[month_name].add(inv_num_for_summary)

            fields_to_sum = ['Invoice Value', 'Taxable Value', 'Integrated Tax', 'Central Tax', 'State/UT Tax', 'Cess']
            for field in fields_to_sum:
                value = row.get(field, 0.0)
                if isinstance(value, (int, float)) and not (math.isnan(value) or math.isinf(value)):
                    monthly_summary_aggr[month_name][field] += float(value)
                elif value is not None and str(value).strip() not in ['', '0', '0.0']:
                    logging.warning(
                        f"Non-numeric value '{value}' for '{field}' in summary for invoice '{inv_num_for_summary}', treating as 0.")

        summary_start_row_data = 3
        rows_added_summary = 0
        for month_n_sum in months_order:
            if month_n_sum in monthly_summary_aggr:
                m_data = monthly_summary_aggr[month_n_sum]
                ws_summary.append([
                    month_n_sum, m_data['count'], m_data['Invoice Value'], m_data['Taxable Value'],
                    m_data['Integrated Tax'], m_data['Central Tax'], m_data['State/UT Tax'], m_data['Cess']
                ])
                rows_added_summary += 1

        if rows_added_summary > 0:
            total_row_summary_vals = ['Total'] + [0.0] * (len(summary_headers) - 1)
            for col_i_sum in range(1, len(summary_headers)):
                curr_col_total_sum = 0.0
                for r_i_sum in range(summary_start_row_data, summary_start_row_data + rows_added_summary):
                    cell_val_sum = ws_summary.cell(row=r_i_sum, column=col_i_sum + 1).value
                    if isinstance(cell_val_sum, (int, float)) and not (
                            math.isnan(cell_val_sum) or math.isinf(cell_val_sum)):
                        curr_col_total_sum += float(cell_val_sum)
                total_row_summary_vals[col_i_sum] = curr_col_total_sum

            ws_summary.append(total_row_summary_vals)
            total_row_num_sum_sheet = summary_start_row_data + rows_added_summary
            for c_idx_sum, val_sum in enumerate(total_row_summary_vals, start=1):
                total_cell_sum = ws_summary.cell(row=total_row_num_sum_sheet, column=c_idx_sum)
                total_cell_sum.font = Font(bold=True, color="FF0000")
                if isinstance(val_sum, (int, float)) and summary_headers[c_idx_sum - 1] not in ['Month',
                                                                                                'No. of Records']:
                    total_cell_sum.number_format = INDIAN_NUMBER_FORMAT

        apply_format_and_autofit(ws_summary, summary_headers, col_format_map=summary_col_format_map,
                                 start_row=summary_start_row_data)
        logging.info(f"Created summary sheet {summary_sheet_name}")

    logging.info("Purchase data processing completed")
    return output_wb
