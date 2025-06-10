import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from openpyxl import Workbook, load_workbook
import logging
import traceback
from collections import Counter

# Attempt to import processor functions
try:
    from credit_note_processor import process_credit_data
except ImportError:
    # Log error, but allow UI to load with functionality disabled
    logging.error("credit_debit_ui: Could not import 'process_credit_data' from 'credit_note_processor.py'.")
    messagebox.showerror("Import Error",
                         "Could not import 'process_credit_data'. Credit Note processing will be affected.")
    process_credit_data = None

try:
    from debit_note_processor import process_debit_note_data
except ImportError:
    logging.error("credit_debit_ui: Could not import 'process_debit_note_data' from 'debit_note_processor.py'.")
    messagebox.showerror("Import Error",
                         "Could not import 'process_debit_note_data'. Debit Note processing will be affected.")
    process_debit_note_data = None

# Telemetry import block
try:
    from utils.telemetry import send_event
    # Using __name__ for logger is a common Python practice
    log_cd = logging.getLogger(__name__) # Or os.path.basename(__file__)
    log_cd.info("Successfully imported send_event from utils.telemetry")
except ImportError as e_import_telemetry:
    log_cd = logging.getLogger(__name__) # Or os.path.basename(__file__)
    log_cd.warning(f"Could not import send_event from utils.telemetry (error: {e_import_telemetry}). Telemetry will be disabled for this module.")
    # The print below is what you're seeing. If logging is configured for console, this print can be removed.
    # print(f"[WARN] {os.path.basename(__file__)}: Failed to import telemetry from utils.telemetry. Error: {e_import_telemetry}")
    def send_event(event_name, payload):  # Dummy function
        pass

PLACEHOLDER_TEXT = "Code"

class CustomErrorDialog(tk.Toplevel): # Your CustomErrorDialog code remains the same
    def __init__(self, parent, title, message, error_details_to_copy):
        super().__init__(parent)
        self.title(title)
        self.error_details = error_details_to_copy
        self.parent = parent
        self.transient(parent)
        self.grab_set()
        main_frame = tk.Frame(self, padx=10, pady=10)
        main_frame.pack(expand=True, fill=tk.BOTH)
        icon_label = tk.Label(main_frame, text="❌", font=("Arial", 24), fg="red")
        icon_label.pack(side=tk.LEFT, padx=(0, 10), anchor='n')
        message_frame = tk.Frame(main_frame)
        message_frame.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)
        tk.Label(message_frame, text=title, font=("Arial", 12, "bold")).pack(anchor="w")
        self.msg_text_widget = tk.Text(message_frame, wrap=tk.WORD, height=5, width=50, borderwidth=0,
                                       bg=self.cget('bg'))
        self.msg_text_widget.insert(tk.END, message)
        self.msg_text_widget.config(state=tk.DISABLED)
        self.msg_text_widget.pack(pady=5, fill=tk.X, expand=True)
        self.copy_status_label = tk.Label(message_frame, text="", fg="green")
        self.copy_status_label.pack(pady=(0, 5))
        button_frame = tk.Frame(self, pady=10)
        button_frame.pack(fill=tk.X)
        def copy_error_to_clipboard_action():
            try:
                self.clipboard_clear()
                self.clipboard_append(self.error_details)
                self.copy_status_label.config(text="Error details copied to clipboard!")
                self.after(2000, lambda: self.copy_status_label.config(text=""))
            except tk.TclError:
                self.copy_status_label.config(text="Could not access clipboard.", fg="red")
                messagebox.showwarning("Clipboard Error", "Could not access the clipboard on this system.", parent=self)
        ok_button = tk.Button(button_frame, text="OK", width=10, command=self.destroy)
        copy_button = tk.Button(button_frame, text="Copy Error Details", width=15,
                                command=copy_error_to_clipboard_action)
        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=0)
        button_frame.columnconfigure(2, weight=0)
        button_frame.columnconfigure(3, weight=1)
        copy_button.grid(row=0, column=1, padx=5)
        ok_button.grid(row=0, column=2, padx=5)
        self.center_window()
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.wait_window(self)

    def center_window(self):
        self.update_idletasks()
        parent_x = self.parent.winfo_x()
        parent_y = self.parent.winfo_y()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()
        dialog_width = self.winfo_reqwidth()
        dialog_height = self.winfo_reqheight()
        x = parent_x + (parent_width - dialog_width) // 2
        y = parent_y + (parent_height - dialog_height) // 2
        self.geometry(f'+{x}+{y}')

class CreditDebitNoteProcessorUI: # Your UI class code remains largely the same
    def __init__(self, root_window):
        self.root = root_window
        self.root.title("Process Credit / Debit Notes")
        self.credit_note_files = []
        self.debit_note_files = []
        self.template_file = None
        self.base_height = 500
        # ... (rest of your __init__ method as provided) ...
        tk.Label(self.root, text="Process Credit / Debit Notes", font=("Arial", 16, "bold")).pack(pady=5)
        main_frame = tk.Frame(self.root)
        main_frame.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)
        left_frame = tk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        tk.Label(left_frame, text="Credit Note Files", font=("Arial", 10, "bold")).pack()
        self.credit_note_tree = ttk.Treeview(left_frame, columns=("File Name", "Branch Code"),
                                             show="headings", height=13)
        self.credit_note_tree.heading("File Name", text="File Name")
        self.credit_note_tree.heading("Branch Code", text="Branch Code")
        self.credit_note_tree.column("File Name", width=140, stretch=True)
        self.credit_note_tree.column("Branch Code", width=85, stretch=True)
        self.credit_note_tree.pack(pady=2, fill=tk.BOTH, expand=True)
        self.credit_note_tree.bind("<Double-1>", self.edit_credit_note_branch_code)
        btn_frame_credit = tk.Frame(left_frame)
        btn_frame_credit.pack(pady=5)
        tk.Button(btn_frame_credit, text="+ Add", command=self.add_credit_note_file).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame_credit, text="- Remove", command=self.delete_credit_note_file).pack(side=tk.LEFT, padx=5)
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        tk.Label(right_frame, text="Debit Note Files", font=("Arial", 10, "bold")).pack()
        self.debit_note_tree = ttk.Treeview(right_frame, columns=("File Name", "Branch Code"),
                                            show="headings", height=13)
        self.debit_note_tree.heading("File Name", text="File Name")
        self.debit_note_tree.heading("Branch Code", text="Branch Code")
        self.debit_note_tree.column("File Name", width=140, stretch=True)
        self.debit_note_tree.column("Branch Code", width=85, stretch=True)
        self.debit_note_tree.pack(pady=2, fill=tk.BOTH, expand=True)
        self.debit_note_tree.bind("<Double-1>", self.edit_debit_note_branch_code)
        btn_frame_debit = tk.Frame(right_frame)
        btn_frame_debit.pack(pady=5)
        tk.Button(btn_frame_debit, text="+ Add", command=self.add_debit_note_file).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame_debit, text="- Remove", command=self.delete_debit_note_file).pack(side=tk.LEFT, padx=5)
        self.template_frame = tk.Frame(self.root)
        self.template_frame.pack(pady=5)
        tk.Label(self.template_frame, text="Template Excel File (Optional):").pack(side=tk.LEFT)
        self.template_label = tk.Label(self.template_frame, text="No file selected")
        self.template_label.pack(side=tk.LEFT, padx=5)
        tk.Button(self.template_frame, text="Select", command=self.select_template).pack(side=tk.LEFT, padx=2)
        tk.Button(self.template_frame, text="Clear", command=self.clear_template).pack(side=tk.LEFT, padx=2)
        self.process_btn = tk.Button(self.root, text="Process Credit / Debit Notes", font=("Arial", 12),
                                     command=self.process_files, state=tk.DISABLED, bg="light grey", width=25)
        self.process_btn.pack(pady=10)
        self.warning_frame = tk.Frame(self.root, borderwidth=1, relief="solid")
        self.warning_title = tk.Label(self.warning_frame, text="Warning!", fg="red",
                                      font=("Arial", 10, "underline"))
        self.warning_text = tk.Label(self.warning_frame, text="", fg="red",
                                     justify=tk.LEFT, wraplength=450)
        self.ignore_var = tk.BooleanVar()
        self.ignore_check = tk.Checkbutton(self.warning_frame, text="Ignore Warning and Proceed",
                                           variable=self.ignore_var,
                                           command=self.update_process_button_state)
        self.update_process_button_state()

    def _add_file_to_tree(self, file_list, tree_widget, file_type_name): # Your method
        files = filedialog.askopenfilenames(
            title=f"Select {file_type_name} Excel Files",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        for file_path in files:
            if not any(f[0] == file_path for f in file_list):
                base = os.path.basename(file_path)
                extracted_branch_key = base.split('_')[0].split('-')[0].split('.')[0] if base else ""
                if extracted_branch_key == os.path.splitext(base)[0]:
                    stored_branch_code = ""
                else:
                    stored_branch_code = extracted_branch_key.strip()
                file_list.append((file_path, stored_branch_code))
                display_branch_code = stored_branch_code if stored_branch_code else PLACEHOLDER_TEXT
                tree_widget.insert("", tk.END, values=(os.path.basename(file_path), display_branch_code))
        self.update_process_button_state()

    def _delete_file_from_tree(self, file_list, tree_widget): # Your method
        selected_items = tree_widget.selection()
        if not selected_items: return
        items_to_remove_from_list = []
        for item_id in selected_items:
            filename_in_tree = tree_widget.item(item_id, "values")[0]
            for i, (fp, stored_bc) in enumerate(file_list):
                if os.path.basename(fp) == filename_in_tree:
                    items_to_remove_from_list.append(file_list[i])
                    break
            tree_widget.delete(item_id)
        for item_tuple in items_to_remove_from_list:
            if item_tuple in file_list: file_list.remove(item_tuple)
        self.update_process_button_state()

    def _edit_branch_code(self, event, file_list, tree_widget): # Your method
        item_id = tree_widget.identify_row(event.y)
        column_id = tree_widget.identify_column(event.x)
        if not item_id or column_id != "#2": return
        current_filename_display = tree_widget.item(item_id, "values")[0]
        list_idx = -1
        try:
            for i, (fp, bc) in enumerate(file_list):
                if os.path.basename(fp) == current_filename_display: list_idx = i; break
        except ValueError: logging.error(f"Item ID {item_id} not found in tree. This should not happen."); return
        if list_idx == -1: logging.error(f"Could not find {current_filename_display} in list for editing."); return
        stored_branch_code = file_list[list_idx][1]
        entry = tk.Entry(tree_widget)
        entry.insert(0, stored_branch_code if stored_branch_code else "")
        if stored_branch_code: entry.select_range(0, tk.END)
        x, y, width, height = tree_widget.bbox(item_id, column_id)
        entry.place(x=x, y=y, width=width, height=height)
        entry.focus_set()
        def save_branch_code(evt):
            new_branch_code_input = entry.get().strip()
            file_list[list_idx] = (file_list[list_idx][0], new_branch_code_input)
            display_text_for_tree = new_branch_code_input if new_branch_code_input else PLACEHOLDER_TEXT
            tree_widget.item(item_id, values=(current_filename_display, display_text_for_tree))
            entry.destroy()
            self.update_process_button_state()
        entry.bind("<Return>", save_branch_code)
        entry.bind("<FocusOut>", save_branch_code)
        entry.bind("<Escape>", lambda e: entry.destroy())

    def add_credit_note_file(self): self._add_file_to_tree(self.credit_note_files, self.credit_note_tree, "Credit Note")
    def delete_credit_note_file(self): self._delete_file_from_tree(self.credit_note_files, self.credit_note_tree)
    def edit_credit_note_branch_code(self, event): self._edit_branch_code(event, self.credit_note_files, self.credit_note_tree)
    def add_debit_note_file(self): self._add_file_to_tree(self.debit_note_files, self.debit_note_tree, "Debit Note")
    def delete_debit_note_file(self): self._delete_file_from_tree(self.debit_note_files, self.debit_note_tree)
    def edit_debit_note_branch_code(self, event): self._edit_branch_code(event, self.debit_note_files, self.debit_note_tree)
    def select_template(self):
        file = filedialog.askopenfilename(title="Select Template Excel File", filetypes=[("Excel Files", "*.xlsx *.xls")])
        if file: self.template_file = file; self.template_label.config(text=os.path.basename(file))
    def clear_template(self): self.template_file = None; self.template_label.config(text="No file selected")
    def update_process_button_state(self): # Your method
        has_any_files = bool(self.credit_note_files or self.debit_note_files)
        all_files_with_codes = self.credit_note_files + self.debit_note_files
        missing_branch_code = any(not code.strip() or code == PLACEHOLDER_TEXT for _, code in all_files_with_codes)
        if missing_branch_code and has_any_files:
            if not self.warning_frame.winfo_ismapped(): self.warning_frame.pack(pady=5, padx=10, fill=tk.X, after=self.process_btn)
            self.warning_title.pack(pady=(5, 0))
            self.warning_text.config(text="Warning: Branch Code is missing. Please double-click to edit or check 'Ignore Warning'.")
            self.warning_text.pack(pady=2); self.ignore_check.pack(pady=2)
            self.root.update_idletasks()
            current_warning_frame_height = self.warning_frame.winfo_reqheight()
            self.root.geometry(f"500x{self.base_height + current_warning_frame_height + 10}")
            if self.ignore_var.get(): self.process_btn.config(state=tk.NORMAL, bg="light green")
            else: self.process_btn.config(state=tk.DISABLED, bg="light grey")
        else:
            if self.warning_frame.winfo_ismapped():
                self.warning_frame.pack_forget(); self.warning_title.pack_forget()
                self.warning_text.pack_forget(); self.ignore_check.pack_forget()
                self.root.update_idletasks()
            self.root.geometry(f"500x{self.base_height}")
            if has_any_files: self.process_btn.config(state=tk.NORMAL, bg="light green")
            else: self.process_btn.config(state=tk.DISABLED, bg="light grey")

    def process_files(self):
        if not (self.credit_note_files or self.debit_note_files):
            messagebox.showerror("Error", "No Credit Note or Debit Note files selected for processing.")
            return

        all_files_with_codes = self.credit_note_files + self.debit_note_files
        missing_branch_code_strict = any(not code.strip() or code == PLACEHOLDER_TEXT for _, code in all_files_with_codes)
        if missing_branch_code_strict and not self.ignore_var.get():
            messagebox.showwarning("Branch Code Missing", "Please review branch codes (cannot be empty or placeholder) or check 'Ignore Warning' to proceed.")
            self.update_process_button_state()
            return

        default_branch = "Default_Branch" # Renamed for clarity
        credit_notes_to_process = [(f, c.strip() if (c.strip() and c != PLACEHOLDER_TEXT) else default_branch) for f, c in self.credit_note_files]
        debit_notes_to_process = [(f, c.strip() if (c.strip() and c != PLACEHOLDER_TEXT) else default_branch) for f, c in self.debit_note_files]

        output_save_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")],
                                                 title="Save Combined Report As", initialfile="Processed_Credit_Debit_Notes.xlsx")
        if not output_save_path:
            send_event("credit_debit_process_cancelled", {"reason": "output_file_not_selected"})
            return

        self.process_btn.config(text="Processing...", state=tk.DISABLED, bg="light grey")
        self.process_btn.update()
        send_event("credit_debit_process_started", {
            "credit_notes_count": len(credit_notes_to_process),
            "debit_notes_count": len(debit_notes_to_process),
            "template_used": bool(self.template_file)
        })

        try:
            if self.template_file: wb = load_workbook(self.template_file)
            else:
                wb = Workbook()
                if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1: del wb['Sheet']

            if credit_notes_to_process and process_credit_data:
                wb = process_credit_data(credit_notes_to_process, template_file=None, existing_wb=wb)
            if debit_notes_to_process and process_debit_note_data:
                wb = process_debit_note_data(debit_notes_to_process, template_file=None, existing_wb=wb)

            wb.save(output_save_path)
            messagebox.showinfo("Success", f"Report saved successfully at:\n{output_save_path}")
            send_event("credit_debit_process_complete", {
                "credit_notes_processed": len(credit_notes_to_process),
                "debit_notes_processed": len(debit_notes_to_process),
                "output_file": os.path.basename(output_save_path),
                "status": "success"
            })
            self.credit_note_files.clear(); self.debit_note_files.clear()
            self.credit_note_tree.delete(*self.credit_note_tree.get_children())
            self.debit_note_tree.delete(*self.debit_note_tree.get_children())
            self.ignore_var.set(False); self.clear_template()
        except Exception as e:
            detailed_error_info = traceback.format_exc()
            logging.error(f"An error occurred during C/D note processing: {str(e)}", exc_info=True)
            send_event("error", {
                "module": "credit_debit_ui.process_files",
                "error_type": type(e).__name__,
                "error_message": str(e),
                "traceback": detailed_error_info # Send full traceback
            })
            self.show_error_with_copy("Processing Error", f"An error occurred:\n{type(e).__name__}: {str(e)}", detailed_error_info)
        finally:
            self.process_btn.config(text="Process Credit / Debit Notes")
            self.update_process_button_state()

    # Keeping your existing show_error_with_copy method from the uploaded file
    def show_error_with_copy(self, title, short_message, detailed_message_to_copy):
        CustomErrorDialog(self.root, title, short_message, detailed_message_to_copy)


if __name__ == "__main__":
    if not logging.getLogger().hasHandlers():
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - CreditDebitUI - %(levelname)s - %(message)s')

    if process_credit_data is None or process_debit_note_data is None:
        logging.critical("Credit/Debit Note UI: One or both processor functions (credit_note_processor, debit_note_processor) could not be imported. UI may not function as expected.")
    root = tk.Tk()
    app = CreditDebitNoteProcessorUI(root)
    root.mainloop()
