# sales_purchase_ui.py

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
from openpyxl import Workbook, load_workbook
import logging
import traceback  # For detailed error reporting
from collections import Counter  # For update_process_button_state logic if needed

# Attempt to import processor functions and telemetry
try:
    from telemetry import send_event
except ImportError:
    logging.warning("telemetry module not found. Event sending will be disabled.")


    def send_event(event_name, payload):
        pass  # Dummy function

try:
    from sales_processor import process_excel_data
except ImportError:
    logging.error("Could not import 'process_excel_data' from 'sales_processor.py'. Sales processing will fail.")
    messagebox.showerror("Import Error",
                         "Could not import 'process_excel_data' from 'sales_processor.py'. Make sure the file exists.")
    process_excel_data = None

try:
    from purchase_processor import process_purchase_data
except ImportError:
    logging.error(
        "Could not import 'process_purchase_data' from 'purchase_processor.py'. Purchase processing will fail.")
    messagebox.showerror("Import Error",
                         "Could not import 'process_purchase_data' from 'purchase_processor.py'. Make sure the file exists.")
    process_purchase_data = None

PLACEHOLDER_TEXT = "Code"  # Placeholder for branch code


class CustomErrorDialog(tk.Toplevel):
    def __init__(self, parent, title, message, error_details_to_copy):
        super().__init__(parent)
        self.title(title)
        self.error_details = error_details_to_copy
        self.parent = parent

        self.transient(parent)
        self.grab_set()

        main_frame = tk.Frame(self, padx=10, pady=10)
        main_frame.pack(expand=True, fill=tk.BOTH)

        icon_label = tk.Label(main_frame, text="❌", font=("Arial", 24), fg="red")
        icon_label.pack(side=tk.LEFT, padx=(0, 10), anchor='n')

        message_frame = tk.Frame(main_frame)
        message_frame.pack(side=tk.LEFT, expand=True, fill=tk.BOTH)

        tk.Label(message_frame, text=title, font=("Arial", 12, "bold")).pack(anchor="w")

        self.msg_text_widget = tk.Text(message_frame, wrap=tk.WORD, height=5, width=50, borderwidth=0,
                                       bg=self.cget('bg'))
        self.msg_text_widget.insert(tk.END, message)
        self.msg_text_widget.config(state=tk.DISABLED)
        self.msg_text_widget.pack(pady=5, fill=tk.X, expand=True)

        self.copy_status_label = tk.Label(message_frame, text="", fg="green")
        self.copy_status_label.pack(pady=(0, 5))

        button_frame = tk.Frame(self, pady=10)
        button_frame.pack(fill=tk.X)

        def copy_error_to_clipboard_action():
            try:
                self.clipboard_clear()
                self.clipboard_append(self.error_details)
                self.copy_status_label.config(text="Error details copied to clipboard!")
                self.after(2000, lambda: self.copy_status_label.config(text=""))
            except tk.TclError:
                self.copy_status_label.config(text="Could not access clipboard.", fg="red")
                messagebox.showwarning("Clipboard Error", "Could not access the clipboard on this system.", parent=self)

        ok_button = tk.Button(button_frame, text="OK", width=10, command=self.destroy)
        copy_button = tk.Button(button_frame, text="Copy Error Details", width=15,
                                command=copy_error_to_clipboard_action)

        button_frame.columnconfigure(0, weight=1)
        button_frame.columnconfigure(1, weight=0)
        button_frame.columnconfigure(2, weight=0)
        button_frame.columnconfigure(3, weight=1)

        copy_button.grid(row=0, column=1, padx=5)
        ok_button.grid(row=0, column=2, padx=5)

        self.center_window()
        self.protocol("WM_DELETE_WINDOW", self.destroy)
        self.wait_window(self)

    def center_window(self):
        self.update_idletasks()
        parent_x = self.parent.winfo_x()
        parent_y = self.parent.winfo_y()
        parent_width = self.parent.winfo_width()
        parent_height = self.parent.winfo_height()

        dialog_width = self.winfo_reqwidth()
        dialog_height = self.winfo_reqheight()

        x = parent_x + (parent_width - dialog_width) // 2
        y = parent_y + (parent_height - dialog_height) // 2
        self.geometry(f'+{x}+{y}')


class SalesPurchaseProcessorUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Process Sales / Purchase")

        self.sales_files = []  # List of tuples: (filepath, branch_code)
        self.purchase_files = []  # List of tuples: (filepath, branch_code)
        self.template_file = None
        self.base_height = 500
        self.warning_frame_height_addition = 100

        self.root.geometry(f"500x{self.base_height}")

        tk.Label(root, text="Process Sales / Purchase", font=("Arial", 16, "bold")).pack(pady=5)
        main_frame = tk.Frame(root)
        main_frame.pack(pady=5, padx=10, fill=tk.BOTH, expand=True)

        # Sales Section
        left_frame = tk.Frame(main_frame)
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        tk.Label(left_frame, text="Sales Ledger Excel Files", font=("Arial", 10, "bold")).pack()
        self.sales_tree = ttk.Treeview(left_frame, columns=("File Name", "Branch Code"),
                                       show="headings", height=13)
        self.sales_tree.heading("File Name", text="File Name")
        self.sales_tree.heading("Branch Code", text="Branch Code")
        self.sales_tree.column("File Name", width=140, stretch=True)
        self.sales_tree.column("Branch Code", width=85, stretch=True)
        self.sales_tree.pack(pady=2, fill=tk.BOTH, expand=True)
        self.sales_tree.bind("<Double-1>", self.edit_sales_branch_code)
        btn_frame_sales = tk.Frame(left_frame)
        btn_frame_sales.pack(pady=5)
        tk.Button(btn_frame_sales, text="+ Add", command=self.add_sales_file).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame_sales, text="- Remove", command=self.delete_sales_file).pack(side=tk.LEFT, padx=5)

        # Purchase Section
        right_frame = tk.Frame(main_frame)
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        tk.Label(right_frame, text="Purchase Register Excel Files", font=("Arial", 10, "bold")).pack()
        self.purchase_tree = ttk.Treeview(right_frame, columns=("File Name", "Branch Code"),
                                          show="headings", height=13)
        self.purchase_tree.heading("File Name", text="File Name")
        self.purchase_tree.heading("Branch Code", text="Branch Code")
        self.purchase_tree.column("File Name", width=140, stretch=True)
        self.purchase_tree.column("Branch Code", width=85, stretch=True)
        self.purchase_tree.pack(pady=2, fill=tk.BOTH, expand=True)
        self.purchase_tree.bind("<Double-1>", self.edit_purchase_branch_code)
        btn_frame_purchase = tk.Frame(right_frame)
        btn_frame_purchase.pack(pady=5)
        tk.Button(btn_frame_purchase, text="+ Add", command=self.add_purchase_file).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame_purchase, text="- Remove", command=self.delete_purchase_file).pack(side=tk.LEFT, padx=5)

        # Template File Section
        self.template_frame = tk.Frame(root)  # Made template_frame an instance variable
        self.template_frame.pack(pady=5, fill=tk.X, padx=10)
        tk.Label(self.template_frame, text="Template Excel File (Optional):").pack(side=tk.LEFT)
        self.template_label = tk.Label(self.template_frame, text="No file selected", width=25, anchor="w")
        self.template_label.pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        tk.Button(self.template_frame, text="Select", command=self.select_template).pack(side=tk.LEFT, padx=2)
        tk.Button(self.template_frame, text="Clear", command=self.clear_template).pack(side=tk.LEFT, padx=2)

        # Process Button
        self.process_btn = tk.Button(root, text="Process Sales / Purchase", font=("Arial", 12),
                                     command=self.process_files, state=tk.DISABLED, bg="light grey")
        self.process_btn.pack(pady=10)

        # Warning Frame - will be packed by update_process_button_state if needed
        self.warning_frame = tk.Frame(root, borderwidth=1, relief="solid")
        self.warning_title = tk.Label(self.warning_frame, text="Warning!", fg="red",
                                      font=("Arial", 10, "underline"))
        self.warning_text = tk.Label(self.warning_frame, text="", fg="red",
                                     justify=tk.LEFT, wraplength=450)
        self.ignore_var = tk.BooleanVar()
        self.ignore_check = tk.Checkbutton(self.warning_frame, text="Ignore Warning and Proceed",
                                           variable=self.ignore_var,
                                           command=self.update_process_button_state)

        self.update_process_button_state()

    def _add_file_to_list_and_tree(self, file_list, tree_widget, file_type_name_for_dialog):
        files = filedialog.askopenfilenames(
            title=f"Select {file_type_name_for_dialog} Excel Files",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        for file_path in files:
            if not any(f[0] == file_path for f in file_list):
                base = os.path.basename(file_path)
                extracted_branch_key = base.split('_')[0].split('-')[0].split('.')[0] if base else ""

                if extracted_branch_key == os.path.splitext(base)[0]:
                    stored_branch_code = ""
                else:
                    stored_branch_code = extracted_branch_key.strip()

                file_list.append((file_path, stored_branch_code))
                display_code = stored_branch_code if stored_branch_code else PLACEHOLDER_TEXT
                tree_widget.insert("", tk.END, values=(os.path.basename(file_path), display_code))
        self.update_process_button_state()

    def _delete_selected_from_list_and_tree(self, file_list, tree_widget):
        selected_tree_items = tree_widget.selection()
        items_to_remove_from_data = []

        for item_id in selected_tree_items:
            filename_in_tree = tree_widget.item(item_id, "values")[0]
            for i, (fp, bc) in enumerate(file_list):
                if os.path.basename(fp) == filename_in_tree:
                    items_to_remove_from_data.append(file_list[i])
                    break
            tree_widget.delete(item_id)

        for item_data in items_to_remove_from_data:
            if item_data in file_list:
                file_list.remove(item_data)
            else:
                logging.warning(f"Tried to remove non-existent item: {item_data}")

        self.update_process_button_state()

    def _edit_branch_code_in_tree(self, event, file_list, tree_widget):
        item_id = tree_widget.identify_row(event.y)
        column_id = tree_widget.identify_column(event.x)

        if not item_id or column_id != "#2":
            return

        current_filename_display = tree_widget.item(item_id, "values")[0]
        list_idx = -1
        for i, (fp, bc) in enumerate(file_list):
            if os.path.basename(fp) == current_filename_display:
                list_idx = i  # Simpler direct match, assuming basename is unique enough in the list for UI
                break

        if list_idx == -1:
            logging.error(f"Could not find {current_filename_display} in internal file list for branch code edit.")
            return

        stored_branch_code = file_list[list_idx][1]

        entry = tk.Entry(tree_widget)
        entry.insert(0, stored_branch_code if stored_branch_code else "")
        if stored_branch_code:
            entry.select_range(0, tk.END)

        x, y, width, height = tree_widget.bbox(item_id, column_id)
        entry.place(x=x, y=y, width=width, height=height)
        entry.focus_set()

        def save_edited_branch_code(evt):
            new_code_input = entry.get().strip()
            file_list[list_idx] = (file_list[list_idx][0], new_code_input)

            display_text = new_code_input if new_code_input else PLACEHOLDER_TEXT
            tree_widget.item(item_id, values=(current_filename_display, display_text))
            entry.destroy()
            self.update_process_button_state()

        entry.bind("<Return>", save_edited_branch_code)
        entry.bind("<FocusOut>", save_edited_branch_code)
        entry.bind("<Escape>", lambda e: entry.destroy())

    def add_sales_file(self):
        self._add_file_to_list_and_tree(self.sales_files, self.sales_tree, "Sales Ledger")

    def delete_sales_file(self):
        self._delete_selected_from_list_and_tree(self.sales_files, self.sales_tree)

    def edit_sales_branch_code(self, event):
        self._edit_branch_code_in_tree(event, self.sales_files, self.sales_tree)

    def add_purchase_file(self):
        self._add_file_to_list_and_tree(self.purchase_files, self.purchase_tree, "Purchase Register")

    def delete_purchase_file(self):
        self._delete_selected_from_list_and_tree(self.purchase_files, self.purchase_tree)

    def edit_purchase_branch_code(self, event):
        self._edit_branch_code_in_tree(event, self.purchase_files, self.purchase_tree)

    def select_template(self):
        file = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx *.xls")])
        if file:
            self.template_file = file
            self.template_label.config(text=os.path.basename(file))

    def clear_template(self):
        self.template_file = None
        self.template_label.config(text="No file selected")

    def update_process_button_state(self):
        has_files = bool(self.sales_files or self.purchase_files)
        # Check if any branch code is empty OR is still the placeholder
        missing_branch_code = any(not code.strip() or code == PLACEHOLDER_TEXT
                                  for _, code in self.sales_files + self.purchase_files)

        if missing_branch_code and has_files:
            if not self.warning_frame.winfo_ismapped():
                # Pack warning frame *after* the process button to match original intent
                self.warning_frame.pack(pady=5, padx=10, fill=tk.X, after=self.process_btn)
            self.warning_title.pack(pady=(5, 0))
            self.warning_text.config(
                text="Warning: Branch Code is missing. Please double-click to edit or check 'Ignore Warning'.")
            self.warning_text.pack(pady=2)
            self.ignore_check.pack(pady=2)

            self.root.update_idletasks()
            current_warning_frame_height = self.warning_frame.winfo_reqheight()
            self.root.geometry(f"500x{self.base_height + current_warning_frame_height + 10}")

            if has_files and self.ignore_var.get():
                self.process_btn.config(state=tk.NORMAL, bg="light green")
            else:
                self.process_btn.config(state=tk.DISABLED, bg="light grey")
        else:
            if self.warning_frame.winfo_ismapped():
                self.warning_frame.pack_forget()
                self.warning_title.pack_forget()
                self.warning_text.pack_forget()
                self.ignore_check.pack_forget()
                self.root.update_idletasks()
            self.root.geometry(f"500x{self.base_height}")

            if has_files:
                self.process_btn.config(state=tk.NORMAL, bg="light green")
            else:
                self.process_btn.config(state=tk.DISABLED, bg="light grey")

    def process_files(self):
        if not (self.sales_files or self.purchase_files):
            messagebox.showerror("Error", "No files selected for processing.")
            return

        if process_excel_data is None and self.sales_files:
            messagebox.showerror("Error", "Sales processor is not available. Cannot process sales files.")
            return
        if process_purchase_data is None and self.purchase_files:
            messagebox.showerror("Error", "Purchase processor is not available. Cannot process purchase files.")
            return

        missing_branch_code_strict = any(not code.strip() or code == PLACEHOLDER_TEXT
                                         for _, code in self.sales_files + self.purchase_files)
        if missing_branch_code_strict and not self.ignore_var.get():
            messagebox.showwarning("Branch Code Missing",
                                   "Please review branch codes (cannot be empty or placeholder) or check 'Ignore Warning' to proceed.")
            self.update_process_button_state()
            return

        default_branch_for_processor = "Default_Branch"
        sales_to_process = [(f, c.strip() if (c.strip() and c != PLACEHOLDER_TEXT) else default_branch_for_processor)
                            for f, c in self.sales_files]
        purchase_to_process = [(f, c.strip() if (c.strip() and c != PLACEHOLDER_TEXT) else default_branch_for_processor)
                               for f, c in self.purchase_files]

        save_file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel Files", "*.xlsx")],
                                                 title="Save Sales/Purchase Report As",
                                                 initialfile="Processed_Sales_Purchase.xlsx")
        if not save_file:
            return

        self.process_btn.config(text="Processing...", state=tk.DISABLED, bg="light grey")
        self.root.update_idletasks()

        try:
            if self.template_file:
                wb = load_workbook(self.template_file)
            else:
                wb = Workbook()
                if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1:
                    del wb['Sheet']

            if sales_to_process and process_excel_data:
                wb = process_excel_data(sales_to_process, template_file=None, existing_wb=wb)
            if purchase_to_process and process_purchase_data:
                wb = process_purchase_data(purchase_to_process, template_file=None, existing_wb=wb)

            wb.save(save_file)

            send_event("sales_purchase_complete", {
                "sales_files_count": len(sales_to_process),
                "purchase_files_count": len(purchase_to_process),
                "output_file": save_file,
                "template_used": bool(self.template_file)
            })

            messagebox.showinfo("Success", f"Report saved successfully at:\n{save_file}")
            self.sales_files.clear()
            self.purchase_files.clear()
            self.sales_tree.delete(*self.sales_tree.get_children())
            self.purchase_tree.delete(*self.purchase_tree.get_children())
            self.ignore_var.set(False)
            self.clear_template()
        except Exception as e:
            detailed_error_info = traceback.format_exc()
            logging.error(f"An error occurred during processing: {str(e)}", exc_info=True)
            print("--- SALES/PURCHASE UI ERROR DETAILS ---")
            print(detailed_error_info)
            print("---------------------------------------")

            send_event("error", {
                "module": "sales_purchase_ui.process_files",
                "error_type": type(e).__name__,
                "error_message": str(e),
                "sales_files_count": len(self.sales_files),
                "purchase_files_count": len(self.purchase_files)
            })
            # Use the new CustomErrorDialog
            CustomErrorDialog(self.root,
                              "Processing Error",
                              f"An error occurred during processing:\n\n{type(e).__name__}: {str(e)}\n\nSee console for full traceback if run from command line.",
                              detailed_error_info)
        finally:
            self.process_btn.config(text="Process Sales / Purchase")  # Reset button text
            self.update_process_button_state()


if __name__ == "__main__":
    if not logging.getLogger().hasHandlers():
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - UI - %(levelname)s - %(message)s')

    if process_excel_data is None and process_purchase_data is None:  # Adjusted check
        logging.critical(
            "Both core processor functions (sales & purchase) could not be imported. UI may not function fully.")
    elif process_excel_data is None:
        logging.warning("Sales processor function could not be imported. Sales processing will be unavailable.")
    elif process_purchase_data is None:
        logging.warning("Purchase processor function could not be imported. Purchase processing will be unavailable.")

    root = tk.Tk()
    app = SalesPurchaseProcessorUI(root)
    root.mainloop()
