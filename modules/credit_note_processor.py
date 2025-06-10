import tkinter as tk # Kept for messagebox, consider removing if UI fully decouples
from tkinter import messagebox # Used for error popups directly in processor
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import datetime
import os
import logging
import math

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# --- Constants ---
INDIAN_NUMBER_FORMAT = r"[>=10000000]##\,##\,##\,##0.00;[>=100000]##\,##\,##0.00;##,##0.00;-;"

# Standardized Fixed Headers for Credit/Debit Notes
# These are the headers that will appear in the output Excel sheets.
NOTE_FIXED_HEADERS = [
    'GSTIN/UIN of Recipient',
    'Receiver Name',
    'Branch',
    'Note Number',       # Original: Voucher No from Tally
    'Note Date',         # Original: Date from Tally
    'Note Ref. No.',     # Original: Voucher Ref. No. from Tally (if present)
    'Note Ref. Date',    # Original: Voucher Ref. Date from Tally (if present)
    'Note Value',        # Original: Gross Total from Tally
    'Taxable Value',     # Original: Value from Tally
    'Integrated Tax',
    'Central Tax',
    'State/UT Tax',
    'Cess',
    'Round Off'
]

# Headers to exclude from the 'Total' row calculation (using standardized names)
NOTE_EXCLUDE_FROM_TOTAL_HEADERS = [
    'GSTIN/UIN of Recipient',
    'Receiver Name',
    'Branch',
    'Note Number',
    'Note Date',
    'Note Ref. No.',
    'Note Ref. Date',
]

# Headers that are strictly textual and should be an empty string if None/empty in source
# (using standardized names)
NOTE_STRICTLY_TEXTUAL_HEADERS = [
    'GSTIN/UIN of Recipient',
    'Receiver Name',
    'Branch', # Assigned directly, but good to list
    'Note Number',
    'Note Ref. No.'
    # Note Date and Note Ref. Date are handled separately for datetime conversion.
]

# Headers that are numeric but excluded from Cr/Dr suffix logic (e.g., already a direct value)
# (using standardized names)
NOTE_NUMERIC_NO_CRDR_SUFFIX_CHECK = [
    'Note Value'
]


# Sheet titles with desired order for CREDIT processing
CREDIT_SECTION_TITLES = [
    ("CREDIT-Total", "Credit Register - Total"),
    ("CREDIT-Total_sws", "Credit Register - Total - Receiver wise"),
    ("CREDIT-R", "Credit Register - Registered (GSTIN Available)"),
    ("CREDIT-UR", "Credit Register - Unregistered (GSTIN Not Available)"),
    ("CREDIT-Summary-Total", "Credit Register - Total - Summary"),
    ("CREDIT-Summary-R", "Credit Register - Registered - Summary"),
    ("CREDIT-Summary-UR", "Credit Register - Unregistered - Summary"),
]
# --- Helper Functions ---

def find_header_row(worksheet):
    logging.debug(f"Searching for header row in sheet: {worksheet.title}")
    for row in worksheet.iter_rows(): # Iterate all rows
        first_cell_val = row[0].value
        if isinstance(first_cell_val, str) and "date" in first_cell_val.lower(): # Case-insensitive check for "date"
            logging.debug(f"Header row potentially found at Excel row: {row[0].row} (1-indexed)")
            return row[0].row # Return 1-indexed row number
    logging.warning(f"Header row containing 'Date' not found in sheet: {worksheet.title}")
    return None


def create_or_replace_sheet(wb, sheet_name, title_text, columns):
    logging.info(f"Creating or replacing sheet: {sheet_name}")
    if sheet_name in wb.sheetnames:
        logging.debug(f"Sheet '{sheet_name}' already exists, deleting.")
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columns))
    cell = ws.cell(row=1, column=1)
    cell.value = title_text
    cell.font = Font(bold=True, size=12)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, col in enumerate(columns, start=1):
        header_cell = ws.cell(row=2, column=idx, value=col)
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "B3" # Freeze above data rows, assuming A column might be Sr.No. or similar
    logging.info(f"Finished creating sheet: {sheet_name}")
    return ws


def apply_format_and_autofit(ws, columns, start_row=3, col_format_map=None):
    logging.debug(f"Applying format and autofit for sheet: {ws.title}")
    for col_idx, col_name in enumerate(columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name)) # Start with header length
        if col_format_map and col_name in col_format_map:
            for row_num in range(start_row, ws.max_row + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                if isinstance(cell.value, (int, float)): # Apply format only if numeric
                    cell.number_format = col_format_map[col_name]
        for row_num in range(2, ws.max_row + 1): # Check from header row for max length
            cell_value = ws.cell(row=row_num, column=col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[col_letter].width = max(15, max_len + 2) # Add padding
    logging.debug(f"Finished applying format and autofit for sheet: {ws.title}")


def add_total_row(ws, columns, start_row, end_row):
    logging.debug(f"Adding total row for sheet: {ws.title}")
    total_row_data = ['Total'] + [''] * (len(columns) - 1)

    for col_idx in range(1, len(columns) + 1):
        if col_idx == 1: continue # Skip 'Total' label column

        header_name = columns[col_idx - 1] # Standardized header name
        if header_name in NOTE_EXCLUDE_FROM_TOTAL_HEADERS:
            logging.debug(f"Excluding column '{header_name}' from total.")
            continue
        total = 0
        has_numeric_data = False
        try:
            for row in range(start_row, end_row + 1):
                value = ws.cell(row=row, column=col_idx).value
                if isinstance(value, (int, float)) and not (math.isnan(value) or math.isinf(value)):
                    total += float(value)
                    has_numeric_data = True
                elif value is not None and str(value).strip() != '':
                    logging.debug(f"Non-numeric value encountered in column '{header_name}' at row {row}: {value}")
            if has_numeric_data:
                total_row_data[col_idx - 1] = total
        except Exception as e:
            logging.error(f"Error totaling column '{header_name}': {e}")
            total_row_data[col_idx - 1] = "Error"

    row_num = end_row + 1
    ws.append(total_row_data)
    for col_idx, value in enumerate(total_row_data, start=1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = Font(bold=True, color="FF0000")
        if isinstance(value, (int, float)):
            cell.number_format = INDIAN_NUMBER_FORMAT # Use global format
    logging.debug(f"Finished adding total row for sheet: {ws.title}")


def safe_note_value_conversion(value, header_standard_name, cell_obj=None, note_type="credit"):
    """
    Converts a value to its appropriate type for note processing.
    - Handles textual fields, ensuring blanks for None/empty.
    - Converts numeric fields, applying sign logic for credit/debit notes.
    - Parses Cr/Dr suffixes for fields like 'Round Off'.
    """
    # 1. Handle strictly textual headers
    if header_standard_name in NOTE_STRICTLY_TEXTUAL_HEADERS:
        return str(value).strip() if value is not None else ''

    # 2. Handle None for potentially numeric fields (default to 0.0 before further processing)
    if value is None:
        return 0.0

    # 3. Handle specific numeric fields that don't use Cr/Dr suffixes (e.g., 'Note Value')
    if header_standard_name in NOTE_NUMERIC_NO_CRDR_SUFFIX_CHECK:
        try:
            numeric_val = float(value)
            # Apply sign convention for credit/debit
            return -abs(numeric_val) if note_type == "credit" else abs(numeric_val)
        except (ValueError, TypeError):
            # If conversion fails for these specific numeric fields, log and return 0.0 or original string
            logging.warning(f"Could not convert '{value}' to float for '{header_standard_name}'. Returning 0.0.")
            return 0.0 # Or decide to return str(value).strip() if non-conversion should be text

    # 4. Handle general numeric fields (Taxable Value, Taxes, Cess, Round Off, Extra Numeric Headers)
    #    These might have Cr/Dr suffixes or need sign adjustment based on note_type.
    numeric_value = 0.0
    is_string_value = isinstance(value, str)
    value_str_stripped = str(value).strip() if is_string_value else ""

    if not value_str_stripped and not isinstance(value, (int, float)): # Empty string or non-numeric non-string
        return 0.0

    try:
        numeric_value = float(value) # Try direct float conversion first
    except (ValueError, TypeError):
        if is_string_value:
            if value_str_stripped.endswith(' Cr'):
                try: numeric_value = float(value_str_stripped[:-3])
                except ValueError: return value_str_stripped # Return original if "123 Cr" fails
            elif value_str_stripped.endswith(' Dr'):
                try: numeric_value = float(value_str_stripped[:-3])
                except ValueError: return value_str_stripped
            else: # Not Cr/Dr, but failed direct float - could be non-numeric text
                return value_str_stripped # Return as text if it's not a recognized numeric format
        else: # Not a string, not float/int directly - e.g. some other object
            return str(value) # Return string representation

    # Apply sign logic for Credit/Debit notes for core financial values
    # 'Round Off' is special: its sign comes from Cr/Dr or cell format, not note_type directly.
    if header_standard_name not in ['Round Off'] + NOTE_STRICTLY_TEXTUAL_HEADERS + NOTE_NUMERIC_NO_CRDR_SUFFIX_CHECK:
        numeric_value = -abs(numeric_value) if note_type == "credit" else abs(numeric_value)

    # For 'Round Off' or other fields with potential Dr/Cr in cell format (if not already handled by suffix)
    if cell_obj and cell_obj.number_format and (header_standard_name == 'Round Off' or not value_str_stripped.endswith((' Cr', ' Dr'))):
        format_str = str(cell_obj.number_format)
        # For Debit notes, Dr format means positive. If value is negative, make positive.
        # For Credit notes, Dr format means negative. If value is positive, make negative.
        if 'Dr' in format_str:
            if note_type == "debit" and numeric_value < 0: numeric_value = abs(numeric_value)
            elif note_type == "credit" and numeric_value > 0: numeric_value = -abs(numeric_value)
        # For Debit notes, Cr format means negative. If value is positive, make negative.
        # For Credit notes, Cr format means positive. If value is negative, make positive.
        elif 'Cr' in format_str:
            if note_type == "debit" and numeric_value > 0: numeric_value = -abs(numeric_value)
            elif note_type == "credit" and numeric_value < 0: numeric_value = abs(numeric_value)
    return numeric_value

# --- Main Processing Function ---
def process_credit_data(input_files, template_file=None, existing_wb=None):
    logging.info("Starting credit note data processing")
    all_credit_data = []

    # Mapping from lowercase source header to standardized header
    source_key_to_standard_header = {
        # Common Tally Fields for Credit Notes (Sales Returns)
        'date': 'Note Date',
        'particulars': 'Receiver Name',
        'voucher type': 'Voucher Type', # Retain for categorization if needed, not in NOTE_FIXED_HEADERS
        'voucher no': 'Note Number',
        'voucher no.': 'Note Number',
        'gstin/uin': 'GSTIN/UIN of Recipient',
        'gross total': 'Note Value',
        'value': 'Taxable Value', # Taxable amount column
        'taxable amount': 'Taxable Value',
        'igst': 'Integrated Tax',
        'cgst': 'Central Tax',
        'sgst': 'State/UT Tax',
        'state/ut tax': 'State/UT Tax',
        'cess': 'Cess',
        'round off': 'Round Off',
        'voucher ref. no.': 'Note Ref. No.',
        'voucher ref no.': 'Note Ref. No.',
        'voucher ref. no': 'Note Ref. No.',
        'voucher ref no': 'Note Ref. No.',
        'voucher ref. date': 'Note Ref. Date',
        'voucher ref date': 'Note Ref. Date',
        # Add other common variations if known
    }
    note_fixed_headers_lower = {h.lower() for h in NOTE_FIXED_HEADERS}
    extra_headers_set = set() # Stores original casing of extra headers

    logging.debug("Starting file processing loop for credit data")
    for filepath, branch_key in input_files:
        logging.info(f"Processing credit file: {filepath} with branch key: {branch_key}")
        try:
            wb = load_workbook(filepath, data_only=True) # data_only=True to get values
            # Determine which sheet to process (e.g., "Sales Register" or "Credit Note Register")
            sheet_to_process_name = "Sales Register" # Default as per original
            ws = None
            if sheet_to_process_name in wb.sheetnames:
                ws = wb[sheet_to_process_name]
            elif "Credit Note Register" in wb.sheetnames: # Common alternative
                ws = wb["Credit Note Register"]
                logging.info(f"Found and using 'Credit Note Register' sheet in {filepath}")
            else: # Fallback logic
                logging.warning(f"'{sheet_to_process_name}' or 'Credit Note Register' not found in {filepath}. Trying first sheet with 'register' or 'credit' in name.")
                possible_sheets = [s_name for s_name in wb.sheetnames if "register" in s_name.lower() or "credit" in s_name.lower()]
                if possible_sheets:
                    ws = wb[possible_sheets[0]]
                    logging.info(f"Using first suitable sheet found: {ws.title}")
                else:
                    ws = wb.active
                    logging.warning(f"No specific sheet found. Falling back to active sheet: {ws.title}")

            if not ws:
                logging.error(f"No suitable sheet could be determined in {filepath}")
                messagebox.showerror("Sheet Error", f"Could not find a suitable sheet in {os.path.basename(filepath)}.")
                continue # Skip this file

            logging.info(f"Processing sheet: {ws.title} from file {filepath}")
            header_row_num = find_header_row(ws)
            if not header_row_num:
                logging.error(f"Header row (containing 'Date') not found in {filepath} (sheet: {ws.title})")
                messagebox.showerror("Header Error", f"Header row not found in {os.path.basename(filepath)} (sheet: {ws.title}).")
                continue

            original_headers_from_sheet = [cell.value for cell in ws[header_row_num] if cell.value is not None]
            original_headers_lower_map = {str(h).lower(): h for h in original_headers_from_sheet if h} # lowercase -> original case

            # Identify extra headers
            for original_header in original_headers_from_sheet:
                if original_header is None: continue
                header_l = str(original_header).lower()
                is_mapped = header_l in source_key_to_standard_header
                is_fixed = False
                if is_mapped:
                    standard_h = source_key_to_standard_header[header_l]
                    if standard_h.lower() in note_fixed_headers_lower: is_fixed = True
                elif header_l in note_fixed_headers_lower: is_fixed = True # Original is already a fixed header (e.g. "Branch")

                if not is_mapped and not is_fixed:
                    extra_headers_set.add(original_header) # Add original casing

            logging.debug(f"Original headers from '{filepath}': {original_headers_from_sheet}")
            logging.debug(f"Extra headers identified so far: {extra_headers_set}")

            for row_idx, row_cells_tuple in enumerate(ws.iter_rows(min_row=header_row_num + 1, values_only=False), start=header_row_num + 1):
                row_data_orig_case = {} # Original header (as in sheet) -> value
                current_row_cells_map = {} # Original header (as in sheet) -> cell object
                for i, cell_obj in enumerate(row_cells_tuple):
                    if i < len(original_headers_from_sheet):
                        header_val_sheet = original_headers_from_sheet[i]
                        row_data_orig_case[header_val_sheet] = cell_obj.value
                        current_row_cells_map[header_val_sheet] = cell_obj

                # Skip "Grand Total" rows or rows without a date
                raw_note_date_value = row_data_orig_case.get(original_headers_lower_map.get('date'))
                particulars_value = str(row_data_orig_case.get(original_headers_lower_map.get('particulars', ''), '')).lower()
                if not raw_note_date_value or 'grand total' in particulars_value:
                    continue

                processed_row_data = {'Branch': branch_key}

                # Handle Note Date
                try:
                    if isinstance(raw_note_date_value, str):
                        try: processed_row_data['Note Date'] = datetime.datetime.strptime(raw_note_date_value, '%Y-%m-%d %H:%M:%S')
                        except ValueError: processed_row_data['Note Date'] = datetime.datetime.strptime(raw_note_date_value.split()[0], '%d-%m-%Y')
                    elif isinstance(raw_note_date_value, datetime.datetime): processed_row_data['Note Date'] = raw_note_date_value
                    elif isinstance(raw_note_date_value, (int, float)): # Excel date serial number
                        processed_row_data['Note Date'] = datetime.datetime.fromordinal(datetime.datetime(1900, 1, 1).toordinal() + int(raw_note_date_value) - 2)
                    else: raise ValueError("Invalid date type")
                except Exception as e_date:
                    logging.debug(f"Row {row_idx}: Skipping due to 'Note Date' parsing error: {raw_note_date_value}, Error: {e_date}")
                    continue # Skip row if main date is unparseable

                # Handle Note Ref. Date (optional)
                raw_note_ref_date_value = row_data_orig_case.get(original_headers_lower_map.get('voucher ref. date'))
                if raw_note_ref_date_value:
                    try:
                        if isinstance(raw_note_ref_date_value, str):
                            try: processed_row_data['Note Ref. Date'] = datetime.datetime.strptime(raw_note_ref_date_value, '%Y-%m-%d %H:%M:%S')
                            except ValueError: processed_row_data['Note Ref. Date'] = datetime.datetime.strptime(raw_note_ref_date_value.split()[0], '%d-%m-%Y')
                        elif isinstance(raw_note_ref_date_value, datetime.datetime): processed_row_data['Note Ref. Date'] = raw_note_ref_date_value
                        elif isinstance(raw_note_ref_date_value, (int, float)):
                            processed_row_data['Note Ref. Date'] = datetime.datetime.fromordinal(datetime.datetime(1900, 1, 1).toordinal() + int(raw_note_ref_date_value) - 2)
                        else: processed_row_data['Note Ref. Date'] = str(raw_note_ref_date_value) # Store as string if unparseable
                    except Exception as e_ref_date:
                        logging.warning(f"Row {row_idx}: Could not parse 'Note Ref. Date': {raw_note_ref_date_value}, Error: {e_ref_date}. Storing as string.")
                        processed_row_data['Note Ref. Date'] = str(raw_note_ref_date_value)
                else:
                    processed_row_data['Note Ref. Date'] = '' # Blank if not present

                # Process other columns
                for original_header, original_value in row_data_orig_case.items():
                    if original_header is None: continue
                    header_lower = str(original_header).lower()
                    standard_header = source_key_to_standard_header.get(header_lower)

                    if standard_header: # Mapped to a standard fixed header
                        if standard_header not in ['Note Date', 'Note Ref. Date']: # Dates already handled
                            cell_obj = current_row_cells_map.get(original_header)
                            processed_row_data[standard_header] = safe_note_value_conversion(original_value, standard_header, cell_obj, "credit")
                    elif original_header in extra_headers_set: # Is an identified extra header
                        cell_obj = current_row_cells_map.get(original_header)
                        # For extra headers, assume they might be numeric and apply conversion logic
                        # The 'note_type' might not be directly applicable unless we know the nature of extra headers.
                        # Defaulting to "credit" for sign if it's a value, but it's safer if extra headers are named distinctively.
                        processed_row_data[original_header] = safe_note_value_conversion(original_value, original_header, cell_obj, "credit")

                # Ensure all NOTE_FIXED_HEADERS are present, default to blank or 0.0
                for fixed_h in NOTE_FIXED_HEADERS:
                    if fixed_h not in processed_row_data:
                        if fixed_h in NOTE_STRICTLY_TEXTUAL_HEADERS or fixed_h in ['Note Date', 'Note Ref. Date']:
                            processed_row_data[fixed_h] = ''
                        else: # Assumed numeric
                            processed_row_data[fixed_h] = 0.0
                # Ensure all collected extra headers are present
                for extra_h in extra_headers_set:
                    if extra_h not in processed_row_data:
                         processed_row_data[extra_h] = safe_note_value_conversion(None, extra_h, None, "credit") # Will default to 0.0 or ''

                all_credit_data.append(processed_row_data)

        except Exception as e_file:
            logging.error(f"Error processing file {filepath}: {e_file}", exc_info=True)
            messagebox.showerror("File Processing Error", f"Error processing file {os.path.basename(filepath)}: {e_file}")
        logging.info(f"Finished processing credit file: {filepath}")

    logging.debug("Finished file processing loop for credit data")

    final_headers_list = NOTE_FIXED_HEADERS + sorted(list(extra_headers_set))
    logging.info(f"Final headers for credit note detail sheets: {final_headers_list}")

    # Categorize data (Registered vs Unregistered)
    data_registered = [row for row in all_credit_data if str(row.get('GSTIN/UIN of Recipient', '')).strip()]
    data_unregistered = [row for row in all_credit_data if not str(row.get('GSTIN/UIN of Recipient', '')).strip()]
    logging.debug(f"Categorization: Registered={len(data_registered)}, Unregistered={len(data_unregistered)}")

    # Sort data
    sort_key_notes = lambda x: (x.get('Note Date', datetime.datetime.min) if isinstance(x.get('Note Date'), datetime.datetime) else datetime.datetime.min,
                                str(x.get('Note Number', '')))
    all_credit_data.sort(key=sort_key_notes)
    data_registered.sort(key=sort_key_notes)
    data_unregistered.sort(key=sort_key_notes)

    def sort_key_receiver_wise(row):
        receiver = str(row.get('Receiver Name', '')).strip().lower()
        date_val = row.get('Note Date', datetime.datetime.min)
        note_num_val = str(row.get('Note Number', ''))
        if receiver in ['cash', '(cancelled )'] or not receiver: return (1, receiver, date_val, note_num_val)
        return (0, receiver, date_val, note_num_val)
    all_credit_data_sws = sorted(all_credit_data, key=sort_key_receiver_wise)

    # --- Workbook Generation ---
    if existing_wb is not None: output_wb = existing_wb
    elif template_file: output_wb = load_workbook(template_file)
    else:
        output_wb = Workbook()
        if 'Sheet' in output_wb.sheetnames and len(output_wb.sheetnames) == 1: del output_wb['Sheet']

    # Column formatting map for detail sheets
    detail_col_format_map = {h: INDIAN_NUMBER_FORMAT for h in NOTE_FIXED_HEADERS if h not in ['GSTIN/UIN of Recipient', 'Receiver Name', 'Branch', 'Note Number', 'Note Date', 'Note Ref. No.', 'Note Ref. Date']}
    detail_col_format_map['Note Date'] = 'DD-MM-YYYY'
    detail_col_format_map['Note Ref. Date'] = 'DD-MM-YYYY'
    for eh in extra_headers_set: # Assume extra headers are numeric for formatting
        if eh not in detail_col_format_map: detail_col_format_map[eh] = INDIAN_NUMBER_FORMAT

    # Create Detail Sheets
    detail_sheets_to_create = [
        ("CREDIT-Total", all_credit_data),
        ("CREDIT-Total_sws", all_credit_data_sws),
        ("CREDIT-R", data_registered),
        ("CREDIT-UR", data_unregistered),
    ]
    for sheet_key, data_list in detail_sheets_to_create:
        if not data_list:
            logging.info(f"No data for detail sheet: {sheet_key}, skipping.")
            continue
        display_title = next((t for n, t in CREDIT_SECTION_TITLES if n == sheet_key), sheet_key)
        ws = create_or_replace_sheet(output_wb, sheet_key, display_title, final_headers_list)
        data_start_row = 3
        for row_item in data_list:
            row_values = []
            for header in final_headers_list:
                value = row_item.get(header, '') # Default to blank if header not in row
                if header in ['Note Date', 'Note Ref. Date'] and isinstance(value, datetime.datetime):
                    value = value.strftime('%d-%m-%Y')
                row_values.append(value)
            ws.append(row_values)
        add_total_row(ws, final_headers_list, data_start_row, ws.max_row)
        apply_format_and_autofit(ws, final_headers_list, col_format_map=detail_col_format_map, start_row=data_start_row)

    # Create Summary Sheets
    summary_headers = ['Month', 'No. of Notes', 'Note Value', 'Taxable Value', 'Integrated Tax', 'Central Tax', 'State/UT Tax', 'Cess']
    summary_col_format_map = {h: INDIAN_NUMBER_FORMAT for h in summary_headers if h not in ['Month', 'No. of Notes']}
    months_order = ['April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December', 'January', 'February', 'March']

    summary_sheets_to_create = [
        ("CREDIT-Summary-Total", all_credit_data),
        ("CREDIT-Summary-R", data_registered),
        ("CREDIT-Summary-UR", data_unregistered),
    ]
    for sheet_key, data_list_summary in summary_sheets_to_create:
        if not data_list_summary:
            logging.info(f"No data for summary sheet: {sheet_key}, skipping.")
            continue
        display_title_summary = next((t for n, t in CREDIT_SECTION_TITLES if n == sheet_key), sheet_key)
        ws_summary = create_or_replace_sheet(output_wb, sheet_key, display_title_summary, summary_headers)
        monthly_summary_aggr = {}
        unique_notes_by_month = {}

        for row_item_sum in data_list_summary:
            date_obj_sum = row_item_sum.get('Note Date')
            note_num_sum = str(row_item_sum.get('Note Number', '')).strip()
            if not isinstance(date_obj_sum, datetime.datetime) or not note_num_sum: continue

            month_name_sum = months_order[(date_obj_sum.month - 4 + 12) % 12]
            if month_name_sum not in monthly_summary_aggr:
                monthly_summary_aggr[month_name_sum] = {h: 0.0 for h in summary_headers if h not in ['Month', 'No. of Notes']}
                monthly_summary_aggr[month_name_sum]['count'] = 0 # For 'No. of Notes'
                unique_notes_by_month[month_name_sum] = set()

            if note_num_sum not in unique_notes_by_month[month_name_sum]:
                monthly_summary_aggr[month_name_sum]['count'] += 1
                unique_notes_by_month[month_name_sum].add(note_num_sum)

            # The headers to be summed in the summary sheet (excluding non-numeric ones)
            numeric_summary_fields = [h for h in summary_headers if h not in ['Month', 'No. of Notes']]

            for field_to_sum in numeric_summary_fields:
                val_add = row_item_sum.get(field_to_sum, 0.0) # Get value using the summary header name directly
                if isinstance(val_add, (int, float)) and not (math.isnan(val_add) or math.isinf(val_add)):
                    monthly_summary_aggr[month_name_sum][field_to_sum] += float(val_add)

        summary_data_start_row = 3
        rows_added_sum = 0
        for month_iter in months_order:
            if month_iter in monthly_summary_aggr:
                month_data_iter = monthly_summary_aggr[month_iter]
                row_to_append_sum = [month_iter, month_data_iter.get('count',0)] + \
                                    [month_data_iter.get(h, 0.0) for h in summary_headers if h not in ['Month', 'No. of Notes']]
                ws_summary.append(row_to_append_sum)
                rows_added_sum +=1

        if rows_added_sum > 0:
            total_row_sum_vals = ['Total', sum(monthly_summary_aggr[m].get('count', 0) for m in monthly_summary_aggr)]
            for h_idx_sum in range(2, len(summary_headers)): # Start from 'Note Value'
                col_h_sum = summary_headers[h_idx_sum]
                total_row_sum_vals.append(sum(monthly_summary_aggr[m].get(col_h_sum, 0.0) for m in monthly_summary_aggr))
            ws_summary.append(total_row_sum_vals)
            total_row_num_on_sheet = summary_data_start_row + rows_added_sum
            for c_idx_sum, val_iter in enumerate(total_row_sum_vals, 1):
                cell_sum_total = ws_summary.cell(row=total_row_num_on_sheet, column=c_idx_sum)
                cell_sum_total.font = Font(bold=True, color="FF0000")
                if isinstance(val_iter, (int, float)) and summary_headers[c_idx_sum -1] not in ['Month', 'No. of Notes']:
                    cell_sum_total.number_format = INDIAN_NUMBER_FORMAT
        apply_format_and_autofit(ws_summary, summary_headers, col_format_map=summary_col_format_map, start_row=summary_data_start_row)

    logging.info("Completed credit note data processing")
    return output_wb
